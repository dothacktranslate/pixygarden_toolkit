#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Set

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("This script requires openpyxl. Install with: python -m pip install openpyxl") from e


def norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(s, 0)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return default


def parse_byte(v: Any, default: Optional[int] = None, bare_two_digit_hex: bool = True) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, int):
        return v & 0xFF
    if isinstance(v, float):
        return int(v) & 0xFF
    s = str(v).strip()
    if not s:
        return default
    if s.lower().startswith("0x"):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    if bare_two_digit_hex and re.fullmatch(r"[0-9A-Fa-f]{2}", s):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    try:
        return int(s, 0) & 0xFF
    except ValueError:
        return default


def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column letter: {col!r}")
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_col(headers: list[str], value: Optional[str], aliases: Iterable[str], required: bool = True) -> Optional[int]:
    if value:
        s = value.strip()
        if re.fullmatch(r"[A-Za-z]+", s):
            return col_letter_to_index(s)
        lowered = {h.lower(): i + 1 for i, h in enumerate(headers) if h}
        if s.lower() in lowered:
            return lowered[s.lower()]
        raise ValueError(f"Could not resolve column {value!r}. Use a column letter or exact header.")
    lowered = {h.lower(): i + 1 for i, h in enumerate(headers) if h}
    for alias in aliases:
        if alias.lower() in lowered:
            return lowered[alias.lower()]
    if required:
        raise ValueError(f"Could not find any of these columns: {', '.join(aliases)}")
    return None


def to_fullwidth_ascii(ch: str) -> str:
    """Map one-byte ASCII printable characters to full-width CP932/Unicode equivalents."""
    if ch == " ":
        return "　"  # U+3000 ideographic space, CP932 81 40
    o = ord(ch)
    if 0x21 <= o <= 0x7E:
        return chr(0xFF00 + o - 0x20)
    return ch


def encode_tree_txt(text: str, text_mode: str = "ascii", encoding: str = "cp932", errors: str = "replace") -> bytes:
    """Encode one TREE INFO text body.

    Original TREE INFO TXT files use literal backslash+n bytes (5C 6E) for line breaks,
    and appear to use CP932 two-byte Japanese/full-width glyphs for body text.

    text_mode:
      ascii             = encode text as given, preserving one-byte ASCII where present.
      cp932             = encode text as given in CP932; non-ASCII remains CP932.
      cp932_fullwidth   = convert printable ASCII to full-width CP932, but keep line breaks
                          as literal 5C 6E control bytes.
    """
    text = norm_text(text)
    out = bytearray()
    for line_index, line in enumerate(text.split("\n")):
        if line_index:
            out += b"\\n"  # original line-break control bytes
        if text_mode in ("ascii", "cp932"):
            out += line.encode(encoding, errors=errors)
        elif text_mode == "cp932_fullwidth":
            converted = "".join(to_fullwidth_ascii(ch) for ch in line)
            out += converted.encode("cp932", errors=errors)
        else:
            raise ValueError(f"Unknown text mode: {text_mode}")
    return bytes(out)


def align(value: int, n: int) -> int:
    if n <= 1:
        return value
    return ((value + n - 1) // n) * n


def safe_rel_path(path_text: str) -> Path:
    p = Path(*[part for part in path_text.replace("\\", "/").split("/") if part and part not in (".", "..")])
    if p.is_absolute() or ".." in p.parts:
        raise ValueError(f"Unsafe archive path: {path_text!r}")
    return p


def load_manifest(path: Optional[Path]) -> Dict[str, Dict[str, str]]:
    if not path:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            p = row.get("path") or row.get("archive_path") or ""
            if p:
                out[p] = row
    return out


def parse_row_set(spec: Optional[str]) -> Optional[Set[int]]:
    if not spec:
        return None
    out: Set[int] = set()
    for part in re.split(r"[, ]+", spec.strip()):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            ia = int(a, 0); ib = int(b, 0)
            if ib < ia:
                ia, ib = ib, ia
            out.update(range(ia, ib + 1))
        else:
            out.add(int(part, 0))
    return out


def parse_path_set(spec: Optional[str]) -> Optional[Set[str]]:
    if not spec:
        return None
    return {p.strip().replace("\\", "/") for p in spec.split(",") if p.strip()}


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Build translated TREE.CDF TXT replacement files from pixygarden_tree_cdf_text.xlsx and write a size/fit report. Supports ASCII and CP932-fullwidth probe modes."
    )
    ap.add_argument("--xlsx", required=True, help="Workbook containing TREE Strings sheet")
    ap.add_argument("--sheet", default="TREE Strings")
    ap.add_argument("--out-dir", required=True, help="Directory where translated TXT files will be written")
    ap.add_argument("--report", required=True, help="CSV report path")
    ap.add_argument("--manifest", help="Optional recursive CDF manifest CSV from pixygarden_cdf_inspector_v2.py")

    ap.add_argument("--path-column", default="archive_path")
    ap.add_argument("--translation-column", default="english_translation")
    ap.add_argument("--fallback-translation-column", default="english_translation")
    ap.add_argument("--row-column", default="row")
    ap.add_argument("--original-size-column", default="original_file_size")
    ap.add_argument("--text-byte-length-column", default="text_byte_length")
    ap.add_argument("--terminator-column", default="terminator_byte")

    ap.add_argument("--text-mode", choices=["ascii", "cp932", "cp932_fullwidth"], default="ascii", help="Text encoding mode. cp932_fullwidth converts printable ASCII to full-width CP932 while preserving \\n controls.")
    ap.add_argument("--encoding", default="cp932", help="Encoding for ascii/cp932 modes; default cp932")
    ap.add_argument("--encoding-errors", default="replace", choices=["strict", "replace", "ignore"])
    ap.add_argument("--default-terminator", default="0x39", help="Default terminator byte if the sheet has none; bare two digits are read as hex; default 0x39")
    ap.add_argument("--force-terminator", default=None, help="Override workbook terminator_byte for every output file, e.g. 0x00 for raw ASCII TREE text")
    ap.add_argument("--align", type=lambda x: int(x, 0), default=4, help="Pad generated files to this byte alignment; default 4")
    ap.add_argument("--pad-byte", type=lambda x: int(x, 0), default=0xFF, help="Padding byte after the 0x39 TXT terminator and for unused slot space; default 0xFF")
    ap.add_argument("--pad-to-original", action="store_true", help="If generated file is smaller than the original file size, pad it out to original_file_size")
    ap.add_argument("--allow-empty", action="store_true", help="Allow empty translation cells; otherwise fallback/original behavior is reported")
    ap.add_argument("--rows", help="Only build these workbook row IDs, e.g. 1 or 1,3,5-8. Uses the workbook's row column, not Excel row numbers.")
    ap.add_argument("--paths", help="Only build these archive paths, comma-separated, e.g. DETAILS.FAT/INFO.FAT/INFO_01A.TXT")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest(Path(args.manifest) if args.manifest else None)
    row_filter = parse_row_set(args.rows)
    path_filter = parse_path_set(args.paths)

    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {args.sheet!r} not found. Available: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    row_col = resolve_col(headers, args.row_column, ["row"])
    path_col = resolve_col(headers, args.path_column, ["archive_path", "path"])
    trans_col = resolve_col(headers, args.translation_column, [args.translation_column])
    fallback_col = resolve_col(headers, args.fallback_translation_column, [args.fallback_translation_column], required=False)
    size_col = resolve_col(headers, args.original_size_column, ["original_file_size", "size"])
    text_len_col = resolve_col(headers, args.text_byte_length_column, ["text_byte_length"], required=False)
    term_col = resolve_col(headers, args.terminator_column, ["terminator_byte"], required=False)
    max_lines_col = resolve_col(headers, None, ["max_allowed_lines"], required=False)
    max_chars_col = resolve_col(headers, None, ["max_english_chars"], required=False)

    default_term = parse_byte(args.default_terminator, 0x39)
    force_term = parse_byte(args.force_terminator, None) if args.force_terminator is not None else None
    pad_byte = args.pad_byte & 0xFF

    fields = [
        "row", "archive_path", "output_path", "translation_column_used", "text_mode",
        "original_size", "original_text_byte_length", "terminator_byte",
        "line_count", "max_allowed_lines", "char_count", "max_english_chars",
        "encoded_text_len", "generated_min_size", "written_size", "delta_vs_original",
        "fits_original_file_size", "requires_growth_bytes",
        "manifest_size", "manifest_abs_off", "manifest_kind", "notes",
    ]

    report_rows = []
    total = written = fit = grow = empty = skipped = 0
    for r in range(2, ws.max_row + 1):
        path_text = norm_text(ws.cell(r, path_col).value).strip().replace("\\", "/")
        if not path_text:
            continue
        row_id = parse_int(ws.cell(r, row_col).value, r - 1)
        if row_filter is not None and row_id not in row_filter:
            skipped += 1; continue
        if path_filter is not None and path_text not in path_filter:
            skipped += 1; continue
        total += 1
        original_size = parse_int(ws.cell(r, size_col).value, 0) or 0
        original_text_len = parse_int(ws.cell(r, text_len_col).value, None) if text_len_col else None
        if force_term is not None:
            term = force_term
        else:
            term = parse_byte(ws.cell(r, term_col).value, default_term) if term_col else default_term
            if term is None:
                term = default_term
        term &= 0xFF

        text = norm_text(ws.cell(r, trans_col).value)
        used_col = args.translation_column
        if not text and fallback_col:
            text = norm_text(ws.cell(r, fallback_col).value)
            used_col = args.fallback_translation_column
        notes = []
        if not text:
            empty += 1
            notes.append("empty_translation")
            if not args.allow_empty:
                text = ""

        body = encode_tree_txt(text, args.text_mode, args.encoding, args.encoding_errors)
        min_size = align(len(body) + 1, args.align)
        file_bytes = bytearray(body)
        file_bytes.append(term)
        if len(file_bytes) < min_size:
            file_bytes.extend(bytes([pad_byte]) * (min_size - len(file_bytes)))

        if args.pad_to_original and original_size and len(file_bytes) <= original_size:
            file_bytes.extend(bytes([pad_byte]) * (original_size - len(file_bytes)))

        rel = safe_rel_path(path_text)
        out_path = out_dir / rel
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(bytes(file_bytes))
        written += 1

        fits = (len(file_bytes) <= original_size) if original_size else False
        if fits:
            fit += 1
        else:
            grow += 1
            notes.append("grows")

        max_lines = parse_int(ws.cell(r, max_lines_col).value, None) if max_lines_col else None
        max_chars = parse_int(ws.cell(r, max_chars_col).value, None) if max_chars_col else None
        line_count = 0 if text == "" else text.count("\n") + 1
        char_count = len(text.replace("\n", ""))
        if max_lines is not None and line_count > max_lines:
            notes.append("too_many_lines")
        if max_chars is not None and char_count > max_chars:
            notes.append("too_many_chars")

        mf = manifest.get(path_text, {})
        report_rows.append({
            "row": row_id,
            "archive_path": path_text,
            "output_path": str(out_path),
            "translation_column_used": used_col,
            "text_mode": args.text_mode,
            "original_size": original_size,
            "original_text_byte_length": original_text_len if original_text_len is not None else "",
            "terminator_byte": f"0x{term:02X}",
            "line_count": line_count,
            "max_allowed_lines": max_lines if max_lines is not None else "",
            "char_count": char_count,
            "max_english_chars": max_chars if max_chars is not None else "",
            "encoded_text_len": len(body),
            "generated_min_size": min_size,
            "written_size": len(file_bytes),
            "delta_vs_original": len(file_bytes) - original_size if original_size else "",
            "fits_original_file_size": "YES" if fits else "NO",
            "requires_growth_bytes": max(0, len(file_bytes) - original_size) if original_size else "",
            "manifest_size": mf.get("size", ""),
            "manifest_abs_off": mf.get("abs_off", ""),
            "manifest_kind": mf.get("kind", ""),
            "notes": ";".join(notes),
        })

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(report_rows)

    print("PixyGarden TREE TXT replacement build")
    print("--------------------------------------")
    print(f"Workbook:             {args.xlsx}")
    print(f"Text mode:            {args.text_mode}")
    print(f"Output directory:     {out_dir}")
    print(f"Report:               {report_path}")
    print(f"Rows processed:       {total}")
    print(f"Rows skipped:         {skipped}")
    print(f"Files written:        {written}")
    print(f"Fit original size:    {fit}")
    print(f"Need growth/repack:   {grow}")
    print(f"Empty translations:   {empty}")
    if args.pad_to_original:
        print("Padding mode:         smaller files padded to original_file_size")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
