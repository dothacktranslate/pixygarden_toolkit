#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Set

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("This script requires openpyxl. Install with: python -m pip install openpyxl") from e


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(s, 0)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return default


def parse_byte(v: Any, default: Optional[int] = None, bare_two_digit_hex: bool = True) -> Optional[int]:
    """Parse a byte value.

    Important PixyGarden convention: spreadsheet cells often contain bare values
    like "39" to mean hex 0x39, not decimal 39. Therefore, bare two hex digits
    are treated as hex by default.
    """
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v) & 0xFF
    if isinstance(v, int):
        return v & 0xFF
    if isinstance(v, float):
        return int(v) & 0xFF
    s = str(v).strip()
    if not s:
        return default
    if s.lower().startswith("0x"):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    if bare_two_digit_hex and re.fullmatch(r"[0-9A-Fa-f]{2}", s):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    try:
        return int(s, 0) & 0xFF
    except ValueError:
        return default


def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column letter: {col!r}")
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_col(
    headers: list[str],
    value: Optional[str],
    aliases: Iterable[str],
    required: bool = True,
) -> Optional[int]:
    """Resolve either a column letter, an exact header, or the first matching alias."""
    lowered = {h.lower(): i + 1 for i, h in enumerate(headers) if h}

    if value:
        s = value.strip()
        # Treat pure letters as Excel letters only if no exact header exists.
        if s.lower() in lowered:
            return lowered[s.lower()]
        if re.fullmatch(r"[A-Za-z]+", s):
            return col_letter_to_index(s)
        raise ValueError(f"Could not resolve column {value!r}. Use a column letter or exact header.")

    for alias in aliases:
        if alias.lower() in lowered:
            return lowered[alias.lower()]

    if required:
        raise ValueError(f"Could not find any of these columns: {', '.join(aliases)}")
    return None


def safe_rel_path(path_text: str) -> Path:
    parts = []
    for part in path_text.replace("\\", "/").split("/"):
        if part in ("", "."):
            continue
        if part == "..":
            raise ValueError(f"Unsafe archive path: {path_text!r}")
        # Keep normal PixyGarden filenames intact but remove filesystem-hostile chars.
        part = re.sub(r'[<>:"|?*\x00-\x1F]+', "_", part)
        parts.append(part)
    if not parts:
        raise ValueError("Empty archive path")
    p = Path(*parts)
    if p.is_absolute() or ".." in p.parts:
        raise ValueError(f"Unsafe archive path: {path_text!r}")
    return p


def align(value: int, n: int) -> int:
    if n <= 1:
        return value
    return ((value + n - 1) // n) * n


def parse_row_set(spec: Optional[str]) -> Optional[Set[int]]:
    if not spec:
        return None
    out: Set[int] = set()
    for part in re.split(r"[, ]+", spec.strip()):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            ia = int(a, 0)
            ib = int(b, 0)
            if ib < ia:
                ia, ib = ib, ia
            out.update(range(ia, ib + 1))
        else:
            out.add(int(part, 0))
    return out


def parse_path_set(spec: Optional[str]) -> Optional[Set[str]]:
    if not spec:
        return None
    return {p.strip().replace("\\", "/") for p in spec.split(",") if p.strip()}


def load_manifest(path: Optional[Path]) -> Dict[str, Dict[str, str]]:
    if not path:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            p = (row.get("path") or row.get("archive_path") or "").replace("\\", "/")
            if p:
                out[p] = row
    return out


def manifest_int(row: Dict[str, str], *names: str) -> Optional[int]:
    for name in names:
        v = row.get(name)
        if v not in (None, ""):
            parsed = parse_int(v, None)
            if parsed is not None:
                return parsed
    return None


# ---------------------------------------------------------------------------
# Text encoding
# ---------------------------------------------------------------------------

def to_fullwidth_ascii(ch: str) -> str:
    """Map printable one-byte ASCII to full-width Unicode/CP932 equivalents."""
    if ch == " ":
        return "　"  # U+3000 / CP932 81 40
    o = ord(ch)
    if 0x21 <= o <= 0x7E:
        return chr(0xFF00 + o - 0x20)
    return ch


def encode_linebreak(mode: str) -> bytes:
    if mode == "backslash_n":
        return b"\\n"  # 5C 6E, used by TREE/PLANET-style TXT files
    if mode == "lf":
        return b"\n"
    if mode == "crlf":
        return b"\r\n"
    if mode == "none":
        return b""
    raise ValueError(f"Unknown linebreak mode: {mode}")


def encode_body(
    text: str,
    text_mode: str,
    linebreak_mode: str,
    encoding: str,
    errors: str,
    escape_terminator_byte: Optional[int] = None,
    escape_terminator_strategy: str = "fullwidth",
) -> bytes:
    """Encode a PixyGarden TXT body.

    text_mode:
      ascii           = keep normal one-byte ASCII wherever possible.
      cp932           = encode text directly as CP932.
      cp932_fullwidth = convert printable ASCII to full-width CP932.

    linebreak_mode:
      backslash_n     = workbook newline -> literal 5C 6E control.
      lf              = workbook newline -> 0A.
      crlf            = workbook newline -> 0D 0A.
      none            = remove workbook newlines.

    escape_terminator_byte:
      If set, and text_mode is ascii/cp932, any one-byte character whose encoded
      byte equals the terminator is escaped. This is useful for TREE/PLANET
      0x39 terminator + ASCII digit '9'. Default strategy emits full-width CP932.
    """
    text = norm_text(text)
    lb = encode_linebreak(linebreak_mode)
    out = bytearray()

    for line_index, line in enumerate(text.split("\n")):
        if line_index:
            out += lb

        if text_mode == "cp932_fullwidth":
            converted = "".join(to_fullwidth_ascii(ch) for ch in line)
            out += converted.encode("cp932", errors=errors)
            continue

        if text_mode not in ("ascii", "cp932"):
            raise ValueError(f"Unknown text mode: {text_mode}")

        for ch in line:
            enc = ch.encode(encoding, errors=errors)
            if (
                escape_terminator_byte is not None
                and len(enc) == 1
                and enc[0] == (escape_terminator_byte & 0xFF)
            ):
                if escape_terminator_strategy == "fullwidth":
                    out += to_fullwidth_ascii(ch).encode("cp932", errors=errors)
                elif escape_terminator_strategy == "drop":
                    continue
                elif escape_terminator_strategy == "space":
                    out += b" "
                else:
                    raise ValueError(f"Unknown escape terminator strategy: {escape_terminator_strategy}")
            else:
                out += enc

    return bytes(out)


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(
        description=(
            "Generic PixyGarden nested-CDF TXT replacement builder. "
            "Works with TREE.CDF, PLANET.CDF, and similar workbook-driven TXT paths."
        )
    )
    ap.add_argument("--xlsx", required=True, help="Workbook containing archive paths and translations")
    ap.add_argument("--sheet", default=None, help="Worksheet name. If omitted, the first sheet is used.")
    ap.add_argument("--out-dir", required=True, help="Output directory mirroring archive-relative paths")
    ap.add_argument("--report", required=True, help="CSV report path")
    ap.add_argument("--manifest", help="Optional recursive manifest CSV; used as fallback for original sizes")

    ap.add_argument("--path-column", default=None, help="Archive path column letter or header. Auto-detects archive_path/path.")
    ap.add_argument("--translation-column", default=None, help="Translation column letter or header. Auto-prefers Better Translation.")
    ap.add_argument("--fallback-translation-column", default=None, help="Fallback translation column if primary is empty.")
    ap.add_argument("--row-column", default=None, help="Row/id column. Auto-detects row/id; otherwise Excel row is used.")
    ap.add_argument("--original-size-column", default=None, help="Original file size column. Auto-detects original_file_size/size.")
    ap.add_argument("--text-byte-length-column", default=None, help="Original text byte length column, optional.")
    ap.add_argument("--terminator-column", default=None, help="Terminator byte column, optional.")

    ap.add_argument("--text-mode", choices=["ascii", "cp932", "cp932_fullwidth"], default="ascii")
    ap.add_argument("--encoding", default="cp932", help="Encoding used for ascii/cp932 modes; default cp932")
    ap.add_argument("--encoding-errors", choices=["strict", "replace", "ignore"], default="replace")
    ap.add_argument("--linebreak-mode", choices=["backslash_n", "lf", "crlf", "none"], default="backslash_n",
                    help="How workbook line breaks are written. Default backslash_n = literal 5C 6E.")
    ap.add_argument("--default-terminator", default="0x39",
                    help="Default terminator byte if sheet has none. Bare two digits are interpreted as hex.")
    ap.add_argument("--force-terminator", default=None,
                    help="Override terminator for every file, e.g. 0x00 for TREE ASCII workflow.")
    ap.add_argument("--pad-byte", type=lambda x: int(x, 0), default=0xFF,
                    help="Padding byte after terminator / to original size. Default 0xFF.")
    ap.add_argument("--align", type=lambda x: int(x, 0), default=4,
                    help="Pad each generated TXT to this alignment before optional original-size padding. Default 4.")
    ap.add_argument("--pad-to-original", action="store_true",
                    help="Pad smaller replacement files to their original file size.")
    ap.add_argument("--escape-terminator-byte", action="store_true",
                    help="Escape one-byte text chars equal to the terminator, e.g. ASCII '9' when terminator is 0x39.")
    ap.add_argument("--escape-terminator-strategy", choices=["fullwidth", "drop", "space"], default="fullwidth")

    ap.add_argument("--include-ext", default=".TXT",
                    help="Only build archive paths ending in this extension. Use empty string to disable. Default .TXT.")
    ap.add_argument("--include-regex", default=None, help="Only build paths matching this regex.")
    ap.add_argument("--exclude-regex", default=None, help="Skip paths matching this regex, e.g. '(^|/)HELP\\.TXT$'.")
    ap.add_argument("--rows", help="Only build these workbook row IDs, e.g. 1,3,5-8")
    ap.add_argument("--paths", help="Only build these archive paths, comma-separated")
    ap.add_argument("--allow-empty", action="store_true", help="Allow empty translation cells without warning")

    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest(Path(args.manifest) if args.manifest else None)
    row_filter = parse_row_set(args.rows)
    path_filter = parse_path_set(args.paths)
    include_re = re.compile(args.include_regex) if args.include_regex else None
    exclude_re = re.compile(args.exclude_regex) if args.exclude_regex else None

    wb = load_workbook(args.xlsx, data_only=True)
    sheet = args.sheet or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} not found. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    path_col = resolve_col(
        headers, args.path_column,
        ["archive_path", "path", "file_path", "archive file", "archive_path_normalized", "cdf_path"]
    )

    trans_col = resolve_col(
        headers, args.translation_column,
        [
            "Better Translation",
            "QA Clean Translation",
            "Final Translation",
            "English Translation",
            "english_translation",
            "Clean Translation",
            "Translation",
            "Insertion Text",
            "Final Insertion Text",
        ]
    )

    fallback_col = resolve_col(
        headers, args.fallback_translation_column,
        [
            "english_translation",
            "English Translation",
            "Translation",
            "Clean Translation",
            "Original Translation",
        ],
        required=False,
    )

    row_col = resolve_col(headers, args.row_column, ["row", "id", "index"], required=False)
    size_col = resolve_col(headers, args.original_size_column, ["original_file_size", "file_size", "size"], required=False)
    text_len_col = resolve_col(headers, args.text_byte_length_column, ["text_byte_length", "original_text_byte_length"], required=False)
    term_col = resolve_col(headers, args.terminator_column, ["terminator_byte", "terminator"], required=False)
    max_lines_col = resolve_col(headers, None, ["max_allowed_lines"], required=False)
    max_chars_col = resolve_col(headers, None, ["max_english_chars"], required=False)

    default_term = parse_byte(args.default_terminator, 0x39)
    force_term = parse_byte(args.force_terminator, None) if args.force_terminator is not None else None
    pad_byte = args.pad_byte & 0xFF

    fields = [
        "excel_row", "row_id", "archive_path", "output_path", "translation_column_used",
        "text_mode", "linebreak_mode", "original_size", "original_text_byte_length",
        "terminator_byte", "pad_byte", "line_count", "char_count",
        "encoded_text_len", "generated_min_size", "written_size", "delta_vs_original",
        "fits_original_file_size", "requires_growth_bytes", "escaped_terminator_byte",
        "manifest_size", "manifest_abs_off", "manifest_kind", "notes",
    ]

    report_rows = []
    total_seen = total_built = skipped = fit = grow = empty = 0

    for excel_row in range(2, ws.max_row + 1):
        raw_path = norm_text(ws.cell(excel_row, path_col).value).strip().replace("\\", "/")
        if not raw_path:
            continue

        total_seen += 1
        row_id = parse_int(ws.cell(excel_row, row_col).value, excel_row - 1) if row_col else excel_row - 1

        if row_filter is not None and row_id not in row_filter:
            skipped += 1
            continue
        if path_filter is not None and raw_path not in path_filter:
            skipped += 1
            continue
        if args.include_ext and not raw_path.upper().endswith(args.include_ext.upper()):
            skipped += 1
            continue
        if include_re and not include_re.search(raw_path):
            skipped += 1
            continue
        if exclude_re and exclude_re.search(raw_path):
            skipped += 1
            continue

        mf = manifest.get(raw_path, {})
        original_size = parse_int(ws.cell(excel_row, size_col).value, None) if size_col else None
        if original_size is None:
            original_size = manifest_int(mf, "size", "file_size", "new_size", "source_size")
        if original_size is None:
            original_size = 0

        original_text_len = parse_int(ws.cell(excel_row, text_len_col).value, None) if text_len_col else None

        if force_term is not None:
            term = force_term
        elif term_col:
            term = parse_byte(ws.cell(excel_row, term_col).value, default_term)
        else:
            term = default_term
        if term is None:
            term = default_term
        term &= 0xFF

        text = norm_text(ws.cell(excel_row, trans_col).value)
        used_col = headers[trans_col - 1] or str(trans_col)

        if not text and fallback_col:
            text = norm_text(ws.cell(excel_row, fallback_col).value)
            used_col = headers[fallback_col - 1] or str(fallback_col)

        notes = []
        if not text:
            empty += 1
            notes.append("empty_translation")
            if not args.allow_empty:
                # Still write a terminator-only file; report makes this visible.
                text = ""

        escape_byte = term if args.escape_terminator_byte else None
        body = encode_body(
            text=text,
            text_mode=args.text_mode,
            linebreak_mode=args.linebreak_mode,
            encoding=args.encoding,
            errors=args.encoding_errors,
            escape_terminator_byte=escape_byte,
            escape_terminator_strategy=args.escape_terminator_strategy,
        )

        # Report whether the raw body would have contained the terminator byte if not escaped.
        raw_body_for_check = encode_body(
            text=text,
            text_mode=args.text_mode,
            linebreak_mode=args.linebreak_mode,
            encoding=args.encoding,
            errors=args.encoding_errors,
            escape_terminator_byte=None,
        )
        escaped = (
            args.escape_terminator_byte
            and term in raw_body_for_check
            and raw_body_for_check != body
        )

        file_bytes = bytearray(body)
        file_bytes.append(term)
        min_size = align(len(file_bytes), args.align)
        if len(file_bytes) < min_size:
            file_bytes.extend(bytes([pad_byte]) * (min_size - len(file_bytes)))

        if args.pad_to_original and original_size and len(file_bytes) <= original_size:
            file_bytes.extend(bytes([pad_byte]) * (original_size - len(file_bytes)))

        rel = safe_rel_path(raw_path)
        out_path = out_dir / rel
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(bytes(file_bytes))
        total_built += 1

        fits = (len(file_bytes) <= original_size) if original_size else False
        if fits:
            fit += 1
        else:
            grow += 1
            notes.append("grows_or_unknown_original_size" if not original_size else "grows")

        line_count = 0 if text == "" else text.count("\n") + 1
        char_count = len(text.replace("\n", ""))

        max_lines = parse_int(ws.cell(excel_row, max_lines_col).value, None) if max_lines_col else None
        max_chars = parse_int(ws.cell(excel_row, max_chars_col).value, None) if max_chars_col else None
        if max_lines is not None and line_count > max_lines:
            notes.append("too_many_lines")
        if max_chars is not None and char_count > max_chars:
            notes.append("too_many_chars")
        if term in raw_body_for_check and not args.escape_terminator_byte:
            notes.append(f"body_contains_terminator_byte_0x{term:02X}")

        report_rows.append({
            "excel_row": excel_row,
            "row_id": row_id,
            "archive_path": raw_path,
            "output_path": str(out_path),
            "translation_column_used": used_col,
            "text_mode": args.text_mode,
            "linebreak_mode": args.linebreak_mode,
            "original_size": original_size if original_size else "",
            "original_text_byte_length": original_text_len if original_text_len is not None else "",
            "terminator_byte": f"0x{term:02X}",
            "pad_byte": f"0x{pad_byte:02X}",
            "line_count": line_count,
            "char_count": char_count,
            "encoded_text_len": len(body),
            "generated_min_size": min_size,
            "written_size": len(file_bytes),
            "delta_vs_original": len(file_bytes) - original_size if original_size else "",
            "fits_original_file_size": "YES" if fits else "NO",
            "requires_growth_bytes": max(0, len(file_bytes) - original_size) if original_size else "",
            "escaped_terminator_byte": "YES" if escaped else "NO",
            "manifest_size": mf.get("size", ""),
            "manifest_abs_off": mf.get("abs_off", ""),
            "manifest_kind": mf.get("kind", ""),
            "notes": ";".join(notes),
        })

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(report_rows)

    print("PixyGarden generic TXT replacement build")
    print("----------------------------------------")
    print(f"Workbook:             {args.xlsx}")
    print(f"Sheet:                {sheet}")
    print(f"Text mode:            {args.text_mode}")
    print(f"Linebreak mode:       {args.linebreak_mode}")
    print(f"Output directory:     {out_dir}")
    print(f"Report:               {report_path}")
    print(f"Rows seen:            {total_seen}")
    print(f"Rows skipped:         {skipped}")
    print(f"Files written:        {total_built}")
    print(f"Fit original size:    {fit}")
    print(f"Need growth/repack:   {grow}")
    print(f"Empty translations:   {empty}")
    if args.pad_to_original:
        print("Padding mode:         smaller files padded to original file size")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
