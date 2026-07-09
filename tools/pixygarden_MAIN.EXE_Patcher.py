#!/usr/bin/env python3
from __future__ import annotations

# v35 broadens the PLANET/Modus FT4/opening-frame element repair: accepts small X drift, matches already-patched element U aliases, and replaces older FT4 source hooks automatically.
# v36 updates REPORT text bytes: Pixies/Objects spacing and MODUS -> MODI.
# v37 makes REPORT_TEXT_PATCH_BYTES an explicit full 120-byte block, including final MODI MADE null padding.
# v40 rolls back v39 steady/template X normalization after it shifted other element graphics; keeps the narrow FT4/source flicker repair. FT4 X adjust default is 0.
# v38 normalizes PLANET/Modus FT4 opening element X to prevent the final 2-4px left snap.
# v41 makes Object Overlay element remaps value-aware instead of forcing N/A.
# v47 rollback/stabilization: based on v43 to restore the verified PLANET/Modus element UV/width path; removes v44-v46 secondary Object Overlay FT4 overlap that corrupted Modus graphics.
# v48 narrowly repairs the observed PLANET Modus/Point element template at x=0x7D/w=0x12 before it renders as clipped Eart/ndFi.
# v50 rolls back v49's packet-level SPRT source-word repair after it also
# affected Neredy/main-menu sprites. Based on v48: keeps the narrow
# Modus/Point bad-template repair and 120-byte REPORT label block, but
# does not patch the shared already-built SPRT packet store.
# v52 rolls back the v51 shared-SPRT/template-guard experiment because it still corrupted main-menu/Neredy graphics; based on safer v50.
# v58 adds a local PLANET/Modus record repair at 80079998, based on No$PSX read/write breaks that identified active element record BASE+0xAD4 (8012A3BC). This avoids the unsafe shared SPRT path.
# v59 adds an earlier local PLANET/Modus record repair at 8007972C, inside the record-read loop, because v58 shortened but did not eliminate the Modus/Point element flicker.
# v60 adds exact active-record FT4 edge-read hooks at 800A3CAC and 800A3CC8. Earth already appeared clean while Fire still flashed briefly, so these hooks correct the current U/W registers immediately when the FT4 path reads the active Modus record 8012A3BC.
# v64 adds an exact Y/shape-specific Proceed without saving? selector fix. Hook-entry tests showed the real selected boxes are normal Yes/No shapes at y=0x41, not the v62 raw8 chunks; v63's RA-only match was too broad and touched Vibration Function.
# v34 adds a narrow Results/Clear Data save-prompt selector-box Y adjustment.
# Renderer metric defaults:
# v110 tunes the confirmed FUN_80051780 hook: OFF remains as v109; ON box shifted 1px left. Yes/No remains as v108/v109.
# v105 adds a guarded FUN_80051780 hook for small grey Yes/No/ON/OFF word boxes.
# v102 reverts the bad v101 selector-word-box experiment; based on safe v100/v97 with confirmed fixes only.
# v95 is based on v94 and adds full-screen memory-card/status text centering for rows 2-38.
# v97 makes the confirmed Pixy-name suffix spacing fix cover the Modus-created suffix cases by default.
# v94 adds explicit CP932 advance/x-offset handling for PlayStation controller symbols: × ○ △ □.
#   i/l/I left 2px, T/G right 1px, apostrophe right 3px, comma/period right 2px,
#   m/w right 1px, capital R advance -1px. Final advance deltas: A-Z -1,
#   hyphen +3, slash +1. Use --draw-shift CHARS PIXELS for custom tuning. Includes PLSEL TIM14 draw patch defaults, including exact following/circle source fixes for both static and dynamic draw paths.
#
# v93 fixes the STAGE moved-UV patch: restore the v82 mistaken planet-name U patch and patch the real TIM04 N/A-line slot U.
# v23 reverts the PLMENU01 planet-icon CLUT default to the original selector after 0x01E1 proved incorrect.
# v22 added an optional PLMENU01 planet-icon palette/CLUT selector patch.
# v21 adds the PLMENU01 element-info draw patch: the dynamic element label
# selector now samples the widened English source rectangles for Earth/Water/
# Wind/Fire/None and changes the sprite width per element. It also updates
# the nearby planet icon source U by default.
#
# Reporting behavior:
#   No CSV/JSON reports are written by default.
#   Pass --reports-dir DIR to write the old CSV/JSON report set.
#
import argparse,csv,json,math,re,shutil,struct,subprocess,tempfile,collections
from pathlib import Path
from dataclasses import dataclass
from typing import Optional
from openpyxl import load_workbook

PSX_HEADER=0x800; SECTOR=0x800
TEXT_LOOP=0x80037D44
TEXT_EXPECT=bytes.fromhex("00000292 01000392")
FALLBACK=0x80037D4C; GLYPH=0x80038038; LOOP_TAIL=0x800381D8; RET=0x800381E8
BSS_END=0x80110FD0
HEAP_PAIR=0x800C60F4
HEAP_EXPECT=bytes.fromhex("1180043c d00f8424")
GP_X=0x538; GP_Y=0x53C; GP_START_X=0x524; GP_HADV=0x50; GP_VADV=0x54
SECTIONS=[("rdata",0x80020000,0x800288FC),("text",0x800288FC,0x800CCC54),("data",0x800CCC54,0x800DA4FC),("sdata",0x800DA4FC,0x800DA898),("sbss",0x800DA898,0x800DAED8),("bss",0x800DAED8,0x80110FD0)]
REG={"zero":0,"at":1,"v0":2,"v1":3,"a0":4,"a1":5,"a2":6,"a3":7,"t0":8,"t1":9,"t2":10,"t3":11,"t4":12,"t5":13,"t6":14,"t7":15,"s0":16,"s1":17,"s2":18,"s3":19,"s4":20,"s5":21,"s6":22,"s7":23,"t8":24,"t9":25,"gp":28,"sp":29,"fp":30,"ra":31}

def s16(x): x&=0xffff; return x-0x10000 if x&0x8000 else x
def hi(addr): return ((addr+0x8000)>>16)&0xffff
def lo(addr): return addr&0xffff
def it(op,rs,rt,imm): return (op<<26)|(rs<<21)|(rt<<16)|(imm&0xffff)
def rt(rs,rt,rd,sh,fn): return (rs<<21)|(rt<<16)|(rd<<11)|(sh<<6)|fn
def jt(op,addr): return (op<<26)|((addr>>2)&0x3ffffff)
def w(x): return struct.pack("<I",x&0xffffffff)
def align(v,n): return (v+n-1)//n*n


# TREE.CDF details text-buffer patch.
# The TREE details/info screen copies INFO_*.TXT bytes into a local stack
# buffer before printing. Original routine uses roughly:
#   stack frame 0x118, buffer clear 0xFC, copy limit 0xFA.
# Full-width CP932 English entries can exceed that, so this expands the
# stack frame and raises the byte copy limit.
TREEBUF_PATCH_SITES = {
    "frame_sub": 0x69E94,   # addiu sp,sp,-frame
    "save_ra":   0x69EA0,   # sw ra,ra_off(sp)
    "clear_len": 0x69EA8,   # addiu a2,zero,clear_size
    "copy_limit":0x69F0C,   # slti v0,a0,copy_limit
    "load_ra":   0x69F40,   # lw ra,ra_off(sp)
    "frame_add": 0x69F4C,   # addiu sp,sp,frame
}
TREEBUF_ORIGINAL_WORDS = {
    "frame_sub": 0x27BDFEE8,
    "save_ra":   0xAFBF0110,
    "clear_len": 0x240600FC,
    "copy_limit":0x288200FA,
    "load_ra":   0x8FBF0110,
    "frame_add": 0x27BD0118,
}
TREEBUF_REG = {"zero":0, "v0":2, "a0":4, "a2":6, "sp":29, "ra":31}

TREE_COPY_TERMINATOR_SITE = 0x69EEC  # addiu a3, zero, 0x0039 in TREE details TXT copy routine
TREE_COPY_TERMINATOR_ORIGINAL = 0x24070039

def patch_tree_copy_terminator(exe, terminator=0x39, force=False, dry_run=False):
    terminator &= 0xFF
    cur = read32(exe, TREE_COPY_TERMINATOR_SITE)
    op=(cur>>26)&0x3F; rs=(cur>>21)&31; rt=(cur>>16)&31
    shape_ok = (op == 0x09 and rs == TREEBUF_REG["zero"] and rt == 7)  # addiu a3,zero,imm
    if not force and not (cur == TREE_COPY_TERMINATOR_ORIGINAL or shape_ok):
        raise RuntimeError(
            f"TREE terminator patch site at 0x{TREE_COPY_TERMINATOR_SITE:X} has unexpected word 0x{cur:08X}; "
            "expected addiu a3,zero,imm. Use --tree-terminator-force to patch anyway."
        )
    new_word = itype_word(0x09, TREEBUF_REG["zero"], 7, terminator)
    if not dry_run:
        exe[TREE_COPY_TERMINATOR_SITE:TREE_COPY_TERMINATOR_SITE+4] = struct.pack("<I", new_word & 0xffffffff)
    return {
        "enabled": True,
        "file_offset": f"0x{TREE_COPY_TERMINATOR_SITE:X}",
        "old_word": f"0x{cur:08X}",
        "new_word": f"0x{new_word:08X}",
        "terminator": f"0x{terminator:02X}",
        "status": "would_patch" if dry_run else "patched",
    }


# PLANET.CDF / STAGE HELP.FAT copy routines also used 0x39 as a text terminator.
# When translated text uses raw ASCII or CP932 full-width digits, 0x39 is unsafe
# because it is also ASCII '9'. These sites mirror the standalone post-patcher
# used earlier, but are now integrated into the primary patcher.
PLANET_STAGE_TERMINATOR_SITES = {
    "planet_help_terminator": {
        "offset": 0x61248,
        "expected": 0x24070039,  # addiu/ori-style imm 0x0039 into a3
        "rt": 7,
        "description": "PLANET.CDF HELP.FAT text copy terminator",
    },
    "stage_help_terminator": {
        "offset": 0x6CD78,
        "expected": 0x24060039,  # addiu/ori-style imm 0x0039 into a2
        "rt": 6,
        "description": "DATA/STAGE/HELP.FAT text copy terminator",
    },
}

def patch_planet_stage_help_terminators(exe, terminator=0x00, force=False, dry_run=False):
    terminator &= 0xFF
    rows=[]
    for name, info in PLANET_STAGE_TERMINATOR_SITES.items():
        off=info["offset"]; expected=info["expected"]; rt_expected=info["rt"]
        cur=read32(exe, off)
        op=(cur>>26)&0x3F; rs=(cur>>21)&31; rt=(cur>>16)&31
        shape_ok = (op == 0x09 and rs == TREEBUF_REG["zero"] and rt == rt_expected)
        if not force and not (cur == expected or shape_ok):
            raise RuntimeError(
                f"{name} at 0x{off:X} has unexpected word 0x{cur:08X}; "
                f"expected compatible addiu r{rt_expected},zero,imm. Use --planet-stage-terminator-force if sure."
            )
        new_word = itype_word(0x09, TREEBUF_REG["zero"], rt_expected, terminator)
        if not dry_run:
            exe[off:off+4]=struct.pack("<I", new_word & 0xffffffff)
        rows.append({
            "site": name,
            "description": info["description"],
            "file_offset": f"0x{off:X}",
            "old_word": f"0x{cur:08X}",
            "new_word": f"0x{new_word:08X}",
            "terminator": f"0x{terminator:02X}",
            "status": "would_patch" if dry_run else "patched",
        })
    return {"enabled": True, "terminator": f"0x{terminator:02X}", "report": rows}


# REPORT.CDF / report-screen label text block patch.
# Keeps the manually fixed report labels in MAIN.EXE from being wiped by the
# main patcher. Same-size replacement; no data shifts.
REPORT_TEXT_PATCH_OFFSET = 0x3504
REPORT_TEXT_PATCH_BYTES = (
    b"  MISSION COMPLETE\x00\x00"
    b"  MISSION FAILED\x00\x00\x00\x00"
    b"EVENTS ENCOUNTERED\x00\x00"
    b"PIXIES EVOLVED    \x00\x00"
    b"OBJECTS FOUND     \x00\x00"
    b"MODI MADE         \x00\x00"
)

def patch_report_text_block(exe, force=False, dry_run=False):
    off = REPORT_TEXT_PATCH_OFFSET
    end = off + len(REPORT_TEXT_PATCH_BYTES)
    if end > len(exe):
        raise RuntimeError(
            f"REPORT text patch range 0x{off:X}-0x{end-1:X} exceeds EXE size 0x{len(exe):X}"
        )
    old = bytes(exe[off:end])
    if old == REPORT_TEXT_PATCH_BYTES:
        return {
            "enabled": True,
            "file_offset": f"0x{off:X}",
            "range": f"0x{off:X}-0x{end-1:X}",
            "length": len(REPORT_TEXT_PATCH_BYTES),
            "status": "already_applied",
        }

    # This area should contain report/menu label text. Be conservative so a
    # wrong EXE or unexpected version does not get silently patched.
    sanity_needles = (b"MISSION", b"OBJECT", b"FOUND", b"MODUS", b"MADE")
    if not force and not any(n in old for n in sanity_needles):
        preview = old[:64].hex(" ")
        raise RuntimeError(
            "REPORT text patch sanity check failed at 0x3504. "
            "The current bytes do not look like the expected report label block. "
            f"Preview: {preview}. Use --report-text-force if this is intentional."
        )

    if not dry_run:
        exe[off:end] = REPORT_TEXT_PATCH_BYTES
    return {
        "enabled": True,
        "file_offset": f"0x{off:X}",
        "range": f"0x{off:X}-0x{end-1:X}",
        "length": len(REPORT_TEXT_PATCH_BYTES),
        "status": "would_patch" if dry_run else "patched",
    }


def read32(buf, off):
    return struct.unpack_from("<I", buf, off)[0]

def itype_word(op, rs, rt, imm):
    return ((op & 0x3F) << 26) | ((rs & 31) << 21) | ((rt & 31) << 16) | (imm & 0xFFFF)

def neg_imm(value):
    return (-value) & 0xFFFF

def treebuf_word_shape_ok(name, cur):
    """Allow original words and already-patched words with the expected opcode/register shape."""
    if cur == TREEBUF_ORIGINAL_WORDS[name]:
        return True
    op=(cur>>26)&0x3F; rs=(cur>>21)&31; rt=(cur>>16)&31
    if name == "frame_sub":
        return op==0x09 and rs==TREEBUF_REG["sp"] and rt==TREEBUF_REG["sp"]
    if name == "save_ra":
        return op==0x2B and rs==TREEBUF_REG["sp"] and rt==TREEBUF_REG["ra"]
    if name == "clear_len":
        return op==0x09 and rs==TREEBUF_REG["zero"] and rt==TREEBUF_REG["a2"]
    if name == "copy_limit":
        return op==0x0A and rs==TREEBUF_REG["a0"] and rt==TREEBUF_REG["v0"]
    if name == "load_ra":
        return op==0x23 and rs==TREEBUF_REG["sp"] and rt==TREEBUF_REG["ra"]
    if name == "frame_add":
        return op==0x09 and rs==TREEBUF_REG["sp"] and rt==TREEBUF_REG["sp"]
    return False


# PLANET/MODOUS Crest text X-position candidate patcher.
#
# Several different places draw the same Underground Crest string/pointer.
# v10/v12 moved the list-entry instances. The target is the currently
# selected Crest-name box, drawn by FUN_800796f0. v14 makes that the default
# and shifts it left by 14 px:
#
#   --modus-crest-text-site detail_800796f0
#   --modus-crest-text-site list_8007cd18_crest
#   --modus-crest-text-site list_80098dc8_crest
#   --modus-crest-text-site list_8009932c_a
#   --modus-crest-text-site list_8009932c_b
#   --modus-crest-text-site list_8009932c_c
#
# Use --modus-crest-text-site none to disable this patch.
MODUS_CREST_TEXT_SITE_CHOICES = {
    # Remaining untried direct Crest detail path:
    #   FUN_800796f0:
    #     FUN_80038474(0x1e,0xffffffc3,0x60,0x24)
    #     FUN_8003861c(&DAT_800da7a8)
    #     FUN_8003861c(CrestName)
    "detail_800796f0": {
        "mode": "range",
        "ranges": [(0x800796F0, 0x8007A32C, "FUN_800796f0_current_selected_crest_name")],
        "old_x": 0x1E,
        "default_new_x": 0x10,  # -14
        "description": "Currently selected Crest-name text window in FUN_800796f0, x 0x1E -> 0x10.",
    },

    # Known wrong for user's target, but kept as an explicit diagnostic only.
    "list_8007cd18_crest": {
        "mode": "fixed_call",
        "call_file": 0x5E8B0,
        "old_x": -130,
        "default_new_x": -118,
        "description": "FUN_8007cd18 Crest list call from v12; known to affect one non-target Underground Crest box.",
    },

    # Candidate found between FUN_80098dc8 and FUN_8009932c.
    "list_80098dc8_crest": {
        "mode": "range",
        "ranges": [(0x80098DC8, 0x8009932C, "FUN_80098dc8_crest_list")],
        "old_x": -130,
        "default_new_x": -118,
        "description": "Crest list candidate immediately before FUN_8009932c.",
    },

    # Candidates v10 could touch inside FUN_8009932c. These are opt-in only.
    "list_8009932c_a": {
        "mode": "fixed_call",
        "call_file": 0x7A0F4,
        "old_x": -130,
        "default_new_x": -118,
        "description": "FUN_8009932c candidate A.",
    },
    "list_8009932c_b": {
        "mode": "fixed_call",
        "call_file": 0x7A874,
        "old_x": -130,
        "default_new_x": -118,
        "description": "FUN_8009932c candidate B.",
    },
    "list_8009932c_c": {
        "mode": "fixed_call",
        "call_file": 0x7B5EC,
        "old_x": -130,
        "default_new_x": -118,
        "description": "FUN_8009932c candidate C.",
    },
}

def _mips_word(buf, off):
    return struct.unpack_from("<I", buf, off)[0]

def _mips_jal_word(addr):
    return jt(3, addr) & 0xFFFFFFFF

def _mips_is_li_reg_imm(word, reg, imm):
    op = (word >> 26) & 0x3F
    rs = (word >> 21) & 31
    rt = (word >> 16) & 31
    return op in (0x09, 0x0D) and rs == 0 and rt == reg and (word & 0xFFFF) == (imm & 0xFFFF)

def _mips_find_last_li_a0_before_call(buf, call_off, x_value):
    candidates = []
    for off in range(max(0, call_off - 0xA0) & ~3, call_off, 4):
        word = _mips_word(buf, off)
        if _mips_is_li_reg_imm(word, 4, x_value):
            candidates.append(off)
    return candidates[-1] if candidates else None

def _mips_has_word_in_range(buf, word, start, end):
    start = max(0, start & ~3)
    end = min(len(buf), end & ~3)
    for off in range(start, end, 4):
        if _mips_word(buf, off) == word:
            return True
    return False

def _mips_candidate_has_text_draw_submit(buf, call_off, force=False):
    has_draw = _mips_has_word_in_range(buf, _mips_jal_word(0x8003861C), call_off, call_off + 0xA0)
    has_submit = _mips_has_word_in_range(buf, _mips_jal_word(0x800384C0), call_off, call_off + 0x120)
    if not (has_draw and has_submit) and not force:
        return False, "missing_nearby_text_draw_submit"
    return True, "matched"

def _patch_one_crest_call(exe, load, call_off, old_x, new_x, label, force=False, dry_run=False):
    if call_off < 0 or call_off + 4 > len(exe):
        raise RuntimeError(f"Crest candidate {label} call offset 0x{call_off:X} is outside EXE size 0x{len(exe):X}.")

    cur_call = _mips_word(exe, call_off)
    expected_call = _mips_jal_word(0x80038474)
    if cur_call != expected_call and not force:
        raise RuntimeError(
            f"Crest candidate {label} call mismatch at 0x{call_off:X}: found 0x{cur_call:08X}, "
            f"expected JAL FUN_80038474 word 0x{expected_call:08X}. "
            "Use --modus-crest-text-x-force only if this MAIN.EXE layout is intentional."
        )

    ok, reason = _mips_candidate_has_text_draw_submit(exe, call_off, force=force)
    if not ok:
        raise RuntimeError(
            f"Crest candidate {label} at call 0x{call_off:X} rejected: {reason}. "
            "Use --modus-crest-text-x-force only if this is intentional."
        )

    a0_old_off = _mips_find_last_li_a0_before_call(exe, call_off, old_x)
    a0_new_off = _mips_find_last_li_a0_before_call(exe, call_off, new_x)

    if a0_old_off is not None:
        old_word = _mips_word(exe, a0_old_off)
        new_word = patch_low16_immediate(old_word, new_x)
        if not dry_run:
            struct.pack_into("<I", exe, a0_old_off, new_word & 0xFFFFFFFF)
        return {
            "name": "modus_crest_text_x_candidate",
            "candidate": label,
            "file_offset": f"0x{a0_old_off:X}",
            "ram_address": f"0x{off2ram(a0_old_off, load):08X}",
            "call_file": f"0x{call_off:X}",
            "call_ram": f"0x{off2ram(call_off, load):08X}",
            "old_word": f"0x{old_word:08X}",
            "new_word": f"0x{new_word:08X}",
            "old_x": old_x,
            "new_x": new_x,
            "status": "would_patch" if dry_run else "patched",
        }

    if a0_new_off is not None:
        word = _mips_word(exe, a0_new_off)
        return {
            "name": "modus_crest_text_x_candidate",
            "candidate": label,
            "file_offset": f"0x{a0_new_off:X}",
            "ram_address": f"0x{off2ram(a0_new_off, load):08X}",
            "call_file": f"0x{call_off:X}",
            "call_ram": f"0x{off2ram(call_off, load):08X}",
            "old_word": f"0x{word:08X}",
            "new_word": f"0x{word:08X}",
            "old_x": old_x,
            "new_x": new_x,
            "status": "already_applied",
        }

    raise RuntimeError(
        f"Crest candidate {label}: could not find a0={old_x} or a0={new_x} before call 0x{call_off:X}."
    )

def patch_modus_crest_text_x(exe, args, dry_run=False):
    if getattr(args, "disable_modus_crest_text_x_patch", False):
        return {"enabled": False, "status": "disabled"}

    site = getattr(args, "modus_crest_text_site", "detail_800796f0")
    if site in (None, "", "none"):
        return {
            "enabled": False,
            "status": "disabled_site_none",
            "available_sites": sorted(MODUS_CREST_TEXT_SITE_CHOICES.keys()),
        }

    if site not in MODUS_CREST_TEXT_SITE_CHOICES:
        raise RuntimeError(
            f"Unknown --modus-crest-text-site {site!r}. Valid choices: "
            + ", ".join(["none"] + sorted(MODUS_CREST_TEXT_SITE_CHOICES.keys()))
        )

    if bytes(exe[:8]) != b"PS-X EXE":
        raise RuntimeError("Modus Crest text X patch expected a PS-X EXE buffer.")
    load = struct.unpack_from("<I", exe, 0x18)[0]

    cfg = MODUS_CREST_TEXT_SITE_CHOICES[site]
    old_x = int(cfg["old_x"])
    # Backward-compatible behavior: the old global default was -118. Treat an
    # unchanged parser value of -118 as "use this site's built-in default";
    # explicit --modus-crest-text-x still overrides.
    requested_x = int(getattr(args, "modus_crest_text_x", cfg["default_new_x"]))
    new_x = int(cfg["default_new_x"]) if requested_x == -118 and cfg["default_new_x"] != -118 else requested_x
    force = bool(getattr(args, "modus_crest_text_x_force", False))

    report = []
    if cfg["mode"] == "fixed_call":
        report.append(_patch_one_crest_call(
            exe, load, int(cfg["call_file"]), old_x, new_x, site,
            force=force, dry_run=dry_run
        ))
    elif cfg["mode"] == "range":
        for start_ram, end_ram, label in cfg["ranges"]:
            start_off = max(0, ram2off(start_ram, load))
            end_off = min(len(exe), ram2off(end_ram, load))
            for call_off in range(start_off & ~3, end_off - 4, 4):
                if _mips_word(exe, call_off) != _mips_jal_word(0x80038474):
                    continue
                # Candidate must have old or already-new x close before call.
                if (_mips_find_last_li_a0_before_call(exe, call_off, old_x) is None and
                    _mips_find_last_li_a0_before_call(exe, call_off, new_x) is None):
                    continue
                try:
                    report.append(_patch_one_crest_call(
                        exe, load, call_off, old_x, new_x, f"{site}:{label}",
                        force=force, dry_run=dry_run
                    ))
                except RuntimeError:
                    if force:
                        raise
                    continue
        if not report:
            raise RuntimeError(f"Crest candidate {site} found no matching FUN_80038474 call in its function range.")
    else:
        raise RuntimeError(f"Internal error: unsupported Crest candidate mode {cfg['mode']!r}")

    patched = sum(1 for r in report if r.get("status") in ("patched", "would_patch"))
    already = sum(1 for r in report if r.get("status") == "already_applied")

    return {
        "enabled": True,
        "status": "would_patch" if dry_run and patched else ("patched" if patched else "already_applied"),
        "candidate": site,
        "old_x": old_x,
        "new_x": new_x,
        "patched_sites": patched,
        "already_applied_sites": already,
        "description": cfg.get("description", ""),
        "report": report,
    }


# PLSEL.CDF / PLSEL00.DAT TIM14 drawing patch.
# Known safe candidate from the PLSEL quad analysis:
#   combined strip UV:      (73,48) to (157,64)
#   combined strip X range: -42 to 42
# Default patch expands the right edge so the edited atlas graphic can extend
# to x=162 without being clipped/squeezed:
#   right U 157 -> 162
#   right X 42  -> 47
#
# The following/planet/circle graphic originally at atlas x=160,y=48 is drawn as
# two textured strips:
#   strip A UV: 160,48 -> 186,112   local X: -42 -> -16
#   strip B UV: 198,48 -> 224,112   local X: -24 ->   2
#
# In the final edited TIM14, that source art is two separated circle halves:
#   left half:  U 180..205, V 48..110
#   right half: U 219..244, V 48..110
#
# The built-in default patch now updates the source U coordinates to those exact
# values and keeps the graphic in the original screen position.
PLSEL_DRAW_PATCH_SITES = {
    "combined_strip_left_u":   (0x3AB40, 0x24050049, "addiu a1,zero,73  ; combined strip left U"),
    "combined_strip_top_v":    (0x3AB50, 0x24030030, "addiu v1,zero,48  ; combined strip top V"),
    "combined_strip_right_u":  (0x3AB5C, 0x2404009D, "addiu a0,zero,157 ; combined strip right U"),
    "combined_strip_right_x":  (0x3AB8C, 0x2403002A, "addiu v1,zero,42  ; combined strip right local X"),
    "combined_strip_left_x":   (0x3AB98, 0x2405FFD6, "addiu a1,zero,-42 ; combined strip left local X"),
    "combined_strip_bottom_v": (0x3AAA0, 0x24120040, "addiu s2,zero,64  ; combined strip bottom V"),

    # Following/planet graphic moved from atlas x=160 to x=180.
    # It is split into two strips, both using TIM14 texture page x=960.
    "following_strip_a_left_u":  (0x430FC, 0x240400A0, "addiu a0,zero,160 ; following graphic strip A left U"),
    "following_strip_a_right_u": (0x43114, 0x240300BA, "addiu v1,zero,186 ; following graphic strip A right U"),
    "following_strip_b_left_u":  (0x43150, 0x240800C6, "addiu t0,zero,198 ; following graphic strip B left U"),
    "following_strip_b_right_u": (0x43194, 0x240300E0, "addiu v1,zero,224 ; following graphic strip B right U"),

    "following_strip_a_left_x":  (0x4315C, 0x2404FFD6, "addiu a0,zero,-42 ; following graphic strip A left local X"),
    "following_strip_a_right_x": (0x43168, 0x2403FFF0, "addiu v1,zero,-16 ; following graphic strip A right local X"),
    "following_strip_b_left_x":  (0x431DC, 0x2405FFE8, "addiu a1,zero,-24 ; following graphic strip B left local X"),
    "following_strip_b_right_x": (0x431D0, 0x24030002, "addiu v1,zero,2   ; following graphic strip B right local X"),

    # Dynamic/state-dependent following/circle draw path.
    # Ugly-test confirmed the static 0x430FC path alone does not affect the
    # broken in-game graphic. This later path uses the final tuned circle-half UVs and
    # is the one likely active in-game.
    "following_dyn_a_left_u_delay": (0x446CC, 0x240400A0, "addiu a0,zero,160 ; dynamic circle strip A left U, branch-delay path"),
    "following_dyn_a_left_u":       (0x44708, 0x240400A0, "addiu a0,zero,160 ; dynamic circle strip A left U"),
    "following_dyn_top_v":          (0x4470C, 0x2408002E, "addiu t0,zero,46  ; dynamic circle top V"),
    "following_dyn_a_right_u":      (0x44710, 0x240300BA, "addiu v1,zero,186 ; dynamic circle strip A right U"),
    "following_dyn_b_left_u":       (0x44748, 0x240A00C6, "addiu t2,zero,198 ; dynamic circle strip B left U"),
    "following_dyn_b_right_u":      (0x447B4, 0x240300E0, "addiu v1,zero,224 ; dynamic circle strip B right U"),
}


def parse_num(v):
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        return int(v, 0)
    raise TypeError(f"Unsupported numeric value: {v!r}")


def patch_low16_immediate(word: int, new_imm: int) -> int:
    return (word & 0xFFFF0000) | (new_imm & 0xFFFF)


def patch_one_plsel_imm(exe, *, name, offset, expected_word, new_imm, dry_run=False, force=False):
    off = parse_num(offset)
    expected = parse_num(expected_word)
    imm = parse_num(new_imm)
    actual = struct.unpack_from('<I', exe, off)[0]
    new_word = patch_low16_immediate(actual if force and actual != expected else expected, imm)

    if actual == new_word:
        return {
            "name": name,
            "file_offset": f"0x{off:X}",
            "expected_word": f"0x{expected:08X}",
            "old_word": f"0x{actual:08X}",
            "new_word": f"0x{new_word:08X}",
            "new_imm": imm,
            "status": "already_applied",
        }

    if actual != expected and not force:
        raise RuntimeError(
            f"PLSEL patch {name} failed verification at 0x{off:X}. "
            f"Expected 0x{expected:08X}, found 0x{actual:08X}. "
            "Use --plsel-force only if you intentionally want to patch this file."
        )

    if not dry_run:
        struct.pack_into('<I', exe, off, new_word & 0xFFFFFFFF)

    return {
        "name": name,
        "file_offset": f"0x{off:X}",
        "expected_word": f"0x{expected:08X}",
        "old_word": f"0x{actual:08X}",
        "new_word": f"0x{new_word:08X}",
        "new_imm": imm,
        "status": "would_patch" if dry_run else "patched",
    }


def patch_plsel_graphic_draw(exe, args, dry_run=False):
    report = []

    # Built-in known-good combined-strip expansion.
    builtins = {
        "combined_strip_right_u": args.plsel_main_right_u,
    }
    if not args.plsel_main_uv_only:
        builtins["combined_strip_right_x"] = args.plsel_main_right_x

    if args.plsel_left_u is not None:
        builtins["combined_strip_left_u"] = args.plsel_left_u
    if args.plsel_left_x is not None and not args.plsel_main_uv_only:
        builtins["combined_strip_left_x"] = args.plsel_left_x
    if args.plsel_top_v is not None:
        builtins["combined_strip_top_v"] = args.plsel_top_v
    if args.plsel_bottom_v is not None:
        builtins["combined_strip_bottom_v"] = args.plsel_bottom_v

    # Built-in following/planet/circle graphic source fix.
    #
    # The edited TIM14 does not use a simple +20 shift for both strips.
    # It is actually two separated circle halves:
    #   left half:  U 180..205, V 48..110
    #   right half: U 219..244, V 48..110
    #
    # Default behavior:
    #   - sample those exact source U coordinates from TIM14
    #   - keep the graphic in its original screen position
    #
    # If you intentionally want to move it on screen, pass
    # --plsel-following-x-shift N and do not use --plsel-following-uv-only.
    if not args.disable_plsel_following_graphic_patch:
        builtins["following_strip_a_left_u"] = args.plsel_following_a_left_u
        builtins["following_strip_a_right_u"] = args.plsel_following_a_right_u
        builtins["following_strip_b_left_u"] = args.plsel_following_b_left_u
        builtins["following_strip_b_right_u"] = args.plsel_following_b_right_u

        # Also patch the dynamic/state-dependent draw path for the same circle
        # halves. This is the path the game appears to use for the visible
        # broken graphic, since ugly-testing the static path did not alter it.
        if not args.disable_plsel_following_dynamic_patch:
            builtins["following_dyn_a_left_u_delay"] = args.plsel_following_a_left_u
            builtins["following_dyn_a_left_u"] = args.plsel_following_a_left_u
            builtins["following_dyn_top_v"] = args.plsel_following_top_v
            builtins["following_dyn_a_right_u"] = args.plsel_following_dynamic_a_right_u
            builtins["following_dyn_b_left_u"] = args.plsel_following_b_left_u
            builtins["following_dyn_b_right_u"] = args.plsel_following_b_right_u

        if not args.plsel_following_uv_only:
            x_shift = args.plsel_following_x_shift
            builtins["following_strip_a_left_x"] = -42 + x_shift
            builtins["following_strip_a_right_x"] = -16 + x_shift
            builtins["following_strip_b_left_x"] = -24 + x_shift
            builtins["following_strip_b_right_x"] = 2 + x_shift

    for key, imm in builtins.items():
        off, expected, desc = PLSEL_DRAW_PATCH_SITES[key]
        rec = patch_one_plsel_imm(
            exe,
            name=key,
            offset=off,
            expected_word=expected,
            new_imm=imm,
            dry_run=dry_run,
            force=args.plsel_force,
        )
        rec["description"] = desc
        report.append(rec)

    # Optional external patch list for additional verified PLSEL draw constants.
    # JSON formats accepted:
    #   {"exe_patches":[{"name":"...","offset":"0x...","expected_word":"0x...","new_imm":123}]}
    #   [{"name":"...","offset":"0x...","expected_word":"0x...","new_imm":123}]
    if args.plsel_patch_json:
        cfg = json.loads(Path(args.plsel_patch_json).read_text(encoding='utf-8'))
        patches = cfg.get("exe_patches", cfg) if isinstance(cfg, dict) else cfg
        for i, p in enumerate(patches):
            name = p.get("name", f"plsel_extra_{i}")
            rec = patch_one_plsel_imm(
                exe,
                name=name,
                offset=p["offset"],
                expected_word=p["expected_word"],
                new_imm=p["new_imm"],
                dry_run=dry_run,
                force=args.plsel_force,
            )
            rec["description"] = p.get("description", "external PLSEL patch-json entry")
            report.append(rec)

    return {"enabled": True, "report": report}

# PLMENU01.TIM element-info label draw patch.
#
# The Pixy information panel uses a dynamic selector for the current element
# symbol/label. In the original game, each option is a small 16x16 source slot:
#   None/Dash U=0x58, Earth U=0x18, Water U=0x28, Wind U=0x38, Fire U=0x48,
#   all at V=0xC0 and width/height 0x10.
#
# The English TIM edit widens the labels and moves the None dash/planet source
# art to the right. A simple immediate patch is not sufficient because width
# must differ by element. This hook replaces the selector block at 80078C40 and
# writes both source U and source width into the draw template.
PLANET_INFO_ELEMENT_HOOK_SITE = 0x80078C40
PLANET_INFO_ELEMENT_HOOK_RETURN = 0x80078CC8
PLANET_INFO_ELEMENT_EXPECT = bytes.fromhex("12000392 ff000224")  # lbu v1,0x12(s0); addiu v0,zero,0xff
PLANET_INFO_PLANET_U_SITE = 0x57F70  # RAM 80077770: addiu v0,zero,0x68 ; planet source U
PLANET_INFO_PLANET_U_EXPECT = 0x24020068
PLANET_INFO_PLANET_V_SITE = 0x57F0C  # RAM 8007770C: addiu a1,zero,0xB8 ; planet source V
PLANET_INFO_PLANET_V_EXPECT = 0x240500B8
PLANET_INFO_PLANET_CLUT_SITE = 0x57F7C  # RAM 8007777C: addiu v0,zero,0x01E9 ; planet palette/CLUT selector

PLANET_COPY_SLOT_SPLIT_HOOK_SITE = 0x8007787C
PLANET_COPY_SLOT_SPLIT_HOOK_RETURN = 0x80077880
PLANET_COPY_SLOT_SPLIT_EXPECT = 0x3C084000  # lui t0,0x4000
PLANET_INFO_PLANET_CLUT_EXPECT = 0x240201E9




# STAGE moved planet UV patch plus safety restores.
#
# Confirmed by in-game testing:
#   - The STAGE/Garden planet source is shared at the old PLANET_INFO_PLANET_*
#     sites, and moving U/V there fixes the STAGE planet.
#   - The N/A-line candidates patched in v82/v93 were wrong. They affect the
#     planet-name/list graphics instead, producing the "Forless shifted right"
#     problem.
#
# v93 therefore keeps the confirmed planet fix, but restores every known wrong
# N/A-line candidate to its original value and does NOT patch the N/A line by
# default. The real N/A-line UV needs a separate ugly-test/trace pass.
STAGE_TIM04_PLANET_U_SITE = PLANET_INFO_PLANET_U_SITE
STAGE_TIM04_PLANET_U_EXPECT = PLANET_INFO_PLANET_U_EXPECT
STAGE_TIM04_PLANET_V_SITE = PLANET_INFO_PLANET_V_SITE
STAGE_TIM04_PLANET_V_EXPECT = PLANET_INFO_PLANET_V_EXPECT

# False candidate from v83/v85: keep/restores this at 0x58. It did not fix the
# Object-overlay N/A symbol and can affect nearby name/list template data if moved.
STAGE_V83_WRONG_PLANET_NAME_SLOT_U_SITE = 0x6060C
STAGE_V83_WRONG_PLANET_NAME_SLOT_U_GOOD = 0x24030058
STAGE_V83_WRONG_PLANET_NAME_SLOT_U_WRONG = 0x24030075

# Runtime packet filter for the real Object-overlay N/A symbol.
#
# The No$PSX packet trace shows the bad source/CLUT word lives at:
#   packet base 800F8710, word [base+0x0C] = 7852C058
# and needs to become:
#   [base+0x0C] = 7852C075
#
# Important: RAM 800A3AB0 is the source/CLUT-word store for this packet:
#   sw/t0 -> [a0+0x0C]
# RAM 800A3AE4 is the following size-word store ([a0+0x10]) for the same packet,
# so v93 deliberately keeps the hook at 800A3AB0 but removes v86's too-strict
# t0==bad_word condition. The exact packet base is specific enough, and t0 is
# overwritten shortly afterward by the original routine.
STAGE_TIM04_NA_PACKET_FILTER_SITE = 0x800A3AB0
STAGE_TIM04_NA_PACKET_FILTER_RETURN = 0x800A3AB4
STAGE_TIM04_NA_PACKET_FILTER_EXPECT = 0xAC88000C

# v130: POLY_FT4 transition-frame N/A source-word filter.
# The bad snapshot shows the visible flash is not an SPRT packet; it is a
# POLY_FT4 packet whose first UV/CLUT word is 0x7852C058. The corrected stable
# frame later uses a normal SPRT with 0x7852C075. This hook targets the FT4
# builder's first UV store, with a proper two-instruction patch so the MIPS
# delay slot cannot clobber the tested register.
STAGE_TIM04_NA_FT4_SOURCE_FILTER_SITE = 0x800A3DAC
STAGE_TIM04_NA_FT4_SOURCE_FILTER_RETURN = 0x800A3DB4
STAGE_TIM04_NA_FT4_SOURCE_FILTER_EXPECT = bytes.fromhex("25186700 0c0083ac")  # or v1,v1,a3 ; sw v1,0x0C(a0)
STAGE_TIM04_NA_FT4_PACKET_CMD = 0x2C808080

# v132: earlier POLY_FT4 edge-coordinate filter.
# v130 proved the bad frame is the FT4 transition path, but it only changed
# the already-built UV0 word. This hook runs earlier, just after the FT4 path
# has computed the horizontal source edges (t1/t4), before any UV words are
# assembled. It preserves flip state by remapping old left/right independently.
STAGE_TIM04_NA_FT4_EDGE_FILTER_SITE = 0x800A3CDC
STAGE_TIM04_NA_FT4_EDGE_FILTER_RETURN = 0x800A3CE4
STAGE_TIM04_NA_FT4_EDGE_FILTER_EXPECT = bytes.fromhex("4000023c24102202")  # lui v0,0x0040 ; and v0,s1,v0


# v58: local PLANET/Modus record repair.
# Breakpoints showed the active Modus element graphic can be read from the local
# record at BASE+0xAD4 (RAM 8012A3BC when BASE=801298E8), and a PLANET-local
# routine around 80079998 runs when the Modus element graphic is requested.  This
# hook repairs that local record before the shared FT4/SPRT builders consume it,
# avoiding the unsafe shared primitive writer path that corrupted Neredy/menu.
MODUS_LOCAL_RECORD_FIX_SITE = 0x80079998
MODUS_LOCAL_RECORD_FIX_RETURN = 0x800799A0
MODUS_LOCAL_RECORD_FIX_EXPECT = bytes.fromhex("f2 0a 83 a4 ce 0a 83 a4")  # sh v1,0x0AF2(a0); sh v1,0x0ACE(a0)

# v59 earlier PLANET/Modus record-read hook.  Breakpoints showed the first
# useful read of the active Modus record occurs in a local loop at 8007972C:
#   lhu v0,0x092A(v1)
#   addiu s0,s0,1
# With v1=80129A98, record base is v1+0x924 = 8012A3BC.
MODUS_LOCAL_RECORD_EARLY_FIX_SITE = 0x8007972C
MODUS_LOCAL_RECORD_EARLY_FIX_RETURN = 0x80079734
MODUS_LOCAL_RECORD_EARLY_FIX_EXPECT = bytes.fromhex("2a 09 62 94 01 00 10 26")  # lhu v0,0x092A(v1); addiu s0,s0,1

# v60: exact active-record FT4 edge-read hooks.
# These are later than v59's record repair but earlier than the FT4 edge math for
# the current primitive. They avoid the unsafe shared SPRT writer and require the
# observed active Modus record pointer s0 == 8012A3BC.
MODUS_EXACT_FT4_EDGE_READ_A_SITE = 0x800A3CAC
MODUS_EXACT_FT4_EDGE_READ_A_RETURN = 0x800A3CB4
MODUS_EXACT_FT4_EDGE_READ_A_EXPECT = bytes.fromhex("0e 00 03 92 08 00 02 92")  # lbu v1,0x0E(s0); lbu v0,0x08(s0)

MODUS_EXACT_FT4_EDGE_READ_B_SITE = 0x800A3CC8
MODUS_EXACT_FT4_EDGE_READ_B_RETURN = 0x800A3CD0
MODUS_EXACT_FT4_EDGE_READ_B_EXPECT = bytes.fromhex("0e 00 09 92 08 00 02 92")  # lbu t1,0x0E(s0); lbu v0,0x08(s0)

MODUS_EXACT_FT4_EDGE_RECORD_RAM = 0x8012A3BC

# v82 mistakenly patched these; restore them to 0x59 by default so the planet
# name/list graphic is not shifted when a v82/v93-patched EXE is used as input.
STAGE_V82_WRONG_PLANET_NAME_U_RESTORE_SITES = {
    "restore_v82_wrong_planet_name_u_static":  (0x66A8C, 0x24020059, 0x24020076, "Restore planet-name/list source U; v82 had mistaken N/A-line patch here"),
    "restore_v82_wrong_planet_name_u_dynamic": (0x66F64, 0x24020059, 0x24020076, "Restore planet-name/list dynamic source U; v82 had mistaken N/A-line patch here"),
}

def build_planet_copy_slot_split_hook(hook_ram, args):
    a = A(hook_ram)

    # Original overwritten instruction at 8007787C.
    a.lui("t0", 0x4000)

    # At 8007787C, after the template copy loop:
    #   v1 = BASE + 0x68
    #   a0 = BASE + 0x164
    #
    # Directly changing the original a1=0xB8 immediate at file 0x57F0C was too
    # broad: that register feeds another nearby template byte before the planet
    # V byte, which appears to be why the planet-name/list graphics shifted.
    #
    # v93 keeps the original template construction intact, then changes only the
    # actual planet source bytes after the copy loop:
    #   original/shared STAGE/Garden planet slot: BASE+0x56 / BASE+0x57
    #   copied PLANET-info planet slot:          BASE+0x152 / BASE+0x153
    #
    # Relative to v1 = BASE+0x68:
    #   BASE+0x56  = v1-0x12
    #   BASE+0x57  = v1-0x11
    #   BASE+0x152 = v1+0xEA
    #   BASE+0x153 = v1+0xEB

    # STAGE/Garden planet source: moved art in STAGE TIM04.
    a.addiu("v0", "zero", int(args.stage_tim03_planet_u) & 0xFF)
    a.word(it(0x28, REG["v1"], REG["v0"], -0x0012))  # sb v0,-0x12(v1) => BASE+0x56
    a.addiu("v0", "zero", int(args.stage_tim03_planet_v) & 0xFF)
    a.word(it(0x28, REG["v1"], REG["v0"], -0x0011))  # sb v0,-0x11(v1) => BASE+0x57

    # PLANET info copied slot: keep PLMENU source separate.
    a.addiu("v0", "zero", int(args.planet_info_planet_u) & 0xFF)
    a.word(it(0x28, REG["v1"], REG["v0"], 0x00EA))  # sb v0,0xEA(v1) => BASE+0x152

    if args.planet_info_planet_v is not None:
        a.addiu("v0", "zero", int(args.planet_info_planet_v) & 0xFF)
        a.word(it(0x28, REG["v1"], REG["v0"], 0x00EB))  # sb v0,0xEB(v1) => BASE+0x153

    a.j(PLANET_COPY_SLOT_SPLIT_HOOK_RETURN); a.nop()
    return a.out()


def patch_planet_copy_slot_split_hook(exe, load, args, dry_run=False):
    if args.disable_planet_copy_slot_split_hook:
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(PLANET_COPY_SLOT_SPLIT_HOOK_SITE, load)
    if site_off < 0 or site_off + 4 > len(exe):
        raise RuntimeError(
            f"PLANET copy-slot split hook site maps outside EXE: RAM 0x{PLANET_COPY_SLOT_SPLIT_HOOK_SITE:08X}, file 0x{site_off:X}"
        )

    cur = read32(exe, site_off)
    replacing_existing_hook = False
    if cur != PLANET_COPY_SLOT_SPLIT_EXPECT:
        already = ((cur >> 26) == 0x02)
        if already:
            # Replace older versions of this hook by default. The old appended
            # code remains in the EXE but becomes unreachable after this site is
            # repointed to the new hook.
            replacing_existing_hook = True
        elif not args.planet_info_force:
            raise RuntimeError(
                f"PLANET copy-slot split hook site mismatch at 0x{site_off:X}: "
                f"found 0x{cur:08X}, expected 0x{PLANET_COPY_SLOT_SPLIT_EXPECT:08X}. "
                "Use --planet-info-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_planet_copy_slot_split_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+4] = w(jt(2, hook_ram))
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{PLANET_COPY_SLOT_SPLIT_HOOK_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "stage_shared_u_addr": "BASE+0x56",
        "stage_shared_v_addr": "BASE+0x57",
        "stage_shared_u": f"0x{int(args.stage_tim03_planet_u) & 0xFF:02X}",
        "stage_shared_v": f"0x{int(args.stage_tim03_planet_v) & 0xFF:02X}",
        "copied_u_addr": "BASE+0x152",
        "copied_v_addr": "BASE+0x153",
        "info_u": f"0x{args.planet_info_planet_u & 0xFF:02X}",
        "info_v": None if args.planet_info_planet_v is None else f"0x{args.planet_info_planet_v & 0xFF:02X}",
    }


def patch_one_main_imm(exe, *, name, offset, expected_word, new_imm, dry_run=False, force=False):
    off = parse_num(offset)
    expected = parse_num(expected_word)
    imm = parse_num(new_imm)
    actual = struct.unpack_from('<I', exe, off)[0]
    new_word = patch_low16_immediate(actual if force and actual != expected else expected, imm)
    if actual == new_word:
        return {
            "name": name,
            "file_offset": f"0x{off:X}",
            "expected_word": f"0x{expected:08X}",
            "old_word": f"0x{actual:08X}",
            "new_word": f"0x{new_word:08X}",
            "new_imm": imm,
            "status": "already_applied",
        }
    if actual != expected and not force:
        raise RuntimeError(
            f"MAIN immediate patch {name} failed verification at 0x{off:X}. "
            f"Expected 0x{expected:08X}, found 0x{actual:08X}. "
            "Use --planet-info-force only if you intentionally want to patch this file."
        )
    if not dry_run:
        struct.pack_into('<I', exe, off, new_word & 0xFFFFFFFF)
    return {
        "name": name,
        "file_offset": f"0x{off:X}",
        "expected_word": f"0x{expected:08X}",
        "old_word": f"0x{actual:08X}",
        "new_word": f"0x{new_word:08X}",
        "new_imm": imm,
        "status": "would_patch" if dry_run else "patched",
    }



def patch_one_main_imm_allowed(exe, *, name, offset, allowed_words, new_imm, dry_run=False, force=False, description=""):
    off = parse_num(offset)
    allowed = [parse_num(x) for x in allowed_words]
    imm = parse_num(new_imm)
    actual = struct.unpack_from('<I', exe, off)[0]
    base_word = actual if (actual in allowed or force) else allowed[0]
    new_word = patch_low16_immediate(base_word, imm)

    if actual == new_word:
        return {
            "name": name,
            "file_offset": f"0x{off:X}",
            "allowed_words": ";".join(f"0x{x:08X}" for x in allowed),
            "old_word": f"0x{actual:08X}",
            "new_word": f"0x{new_word:08X}",
            "new_imm": imm,
            "description": description,
            "status": "already_applied",
        }

    if actual not in allowed and not force:
        raise RuntimeError(
            f"MAIN immediate patch {name} failed verification at 0x{off:X}. "
            f"Expected one of {', '.join(f'0x{x:08X}' for x in allowed)}, found 0x{actual:08X}. "
            "Use --stage-tim03-uv-force only if you intentionally want to patch this file."
        )

    if not dry_run:
        struct.pack_into('<I', exe, off, new_word & 0xFFFFFFFF)

    return {
        "name": name,
        "file_offset": f"0x{off:X}",
        "allowed_words": ";".join(f"0x{x:08X}" for x in allowed),
        "old_word": f"0x{actual:08X}",
        "new_word": f"0x{new_word:08X}",
        "new_imm": imm,
        "description": description,
        "status": "would_patch" if dry_run else "patched",
    }


def patch_stage_tim03_moved_uvs(exe, args, dry_run=False):
    # Function name kept for compatibility with the existing main() call/report keys.
    # v93 no longer patches the shared planet U/V immediates directly. Instead,
    # patch_planet_copy_slot_split_hook() changes only the final planet slot bytes
    # after the template copy loop. This avoids changing nearby template bytes that
    # reused the same a1=0xB8 register and appeared to shift planet-name/list graphics.
    if getattr(args, "disable_stage_tim03_moved_uv_patch", False):
        return {"enabled": False, "status": "disabled"}

    force = bool(getattr(args, "stage_tim03_uv_force", False) or getattr(args, "planet_info_force", False))
    report = []

    # Restore the broad/shared immediate sites to stock values. The copy-slot hook
    # will write the moved STAGE planet U/V into BASE+0x56/+0x57 narrowly.
    report.append(patch_one_main_imm_allowed(
        exe,
        name="restore_shared_planet_u_immediate_for_isolated_hook",
        offset=STAGE_TIM04_PLANET_U_SITE,
        allowed_words=[STAGE_TIM04_PLANET_U_EXPECT, patch_low16_immediate(STAGE_TIM04_PLANET_U_EXPECT, int(args.stage_tim03_planet_u) & 0xFF)],
        new_imm=0x68,
        dry_run=dry_run,
        force=force,
        description="Restore broad shared U immediate; v93 writes moved STAGE planet U only to BASE+0x56 in copy-slot hook",
    ))
    report.append(patch_one_main_imm_allowed(
        exe,
        name="restore_shared_planet_v_immediate_for_isolated_hook",
        offset=STAGE_TIM04_PLANET_V_SITE,
        allowed_words=[STAGE_TIM04_PLANET_V_EXPECT, patch_low16_immediate(STAGE_TIM04_PLANET_V_EXPECT, int(args.stage_tim03_planet_v) & 0xFF)],
        new_imm=0xB8,
        dry_run=dry_run,
        force=force,
        description="Restore broad shared V immediate; avoids changing another nearby template byte that reused a1",
    ))

    # v120: seed the old/static N/A candidate to the moved N/A U before the
    # runtime template hook has a chance to run. The v93+ template hook still
    # handles the active steady-state packet; this early seed is meant to remove
    # the one-frame flash where the opening overlay/screen briefly shows the old
    # U=0x58 source area before being corrected to U=0x75.
    if getattr(args, "disable_stage_tim04_na_static_seed", False):
        static_na_u = 0x58
        static_na_name = "restore_inactive_na_static_candidate_u"
        static_na_desc = "Restore old/static N/A candidate to U=0x58; runtime template hook handles steady-state N/A"
    else:
        static_na_u = int(args.stage_tim03_na_line_u) & 0xFF
        static_na_name = "seed_early_na_static_candidate_u"
        static_na_desc = "Seed old/static N/A candidate to moved U before first visible frame; runtime template hook still handles steady-state N/A"
    report.append(patch_one_main_imm_allowed(
        exe,
        name=static_na_name,
        offset=STAGE_V83_WRONG_PLANET_NAME_SLOT_U_SITE,
        allowed_words=[STAGE_V83_WRONG_PLANET_NAME_SLOT_U_GOOD, STAGE_V83_WRONG_PLANET_NAME_SLOT_U_WRONG],
        new_imm=static_na_u,
        dry_run=dry_run,
        force=force,
        description=static_na_desc,
    ))

    # Undo the two mistaken v82 patches if present. On a clean EXE these are
    # already 0x59 and will report already_applied.
    for name, (off, good_word, wrong_word, desc) in STAGE_V82_WRONG_PLANET_NAME_U_RESTORE_SITES.items():
        report.append(patch_one_main_imm_allowed(
            exe,
            name=name,
            offset=off,
            allowed_words=[good_word, wrong_word],
            new_imm=0x59,
            dry_run=dry_run,
            force=force,
            description=desc,
        ))

    patched = sum(1 for r in report if r.get("status") in {"patched", "would_patch"})
    already = sum(1 for r in report if r.get("status") == "already_applied")
    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "shared_immediates": "restored_to_original; moved planet U/V written by copy-slot hook",
        "planet_u": f"0x{int(args.stage_tim03_planet_u) & 0xFF:02X}",
        "planet_v": f"0x{int(args.stage_tim03_planet_v) & 0xFF:02X}",
        "na_static_candidate_u": f"0x{int(args.stage_tim03_na_line_u) & 0xFF:02X}",
        "patched_sites": patched,
        "already_applied_sites": already,
        "report": report,
    }


def build_stage_tim04_na_packet_filter_hook(hook_ram, args):
    a = A(hook_ram)

    packet_base = int(args.stage_tim04_na_packet_base) & 0xFFFFFFFF
    good_word = int(args.stage_tim04_na_good_word) & 0xFFFFFFFF

    # v86 checked both packet base and the transient t0 word. No$PSX showed the
    # final memory address clearly, but the register value at a memory-change
    # break can be awkward/stale depending on timing. v93 therefore filters only
    # by the exact packet base 800F8710, then forces the source/CLUT word to the
    # desired value before executing the original store.
    #
    # Only t0/r8 is changed, and only for this one packet. The original code
    # reloads t0 from scratchpad a few instructions later, so this does not leak
    # into later primitive construction.
    a.lui("v0", hi(packet_base)); a.addiu("v0", "v0", s16(lo(packet_base)))
    a.bne("a0", "v0", "store"); a.nop()

    a.lui("t0", (good_word >> 16) & 0xFFFF); a.ori("t0", "t0", lo(good_word))

    a.lab("store")
    a.sw("t0", 0x000C, "a0")  # original overwritten instruction: sw t0,0x0C(a0)
    a.j(STAGE_TIM04_NA_PACKET_FILTER_RETURN); a.nop()
    return a.out()


def patch_stage_tim04_na_packet_filter_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_stage_tim04_na_packet_filter_hook", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(STAGE_TIM04_NA_PACKET_FILTER_SITE, load)
    if site_off < 0 or site_off + 4 > len(exe):
        raise RuntimeError(
            f"STAGE TIM04 N/A packet filter hook site maps outside EXE: RAM 0x{STAGE_TIM04_NA_PACKET_FILTER_SITE:08X}, file 0x{site_off:X}"
        )

    cur = read32(exe, site_off)
    replacing_existing_hook = False
    if cur != STAGE_TIM04_NA_PACKET_FILTER_EXPECT:
        already = ((cur >> 26) == 0x02)
        if already:
            # Allow v93 to replace the older v86 hook automatically. This keeps
            # testing convenient if the user accidentally uses a previously
            # patched MAIN.EXE as input. A fresh/clean MAIN.EXE is still preferred.
            replacing_existing_hook = True
        elif not args.stage_tim04_na_packet_filter_force:
            raise RuntimeError(
                f"STAGE TIM04 N/A packet filter hook site mismatch at 0x{site_off:X}: "
                f"found 0x{cur:08X}, expected 0x{STAGE_TIM04_NA_PACKET_FILTER_EXPECT:08X}. "
                "Use --stage-tim04-na-packet-filter-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_stage_tim04_na_packet_filter_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+4] = w(jt(2, hook_ram))
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{STAGE_TIM04_NA_PACKET_FILTER_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "packet_base": f"0x{int(args.stage_tim04_na_packet_base) & 0xFFFFFFFF:08X}",
        "bad_word": f"0x{int(args.stage_tim04_na_bad_word) & 0xFFFFFFFF:08X}",
        "good_word": f"0x{int(args.stage_tim04_na_good_word) & 0xFFFFFFFF:08X}",
        "description": "Runtime filter: for packet base 800F8710, force source/CLUT word [base+0x0C] to 7852C075.",
    }




# v20/v21 note:
# Do NOT re-add the v19 packet-stage source-word filter at 800A3AB0 as a default
# fix. That store is shared by many stable SPRT packets; matching source words
# like 0x7852C018 also catches unrelated graphics such as Neredy/menu/load
# sprites. v20 intentionally rolls back to the last non-breaking branch.
#
def stage_tim04_object_overlay_target_cases(args, include_aliases=False):
    """Build exact Object-overlay element/N-A relocation cases.

    The Object/Pixy overlay slot can contain the normal element symbols or the
    N/A line.  Do not assume Object values are always N/A: late-game objects may
    legitimately carry Earth/Water/Wind/Fire.  v43 treats this as a relocation
    table: only original stock source U values are semantic identities by
    default, and each is mapped to the edited TIM04 source U, width, and
    absolute destination X.
    """
    base_x = int(args.stage_tim04_object_overlay_base_x)
    cases = [
        {
            "name": "earth",
            "match_us": (0x18, int(args.stage_tim04_earth_u) & 0xFF) if include_aliases else (0x18,),
            "new_u": int(args.stage_tim04_earth_u) & 0xFF,
            "new_w": int(args.stage_tim04_earth_w) & 0xFFFF,
            "x": base_x + int(args.stage_tim04_earth_x_shift),
        },
        {
            "name": "water",
            "match_us": (0x28, int(args.stage_tim04_water_u) & 0xFF) if include_aliases else (0x28,),
            "new_u": int(args.stage_tim04_water_u) & 0xFF,
            "new_w": int(args.stage_tim04_water_w) & 0xFFFF,
            "x": base_x + int(args.stage_tim04_water_x_shift),
        },
        {
            "name": "wind",
            "match_us": (0x38, int(args.stage_tim04_wind_u) & 0xFF) if include_aliases else (0x38,),
            "new_u": int(args.stage_tim04_wind_u) & 0xFF,
            "new_w": int(args.stage_tim04_wind_w) & 0xFFFF,
            "x": base_x + int(args.stage_tim04_wind_x_shift),
        },
        {
            "name": "fire",
            "match_us": (0x48, int(args.stage_tim04_fire_u) & 0xFF) if include_aliases else (0x48,),
            "new_u": int(args.stage_tim04_fire_u) & 0xFF,
            "new_w": int(args.stage_tim04_fire_w) & 0xFFFF,
            "x": base_x + int(args.stage_tim04_fire_x_shift),
        },
        {
            "name": "dash",
            "match_us": (int(args.stage_tim04_na_template_old_u) & 0xFF, int(args.stage_tim03_na_line_u) & 0xFF) if include_aliases else (int(args.stage_tim04_na_template_old_u) & 0xFF,),
            "new_u": int(args.stage_tim03_na_line_u) & 0xFF,
            "new_w": int(args.stage_tim04_dash_w) & 0xFFFF,
            "x": int(args.stage_tim04_object_overlay_na_x),
        },
    ]
    return cases


def stage_tim04_object_overlay_candidate_x_values(args):
    vals = {int(args.stage_tim04_object_overlay_na_x)}
    for case in stage_tim04_object_overlay_target_cases(args, include_aliases=False):
        vals.add(int(case["x"]))
    return sorted(vals)


def stage_tim04_object_overlay_ft4_guard(a, args, done_label, prefix):
    """Shared Object-overlay FT4 guard.

    Stable Object repair can use [s0+0x1E] == 0x1000.  Some transition frames
    appear before that halfword is final, so also accept known Object-overlay
    value X positions: N/A plus the per-element absolute X values.  This keeps
    the guard tied to the exact Object/Pixy value slot without assuming the
    object's symbol is always N/A.
    """
    slot_ram = int(args.stage_tim04_object_overlay_na_slot_ram) & 0xFFFFFFFF
    object_guard = int(args.stage_tim04_object_overlay_object_guard_halfword) & 0xFFFF
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    stage_marker = int(args.stage_tim04_object_overlay_marker) & 0xFF

    a.lui("t5", (slot_ram >> 16) & 0xFFFF)
    a.ori("t5", "t5", slot_ram & 0xFFFF)
    a.bne("s0", "t5", done_label); a.nop()

    # Accept either the stable Object discriminator or a known Object-value X.
    accept = f"{prefix}_object_ft4_guard_accept"
    a.lhu("t5", 0x001E, "s0")
    a.addiu("t6", "zero", object_guard)
    a.beq("t5", "t6", accept); a.nop()
    a.lh("t5", 0x0004, "s0")
    for x in stage_tim04_object_overlay_candidate_x_values(args):
        a.addiu("t6", "zero", int(x))
        a.beq("t5", "t6", accept); a.nop()
    a.j(done_label); a.nop()
    a.lab(accept)

    a.lbu("t5", 0x0017, "s0")
    a.addiu("t6", "zero", stage_marker)
    a.bne("t5", "t6", done_label); a.nop()

    a.lbu("t5", 0x000F, "s0")
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", done_label); a.nop()

    a.lhu("t5", 0x000A, "s0")
    a.addiu("t6", "zero", h)
    a.bne("t5", "t6", done_label); a.nop()

    a.lbu("t5", 0x0014, "s0")
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", done_label); a.nop()
    a.lbu("t5", 0x0015, "s0")
    a.bne("t5", "t6", done_label); a.nop()
    a.lbu("t5", 0x0016, "s0")
    a.bne("t5", "t6", done_label); a.nop()


def stage_tim04_select_object_overlay_target_from_u(a, args, done_label, prefix, set_source_word=False):
    """Map the exact Object-overlay value slot by current graphic identity.

    This replaces the older force-N/A behavior.  If the current template says
    Earth/Water/Wind/Fire/N-A, write the edited source U, edited width, and the
    matching absolute X for that symbol.  Unknown symbols are left untouched.
    """
    v = int(args.stage_tim04_na_template_v) & 0xFF
    pfx = int(args.stage_tim04_object_element_packet_prefix) & 0xFFFF0000
    cases = stage_tim04_object_overlay_target_cases(args, include_aliases=False)

    a.lbu("t5", 0x000E, "s0")
    for case in cases:
        lab = f"{prefix}_object_{case['name']}"
        for old_u in sorted(set(int(u) & 0xFF for u in case["match_us"])):
            a.addiu("t6", "zero", old_u)
            a.beq("t5", "t6", lab); a.nop()
    a.j(done_label); a.nop()

    for case in cases:
        lab = f"{prefix}_object_{case['name']}"
        new_u = int(case["new_u"]) & 0xFF
        new_w = int(case["new_w"]) & 0xFFFF
        x = int(case["x"])
        right = (new_u + new_w - 1) & 0xFF
        a.lab(lab)
        a.addiu("t5", "zero", x)
        a.word(it(0x29, REG["s0"], REG["t5"], 0x0004))
        a.addiu("t5", "zero", new_w)
        a.word(it(0x29, REG["s0"], REG["t5"], 0x0008))
        a.addiu("t5", "zero", new_u)
        a.word(it(0x28, REG["s0"], REG["t5"], 0x000E))
        a.addiu("t1", "zero", new_u)
        a.addiu("t4", "zero", right)
        if set_source_word:
            word = pfx | (v << 8) | new_u
            a.lui("v1", (word >> 16) & 0xFFFF)
            a.ori("v1", "v1", word & 0xFFFF)
        a.j(done_label); a.nop()


def stage_tim04_remap_object_overlay_ft4_edges(a, args, done_label):
    """v43: remap Object-overlay FT4/opening edge path by stock old-U identity only."""
    if int(args.stage_tim04_object_overlay_na_slot_ram) == 0:
        return
    stage_tim04_object_overlay_ft4_guard(a, args, done_label, "obj_ft4_edge")
    stage_tim04_select_object_overlay_target_from_u(a, args, done_label, "obj_ft4_edge", set_source_word=False)


def stage_tim04_remap_object_overlay_ft4_source(a, args, done_label):
    """v43: remap Object-overlay FT4 first-UV/source-word path by stock old-U identity only."""
    if int(args.stage_tim04_object_overlay_na_slot_ram) == 0:
        return
    stage_tim04_object_overlay_ft4_guard(a, args, done_label, "obj_ft4_src")
    stage_tim04_select_object_overlay_target_from_u(a, args, done_label, "obj_ft4_src", set_source_word=True)


def stage_tim04_modus_ft4_guard(a, args, done_label, prefix):
    """Guard for PLANET/Modus FT4/opening element templates.

    v35: the FT4/opening quad can be a few pixels away from the steady SPRT
    template while the panel animates. The old exact-X guard missed those
    frames, which let one stale/old element primitive draw briefly before the
    stable template repair corrected it. Keep the rest of the guard narrow
    (Y, marker, V, height, and color), but accept a small configurable X drift.
    """
    x = int(args.modus_stage_element_x)
    x_tol = max(0, int(getattr(args, "modus_stage_ft4_x_tolerance", 0)))
    y = int(args.modus_stage_element_y)
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    marker = int(args.modus_stage_element_marker) & 0xFF

    a.lh("t5", 0x0004, "s0")
    if x_tol:
        x_ok = f"{prefix}_modus_x_ok"
        for cx in range(x - x_tol, x + x_tol + 1):
            a.addiu("t6", "zero", cx)
            a.beq("t5", "t6", x_ok); a.nop()
        a.j(done_label); a.nop()
        a.lab(x_ok)
    else:
        a.addiu("t6", "zero", x)
        a.bne("t5", "t6", done_label); a.nop()

    a.lh("t5", 0x0006, "s0")
    a.addiu("t6", "zero", y)
    a.bne("t5", "t6", done_label); a.nop()

    a.lbu("t5", 0x0017, "s0")
    a.addiu("t6", "zero", marker)
    a.bne("t5", "t6", done_label); a.nop()

    a.lbu("t5", 0x000F, "s0")
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", done_label); a.nop()

    a.lhu("t5", 0x000A, "s0")
    a.addiu("t6", "zero", h)
    a.bne("t5", "t6", done_label); a.nop()

    a.lbu("t5", 0x0014, "s0")
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", done_label); a.nop()
    a.lbu("t5", 0x0015, "s0")
    a.bne("t5", "t6", done_label); a.nop()
    a.lbu("t5", 0x0016, "s0")
    a.bne("t5", "t6", done_label); a.nop()


def stage_tim04_select_modus_target_from_u(a, args, done_label, prefix, set_source_word=False):
    """Map old PLANET/Modus element U during FT4/opening frames.

    Steady-state Modus uses the padded PLMENU01 rectangles from v29. The opening
    POLY_FT4 frame, however, is built after geometry has already been calculated,
    so using padded/wider source rectangles there can make the element appear a
    couple of pixels to the right before the stable SPRT path snaps into place.

    v31 therefore uses visible-ink PLMENU01 bounds only for the FT4/opening path:
      Earth U=0x19 W=0x17
      Water U=0x32 W=0x18
      Wind  U=0x4C W=0x13
      Fire  U=0x61 W=0x12

    v38 additionally normalizes the accepted FT4/opening destination X. v35's
    widened guard allowed the stale/flicker frame to be repaired, but if the
    opening template arrived a few pixels away from the steady SPRT template it
    could still visibly snap left after the panel settled. v40 keeps this narrow
    FT4-only hook but defaults the X adjustment to 0, because v39
    steady/template X normalization proved too broad in testing.

    The stable path still uses the padded v29 defaults.
    """
    v = int(args.stage_tim04_na_template_v) & 0xFF
    pfx = int(args.stage_tim04_object_element_packet_prefix) & 0xFFFF0000
    ft4_x = int(args.modus_stage_element_x) + int(getattr(args, "modus_stage_ft4_x_adjust", 0))
    cases = [
        # Match stock/old U, steady padded U, and FT4 visible-ink U.  The latter
        # two matter when the opening/transition primitive sees a template that
        # was already touched by the steady SPRT repair or by a previous frame.
        ((0x18, int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_u) & 0xFF), int(args.modus_stage_ft4_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_w) & 0xFFFF, "earth"),
        ((0x28, int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_ft4_water_u) & 0xFF), int(args.modus_stage_ft4_water_u) & 0xFF, int(args.modus_stage_ft4_water_w) & 0xFFFF, "water"),
        ((0x38, int(args.modus_stage_wind_u) & 0xFF, int(args.modus_stage_ft4_wind_u) & 0xFF), int(args.modus_stage_ft4_wind_u) & 0xFF, int(args.modus_stage_ft4_wind_w) & 0xFFFF, "wind"),
        ((0x48, int(args.modus_stage_fire_u) & 0xFF, int(args.modus_stage_ft4_fire_u) & 0xFF), int(args.modus_stage_ft4_fire_u) & 0xFF, int(args.modus_stage_ft4_fire_w) & 0xFFFF, "fire"),
    ]
    a.lbu("t5", 0x000E, "s0")
    for match_us, new_u, new_w, name in cases:
        lab = f"{prefix}_modus_{name}"
        for old_u in sorted(set(int(u) & 0xFF for u in match_us)):
            a.addiu("t6", "zero", old_u)
            a.beq("t5", "t6", lab); a.nop()
    a.j(done_label); a.nop()

    for match_us, new_u, new_w, name in cases:
        lab = f"{prefix}_modus_{name}"
        right = (new_u + new_w - 1) & 0xFF
        a.lab(lab)
        a.addiu("t5", "zero", ft4_x)
        a.word(it(0x29, REG["s0"], REG["t5"], 0x0004))
        a.addiu("t5", "zero", new_w)
        a.word(it(0x29, REG["s0"], REG["t5"], 0x0008))
        a.addiu("t5", "zero", new_u)
        a.word(it(0x28, REG["s0"], REG["t5"], 0x000E))
        a.addiu("t1", "zero", new_u)
        a.addiu("t4", "zero", right)
        if set_source_word:
            word = pfx | (v << 8) | new_u
            a.lui("v1", (word >> 16) & 0xFFFF)
            a.ori("v1", "v1", word & 0xFFFF)
        a.j(done_label); a.nop()


def stage_tim04_force_modus_element_ft4_edges(a, args, done_label):
    """v38: make PLANET/Modus FT4/opening edges use corrected UVs and X."""
    stage_tim04_modus_ft4_guard(a, args, done_label, "modus_ft4_edge")
    stage_tim04_select_modus_target_from_u(a, args, done_label, "modus_ft4_edge", set_source_word=False)


def stage_tim04_force_modus_element_ft4_source(a, args, done_label):
    """v38: make PLANET/Modus FT4 first UV/source word use corrected UVs and X."""
    stage_tim04_modus_ft4_guard(a, args, done_label, "modus_ft4_src")
    stage_tim04_select_modus_target_from_u(a, args, done_label, "modus_ft4_src", set_source_word=True)


def build_stage_tim04_na_ft4_edge_filter_hook(hook_ram, args):
    a = A(hook_ram)
    old_u = int(args.stage_tim04_na_template_old_u) & 0xFF
    new_u = int(args.stage_tim03_na_line_u) & 0xFF
    size = int(args.stage_tim04_na_template_size) & 0xFF
    v = int(args.stage_tim04_na_template_v) & 0xFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    old_right = (old_u + size - 1) & 0xFF
    new_right = (new_u + size - 1) & 0xFF

    # Original overwritten pair at 800A3CDC/800A3CE0:
    #   lui v0,0x0040
    #   and v0,s1,v0
    # The original branch at 800A3CE4 consumes v0, so preserve those semantics.
    a.lui("v0", 0x0040)
    a.word(rt(REG["s1"], REG["v0"], REG["v0"], 0, 0x24))  # and v0,s1,v0

    # v41: exact Object-overlay element/N-A plus PLANET/Modus element transition-frame repairs.
    # These mirror the stable template repairs so opening FT4 frames do not
    # flash/snap before the steady SPRT path takes over.
    stage_tim04_remap_object_overlay_ft4_edges(a, args, "done")
    stage_tim04_force_modus_element_ft4_edges(a, args, "done")

    # Narrow to the same gray 16x16 N/A template, but earlier than the packet
    # assembly. This catches the transformed/FT4 opening quad before UV words are
    # created at all.
    a.lbu("t5", 0x000E, "s0")
    a.addiu("t6", "zero", old_u)
    a.bne("t5", "t6", "done"); a.nop()

    a.lbu("t5", 0x000F, "s0")
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", "done"); a.nop()

    a.lbu("t5", 0x0008, "s0")
    a.addiu("t6", "zero", size)
    a.bne("t5", "t6", "done"); a.nop()

    a.lbu("t5", 0x000A, "s0")
    a.bne("t5", "t6", "done"); a.nop()

    a.lbu("t5", 0x0014, "s0")
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", "done"); a.nop()
    a.lbu("t5", 0x0015, "s0")
    a.bne("t5", "t6", "done"); a.nop()
    a.lbu("t5", 0x0016, "s0")
    a.bne("t5", "t6", "done"); a.nop()

    # t1 and t4 are the precomputed left/right U edges. Depending on horizontal
    # flip, they can be old_u/old_right or old_right/old_u. Remap each edge
    # independently so the flip logic remains intact.
    a.addiu("t6", "zero", old_u)
    a.bne("t1", "t6", "check_t1_right"); a.nop()
    a.addiu("t1", "zero", new_u)
    a.j("check_t4_left"); a.nop()

    a.lab("check_t1_right")
    a.addiu("t6", "zero", old_right)
    a.bne("t1", "t6", "check_t4_left"); a.nop()
    a.addiu("t1", "zero", new_right)

    a.lab("check_t4_left")
    a.addiu("t6", "zero", old_u)
    a.bne("t4", "t6", "check_t4_right"); a.nop()
    a.addiu("t4", "zero", new_u)
    a.j("done"); a.nop()

    a.lab("check_t4_right")
    a.addiu("t6", "zero", old_right)
    a.bne("t4", "t6", "done"); a.nop()
    a.addiu("t4", "zero", new_right)

    a.lab("done")
    a.j(STAGE_TIM04_NA_FT4_EDGE_FILTER_RETURN); a.nop()
    return a.out()


def patch_stage_tim04_na_ft4_edge_filter_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_stage_tim04_na_ft4_edge_filter_hook", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(STAGE_TIM04_NA_FT4_EDGE_FILTER_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(
            f"STAGE TIM04 N/A FT4 edge hook site maps outside EXE: "
            f"RAM 0x{STAGE_TIM04_NA_FT4_EDGE_FILTER_SITE:08X}, file 0x{site_off:X}"
        )

    cur = bytes(exe[site_off:site_off+8])
    replacing_existing_hook = False
    if cur != STAGE_TIM04_NA_FT4_EDGE_FILTER_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already:
            # v27: replace older FT4-edge hooks automatically. Otherwise an EXE
            # already patched by v132/v24/v25 can keep the old transition-frame
            # code and still flash the wrong Object-overlay graphic before the
            # stable N/A repair snaps in.
            replacing_existing_hook = True
        elif not getattr(args, "stage_tim04_na_ft4_edge_filter_force", False):
            raise RuntimeError(
                f"STAGE TIM04 N/A FT4 edge hook site mismatch at 0x{site_off:X}: "
                f"{cur.hex(' ')}; expected {STAGE_TIM04_NA_FT4_EDGE_FILTER_EXPECT.hex(' ')}. "
                "Use --stage-tim04-na-ft4-edge-filter-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_stage_tim04_na_ft4_edge_filter_hook(hook_ram, args)
    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{STAGE_TIM04_NA_FT4_EDGE_FILTER_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "old_u": f"0x{int(args.stage_tim04_na_template_old_u) & 0xFF:02X}",
        "new_u": f"0x{int(args.stage_tim03_na_line_u) & 0xFF:02X}",
        "description": "Remaps the FT4 transition quad's source U edges before its UV words are assembled.",
    }


def build_stage_tim04_na_ft4_source_filter_hook(hook_ram, args):
    a = A(hook_ram)
    bad_word = int(args.stage_tim04_na_bad_word) & 0xFFFFFFFF
    good_word = int(args.stage_tim04_na_good_word) & 0xFFFFFFFF
    packet_cmd = int(args.stage_tim04_na_ft4_packet_cmd) & 0xFFFFFFFF

    # Original overwritten pair at 800A3DAC/800A3DB0:
    #   or v1,v1,a3          ; finish U0/V0/CLUT/TPAGE word for POLY_FT4
    #   sw v1,0x0C(a0)       ; store uv0/clut/tpage
    # We install a two-word patch at the site (j hook; nop delay), so the
    # following store cannot run as a delay slot before the filter.
    a.word(rt(REG["v1"], REG["a3"], REG["v1"], 0, 0x25))  # or v1,v1,a3

    # v30: exact transition-frame source-word repairs. These run before the
    # legacy N/A word-only filter so they can handle evolved Object slots and
    # PLANET/Modus old-U element slots without touching unrelated packets.
    stage_tim04_remap_object_overlay_ft4_source(a, args, "store")
    stage_tim04_force_modus_element_ft4_source(a, args, "store")

    a.lui("t0", (bad_word >> 16) & 0xFFFF)
    a.ori("t0", "t0", bad_word & 0xFFFF)
    a.bne("v1", "t0", "store"); a.nop()

    # Narrow to the exact gray POLY_FT4 command/color word seen in the bad
    # snapshot. At this point command/color has already been stored at +0x04.
    a.lw("t0", 0x0004, "a0")
    a.lui("t1", (packet_cmd >> 16) & 0xFFFF)
    a.ori("t1", "t1", packet_cmd & 0xFFFF)
    a.bne("t0", "t1", "store"); a.nop()

    a.lui("v1", (good_word >> 16) & 0xFFFF)
    a.ori("v1", "v1", good_word & 0xFFFF)

    a.lab("store")
    a.sw("v1", 0x000C, "a0")
    a.j(STAGE_TIM04_NA_FT4_SOURCE_FILTER_RETURN); a.nop()
    return a.out()


def patch_stage_tim04_na_ft4_source_filter_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_stage_tim04_na_ft4_source_filter_hook", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(STAGE_TIM04_NA_FT4_SOURCE_FILTER_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(
            f"STAGE TIM04 N/A FT4 source hook site maps outside EXE: "
            f"RAM 0x{STAGE_TIM04_NA_FT4_SOURCE_FILTER_SITE:08X}, file 0x{site_off:X}"
        )

    cur = bytes(exe[site_off:site_off+8])
    replacing_existing_hook = False
    if cur != STAGE_TIM04_NA_FT4_SOURCE_FILTER_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already:
            # v35: replace older FT4 source hooks automatically, mirroring the
            # edge hook behavior. Otherwise an EXE patched by an earlier version
            # can keep stale transition-frame code even after running this patcher.
            replacing_existing_hook = True
        elif not getattr(args, "stage_tim04_na_ft4_source_filter_force", False):
            raise RuntimeError(
                f"STAGE TIM04 N/A FT4 source hook site mismatch at 0x{site_off:X}: "
                f"{cur.hex(' ')}; expected {STAGE_TIM04_NA_FT4_SOURCE_FILTER_EXPECT.hex(' ')}. "
                "Use --stage-tim04-na-ft4-source-filter-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_stage_tim04_na_ft4_source_filter_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)  # j hook ; nop delay-slot
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{STAGE_TIM04_NA_FT4_SOURCE_FILTER_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "bad_word": f"0x{int(args.stage_tim04_na_bad_word) & 0xFFFFFFFF:08X}",
        "good_word": f"0x{int(args.stage_tim04_na_good_word) & 0xFFFFFFFF:08X}",
        "packet_cmd": f"0x{int(args.stage_tim04_na_ft4_packet_cmd) & 0xFFFFFFFF:08X}",
        "description": "Fixes the POLY_FT4 transition-frame N/A source word, with a safe two-word/delay-slot-aware hook.",
    }


# v111 Object-overlay element/N-A template hook.
#
# The shared renderer builds a source/CLUT word from the template at s0:
#   U      = byte [s0+0x0E]
#   V      = byte [s0+0x0F]
#   width  = half [s0+0x08]
#   height = half [s0+0x0A]
#   color  = bytes [s0+0x14..0x16]
#
# Instead of changing all final packets after assembly, this hook runs
# immediately before the source word is built. It narrows the match to the
# gray 16x16 Object-overlay element/N-A templates at V=0xC0 and then remaps
# old Japanese slots to the widened English slots:
#   Earth old U=0x18 -> --stage-tim04-earth-u / --stage-tim04-earth-w
#   Water old U=0x28 -> --stage-tim04-water-u / --stage-tim04-water-w
#   Wind  old U=0x38 -> --stage-tim04-wind-u  / --stage-tim04-wind-w
#   Fire  old U=0x48 -> --stage-tim04-fire-u  / --stage-tim04-fire-w
#   N/A   old U=0x58 -> --stage-tim03-na-line-u / --stage-tim04-dash-w
#
# Then it executes the original overwritten instruction:
#   lbu v1,0x0F(s0)
STAGE_TIM04_NA_TEMPLATE_U_HOOK_SITE = 0x800A3AB4
STAGE_TIM04_NA_TEMPLATE_U_HOOK_RETURN = 0x800A3AB8
STAGE_TIM04_NA_TEMPLATE_U_HOOK_EXPECT = 0x9203000F


def stage_tim04_apply_x_delta(a, delta):
    if int(delta) == 0:
        return
    a.lh("v1", 0x0004, "s0")
    a.addiu("v1", "v1", int(delta))
    a.word(it(0x29, REG["s0"], REG["v1"], 0x0004))  # sh v1,0x04(s0) ; destination X


def stage_tim04_apply_element_x_transition(a, label, target_u, target_shift, element_u_to_shift):
    """Move the persistent STAGE template from its current element shift to
    target_shift without accumulating across frames or N/A transitions."""
    if not any(int(s) for _, s in element_u_to_shift) and int(target_shift) == 0:
        return
    done = f"{label}_x_done"
    default = f"{label}_x_default"
    a.lbu("v0", 0x000E, "s0")
    a.addiu("v1", "zero", int(target_u) & 0xFF)
    a.beq("v0", "v1", done); a.nop()
    for idx, (u, prev_shift) in enumerate(element_u_to_shift):
        if (int(u) & 0xFF) == (int(target_u) & 0xFF):
            continue
        lab = f"{label}_from_{idx}"
        a.addiu("v1", "zero", int(u) & 0xFF)
        a.beq("v0", "v1", lab); a.nop()
    a.j(default); a.nop()
    for idx, (u, prev_shift) in enumerate(element_u_to_shift):
        if (int(u) & 0xFF) == (int(target_u) & 0xFF):
            continue
        lab = f"{label}_from_{idx}"
        a.lab(lab)
        stage_tim04_apply_x_delta(a, int(target_shift) - int(prev_shift))
        a.j(done); a.nop()
    a.lab(default)
    stage_tim04_apply_x_delta(a, int(target_shift))
    a.lab(done)


def stage_tim04_restore_x_for_dash(a, label, element_u_to_shift):
    """Undo any current element shift when returning to the N/A line."""
    if not any(int(s) for _, s in element_u_to_shift):
        return
    done = f"{label}_x_done"
    a.lbu("v0", 0x000E, "s0")
    for idx, (u, shift) in enumerate(element_u_to_shift):
        lab = f"{label}_from_{idx}"
        a.addiu("v1", "zero", int(u) & 0xFF)
        a.beq("v0", "v1", lab); a.nop()
    a.j(done); a.nop()
    for idx, (u, shift) in enumerate(element_u_to_shift):
        lab = f"{label}_from_{idx}"
        a.lab(lab)
        stage_tim04_apply_x_delta(a, -int(shift))
        a.j(done); a.nop()
    a.lab(done)



def stage_tim04_apply_modus_crest_x_fix(a, args):
    """Modus subscreen Crest-label X-position fix.

    The Modus Crest label appears as a rendered STAGE/TexBlend template rather
    than a direct MAIN.EXE text draw call.  The observed final packet is:
      64808080 0036000A 78520018 00100058
    which corresponds to x=0x0A, y=0x36, U=0x18, V=0x00, w=0x58, h=0x10.

    This guard runs inside the existing STAGE TIM04 template hook, before the
    source/CLUT word is built, and changes only the destination X field.
    """
    if getattr(args, "disable_modus_crest_template_x_fix", False):
        return

    old_x = int(args.modus_crest_template_old_x)
    new_x = int(args.modus_crest_template_new_x)
    y = int(args.modus_crest_template_y)
    u = int(args.modus_crest_template_u) & 0xFF
    v = int(args.modus_crest_template_v) & 0xFF
    w_ = int(args.modus_crest_template_w) & 0xFFFF
    h = int(args.modus_crest_template_h) & 0xFFFF
    color = int(args.modus_crest_template_color) & 0xFF

    a.lh("v0", 0x0004, "s0")
    a.addiu("v1", "zero", old_x)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()
    a.lh("v0", 0x0006, "s0")
    a.addiu("v1", "zero", y)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()

    a.lbu("v0", 0x000E, "s0")
    a.addiu("v1", "zero", u)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()
    a.lbu("v0", 0x000F, "s0")
    a.addiu("v1", "zero", v)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()

    a.lhu("v0", 0x0008, "s0")
    a.addiu("v1", "zero", w_)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()
    a.lhu("v0", 0x000A, "s0")
    a.addiu("v1", "zero", h)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()

    a.lbu("v0", 0x0014, "s0")
    a.addiu("v1", "zero", color)
    a.bne("v0", "v1", "modus_crest_done"); a.nop()
    a.lbu("v0", 0x0015, "s0")
    a.bne("v0", "v1", "modus_crest_done"); a.nop()
    a.lbu("v0", 0x0016, "s0")
    a.bne("v0", "v1", "modus_crest_done"); a.nop()

    a.addiu("v0", "zero", new_x)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0004))  # sh v0,0x04(s0) ; destination X

    a.lab("modus_crest_done")


def stage_tim04_apply_modus_element_size_fix(a, args):
    """PLANET/Modus subscreen PLMENU01 element-word source/width/X fix.

    User-verified PLMENU01.TIM visible ink bounds:
      Earth: x=25, W=23, visible y=193, visible h=11
      Water: x=50, W=24
      Wind:  x=76, W=19
      Fire:  x=97, W=18

    v29 uses safe padded source rectangles where the surrounding pixels are
    transparent and do not overlap the neighboring word:
      Earth: U=0x18, W=0x19
      Water: U=0x31, W=0x1A
      Wind:  U=0x4B, W=0x15
      Fire:  U=0x60, W=0x14

    v48 snapshot fix:
      The broken PLANET Select Modus/Point overlay snapshots show a second
      Modus template with marker 0x23 at x=0x7D, y=-37, w=0x12, h=0x10,
      V=0xC0, RGB=0x80/0x80/0x80, and old source U values.  That is why Earth
      appeared clipped as "Eart" and Fire sampled "ndFi".  Accept only this
      observed bad x/w shape in addition to the normal x=0x7A path, then write
      the Modus padded U/W and normalize X back to 0x7A.  This is narrower than
      the earlier broad v39 X normalization.
    """
    if getattr(args, "disable_modus_stage_template_element_fix", False):
        return
    x = int(args.modus_stage_element_x)
    bad_x = int(getattr(args, "modus_stage_element_bad_x", 0x7D))
    bad_w = int(getattr(args, "modus_stage_element_bad_w", 0x12)) & 0xFFFF
    y = int(args.modus_stage_element_y)
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    marker = int(args.modus_stage_element_marker) & 0xFF

    # Normal stable template: exact X=0x7A.
    # Observed broken template: X=0x7D AND W=0x12.  Do not accept a broad X
    # range, because previous broad normalization shifted unrelated element
    # graphics.
    a.lh("v0", 0x0004, "s0")
    a.addiu("v1", "zero", x)
    a.beq("v0", "v1", "modus_elem_x_ok"); a.nop()
    a.addiu("v1", "zero", bad_x)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()
    a.lhu("v0", 0x0008, "s0")
    a.addiu("v1", "zero", bad_w)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()
    a.lab("modus_elem_x_ok")

    a.lh("v0", 0x0006, "s0")
    a.addiu("v1", "zero", y)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()

    a.lbu("v0", 0x0017, "s0")
    a.addiu("v1", "zero", marker)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()

    a.lbu("v0", 0x000F, "s0")
    a.addiu("v1", "zero", v)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()
    a.lhu("v0", 0x000A, "s0")
    a.addiu("v1", "zero", h)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()

    a.lbu("v0", 0x0014, "s0")
    a.addiu("v1", "zero", color)
    a.bne("v0", "v1", "modus_elem_done"); a.nop()
    a.lbu("v0", 0x0015, "s0")
    a.bne("v0", "v1", "modus_elem_done"); a.nop()
    a.lbu("v0", 0x0016, "s0")
    a.bne("v0", "v1", "modus_elem_done"); a.nop()

    # Match old source U only; then write the stable Modus PLMENU01 padded
    # source U/width and normalize destination X.  Matching old U only avoids
    # re-processing already-relocated templates and keeps this as a relocation
    # layer rather than a visual nudge.
    a.lbu("v0", 0x000E, "s0")
    for old_u, lab in ((0x18,"earth"),(0x28,"water"),(0x38,"wind"),(0x48,"fire")):
        a.addiu("v1", "zero", old_u)
        a.beq("v0", "v1", f"modus_elem_{lab}"); a.nop()
    a.j("modus_elem_done"); a.nop()

    a.lab("modus_elem_earth")
    a.addiu("v0", "zero", x)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0004))
    a.addiu("v0", "zero", int(args.modus_stage_earth_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.modus_stage_earth_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("modus_elem_water")
    a.addiu("v0", "zero", x)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0004))
    a.addiu("v0", "zero", int(args.modus_stage_water_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.modus_stage_water_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("modus_elem_wind")
    a.addiu("v0", "zero", x)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0004))
    a.addiu("v0", "zero", int(args.modus_stage_wind_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.modus_stage_wind_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("modus_elem_fire")
    a.addiu("v0", "zero", x)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0004))
    a.addiu("v0", "zero", int(args.modus_stage_fire_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.modus_stage_fire_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("modus_elem_done")


def stage_tim04_remap_object_overlay_value_slot(a, args, done_label, jump_to_orig=True):
    """Remap only the STAGE Object-overlay value slot by actual element/N-A identity.

    v43 replaces the older exact-slot force-N/A behavior with a relocation table.  The guard is still
    narrow enough to avoid Pixy overlay and unrelated STAGE graphics, but once
    the Object slot is confirmed it reads the current source U and maps that
    symbol to the edited Earth/Water/Wind/Fire/N-A source and width.  This means
    late-game elemental objects will keep their correct element instead of being
    collapsed to the N/A line.
    """
    if int(args.stage_tim04_object_overlay_na_slot_ram) == 0:
        return

    slot_ram = int(args.stage_tim04_object_overlay_na_slot_ram) & 0xFFFFFFFF
    object_guard = int(args.stage_tim04_object_overlay_object_guard_halfword) & 0xFFFF
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    stage_marker = int(args.stage_tim04_object_overlay_marker) & 0xFF

    a.lui("v0", (slot_ram >> 16) & 0xFFFF)
    a.ori("v0", "v0", slot_ram & 0xFFFF)
    a.bne("s0", "v0", done_label); a.nop()

    a.lhu("v0", 0x001E, "s0")
    a.addiu("v1", "zero", object_guard)
    a.bne("v0", "v1", done_label); a.nop()

    a.lbu("v0", 0x0017, "s0")
    a.addiu("v1", "zero", stage_marker)
    a.bne("v0", "v1", done_label); a.nop()

    a.lbu("v0", 0x000F, "s0")
    a.addiu("v1", "zero", v)
    a.bne("v0", "v1", done_label); a.nop()

    a.lhu("v0", 0x000A, "s0")
    a.addiu("v1", "zero", h)
    a.bne("v0", "v1", done_label); a.nop()

    a.lbu("v0", 0x0014, "s0")
    a.addiu("v1", "zero", color)
    a.bne("v0", "v1", done_label); a.nop()
    a.lbu("v0", 0x0015, "s0")
    a.bne("v0", "v1", done_label); a.nop()
    a.lbu("v0", 0x0016, "s0")
    a.bne("v0", "v1", done_label); a.nop()

    cases = stage_tim04_object_overlay_target_cases(args, include_aliases=False)
    a.lbu("v0", 0x000E, "s0")
    for case in cases:
        lab = f"object_overlay_slot_{case['name']}"
        for match_u in sorted(set(int(u) & 0xFF for u in case["match_us"])):
            a.addiu("v1", "zero", match_u)
            a.beq("v0", "v1", lab); a.nop()
    a.j(done_label); a.nop()

    for case in cases:
        lab = f"object_overlay_slot_{case['name']}"
        new_u = int(case["new_u"]) & 0xFF
        new_w = int(case["new_w"]) & 0xFFFF
        x = int(case["x"])
        a.lab(lab)
        a.addiu("v1", "zero", x)
        a.word(it(0x29, REG["s0"], REG["v1"], 0x0004))  # sh v1,0x04(s0) ; destination X
        a.addiu("v0", "zero", new_u)
        a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))  # sb v0,0x0E(s0) ; U
        a.addiu("v0", "zero", new_w)
        a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))  # sh v0,0x08(s0) ; width
        if jump_to_orig:
            a.j("orig"); a.nop()
        else:
            a.j(done_label); a.nop()


def stage_tim04_apply_object_overlay_element_restore(a, args, element_u_to_shift):
    """v43 exact Object-overlay old-U relocation repair path.

    Earlier patchers forced the Object value slot to the N/A line because the
    early-game objects all appeared to use N/A.  v43 preserves the original game's current symbol identity: stock
    Earth/Water/Wind/Fire/N-A U values are each remapped to the edited source U,
    width, and absolute X.  Unknown symbols are left alone.
    """
    if getattr(args, "disable_stage_tim04_na_template_u_hook", False):
        return

    stage_tim04_remap_object_overlay_value_slot(a, args, "object_overlay_slot_done", jump_to_orig=True)
    a.lab("object_overlay_slot_done")




def _modus_exact_ft4_edge_cases(args):
    return [
        ("earth", (0x18, int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_u) & 0xFF), int(args.modus_stage_ft4_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_w) & 0xFFFF),
        ("water", (0x28, int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_ft4_water_u) & 0xFF), int(args.modus_stage_ft4_water_u) & 0xFF, int(args.modus_stage_ft4_water_w) & 0xFFFF),
        ("wind",  (0x38, int(args.modus_stage_wind_u) & 0xFF,  int(args.modus_stage_ft4_wind_u) & 0xFF),  int(args.modus_stage_ft4_wind_u) & 0xFF,  int(args.modus_stage_ft4_wind_w) & 0xFFFF),
        ("fire",  (0x48, int(args.modus_stage_fire_u) & 0xFF,  int(args.modus_stage_ft4_fire_u) & 0xFF),  int(args.modus_stage_ft4_fire_u) & 0xFF,  int(args.modus_stage_ft4_fire_w) & 0xFFFF),
    ]


def _build_modus_exact_ft4_edge_read_hook(hook_ram, args, variant):
    """v60 exact active-record FT4 U/W register fix.

    Variant A is the 800A3CAC path:
      lbu v1,0x0E(s0)
      lbu v0,0x08(s0)
    Variant B is the 800A3CC8 path:
      lbu t1,0x0E(s0)
      lbu v0,0x08(s0)

    The hook first executes the original loads, then if s0 is exactly the active
    Modus record observed in No$PSX (8012A3BC), it replaces the current U/W
    registers with the visible FT4 source U/W for Earth/Water/Wind/Fire before
    the right-edge math runs. This is meant to remove the last one-frame Fire
    flash without touching generic menu/Neredy sprites.
    """
    a = A(hook_ram)
    u_reg = "v1" if variant == "A" else "t1"

    # Original overwritten pair.
    a.lbu(u_reg, 0x000E, "s0")
    a.lbu("v0", 0x0008, "s0")

    rec = int(getattr(args, "modus_exact_ft4_edge_record_ram", MODUS_EXACT_FT4_EDGE_RECORD_RAM)) & 0xFFFFFFFF
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    marker = int(args.modus_stage_element_marker) & 0xFF

    # Exact active-record pointer guard. This is intentionally stricter than the
    # older shared FT4 hooks and avoids the failed shared-SPRT path entirely.
    a.lui("t5", (rec >> 16) & 0xFFFF)
    a.ori("t5", "t5", rec & 0xFFFF)
    a.bne("s0", "t5", "mefer_done"); a.nop()

    # Shape guard.
    a.lbu("t5", 0x000F, "s0")
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", "mefer_done"); a.nop()

    a.lhu("t5", 0x000A, "s0")
    a.addiu("t6", "zero", h)
    a.bne("t5", "t6", "mefer_done"); a.nop()

    a.lbu("t5", 0x0017, "s0")
    a.addiu("t6", "zero", marker)
    a.bne("t5", "t6", "mefer_done"); a.nop()

    a.lbu("t5", 0x0014, "s0")
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", "mefer_done"); a.nop()
    a.lbu("t5", 0x0015, "s0")
    a.bne("t5", "t6", "mefer_done"); a.nop()
    a.lbu("t5", 0x0016, "s0")
    a.bne("t5", "t6", "mefer_done"); a.nop()

    # Map old/padded/FT4 U aliases to the FT4-visible U/W used by this path.
    for name, match_us, new_u, new_w in _modus_exact_ft4_edge_cases(args):
        lab = f"mefer_{name}"
        for old_u in sorted(set(int(u) & 0xFF for u in match_us)):
            a.addiu("t5", "zero", old_u)
            a.beq(u_reg, "t5", lab); a.nop()
    a.j("mefer_done"); a.nop()

    for name, match_us, new_u, new_w in _modus_exact_ft4_edge_cases(args):
        a.lab(f"mefer_{name}")
        a.addiu(u_reg, "zero", int(new_u) & 0xFF)
        a.addiu("v0", "zero", int(new_w) & 0xFFFF)
        a.j("mefer_done"); a.nop()

    a.lab("mefer_done")
    ret = MODUS_EXACT_FT4_EDGE_READ_A_RETURN if variant == "A" else MODUS_EXACT_FT4_EDGE_READ_B_RETURN
    a.j(ret); a.nop()
    return a.out()


def patch_modus_exact_ft4_edge_read_hooks(exe, load, args, dry_run=False):
    if getattr(args, "disable_modus_exact_ft4_edge_read_hooks", False):
        return {"enabled": False, "status": "disabled"}

    sites = [
        ("A", MODUS_EXACT_FT4_EDGE_READ_A_SITE, MODUS_EXACT_FT4_EDGE_READ_A_EXPECT, "v1/v0 edge-read path"),
        ("B", MODUS_EXACT_FT4_EDGE_READ_B_SITE, MODUS_EXACT_FT4_EDGE_READ_B_EXPECT, "t1/v0 edge-read path"),
    ]
    rows = []
    for variant, site_ram, expect, desc in sites:
        site_off = ram2off(site_ram, load)
        if site_off < 0 or site_off + 8 > len(exe):
            raise RuntimeError(
                f"Modus exact FT4 edge-read {variant} site maps outside EXE: RAM 0x{site_ram:08X}, file 0x{site_off:X}"
            )
        cur = bytes(exe[site_off:site_off+8])
        replacing_existing_hook = False
        if cur != expect:
            already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
            if already:
                replacing_existing_hook = True
            elif not getattr(args, "modus_exact_ft4_edge_read_force", False):
                raise RuntimeError(
                    f"Modus exact FT4 edge-read {variant} site mismatch at 0x{site_off:X}: {cur.hex(' ')}; "
                    f"expected {expect.hex(' ')}. Use --modus-exact-ft4-edge-read-force only if intentional."
                )

        hook_off = align(len(exe), 4)
        hook_ram = off2ram(hook_off, load)
        code = _build_modus_exact_ft4_edge_read_hook(hook_ram, args, variant)

        if not dry_run:
            if len(exe) < hook_off:
                exe += b"\0" * (hook_off - len(exe))
            exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
            exe[hook_off:hook_off+len(code)] = code

        rows.append({
            "variant": variant,
            "enabled": True,
            "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
            "site_ram": f"0x{site_ram:08X}",
            "site_file": f"0x{site_off:X}",
            "hook_ram": f"0x{hook_ram:08X}",
            "hook_file": f"0x{hook_off:X}",
            "hook_len": len(code),
            "record": f"0x{int(getattr(args, 'modus_exact_ft4_edge_record_ram', MODUS_EXACT_FT4_EDGE_RECORD_RAM)) & 0xFFFFFFFF:08X}",
            "description": desc,
        })

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "report": rows,
        "description": "Exact active-record FT4 U/W register repair at the two edge-read paths before current primitive edge math.",
    }



def build_modus_local_record_fix_hook(hook_ram, args):
    a = A(hook_ram)

    # Original overwritten pair at 80079998/8007999C:
    #   sh v1,0x0AF2(a0)
    #   sh v1,0x0ACE(a0)
    # At this point a0 is the PLANET template base (observed 801298E8).
    a.word(it(0x29, REG["a0"], REG["v1"], 0x0AF2))
    a.word(it(0x29, REG["a0"], REG["v1"], 0x0ACE))

    # Preserve scratch registers: this is a local routine and later code may
    # still expect caller-temporary registers to remain meaningful.
    a.addiu("sp", "sp", -0x20)
    a.sw("t5", 0x00, "sp")
    a.sw("t6", 0x04, "sp")
    a.sw("t7", 0x08, "sp")
    a.sw("t8", 0x0C, "sp")

    # Active record: BASE+0xAD4.  Field offsets match the TexBlend/SPRT template
    # layout used by the existing guarded repairs:
    #   +04 X, +06 Y, +08 W, +0A H, +0E U, +0F V, +14..16 RGB, +17 marker.
    # Relative to BASE, these are +0AD8/+0ADA/+0ADC/+0ADE/+0AE2/+0AE3/+0AE8..+0AEB.
    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    marker = int(args.modus_stage_element_marker) & 0xFF
    x = int(args.modus_stage_element_x)

    # Narrow guard: only the gray Modus/Point element record.
    a.lbu("t5", 0x0AE3, "a0")
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", "mlrf_done"); a.nop()

    a.lhu("t5", 0x0ADE, "a0")
    a.addiu("t6", "zero", h)
    a.bne("t5", "t6", "mlrf_done"); a.nop()

    a.lbu("t5", 0x0AEB, "a0")
    a.addiu("t6", "zero", marker)
    a.bne("t5", "t6", "mlrf_done"); a.nop()

    a.lbu("t5", 0x0AE8, "a0")
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", "mlrf_done"); a.nop()
    a.lbu("t5", 0x0AE9, "a0")
    a.bne("t5", "t6", "mlrf_done"); a.nop()
    a.lbu("t5", 0x0AEA, "a0")
    a.bne("t5", "t6", "mlrf_done"); a.nop()

    # Map by semantic stock U.  Also accept already-padded U aliases so the hook
    # can normalize W/X without oscillating or needing a second pass.
    a.lbu("t5", 0x0AE2, "a0")
    cases = [
        ("earth", (0x18, int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_u) & 0xFF), int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_earth_w) & 0xFFFF),
        ("water", (0x28, int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_ft4_water_u) & 0xFF), int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_water_w) & 0xFFFF),
        ("wind",  (0x38, int(args.modus_stage_wind_u) & 0xFF,  int(args.modus_stage_ft4_wind_u) & 0xFF),  int(args.modus_stage_wind_u) & 0xFF,  int(args.modus_stage_wind_w) & 0xFFFF),
        ("fire",  (0x48, int(args.modus_stage_fire_u) & 0xFF,  int(args.modus_stage_ft4_fire_u) & 0xFF),  int(args.modus_stage_fire_u) & 0xFF,  int(args.modus_stage_fire_w) & 0xFFFF),
    ]
    for name, match_us, new_u, new_w in cases:
        lab = f"mlrf_{name}"
        for old_u in sorted(set(int(u) & 0xFF for u in match_us)):
            a.addiu("t6", "zero", old_u)
            a.beq("t5", "t6", lab); a.nop()
    a.j("mlrf_done"); a.nop()

    for name, match_us, new_u, new_w in cases:
        a.lab(f"mlrf_{name}")
        a.addiu("t7", "zero", x)
        a.word(it(0x29, REG["a0"], REG["t7"], 0x0AD8))  # sh t7,BASE+0xAD8 => record+0x04 X
        a.addiu("t7", "zero", new_u)
        a.word(it(0x28, REG["a0"], REG["t7"], 0x0AE2))  # sb t7,BASE+0xAE2 => record+0x0E U
        a.addiu("t7", "zero", new_w)
        a.word(it(0x29, REG["a0"], REG["t7"], 0x0ADC))  # sh t7,BASE+0xADC => record+0x08 W
        a.j("mlrf_done"); a.nop()

    a.lab("mlrf_done")
    a.lw("t5", 0x00, "sp")
    a.lw("t6", 0x04, "sp")
    a.lw("t7", 0x08, "sp")
    a.lw("t8", 0x0C, "sp")
    a.addiu("sp", "sp", 0x20)
    a.j(MODUS_LOCAL_RECORD_FIX_RETURN); a.nop()
    return a.out()



def build_modus_local_record_early_fix_hook(hook_ram, args):
    a = A(hook_ram)

    # v59: run the same local record repair earlier, while the PLANET/Modus
    # record-read loop is visiting each 0x24-byte record.  At the observed
    # flicker break, v1=80129A98 and the live Modus element record is v1+0x924
    # = 8012A3BC.  Offsets below are therefore relative to v1.
    a.addiu("sp", "sp", -0x20)
    a.sw("t5", 0x00, "sp")
    a.sw("t6", 0x04, "sp")
    a.sw("t7", 0x08, "sp")
    a.sw("t8", 0x0C, "sp")

    v = int(args.stage_tim04_na_template_v) & 0xFF
    h = int(args.modus_stage_element_h) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF
    marker = int(args.modus_stage_element_marker) & 0xFF
    x = int(args.modus_stage_element_x)

    # Guard: only the gray Modus/Point element record.  The same guards are used
    # by v58's later local repair, but this one fires before the record is read.
    a.lbu("t5", 0x0933, "v1")  # record+0x0F V
    a.addiu("t6", "zero", v)
    a.bne("t5", "t6", "mlref_done"); a.nop()

    a.lhu("t5", 0x092E, "v1")  # record+0x0A H
    a.addiu("t6", "zero", h)
    a.bne("t5", "t6", "mlref_done"); a.nop()

    a.lbu("t5", 0x093B, "v1")  # record+0x17 marker
    a.addiu("t6", "zero", marker)
    a.bne("t5", "t6", "mlref_done"); a.nop()

    a.lbu("t5", 0x0938, "v1")  # record+0x14 R
    a.addiu("t6", "zero", color)
    a.bne("t5", "t6", "mlref_done"); a.nop()
    a.lbu("t5", 0x0939, "v1")  # record+0x15 G
    a.bne("t5", "t6", "mlref_done"); a.nop()
    a.lbu("t5", 0x093A, "v1")  # record+0x16 B
    a.bne("t5", "t6", "mlref_done"); a.nop()

    a.lbu("t5", 0x0932, "v1")  # record+0x0E U
    cases = [
        ("earth", (0x18, int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_ft4_earth_u) & 0xFF), int(args.modus_stage_earth_u) & 0xFF, int(args.modus_stage_earth_w) & 0xFFFF),
        ("water", (0x28, int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_ft4_water_u) & 0xFF), int(args.modus_stage_water_u) & 0xFF, int(args.modus_stage_water_w) & 0xFFFF),
        ("wind",  (0x38, int(args.modus_stage_wind_u) & 0xFF,  int(args.modus_stage_ft4_wind_u) & 0xFF),  int(args.modus_stage_wind_u) & 0xFF,  int(args.modus_stage_wind_w) & 0xFFFF),
        ("fire",  (0x48, int(args.modus_stage_fire_u) & 0xFF,  int(args.modus_stage_ft4_fire_u) & 0xFF),  int(args.modus_stage_fire_u) & 0xFF,  int(args.modus_stage_fire_w) & 0xFFFF),
    ]
    for name, match_us, new_u, new_w in cases:
        lab = f"mlref_{name}"
        for old_u in sorted(set(int(u) & 0xFF for u in match_us)):
            a.addiu("t6", "zero", old_u)
            a.beq("t5", "t6", lab); a.nop()
    a.j("mlref_done"); a.nop()

    for name, match_us, new_u, new_w in cases:
        a.lab(f"mlref_{name}")
        a.addiu("t7", "zero", x)
        a.word(it(0x29, REG["v1"], REG["t7"], 0x0928))  # sh t7,record+0x04 X
        a.addiu("t7", "zero", new_u)
        a.word(it(0x28, REG["v1"], REG["t7"], 0x0932))  # sb t7,record+0x0E U
        a.addiu("t7", "zero", new_w)
        a.word(it(0x29, REG["v1"], REG["t7"], 0x092C))  # sh t7,record+0x08 W
        a.j("mlref_done"); a.nop()

    a.lab("mlref_done")
    a.lw("t5", 0x00, "sp")
    a.lw("t6", 0x04, "sp")
    a.lw("t7", 0x08, "sp")
    a.lw("t8", 0x0C, "sp")
    a.addiu("sp", "sp", 0x20)

    # Original overwritten pair at 8007972C/80079730:
    #   lhu v0,0x092A(v1)
    #   addiu s0,s0,1
    a.word(it(0x25, REG["v1"], REG["v0"], 0x092A))
    a.addiu("s0", "s0", 1)
    a.j(MODUS_LOCAL_RECORD_EARLY_FIX_RETURN); a.nop()
    return a.out()


def patch_modus_local_record_early_fix_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_modus_local_record_early_fix", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(MODUS_LOCAL_RECORD_EARLY_FIX_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(
            f"Modus early local record fix site maps outside EXE: RAM 0x{MODUS_LOCAL_RECORD_EARLY_FIX_SITE:08X}, file 0x{site_off:X}"
        )

    cur = bytes(exe[site_off:site_off+8])
    replacing_existing_hook = False
    if cur != MODUS_LOCAL_RECORD_EARLY_FIX_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already:
            replacing_existing_hook = True
        elif not getattr(args, "modus_local_record_early_force", False):
            raise RuntimeError(
                f"Modus early local record fix site mismatch at 0x{site_off:X}: {cur.hex(' ')}; "
                f"expected {MODUS_LOCAL_RECORD_EARLY_FIX_EXPECT.hex(' ')}. "
                "Use --modus-local-record-early-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_modus_local_record_early_fix_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{MODUS_LOCAL_RECORD_EARLY_FIX_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "record": "v1+0x924 / observed RAM 8012A3BC",
        "description": "Earlier local PLANET/Modus record repair inside the record-read loop, before the shared builders see the old element values.",
    }


def patch_modus_local_record_fix_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_modus_local_record_fix", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(MODUS_LOCAL_RECORD_FIX_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(
            f"Modus local record fix site maps outside EXE: RAM 0x{MODUS_LOCAL_RECORD_FIX_SITE:08X}, file 0x{site_off:X}"
        )

    cur = bytes(exe[site_off:site_off+8])
    replacing_existing_hook = False
    if cur != MODUS_LOCAL_RECORD_FIX_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already:
            replacing_existing_hook = True
        elif not getattr(args, "modus_local_record_fix_force", False):
            raise RuntimeError(
                f"Modus local record fix site mismatch at 0x{site_off:X}: {cur.hex(' ')}; "
                f"expected {MODUS_LOCAL_RECORD_FIX_EXPECT.hex(' ')}. "
                "Use --modus-local-record-fix-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_modus_local_record_fix_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{MODUS_LOCAL_RECORD_FIX_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "record": "BASE+0xAD4 / observed RAM 8012A3BC",
        "description": "Local PLANET/Modus record repair before shared FT4/SPRT builders read the element template.",
    }

def build_stage_tim04_na_template_u_hook(hook_ram, args):
    a = A(hook_ram)
    dash_old_u = int(args.stage_tim04_na_template_old_u) & 0xFF
    dash_new_u = int(args.stage_tim03_na_line_u) & 0xFF
    v = int(args.stage_tim04_na_template_v) & 0xFF
    size = int(args.stage_tim04_na_template_size) & 0xFFFF
    color = int(args.stage_tim04_na_template_color) & 0xFF

    element_u_to_shift = (
        (int(args.stage_tim04_earth_u) & 0xFF, int(args.stage_tim04_earth_x_shift)),
        (int(args.stage_tim04_water_u) & 0xFF, int(args.stage_tim04_water_x_shift)),
        (int(args.stage_tim04_wind_u) & 0xFF,  int(args.stage_tim04_wind_x_shift)),
        (int(args.stage_tim04_fire_u) & 0xFF,  int(args.stage_tim04_fire_x_shift)),
    )

    # Modus screen uses the same STAGE.TIM builder but separate templates.
    # Keep these narrowly guarded, then repair the Object-overlay element words
    # separately in case the shared template returns from Modus partially widened.
    stage_tim04_apply_modus_crest_x_fix(a, args)
    stage_tim04_apply_modus_element_size_fix(a, args)
    stage_tim04_apply_object_overlay_element_restore(a, args, element_u_to_shift)

    # Use v0/v1 only; the original code immediately overwrites both at the
    # return site. No persistent caller state is clobbered.
    a.lbu("v0", 0x000F, "s0")
    a.addiu("v1", "zero", v)
    a.bne("v0", "v1", "orig"); a.nop()

    a.lhu("v0", 0x0008, "s0")
    a.addiu("v1", "zero", size)
    a.bne("v0", "v1", "orig"); a.nop()

    a.lhu("v0", 0x000A, "s0")
    a.bne("v0", "v1", "orig"); a.nop()

    # Match the gray TexBlend packet bytes (64808080) so unrelated 16x16
    # symbols that happen to share the same V/size are not remapped.
    a.lbu("v0", 0x0014, "s0")
    a.addiu("v1", "zero", color)
    a.bne("v0", "v1", "orig"); a.nop()
    a.lbu("v0", 0x0015, "s0")
    a.bne("v0", "v1", "orig"); a.nop()
    a.lbu("v0", 0x0016, "s0")
    a.bne("v0", "v1", "orig"); a.nop()

    a.lbu("v0", 0x000E, "s0")
    a.addiu("v1", "zero", 0x18)
    a.beq("v0", "v1", "case_earth"); a.nop()
    a.addiu("v1", "zero", 0x28)
    a.beq("v0", "v1", "case_water"); a.nop()
    a.addiu("v1", "zero", 0x38)
    a.beq("v0", "v1", "case_wind"); a.nop()
    a.addiu("v1", "zero", 0x48)
    a.beq("v0", "v1", "case_fire"); a.nop()
    a.addiu("v1", "zero", dash_old_u)
    a.beq("v0", "v1", "case_dash"); a.nop()
    a.j("orig"); a.nop()

    a.lab("case_earth")
    stage_tim04_apply_element_x_transition(a, "case_earth", args.stage_tim04_earth_u, args.stage_tim04_earth_x_shift, element_u_to_shift)
    a.addiu("v0", "zero", int(args.stage_tim04_earth_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))  # sb v0,0x0E(s0)
    a.addiu("v0", "zero", int(args.stage_tim04_earth_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))  # sh v0,0x08(s0)
    a.j("orig"); a.nop()

    a.lab("case_water")
    stage_tim04_apply_element_x_transition(a, "case_water", args.stage_tim04_water_u, args.stage_tim04_water_x_shift, element_u_to_shift)
    a.addiu("v0", "zero", int(args.stage_tim04_water_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.stage_tim04_water_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("case_wind")
    stage_tim04_apply_element_x_transition(a, "case_wind", args.stage_tim04_wind_u, args.stage_tim04_wind_x_shift, element_u_to_shift)
    a.addiu("v0", "zero", int(args.stage_tim04_wind_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.stage_tim04_wind_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("case_fire")
    stage_tim04_apply_element_x_transition(a, "case_fire", args.stage_tim04_fire_u, args.stage_tim04_fire_x_shift, element_u_to_shift)
    a.addiu("v0", "zero", int(args.stage_tim04_fire_u) & 0xFF)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.stage_tim04_fire_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))
    a.j("orig"); a.nop()

    a.lab("case_dash")
    stage_tim04_restore_x_for_dash(a, "case_dash", element_u_to_shift)
    a.addiu("v0", "zero", dash_new_u)
    a.word(it(0x28, REG["s0"], REG["v0"], 0x000E))
    a.addiu("v0", "zero", int(args.stage_tim04_dash_w) & 0xFFFF)
    a.word(it(0x29, REG["s0"], REG["v0"], 0x0008))

    a.lab("orig")
    a.word(STAGE_TIM04_NA_TEMPLATE_U_HOOK_EXPECT)  # original: lbu v1,0x0F(s0)
    a.j(STAGE_TIM04_NA_TEMPLATE_U_HOOK_RETURN); a.nop()
    return a.out()


def patch_stage_tim04_na_template_u_hook(exe, load, args, dry_run=False):
    if getattr(args, "disable_stage_tim04_na_template_u_hook", False):
        return {"enabled": False, "status": "disabled"}

    site_off = ram2off(STAGE_TIM04_NA_TEMPLATE_U_HOOK_SITE, load)
    if site_off < 0 or site_off + 4 > len(exe):
        raise RuntimeError(
            f"STAGE TIM04 N/A template hook site maps outside EXE: RAM 0x{STAGE_TIM04_NA_TEMPLATE_U_HOOK_SITE:08X}, file 0x{site_off:X}"
        )

    cur = read32(exe, site_off)
    replacing_existing_hook = False
    if cur != STAGE_TIM04_NA_TEMPLATE_U_HOOK_EXPECT:
        already = ((cur >> 26) == 0x02)
        if already:
            replacing_existing_hook = True
        elif not getattr(args, "stage_tim04_na_template_u_force", False):
            raise RuntimeError(
                f"STAGE TIM04 N/A template hook site mismatch at 0x{site_off:X}: "
                f"found 0x{cur:08X}, expected 0x{STAGE_TIM04_NA_TEMPLATE_U_HOOK_EXPECT:08X}. "
                "Use --stage-tim04-na-template-u-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_stage_tim04_na_template_u_hook(hook_ram, args)

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+4] = w(jt(2, hook_ram))
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else ("replaced_existing_hook" if replacing_existing_hook else "patched"),
        "site_ram": f"0x{STAGE_TIM04_NA_TEMPLATE_U_HOOK_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "old_u": f"0x{int(args.stage_tim04_na_template_old_u) & 0xFF:02X}",
        "new_u": f"0x{int(args.stage_tim03_na_line_u) & 0xFF:02X}",
        "v": f"0x{int(args.stage_tim04_na_template_v) & 0xFF:02X}",
        "size": f"0x{int(args.stage_tim04_na_template_size) & 0xFFFF:04X}",
        "description": "Template-level N/A remap: only 16x16 gray TexBlend templates with U/V=58/C0 are changed to U=75 before source-word construction.",
    }



def planet_info_apply_x_delta(a, delta):
    if int(delta) == 0:
        return
    a.lh("v1", 0x0418, "a0")
    a.addiu("v1", "v1", int(delta))
    a.word(it(0x29, REG["a0"], REG["v1"], 0x0418))  # sh v1,0x418(a0) ; destination X


def planet_info_apply_element_x_transition(a, label, target_u, target_shift, element_u_to_shift):
    """Move the persistent PLANET element template from its current element
    shift to target_shift without accumulating across frames."""
    if not any(int(s) for _, s in element_u_to_shift) and int(target_shift) == 0:
        return
    done = f"{label}_x_done"
    default = f"{label}_x_default"
    a.lbu("v0", 0x0422, "a0")
    a.addiu("v1", "zero", int(target_u) & 0xFF)
    a.beq("v0", "v1", done); a.nop()
    for idx, (u, prev_shift) in enumerate(element_u_to_shift):
        if (int(u) & 0xFF) == (int(target_u) & 0xFF):
            continue
        lab = f"{label}_from_{idx}"
        a.addiu("v1", "zero", int(u) & 0xFF)
        a.beq("v0", "v1", lab); a.nop()
    a.j(default); a.nop()
    for idx, (u, prev_shift) in enumerate(element_u_to_shift):
        if (int(u) & 0xFF) == (int(target_u) & 0xFF):
            continue
        lab = f"{label}_from_{idx}"
        a.lab(lab)
        planet_info_apply_x_delta(a, int(target_shift) - int(prev_shift))
        a.j(done); a.nop()
    a.lab(default)
    planet_info_apply_x_delta(a, int(target_shift))
    a.lab(done)


def planet_info_restore_x_for_dash(a, label, element_u_to_shift):
    """Undo any current element shift when returning to N/A/dash."""
    if not any(int(s) for _, s in element_u_to_shift):
        return
    done = f"{label}_x_done"
    a.lbu("v0", 0x0422, "a0")
    for idx, (u, shift) in enumerate(element_u_to_shift):
        lab = f"{label}_from_{idx}"
        a.addiu("v1", "zero", int(u) & 0xFF)
        a.beq("v0", "v1", lab); a.nop()
    a.j(done); a.nop()
    for idx, (u, shift) in enumerate(element_u_to_shift):
        lab = f"{label}_from_{idx}"
        a.lab(lab)
        planet_info_apply_x_delta(a, -int(shift))
        a.j(done); a.nop()
    a.lab(done)


def planet_info_store_element_sprite(a, label, source_u, width, target_shift=0, element_u_to_shift=(), restore_dash=False):
    a.lab(label)
    if restore_dash:
        planet_info_restore_x_for_dash(a, label, element_u_to_shift)
    else:
        planet_info_apply_element_x_transition(a, label, source_u, target_shift, element_u_to_shift)
    a.addiu("v0", "zero", int(source_u) & 0xFF)
    a.word(it(0x28, REG["a0"], REG["v0"], 0x0422))  # sb v0,0x422(a0) ; source U
    a.addiu("v0", "zero", int(width) & 0xFFFF)
    a.word(it(0x29, REG["a0"], REG["v0"], 0x041C))  # sh v0,0x41C(a0) ; width
    a.j(PLANET_INFO_ELEMENT_HOOK_RETURN); a.nop()


def build_planet_info_element_hook(hook_ram, args):
    a = A(hook_ram)
    # Original block only clobbered v0/v1 and used a0/v1 as the draw-template
    # pointer. At the return point the game reloads a0 from gp+0x9AC, so no
    # register save is needed here.
    a.lbu("v1", 0x12, "s0")      # current element selector: FF none, 1 Earth, 2 Water, 3 Wind, 4 Fire
    a.lw("a0", 0x09AC, "gp")     # draw-template base

    a.addiu("v0", "zero", 0x00FF)
    a.beq("v1", "v0", "case_dash"); a.nop()
    a.addiu("v0", "zero", 0x0001)
    a.beq("v1", "v0", "case_earth"); a.nop()
    a.addiu("v0", "zero", 0x0002)
    a.beq("v1", "v0", "case_water"); a.nop()
    a.addiu("v0", "zero", 0x0003)
    a.beq("v1", "v0", "case_wind"); a.nop()
    a.addiu("v0", "zero", 0x0004)
    a.beq("v1", "v0", "case_fire"); a.nop()
    # Unknown selector: preserve the original behavior as closely as possible
    # by leaving the existing template U/width untouched.
    a.j(PLANET_INFO_ELEMENT_HOOK_RETURN); a.nop()

    element_u_to_shift = (
        (int(args.planet_info_earth_u) & 0xFF, int(args.planet_info_earth_x_shift)),
        (int(args.planet_info_water_u) & 0xFF, int(args.planet_info_water_x_shift)),
        (int(args.planet_info_wind_u) & 0xFF,  int(args.planet_info_wind_x_shift)),
        (int(args.planet_info_fire_u) & 0xFF,  int(args.planet_info_fire_x_shift)),
    )
    planet_info_store_element_sprite(a, "case_dash",  args.planet_info_dash_u,  args.planet_info_dash_w, 0, element_u_to_shift, restore_dash=True)
    planet_info_store_element_sprite(a, "case_earth", args.planet_info_earth_u, args.planet_info_earth_w, args.planet_info_earth_x_shift, element_u_to_shift)
    planet_info_store_element_sprite(a, "case_water", args.planet_info_water_u, args.planet_info_water_w, args.planet_info_water_x_shift, element_u_to_shift)
    planet_info_store_element_sprite(a, "case_wind",  args.planet_info_wind_u,  args.planet_info_wind_w, args.planet_info_wind_x_shift, element_u_to_shift)
    planet_info_store_element_sprite(a, "case_fire",  args.planet_info_fire_u,  args.planet_info_fire_w, args.planet_info_fire_x_shift, element_u_to_shift)
    return a.out()


def patch_planet_info_element_draw(exe, load, args, dry_run=False):
    report = []

    # Do not directly patch PLANET_INFO_PLANET_U_SITE by default. That source-U site is
    # shared by the Garden stage/music-title screen, which must keep U=0x68.
    # The copy-slot split hook changes only the copied PLANET info slot to U=0x86.
    if args.enable_shared_planet_icon_patch and (not args.disable_planet_info_planet_icon_patch):
        report.append(patch_one_main_imm(
            exe,
            name="shared_planet_u_LEGACY",
            offset=PLANET_INFO_PLANET_U_SITE,
            expected_word=PLANET_INFO_PLANET_U_EXPECT,
            new_imm=args.planet_info_planet_u,
            dry_run=dry_run,
            force=args.planet_info_force,
        ))
        if args.planet_info_planet_v is not None:
            report.append(patch_one_main_imm(
                exe,
                name="shared_planet_v_LEGACY",
                offset=PLANET_INFO_PLANET_V_SITE,
                expected_word=PLANET_INFO_PLANET_V_EXPECT,
                new_imm=args.planet_info_planet_v,
                dry_run=dry_run,
                force=args.planet_info_force,
            ))
        if (not args.disable_planet_info_planet_clut_patch) and args.planet_info_planet_clut != 0x01E9:
            report.append(patch_one_main_imm(
                exe,
                name="shared_planet_clut_LEGACY",
                offset=PLANET_INFO_PLANET_CLUT_SITE,
                expected_word=PLANET_INFO_PLANET_CLUT_EXPECT,
                new_imm=args.planet_info_planet_clut,
                dry_run=dry_run,
                force=args.planet_info_force,
            ))
        else:
            report.append("shared_planet_clut_LEGACY: kept original 0x01E9")
    else:
        report.append("shared_planet_u: not patched by PLANET info hook; STAGE TIM03 UV patch may update the shared/stage source; copy-slot hook writes info U=0x%02X to BASE+0x152" % (args.planet_info_planet_u & 0xFF))
        report.append("shared_planet_v: not patched by PLANET info hook; STAGE TIM03 UV patch may update the shared/stage source; copy-slot hook writes info V=0x%02X to BASE+0x153" % (args.planet_info_planet_v & 0xFF))
        report.append("planet_info_planet_clut: kept original 0x01E9")

    site_off = ram2off(PLANET_INFO_ELEMENT_HOOK_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(
            f"PLANET info element hook site maps outside EXE: RAM 0x{PLANET_INFO_ELEMENT_HOOK_SITE:08X}, file 0x{site_off:X}"
        )
    cur = bytes(exe[site_off:site_off+8])
    if cur != PLANET_INFO_ELEMENT_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already and not args.planet_info_force:
            return {
                "enabled": True,
                "status": "already_hooked",
                "site_ram": f"0x{PLANET_INFO_ELEMENT_HOOK_SITE:08X}",
                "site_file": f"0x{site_off:X}",
                "report": report,
            }
        if not args.planet_info_force:
            raise RuntimeError(
                f"PLANET info element hook site mismatch at 0x{site_off:X}: {cur.hex(' ')}; "
                f"expected {PLANET_INFO_ELEMENT_EXPECT.hex(' ')}. Use --planet-info-force only if intentional."
            )

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    code = build_planet_info_element_hook(hook_ram, args)
    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
        exe[hook_off:hook_off+len(code)] = code

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "site_ram": f"0x{PLANET_INFO_ELEMENT_HOOK_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "hook_len": len(code),
        "dash":  {"u": args.planet_info_dash_u,  "w": args.planet_info_dash_w,  "x_shift": 0},
        "earth": {"u": args.planet_info_earth_u, "w": args.planet_info_earth_w, "x_shift": args.planet_info_earth_x_shift},
        "water": {"u": args.planet_info_water_u, "w": args.planet_info_water_w, "x_shift": args.planet_info_water_x_shift},
        "wind":  {"u": args.planet_info_wind_u,  "w": args.planet_info_wind_w,  "x_shift": args.planet_info_wind_x_shift},
        "fire":  {"u": args.planet_info_fire_u,  "w": args.planet_info_fire_w,  "x_shift": args.planet_info_fire_x_shift},
        "planet_u": args.planet_info_planet_u,
        "planet_v": args.planet_info_planet_v,
        "planet_clut": args.planet_info_planet_clut,
        "report": report,
    }


def patch_tree_text_buffer(exe, copy_limit=0x1FA, force=False, dry_run=False):
    if copy_limit < 0xFA:
        raise ValueError("TREE copy limit should be >= 0xFA")
    if copy_limit > 0x7FFF:
        raise ValueError("TREE copy limit must fit signed 16-bit SLTI immediate")
    clear_size = copy_limit + 2
    ra_off = align(0x10 + clear_size, 0x10)
    frame_size = ra_off + 8
    if frame_size > 0x7FFF:
        raise ValueError("computed TREE stack frame too large for addiu immediate")

    observed={}
    for name, off in TREEBUF_PATCH_SITES.items():
        cur=read32(exe,off); observed[name]=cur
        if not force and not treebuf_word_shape_ok(name,cur):
            raise RuntimeError(
                f"TREE text buffer patch site {name} at 0x{off:X} has unexpected word 0x{cur:08X}; "
                f"expected original/compatible shape. Use --tree-buffer-force to patch anyway."
            )

    new_words={
        "frame_sub": itype_word(0x09, TREEBUF_REG["sp"], TREEBUF_REG["sp"], neg_imm(frame_size)),
        "save_ra":   itype_word(0x2B, TREEBUF_REG["sp"], TREEBUF_REG["ra"], ra_off),
        "clear_len": itype_word(0x09, TREEBUF_REG["zero"], TREEBUF_REG["a2"], clear_size),
        "copy_limit":itype_word(0x0A, TREEBUF_REG["a0"], TREEBUF_REG["v0"], copy_limit),
        "load_ra":   itype_word(0x23, TREEBUF_REG["sp"], TREEBUF_REG["ra"], ra_off),
        "frame_add": itype_word(0x09, TREEBUF_REG["sp"], TREEBUF_REG["sp"], frame_size),
    }
    if not dry_run:
        for name, word in new_words.items():
            off=TREEBUF_PATCH_SITES[name]
            exe[off:off+4]=struct.pack("<I", word & 0xffffffff)
    report=[]
    for name in TREEBUF_PATCH_SITES:
        report.append({
            "site":name,
            "file_offset":f"0x{TREEBUF_PATCH_SITES[name]:X}",
            "old_word":f"0x{observed[name]:08X}",
            "new_word":f"0x{new_words[name]:08X}",
            "status":"would_patch" if dry_run else "patched",
            "copy_limit":f"0x{copy_limit:X}",
            "clear_size":f"0x{clear_size:X}",
            "ra_off":f"0x{ra_off:X}",
            "frame_size":f"0x{frame_size:X}",
        })
    info={
        "enabled":True,
        "copy_limit":copy_limit,
        "clear_size":clear_size,
        "ra_off":ra_off,
        "frame_size":frame_size,
        "report":report,
    }
    return info
def ram2off(a,load): return a-load+PSX_HEADER
def off2ram(o,load): return load+o-PSX_HEADER
# Results/Clear Data save-prompt selector-box Y adjustment.
#
# The Results/Clear Data prompt draws "Save clear data?\n                 Yes         No"
# through a special results-screen path. The selected Yes/No grey rectangle uses
# the same hardcoded y=0x32 as some other prompts, but the text baseline on this
# screen sits a few pixels lower. Keep this fix narrow: only the two result/clear
# data paths and only the Save-clear prompt y constants, not the shared Yes/No
# box hook or the later "Proceed without saving?" prompt.
CLEAR_DATA_SELECTOR_Y_SITES = [
    (0x8006CA1C, 0x24050032, "FUN_8006c7d8 Save clear data Yes y"),
    (0x8006CA2C, 0x24050032, "FUN_8006c7d8 Save clear data No y"),
    (0x8006F368, 0x24050032, "FUN_8006f0e8 Save clear data Yes y"),
    (0x8006F374, 0x24050032, "FUN_8006f0e8 Save clear data No y"),
]

def patch_results_clear_data_selector_y(exe, load, *, y_value=0x35, force=False, dry_run=False):
    report=[]
    if y_value < -0x8000 or y_value > 0x7FFF:
        raise ValueError(f"clear-data selector Y out of signed 16-bit range: {y_value}")
    new_word = it(0x09, REG["zero"], REG["a1"], y_value & 0xFFFF)  # addiu a1,zero,y
    for ram, expected, desc in CLEAR_DATA_SELECTOR_Y_SITES:
        off = ram2off(ram, load)
        rec={"description":desc,"ram":f"0x{ram:08X}","file_offset":f"0x{off:X}","expected_word":f"0x{expected:08X}","new_word":f"0x{new_word:08X}","new_y":y_value}
        if off < 0 or off + 4 > len(exe):
            rec["status"]="out_of_range"; report.append(rec); continue
        cur = struct.unpack_from("<I", exe, off)[0]
        rec["old_word"] = f"0x{cur:08X}"
        if cur == new_word:
            rec["status"]="already_patched"; report.append(rec); continue
        if cur != expected and not (force and (cur & 0xFFFF0000) == 0x24050000):
            rec["status"]="unexpected_word"; report.append(rec); continue
        if not dry_run:
            struct.pack_into("<I", exe, off, new_word)
        rec["status"]="would_patch" if dry_run else "patched"
        report.append(rec)
    patched=sum(1 for r in report if r.get("status") in {"patched","would_patch","already_patched"})
    return {"enabled":True,"status":"ok" if patched == len(CLEAR_DATA_SELECTOR_Y_SITES) else "partial","new_y":y_value,"sites":len(CLEAR_DATA_SELECTOR_Y_SITES),"patched_or_already":patched,"report":report}

def sec_for(a):
    for n,s,e in SECTIONS:
        if s<=a<e: return n
    return "before_rdata" if a<SECTIONS[0][1] else "after_bss"
def allfind(data,needle):
    d=bytes(data); out=[]; p=0
    while True:
        i=d.find(needle,p)
        if i<0: return out
        out.append(i); p=i+1

class A:
    """Tiny MIPS assembler used by the injected hook builders.

    The PlayStation's R3000A has a load-delay slot: the instruction immediately
    after lb/lbu/lh/lhu/lw cannot safely consume the loaded register.  Older
    emulators often hide this, but stricter emulators such as DuckStation expose
    it.  To keep every generated hook PS1-safe, this assembler automatically
    inserts a NOP after a load unless the next emitted word is already an
    explicit NOP.  Labels are zero-width, so if a label is placed immediately
    after a load it will point at the inserted delay NOP, which is safe for both
    fall-through and branch targets.
    """
    LOAD_OPS = {0x20, 0x21, 0x23, 0x24, 0x25}  # lb, lh, lw, lbu, lhu

    def __init__(self,base):
        self.base=base; self.items=[]; self.labels={}; self._load_delay_pending=False

    @classmethod
    def _is_load_word(cls,x):
        return ((x >> 26) & 0x3f) in cls.LOAD_OPS

    def _before_instruction(self, next_word=None):
        if not self._load_delay_pending:
            return
        # An explicitly emitted NOP satisfies the delay slot; otherwise insert one.
        if next_word == 0:
            self._load_delay_pending=False
        else:
            self.items.append(("w",0))
            self._load_delay_pending=False

    def lab(self,n): self.items.append(("lab",n))
    def word(self,x):
        x &= 0xffffffff
        self._before_instruction(x)
        self.items.append(("w",x))
        self._load_delay_pending = self._is_load_word(x)
    def nop(self): self.word(0)
    def lbu(self,rt,off,rs): self.word(it(0x24,REG[rs],REG[rt],off))
    def lb(self,rt,off,rs): self.word(it(0x20,REG[rs],REG[rt],off))
    def lhu(self,rt,off,rs): self.word(it(0x25,REG[rs],REG[rt],off))
    def lh(self,rt,off,rs): self.word(it(0x21,REG[rs],REG[rt],off))
    def lw(self,rt,off,rs): self.word(it(0x23,REG[rs],REG[rt],off))
    def sw(self,rt,off,rs): self.word(it(0x2b,REG[rs],REG[rt],off))
    def lui(self,rt,imm): self.word(it(0x0f,0,REG[rt],imm))
    def ori(self,rt,rs,imm): self.word(it(0x0d,REG[rs],REG[rt],imm))
    def addiu(self,rt,rs,imm): self.word(it(0x09,REG[rs],REG[rt],imm))
    def sltiu(self,rt,rs,imm): self.word(it(0x0b,REG[rs],REG[rt],imm))
    def sll(self,rd,rt_,sh): self.word(rt(0,REG[rt_],REG[rd],sh,0))
    def addu(self,rd,rs,rt_): self.word(rt(REG[rs],REG[rt_],REG[rd],0,0x21))
    def subu(self,rd,rs,rt_): self.word(rt(REG[rs],REG[rt_],REG[rd],0,0x23))
    def j(self,t): self._before_instruction(); self.items.append(("j",t))
    def beq(self,rs,rt_,lbl): self._before_instruction(); self.items.append(("beq",rs,rt_,lbl))
    def bne(self,rs,rt_,lbl): self._before_instruction(); self.items.append(("bne",rs,rt_,lbl))
    def out(self):
        pc=self.base
        for itx in self.items:
            if itx[0]=="lab": self.labels[itx[1]]=pc
            else: pc+=4
        b=bytearray(); pc=self.base
        for itx in self.items:
            k=itx[0]
            if k=="lab": continue
            if k=="w": b+=w(itx[1])
            elif k=="j":
                t=itx[1]; t=self.labels[t] if isinstance(t,str) else t; b+=w(jt(2,t))
            elif k in ("beq","bne"):
                _,rs,rt_,lbl=itx; imm=(self.labels[lbl]-(pc+4))>>2; op=4 if k=="beq" else 5
                b+=w(it(op,REG[rs],REG[rt_],imm))
            pc+=4
        return bytes(b)

def xlsx_path(p):
    p=Path(p)
    if p.suffix.lower()==".xlsx": return p
    soffice=shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice: raise RuntimeError("Save workbook as .xlsx or install LibreOffice for .xls conversion.")
    od=Path(tempfile.mkdtemp(prefix="pixy_xls_"))
    subprocess.run([soffice,"--headless","--convert-to","xlsx","--outdir",str(od),str(p)],check=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    m=list(od.glob("*.xlsx"))
    if not m: raise RuntimeError("XLS conversion failed")
    return m[0]
def col(c):
    n=0
    for ch in c.strip().upper(): n=n*26+ord(ch)-64
    return n
def hdr(ws,names):
    want={x.lower() for x in names}
    for c in ws[1]:
        if c.value is not None and str(c.value).strip().lower() in want: return c.column
    return None
def intauto(v):
    if v is None: return None
    s=str(v).strip()
    if not s: return None
    try: return int(s,0)
    except:
        try: return int(float(s))
        except: return None
def hexbytes(v): return bytes(int(x,16) for x in re.findall(r"[0-9A-Fa-f]{2}",str(v or "")))
def norm(s): return str(s).replace("\r\n","\n").replace("\r","\n")
def stripctrl(s):
    s=norm(s); s=re.sub(r"v[0-9A-Za-z]c[0-9A-Za-z]","",s); s=re.sub(r"u[*,(+\-]","",s); s=re.sub(r"f[0-9][0-9A-Fa-f]{6}","",s); return s
def enc_text(s):
    out=bytearray()
    for ch in norm(s):
        if ch=="\n": out.append(0x0a)
        elif 0x20<=ord(ch)<=0x7e: out.append(ord(ch))
        else: out+=ch.encode("cp932",errors="replace")
    return bytes(out)


def to_fullwidth_ascii_text(s):
    """Convert plain ASCII-ish text to full-width Unicode characters before CP932 encoding.

    This is useful for fixed-width CP932 name tables that expect two-byte glyphs.
    Printable ASCII 0x21-0x7E is mapped to U+FF01..U+FF5E; space maps to U+3000.
    Newlines are not expected in name slots and are converted to U+3000. Non-ASCII
    characters are preserved and encoded through CP932.
    """
    out=[]
    for ch in norm(s):
        o=ord(ch)
        if ch in (" ", "\n", "\t"):
            out.append("　")
        elif 0x21 <= o <= 0x7E:
            out.append(chr(0xFF00 + o - 0x20))
        else:
            out.append(ch)
    return "".join(out)

def enc_text_fullwidth_cp932(s):
    return to_fullwidth_ascii_text(s).encode("cp932", errors="replace") + b"\0"

def cp932(ch):
    b=ch.encode("cp932"); return b[0] if len(b)==1 else (b[0]<<8)|b[1]
def fw(ch):
    if ch==" ": return "　"
    return chr(0xff00+ord(ch)-0x20) if 0x21<=ord(ch)<=0x7e else "　"
def ascii_table(mapjson=None):
    t=[cp932("　")]*128
    for i in range(0x20,0x7f): t[i]=cp932(fw(chr(i)))
    for a,b in {"'":"’",'"':"”","`":"’"}.items():
        try: t[ord(a)]=cp932(b)
        except: pass
    if mapjson:
        data=json.loads(Path(mapjson).read_text(encoding="utf-8"))
        for k,v in data.items():
            o=ord(k)
            if isinstance(v,int): t[o]=v
            elif isinstance(v,str) and v.lower().startswith("0x"): t[o]=int(v,16)
            elif isinstance(v,str) and len(v)==1: t[o]=cp932(v)
            else:
                bs=bytes(int(x,16) for x in re.findall(r"[0-9A-Fa-f]{2}",str(v)))
                if len(bs)!=2: raise ValueError(f"Bad map {k}:{v}")
                t[o]=(bs[0]<<8)|bs[1]
    return b"".join(struct.pack("<H",x&0xffff) for x in t)

@dataclass
class Row:
    sheet_row:int; old_ptr:int; old_off:int; orig:bytes; old_len:int; text:str; enc:bytes; new_ptr:int=0; new_off:int=0
@dataclass
class Cand:
    cid:int; row:Row; off:int; ram:int; section:str; cluster:int=-1; csize:int=1; cuniq:int=1; selected:bool=False; reason:str=""

@dataclass(frozen=True)
class DirectMipsHit:
    row:Row
    kind:str
    confidence:str
    lui_off:int
    use_off:int
    gap:int
    lui_reg:int
    dest_reg:int
    old_lui_word:int
    old_use_word:int


def load_rows(xlsx,sheet,pc,bc,tc,pbase,do_strip,origexe):
    wb=load_workbook(xlsx_path(xlsx),data_only=True); ws=wb[sheet]
    pc=col(pc) if pc else hdr(ws,["pointer","Pointer"])
    bc=col(bc) if bc else hdr(ws,["string_bytes","String Bytes","original_bytes","Original Bytes"])
    tc=col(tc) if tc else hdr(ws,["QA Clean Translation","Better Translation","English Translation","english_translation","Clean Translation","Translation","Insertion Text","Final Insertion Text"])
    if not pc or not bc or not tc: raise RuntimeError("Could not find pointer/original-bytes/text column. Use explicit column args.")
    rows=[]
    for r in range(2,ws.max_row+1):
        ptr=intauto(ws.cell(r,pc).value); ob=hexbytes(ws.cell(r,bc).value); val=ws.cell(r,tc).value
        if ptr is None or not ob or val is None: continue
        off=ptr-pbase; oldlen=len(ob)
        if 0<=off+len(ob)<len(origexe) and origexe[off+len(ob)]==0: oldlen+=1
        txt=stripctrl(val) if do_strip else norm(val)
        rows.append(Row(r,ptr,off,ob,oldlen,txt,enc_text(txt)+b"\0"))
    rows.sort(key=lambda x:(x.old_off,x.sheet_row))
    return rows

# v75 renderer metric defaults.  These are intended to use a simple rule:
# every Latin glyph is drawn so its visible left edge lands on the cursor, and
# its advance is visible_width + 1 px.  That gives one blank pixel between any
# two neighboring Latin glyphs while keeping line starts visually aligned.
# v75 tuning: keep the v74 proven metric path, then add final bearing/advance
# locks and a small context-aware pair-kern layer for cases like w→i where
# a single per-character advance cannot satisfy every neighboring pair.
DEFAULT_ADVANCE_7_CHARS="ABCDEFGHKLNOPQRSTUVXYZ02356789pm"
DEFAULT_ADVANCE_6_CHARS="Jabcdefghjknoqrstuvxy"
DEFAULT_ADVANCE_5_CHARS="z"
DEFAULT_ADVANCE_9_CHARS="M4w"
DEFAULT_ADVANCE_8_CHARS=""
DEFAULT_ADVANCE_3_CHARS="Iil"
DEFAULT_ADVANCE_2_CHARS=""
DEFAULT_ADVANCE_10_CHARS="W"
DEFAULT_ADVANCE_4_CHARS="1"
# Only punctuation that we intentionally give compact English metrics.
# Other printable ASCII falls back to --ascii-advance, now defaulting to the
# original/Japanese full-width spacing.
DEFAULT_PUNCT_CHARS="""!\"',-.?:;"""
EXPECTED_ALNUM="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
# v75 final locked advances. These are applied after grouped/default
# advance arguments, built-in deltas, and manual --advance-override, so older
# command lines cannot silently revert the in-game spacing test results.
#
# Notes from v74 test pass:
#   p: one pixel too loose -> 6
#   R: needs one more right-side pixel -> 8
#   M: one pixel too much on the right -> 8
#   I: needs one more right-side pixel before m -> 4
#   period: one pixel too much advance -> 2
#   u: needs one more right-side pixel -> 7
#   P: shifted left, with matching advance reduction so its right gap stays sane
DEFAULT_V75_FINAL_ADVANCE_OVERRIDES=[("I",4),("il",3),("p",6),("m",7),("z",5),("R",8),("M",8),(".",2),("u",7),("P",6)]

# v75 final draw-bearing locks. Positive values draw the glyph to the right;
# negative values draw it left. The renderer compensates GP_HADV so the next
# cursor position is still controlled by the advance table.
DEFAULT_V75_FINAL_XOFF_OVERRIDES={"P":-1, ",":1}

# Context-aware pair kerning. Pair deltas adjust the current glyph's advance
# only when the next glyph matches. This is intentionally small and explicit:
# most font tuning stays per-glyph, while awkward pairs like w+i can tighten
# without making w too tight before every other letter.
DEFAULT_V75_PAIR_KERNS=[("wi",-1)]

def iter_v74_final_advance_map(args):
    if getattr(args, "disable_v74_final_spacing_lock", False) or getattr(args, "disable_v75_final_spacing_lock", False):
        return {}
    out={}
    for chars,adv in DEFAULT_V75_FINAL_ADVANCE_OVERRIDES:
        for ch in chars:
            out[ch]=int(adv)
    return out

def iter_v75_final_xoff_map(args):
    if getattr(args, "disable_v75_final_xoff_lock", False):
        return {}
    return dict(DEFAULT_V75_FINAL_XOFF_OVERRIDES)

def iter_v75_pair_kern_map(args):
    if getattr(args, "disable_v75_pair_kern", False):
        return {}
    # With v77 TIM-derived metrics active, the base model is already
    # visible_width + tracking, so default pair kerning is disabled. Manual
    # --pair-kern entries still apply below if the user wants extra polish.
    font_active = False
    try:
        font_active = bool(load_v76_font_tim_metrics(args).get('enabled'))
    except Exception:
        font_active = False
    out={} if font_active else {pair:int(delta) for pair,delta in DEFAULT_V75_PAIR_KERNS}
    for item in getattr(args, "pair_kern", []) or []:
        pair, delta = item
        pair = str(pair)
        if len(pair) != 2:
            raise ValueError(f"Pair kern key must be exactly 2 ASCII characters, got {pair!r}")
        if any(ord(ch) >= 128 for ch in pair):
            raise ValueError(f"Pair kern key must be ASCII, got {pair!r}")
        delta = int(delta)
        if not -128 <= delta <= 127:
            raise ValueError(f"Pair kern delta must fit signed byte, got {delta} for {pair!r}")
        out[pair]=delta
    # A zero delta is a convenient command-line way to neutralize a default pair.
    return {pair:delta for pair,delta in out.items() if delta}

def cp932_fullwidth_ascii_bytes_for_pair(ch):
    try:
        b = fw(ch).encode("cp932")
    except Exception:
        return None
    return b if len(b) == 2 else None

# v76 FONT11Z0.TIM metric support.
#
# This adds an optional build-time path to derive per-glyph draw bearings and
# advances from the edited font TIM instead of relying only on hand-tuned
# grouped advances. The intended model is:
#   xoff    = -first_visible_column
#   advance = visible_width + tracking
#
# FONT11Z0 stores four CP932 glyphs per 11x11 cell using the 4 bits of each
# 4bpp pixel. Horizontally cells are 11 px plus one separator column (step X=12),
# but vertically the TIM is 253 px high = 23 rows * 11 px, so step Y is 11.
# v78 incorrectly used step Y=12, which mixed the next row into some lowercase
# metrics and caused intermittent huge gaps. For full-width ASCII alphanumerics, v80 uses the known CP932/modulo-4 layout directly for printable ASCII.
# Alphanumerics use the compact FONT11Z0 full-width ASCII sequence; punctuation
# uses the normal CP932 table position matching ascii_table(), including curly
# quote mappings for raw ASCII quotes. A JSON map can still override/add custom
# glyph locations.

def _v76_try_int(v):
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        return int(v, 0)
    raise TypeError(f"Unsupported integer-like value: {v!r}")

def resolve_font_tim_path(args):
    if getattr(args, '_resolved_font_tim_path', None) is not None:
        return args._resolved_font_tim_path
    cand=[]
    if getattr(args, 'font_tim', None):
        cand.append(Path(args.font_tim))
    cand.append(Path('FONT11Z0.TIM'))
    try:
        cand.append(Path(__file__).resolve().with_name('FONT11Z0.TIM'))
    except Exception:
        pass
    out=None
    for p in cand:
        try:
            if p and p.exists() and p.is_file():
                out=p.resolve()
                break
        except Exception:
            continue
    args._resolved_font_tim_path=out
    return out

def v77_cp932_ascii_rank(ch):
    """Return normalized FONT11Z0 ASCII rank for full-width CP932 alphanumerics.

    The atlas packs 0-9, A-Z, a-z contiguously. Each packed character advances
    one bitplane slot; four slots share one 11x11 cell.
    """
    if not isinstance(ch, str) or len(ch) != 1:
        return None
    try:
        b = fw(ch).encode("cp932")
    except Exception:
        return None
    if len(b) != 2 or b[0] != 0x82:
        return None
    trail = b[1]
    if "0" <= ch <= "9":
        return trail - 0x4F
    if "A" <= ch <= "Z":
        return 10 + (trail - 0x60)
    if "a" <= ch <= "z":
        return 36 + (trail - 0x81)
    return None


def v80_metric_source_char_for_ascii(ch):
    # Mirror ascii_table() so the metric matches the glyph actually drawn.
    if ch == "'" or ch == "`":
        return "’"
    if ch == '\"':
        return "”"
    if ch == " ":
        return "　"
    return fw(ch)


def v80_normal_cp932_pos_for_char(ch):
    try:
        b = v80_metric_source_char_for_ascii(ch).encode("cp932")
    except Exception:
        return None
    if len(b) != 2:
        return None
    lead, trail = b[0], b[1]
    if not (0x81 <= lead <= 0x9F or 0xE0 <= lead <= 0xEF):
        return None
    if 0x40 <= trail <= 0x7E:
        ti = trail - 0x40
    elif 0x80 <= trail <= 0xFC:
        ti = 63 + (trail - 0x80)
    else:
        return None
    # FONT11Z0 packs normal CP932 order four bitplanes per 11x11 cell.
    # This works for the 0x81 punctuation/quote block. Alphanumerics are handled
    # separately because the edited FONT11Z0 ASCII sequence is compacted.
    idx = (lead - 0x81) * 188 + ti
    cell = idx // 4
    row = cell // 21
    col = cell % 21
    plane = idx % 4
    return (col,row,plane)


def v77_builtin_font_ascii_map(args):
    # Base slot for ASCII '0' in FONT11Z0.TIM: row 1, col 15, bitplane 3.
    # Absolute bitplane slot = row*21*4 + col*4 + plane = 147.
    base_abs = int(getattr(args, "font_ascii_base_abs", 147))
    out={}

    # Compact full-width alphanumeric sequence: 0-9, A-Z, a-z.
    for ch in EXPECTED_ALNUM:
        rank = v77_cp932_ascii_rank(ch)
        if rank is None:
            continue
        pos = base_abs + rank
        row = pos // (21*4)
        rem = pos % (21*4)
        col = rem // 4
        plane = rem % 4
        out[ch]=(col,row,plane)

    # Punctuation/symbols: derive from the actual CP932 glyph that ascii_table()
    # draws. Space intentionally remains controlled by --space-advance.
    for code in range(0x21,0x7F):
        ch = chr(code)
        if ch in out:
            continue
        pos = v80_normal_cp932_pos_for_char(ch)
        if pos is not None:
            out[ch]=pos
    return out


def load_font_map_json(args):
    if hasattr(args, '_font_map_cache'):
        return args._font_map_cache
    # Start with the built-in CP932 modulo-4 full-width ASCII map. JSON, if
    # supplied, can override entries or add punctuation/custom glyphs.
    mp=v77_builtin_font_ascii_map(args)
    src='built-in CP932 modulo-4 printable ASCII map'
    map_path = getattr(args, 'font_map_json', None)
    if map_path:
        p=Path(map_path)
    else:
        p=None
        for q in (Path('FONT11Z0_ascii_map.json'), Path('font_ascii_map.json')):
            if q.exists() and q.is_file():
                p=q; break
        if p is None:
            try:
                for qname in ('FONT11Z0_ascii_map.json','font_ascii_map.json'):
                    q=Path(__file__).resolve().with_name(qname)
                    if q.exists() and q.is_file():
                        p=q; break
            except Exception:
                pass
    if p and p.exists():
        src=str(p)
        data=json.loads(p.read_text(encoding='utf-8'))
        for ch, spec in data.items():
            if not isinstance(ch,str) or len(ch)!=1 or ord(ch)>=128:
                raise ValueError(f"Font map key must be a single ASCII character, got {ch!r}")
            plane=None
            if isinstance(spec, (list, tuple)) and len(spec) >= 2:
                col=_v76_try_int(spec[0]); row=_v76_try_int(spec[1])
                if len(spec) >= 3 and spec[2] is not None:
                    plane=_v76_try_int(spec[2])
            elif isinstance(spec, dict):
                col=_v76_try_int(spec.get('col', spec.get('x', spec.get('cell_x'))))
                row=_v76_try_int(spec.get('row', spec.get('y', spec.get('cell_y'))))
                if any(k in spec for k in ('plane','bit','clut','mod')):
                    plane=_v76_try_int(spec.get('plane', spec.get('bit', spec.get('clut', spec.get('mod')))))
            else:
                raise ValueError(f"Bad font map entry for {ch!r}: {spec!r}")
            if plane is None:
                mp[ch]=(col,row)
            else:
                if not 0 <= int(plane) <= 3:
                    raise ValueError(f"Font map plane for {ch!r} must be 0..3, got {plane}")
                mp[ch]=(col,row,int(plane))
    args._font_map_cache=(mp,src)
    return args._font_map_cache

def decode_tim_4bpp_indices(path):
    data=Path(path).read_bytes()
    if len(data) < 20:
        raise ValueError(f"TIM file too small: {path}")
    magic, flags = struct.unpack_from('<II', data, 0)
    if magic != 0x10:
        raise ValueError(f"Not a TIM file: {path}")
    off=8
    if flags & 0x08:
        clut_len, = struct.unpack_from('<I', data, off)
        off += clut_len
    img_len, = struct.unpack_from('<I', data, off)
    x, y, w_words, h = struct.unpack_from('<HHHH', data, off+4)
    px = data[off+12:off+img_len]
    width = w_words * 4
    rows=[]
    idx=0
    for _ in range(h):
        row=[]
        for _xw in range(w_words):
            b0=px[idx]; b1=px[idx+1]; idx += 2
            row.extend((b0 & 0x0F, (b0 >> 4) & 0x0F, b1 & 0x0F, (b1 >> 4) & 0x0F))
        rows.append(row)
    return {'width': width, 'height': h, 'rows': rows, 'tim_x': x, 'tim_y': y}

def compute_tim_cell_metrics(rows, cell_x, cell_y, cell_w, cell_h, occupied_zero=False, plane=None):
    # Include any visible pixel in the selected bitplane. For FONT11Z0.TIM,
    # four glyphs share each cell: pixel index bit0/bit1/bit2/bit3 represent
    # planes 0..3. Selecting the correct plane is what makes the metrics match
    # the actual CP932 character rather than the union of all four glyphs.
    mask = None if plane is None else (1 << int(plane))
    visible_cols=[]
    for dx in range(cell_w):
        x=cell_x+dx; col_has=False
        for y in range(cell_y, cell_y+cell_h):
            try:
                v=rows[y][x]
            except Exception:
                continue
            if mask is None:
                occupied = (v != 0) or occupied_zero
            else:
                occupied = bool(v & mask)
            if occupied:
                col_has=True; break
        visible_cols.append(col_has)
    first=None; last=None
    for i,flag in enumerate(visible_cols):
        if flag: first=i; break
    for i in range(len(visible_cols)-1, -1, -1):
        if visible_cols[i]: last=i; break
    if first is None or last is None:
        return {'empty': True, 'first': 0, 'last': -1, 'width': 0}
    return {'empty': False, 'first': first, 'last': last, 'width': last-first+1}

def load_v76_font_tim_metrics(args):
    if hasattr(args, '_v76_font_tim_metric_cache'):
        return args._v76_font_tim_metric_cache
    info={'enabled': False,'tim_path': None,'map_path': None,'metrics': {},'warning': None,'tracking': int(getattr(args,'font_tracking',1))}
    if getattr(args, 'disable_font_tim_metrics', False):
        args._v76_font_tim_metric_cache=info
        return info
    tim_path = resolve_font_tim_path(args)
    info['tim_path'] = str(tim_path) if tim_path else None
    if not tim_path:
        info['warning'] = 'FONT TIM not found; using built-in spacing tables only'
        args._v76_font_tim_metric_cache=info
        return info
    font_map, map_src = load_font_map_json(args)
    info['map_path'] = map_src
    if not font_map:
        info['warning'] = 'FONT TIM found; using built-in CP932 modulo-4 printable ASCII map'
        args._v76_font_tim_metric_cache=info
        return info
    tim = decode_tim_4bpp_indices(tim_path)
    rows = tim['rows']
    cw=int(getattr(args,'font_cell_width',11)); ch=int(getattr(args,'font_cell_height',11))
    sx=int(getattr(args,'font_cell_step_x',12)); sy=int(getattr(args,'font_cell_step_y',11))
    tracking=int(getattr(args,'font_tracking',1)); occupied_zero=bool(getattr(args,'font_zero_is_occupied',False))
    metrics={}
    for ch_ascii,spec in font_map.items():
        if len(spec) >= 3:
            col,row,plane = spec[0],spec[1],spec[2]
        else:
            col,row = spec[0],spec[1]
            plane = None
        m=compute_tim_cell_metrics(rows, int(col)*sx, int(row)*sy, cw, ch, occupied_zero=occupied_zero, plane=plane)
        if m['empty']:
            continue
        metrics[ch_ascii]={'advance': max(0, min(255, int(m['width']) + tracking)), 'xoff': max(-128, min(127, -int(m['first']))), 'first': int(m['first']), 'last': int(m['last']), 'width': int(m['width']), 'cell_col': int(col), 'cell_row': int(row), 'plane': None if plane is None else int(plane)}
    info['enabled']=bool(metrics); info['metrics']=metrics
    if not metrics:
        info['warning'] = 'FONT TIM metrics mode was requested, but no mapped glyphs produced visible bounds; using built-in spacing tables only'
    args._v76_font_tim_metric_cache=info
    return info

def iter_v76_font_metric_map(args):
    return load_v76_font_tim_metrics(args).get('metrics', {})

def validate_advance_groups(args):
    groups=[
        (7,args.advance7_chars),
        (6,args.advance6_chars),
        (5,args.advance5_chars),
        (9,args.advance9_chars),
        (8,args.advance8_chars),
        (3,args.advance3_chars),
        (2,args.advance2_chars),
        (10,args.advance10_chars),
        (4,args.advance4_chars),
    ]
    seen={}; dup=[]
    for adv,chars in groups:
        for ch in chars:
            if ch in seen: dup.append((ch,seen[ch],adv))
            else: seen[ch]=adv
    final_map=iter_v74_final_advance_map(args)
    missing=[ch for ch in EXPECTED_ALNUM if ch not in seen and ch not in final_map]
    extra=[ch for ch in seen if ch not in EXPECTED_ALNUM]
    if dup:
        msg=", ".join(f"{ch!r} in {a} and {b}" for ch,a,b in dup)
        raise ValueError(f"Advance character appears in more than one alphanumeric group: {msg}")
    if missing:
        raise ValueError("Advance groups are missing alphanumeric characters: "+"".join(missing))
    if extra:
        raise ValueError("Advance groups contain non-alphanumeric characters; put punctuation in --punct-chars instead: "+"".join(extra))
    return seen

def build_advance_table(args):
    # Table indexed by raw ASCII byte, values are cursor advances.
    validate_advance_groups(args)
    tbl=[args.ascii_advance & 0xff]*128
    for i in range(128):
        ch=chr(i)
        if 0x20 <= i <= 0x7e and ch in args.punct_chars:
            tbl[i]=args.punct_advance & 0xff
    tbl[0x20]=args.space_advance & 0xff
    for adv,chars in [(7,args.advance7_chars),(6,args.advance6_chars),(5,args.advance5_chars),(9,args.advance9_chars),(8,args.advance8_chars),(3,args.advance3_chars),(2,args.advance2_chars),(10,args.advance10_chars),(4,args.advance4_chars)]:
        for ch in chars:
            o=ord(ch)
            if not (0 <= o < 128): raise ValueError(f"Non-ASCII advance-map character: {ch!r}")
            tbl[o]=adv & 0xff
    # Renderer/metrics tweaks. These apply to raw ASCII and mapped full-width
    # CP932 Latin glyphs because both paths use this ASCII-indexed advance table.
    #
    # The old capital-R-only tweak is retained as an optional compatibility knob,
    # but its default is now 0. By default R receives only the same A-Z delta as
    # every other capital letter.
    r=ord("R")
    tbl[r]=max(0, min(255, tbl[r] + args.capital_r_advance_delta))

    # Final requested deltas, applied after grouped/default advances.
    # Defaults: all capitals -1, hyphen +3, slash +1.
    for ch in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        i=ord(ch)
        tbl[i]=max(0, min(255, tbl[i] + args.uppercase_advance_delta))
    for ch,delta in (("-",args.hyphen_advance_delta),("/",args.slash_advance_delta)):
        i=ord(ch)
        tbl[i]=max(0, min(255, tbl[i] + delta))

    # v71 wide punctuation defaults.  Parentheses and slash are drawn from
    # relatively wide full-width glyphs; give them explicit fixed advances so
    # they do not collapse like comma/period punctuation.
    if getattr(args, "paren_advance", None) is not None:
        adv = max(0, min(255, int(args.paren_advance)))
        for ch in "()":
            tbl[ord(ch)] = adv
    slash_fixed = getattr(args, "slash_fixed_advance", None)
    if slash_fixed is not None and int(slash_fixed) >= 0:
        tbl[ord("/")] = max(0, min(255, int(slash_fixed)))

    # User advance overrides. Repeatable:
    #   --advance-override CHARS PIXELS
    # This is absolute, not a delta.
    for item in getattr(args, "advance_override", []) or []:
        chars, adv = item
        adv = int(adv)
        if not 0 <= adv <= 255:
            raise ValueError(f"Advance override {adv} for {chars!r} must be 0..255")
        for ch in chars:
            o=ord(ch)
            if not (0 <= o < 128):
                raise ValueError(f"Non-ASCII advance override character: {ch!r}")
            tbl[o]=adv & 0xff

    # v74/v75 final locked metric fix. This intentionally happens after grouped
    # defaults, built-in deltas, and user --advance-override so old command
    # lines cannot silently undo the requested metric corrections.
    for ch,adv in iter_v74_final_advance_map(args).items():
        o=ord(ch)
        if not (0 <= o < 128):
            raise ValueError(f"Non-ASCII v74 final advance character: {ch!r}")
        tbl[o]=max(0, min(255, int(adv)))

    # v76 optional FONT11Z0.TIM-derived metrics. If a FONT TIM and a matching
    # ASCII glyph map JSON are supplied, these become the final per-glyph
    # advances. This is the consistent "visible_width + tracking" path.
    for ch,m in iter_v76_font_metric_map(args).items():
        o=ord(ch)
        if not (0 <= o < 128):
            raise ValueError(f"Non-ASCII v76 font metric character: {ch!r}")
        tbl[o]=max(0, min(255, int(m['advance'])))
    return bytes(tbl)

def encode_signed_byte(v):
    if not -128 <= v <= 127:
        raise ValueError(f"X offset {v} does not fit signed byte")
    return v & 0xff

def build_x_offset_table(args):
    # Table indexed by ASCII byte. Values are signed pixel offsets loaded via LB.
    # Negative offset moves the glyph left; positive moves it right. The hook
    # compensates GP_HADV so the next glyph's cursor position remains unchanged.
    #
    # Default groups are applied first. --draw-shift overrides are applied last,
    # so you can tune any character without needing a new patcher version:
    #
    #   --draw-shift "ABC" -1
    #   --draw-shift "A1le" 2
    #   --draw-shift "'" 3
    #
    tbl=[0]*128
    owners={}

    def apply_chars(chars, off, label, override=False):
        if chars is None:
            return
        off = int(off)
        enc = encode_signed_byte(off)
        for ch in chars:
            o=ord(ch)
            if not 0 <= o < 128:
                raise ValueError(f"Non-ASCII character in X-offset group {label}: {ch!r}")
            if not override and ch in owners and owners[ch][0] != off:
                raise ValueError(
                    f"Character {ch!r} appears in multiple default X-offset groups: "
                    f"{owners[ch][1]}={owners[ch][0]} and {label}={off}. "
                    f"Use --draw-shift to intentionally override."
                )
            owners[ch]=(off,label)
            tbl[o]=enc

    # Built-in/default groups.
    apply_chars(args.xoff_left2_chars, -2, "--xoff-left2-chars")
    apply_chars(args.xoff_left1_chars, -1, "--xoff-left1-chars")
    apply_chars(args.xoff_right3_chars, 3, "--xoff-right3-chars")
    apply_chars(args.xoff_right2_chars, 2, "--xoff-right2-chars")
    apply_chars(args.xoff_right1_chars, 1, "--xoff-right1-chars")

    # User overrides. Repeatable:
    #   --draw-shift CHARS PIXELS
    # Values may be negative, zero, or positive; range must fit signed byte.
    for item in getattr(args, "draw_shift", []) or []:
        chars, off = item
        apply_chars(chars, int(off), "--draw-shift", override=True)

    # v75 final bearing locks: keep these after old/default groups and
    # user draw-shift so old batch files cannot cancel the latest tested fix.
    for ch, off in iter_v75_final_xoff_map(args).items():
        apply_chars(ch, int(off), "v75-final-xoff", override=True)

    # v76 optional FONT11Z0.TIM-derived draw bearings. These are applied last
    # so the glyph's visible left edge lands on the cursor when a mapped TIM
    # metric is available.
    for ch,m in iter_v76_font_metric_map(args).items():
        apply_chars(ch, int(m['xoff']), "v76-font-tim-xoff", override=True)

    return bytes(tbl)

def fullwidth_punct_pairs(punct_chars):
    # Return (cp932_word, ascii_index) pairs for full-width punctuation variants.
    # Used so two-byte full-width punctuation uses the same advance as one-byte punctuation.
    pairs=[]; seen=set()
    for ch in punct_chars:
        if not (0x21 <= ord(ch) <= 0x7e):
            continue
        try:
            code=cp932(fw(ch))
        except Exception:
            continue
        if code not in seen:
            seen.add(code); pairs.append((code,ord(ch)))
    # Full-width/ideographic space maps to ASCII space.
    pairs.insert(0,(cp932("　"),0x20))
    return pairs

def load_adv_from_index(a,adv_table,index_reg="v0",tmp="t1"):
    a.lui(tmp,hi(adv_table)); a.addiu(tmp,tmp,s16(lo(adv_table))); a.addu(tmp,tmp,index_reg); a.lbu(tmp,0,tmp)

def load_xoff_from_index(a,xoff_table,index_reg="v0",tmp="t3"):
    a.lui(tmp,hi(xoff_table)); a.addiu(tmp,tmp,s16(lo(xoff_table))); a.addu(tmp,tmp,index_reg); a.lb(tmp,0,tmp)

def emit_cp932_word_to_a0(a):
    # a0 = (v0 << 8) + v1, matching the CP932 word convention used by ascii_table().
    a.sll("a0","v0",8); a.addu("a0","a0","v1")

def emit_v75_pair_kern_onebyte(a, args):
    # Current ASCII index is v0. Next raw byte is v1 (= *(s0+1)).
    # t1 holds current advance. t0/t4 are scratch here.
    for n,(pair,delta) in enumerate(iter_v75_pair_kern_map(args).items()):
        left,right = pair[0], pair[1]
        nxt=f"v75_pair1_next_{n}"
        apply=f"v75_pair1_apply_{n}"
        a.ori("t0","zero",ord(left)); a.bne("v0","t0",nxt); a.nop()
        a.ori("t0","zero",ord(right)); a.beq("v1","t0",apply); a.nop()
        fb=cp932_fullwidth_ascii_bytes_for_pair(right)
        if fb:
            a.ori("t0","zero",fb[0]); a.bne("v1","t0",nxt); a.nop()
            a.lbu("t4",2,"s0")
            a.ori("t0","zero",fb[1]); a.bne("t4","t0",nxt); a.nop()
        else:
            a.j(nxt); a.nop()
        a.lab(apply)
        a.addiu("t1","t1",int(delta))
        a.lab(nxt)

def emit_v75_pair_kern_twobyte(a, args):
    # Current ASCII-equivalent index is t2. Next raw byte starts at *(s0+2).
    # t1 holds current advance. t0/t4 are scratch here.
    for n,(pair,delta) in enumerate(iter_v75_pair_kern_map(args).items()):
        left,right = pair[0], pair[1]
        nxt=f"v75_pair2_next_{n}"
        apply=f"v75_pair2_apply_{n}"
        a.ori("t0","zero",ord(left)); a.bne("t2","t0",nxt); a.nop()
        a.lbu("t4",2,"s0")
        a.ori("t0","zero",ord(right)); a.beq("t4","t0",apply); a.nop()
        fb=cp932_fullwidth_ascii_bytes_for_pair(right)
        if fb:
            a.ori("t0","zero",fb[0]); a.bne("t4","t0",nxt); a.nop()
            a.lbu("t4",3,"s0")
            a.ori("t0","zero",fb[1]); a.bne("t4","t0",nxt); a.nop()
        else:
            a.j(nxt); a.nop()
        a.lab(apply)
        a.addiu("t1","t1",int(delta))
        a.lab(nxt)

def build_stub(stub,table,adv_table,xoff_table,args):
    a=A(stub)
    a.lbu("v0",0,"s0"); a.beq("v0","zero","ret"); a.nop()
    a.ori("t0","zero",0x0a); a.beq("v0","t0","lf"); a.nop()
    a.sltiu("t0","v0",0x20); a.bne("t0","zero","fallback"); a.nop()
    # Printable ASCII continues through the one-byte map. Bytes >= 0x7F are handled
    # directly as two-byte CP932 glyphs where possible, so their cursor advance can
    # match the one-byte table instead of using the game's original full-width spacing.
    a.sltiu("t0","v0",0x7f); a.beq("t0","zero","db_try"); a.nop()
    a.lbu("v1",1,"s0")
    a.ori("t0","zero",0x5c); a.bne("v0","t0","legacy" if not args.no_preserve_legacy_controls else "map"); a.nop()
    a.ori("t1","zero",0x6e); a.beq("v1","t1","fallback"); a.nop()
    if not args.no_preserve_legacy_controls:
        a.lab("legacy")
        for ch in (0x63,0x76,0x66,0x25):
            a.ori("t0","zero",ch); a.beq("v0","t0","checkdig"); a.nop()
        if args.preserve_u_controls:
            a.ori("t0","zero",0x75); a.bne("v0","t0","map"); a.nop()
            for ch in (0x2a,0x2c,0x28,0x2b,0x2d):
                a.ori("t1","zero",ch); a.beq("v1","t1","fallback"); a.nop()
            a.j("map"); a.nop()
        a.lab("checkdig"); a.addiu("t1","v1",-0x30); a.sltiu("t1","t1",10); a.bne("t1","zero","fallback"); a.nop()
    a.lab("map")
    a.sll("t0","v0",1); a.lui("a0",hi(table)); a.addiu("a0","a0",s16(lo(table))); a.addu("a0","a0","t0"); a.lhu("a0",0,"a0")
    load_adv_from_index(a,adv_table,"v0","t1")
    load_xoff_from_index(a,xoff_table,"v0","t3")
    emit_v75_pair_kern_onebyte(a,args)
    a.j("store1"); a.nop()

    # Two-byte CP932 direct path. Full-width Latin letters/digits/punctuation use
    # the same advance as the one-byte equivalent. Non-Latin/Japanese CP932
    # falls back to the game's original two-byte parser, preserving the original
    # Japanese/full-width cursor spacing instead of forcing English spacing.
    a.lab("db_try")
    a.lbu("v1",1,"s0")

    # v94: PlayStation controller button symbols used in tutorial text.
    # These are CP932 0x81-block glyphs in FONT11Z0.TIM. The original
    # non-Latin fallback draws them but does not give them the compact English
    # advance path, which can make the next glyph overlap. Handle them here
    # exactly like mapped two-byte glyphs: keep a0 as the real CP932 glyph code,
    # then provide a measured FONT11Z0 advance and draw-bearing. Metrics were
    # derived from the current FONT11Z0.TIM with tracking=1:
    #   × 817E width 7 first 2 -> advance 8,  xoff -2
    #   ○ 819B width 10 first 0 -> advance 11, xoff 0
    #   △ 81A2 width 11 first 0 -> advance 12, xoff 0
    #   □ 81A0 width 9 first 1 -> advance 10, xoff -1
    for n,(code,adv,xoff) in enumerate(((0x817E,8,-2),(0x819B,11,0),(0x81A2,12,0),(0x81A0,10,-1))):
        hi_b=(code>>8)&0xff; lo_b=code&0xff
        nxt=f"db_button_next_{n}"
        a.ori("t0","zero",hi_b); a.bne("v0","t0",nxt); a.nop()
        a.ori("t0","zero",lo_b); a.bne("v1","t0",nxt); a.nop()
        emit_cp932_word_to_a0(a)
        a.ori("t1","zero",int(adv)&0xff)
        if int(xoff) < 0:
            a.addiu("t3","zero",int(xoff))
        else:
            a.ori("t3","zero",int(xoff)&0xff)
        a.j("store2"); a.nop()
        a.lab(nxt)

    # Map exact full-width punctuation/space pairs to their ASCII advance entries.
    for n,(code,ascii_idx) in enumerate(fullwidth_punct_pairs(args.punct_chars)):
        hi_b=(code>>8)&0xff; lo_b=code&0xff
        nxt=f"db_punct_next_{n}"
        a.ori("t0","zero",hi_b); a.bne("v0","t0",nxt); a.nop()
        a.ori("t0","zero",lo_b); a.bne("v1","t0",nxt); a.nop()
        a.ori("t2","zero",ascii_idx); a.j("db_mapped"); a.nop()
        a.lab(nxt)
    # Full-width digits/uppercase/lowercase in CP932 all live under lead byte 0x82.
    a.ori("t0","zero",0x82); a.bne("v0","t0","db_default"); a.nop()
    # digits: 824F..8258 -> ASCII 0x30..0x39
    a.sltiu("t0","v1",0x4f); a.bne("t0","zero","db_check_upper"); a.nop()
    a.sltiu("t0","v1",0x59); a.beq("t0","zero","db_check_upper"); a.nop()
    a.addiu("t2","v1",0x30-0x4f); a.j("db_mapped"); a.nop()
    a.lab("db_check_upper")
    # uppercase: 8260..8279 -> ASCII 0x41..0x5A
    a.sltiu("t0","v1",0x60); a.bne("t0","zero","db_check_lower"); a.nop()
    a.sltiu("t0","v1",0x7a); a.beq("t0","zero","db_check_lower"); a.nop()
    a.addiu("t2","v1",0x41-0x60); a.j("db_mapped"); a.nop()
    a.lab("db_check_lower")
    # lowercase: 8281..829A -> ASCII 0x61..0x7A
    a.sltiu("t0","v1",0x81); a.bne("t0","zero","db_default"); a.nop()
    a.sltiu("t0","v1",0x9b); a.beq("t0","zero","db_default"); a.nop()
    a.addiu("t2","v1",0x61-0x81); a.j("db_mapped"); a.nop()
    a.lab("db_mapped")
    emit_cp932_word_to_a0(a); load_adv_from_index(a,adv_table,"t2","t1"); load_xoff_from_index(a,xoff_table,"t2","t3"); emit_v75_pair_kern_twobyte(a,args); a.j("store2"); a.nop()
    a.lab("db_default")
    # Non-Latin CP932 / Japanese text should keep the game's original two-byte
    # parser and original full-width spacing. The fallback path reloads v0/v1
    # and jumps back into the original code path.
    a.j("fallback"); a.nop()

    # Store paths. t1 = normal advance, t3 = signed draw X offset.
    # If t3 != 0, temporarily shift GP_X for this glyph, then compensate
    # GP_HADV by subtracting the offset. This preserves the next glyph cursor:
    #   final_x = (old_x + xoff) + (advance - xoff) = old_x + advance
    a.lab("store1")
    a.beq("t3","zero","store1_no_xoff"); a.nop()
    a.lw("t0",GP_X,"gp"); a.addu("t0","t0","t3"); a.sw("t0",GP_X,"gp"); a.subu("t1","t1","t3")
    a.lab("store1_no_xoff"); a.sw("t1",GP_HADV,"gp"); a.addiu("s0","s0",1); a.j(GLYPH); a.nop()
    a.lab("store2")
    a.beq("t3","zero","store2_no_xoff"); a.nop()
    a.lw("t0",GP_X,"gp"); a.addu("t0","t0","t3"); a.sw("t0",GP_X,"gp"); a.subu("t1","t1","t3")
    a.lab("store2_no_xoff"); a.sw("t1",GP_HADV,"gp"); a.addiu("s0","s0",2); a.j(GLYPH); a.nop()
    a.lab("lf"); a.lw("v1",GP_Y,"gp"); a.lw("a0",GP_VADV,"gp"); a.lh("v0",GP_START_X,"gp"); a.addu("v1","v1","a0"); a.sw("v0",GP_X,"gp"); a.sw("v1",GP_Y,"gp"); a.addiu("s0","s0",1); a.j(LOOP_TAIL); a.nop()
    a.lab("fallback"); a.lbu("v0",0,"s0"); a.lbu("v1",1,"s0"); a.j(FALLBACK); a.nop()
    a.lab("ret"); a.j(RET); a.nop()
    return a.out()

def ascii_payload(addr,args):
    tab=ascii_table(args.ascii_map_json)
    adv=build_advance_table(args)
    xoff=build_x_offset_table(args)
    adv_addr=addr+len(tab)
    xoff_addr=adv_addr+len(adv)
    code=align(xoff_addr+len(xoff),4)
    pad=b"\0"*(code-(xoff_addr+len(xoff)))
    stub=build_stub(code,addr,adv_addr,xoff_addr,args)
    return tab+adv+xoff+pad+stub,addr,code,xoff_addr
def is_j_word_to_loaded_text(word):
    return (word >> 26) == 0x02

def patch_parser(exe,load,stub,force=False):
    off=ram2off(TEXT_LOOP,load); actual=bytes(exe[off:off+8])
    word0=read32(exe,off); word1=read32(exe,off+4)
    # Accept the clean original parser bytes or an already-patched J/NOP from an
    # earlier renderer hook. This lets v7 be run as a post-patcher on an already
    # extended MAIN.EXE, appending a new hook and repointing the parser to it.
    already_hooked = is_j_word_to_loaded_text(word0) and word1 == 0
    if not force and actual!=TEXT_EXPECT and not already_hooked:
        raise RuntimeError(f"Parser bytes mismatch at 0x{off:X}: {actual.hex(' ')}. Use --parser-force if sure.")
    exe[off:off+8]=w(jt(2,stub))+w(0)
def patch_heap(exe,load,end):
    h=end-4; off=ram2off(HEAP_PAIR,load); actual=bytes(exe[off:off+8])
    word0=struct.unpack_from("<I",actual,0)[0]; word1=struct.unpack_from("<I",actual,4)[0]
    # Accept either the original heap-pair bytes or an already-patched compatible
    # pair. This lets the patcher update the heap again when the EXE grows.
    compatible = (
        actual == HEAP_EXPECT or
        ((word0>>26)==0x0f and ((word0>>21)&31)==0 and ((word0>>16)&31)==REG["a0"] and
         (word1>>26)==0x09 and ((word1>>21)&31)==REG["a0"] and ((word1>>16)&31)==REG["a0"])
    )
    if not compatible: raise RuntimeError(f"Heap pair mismatch at 0x{off:X}: {actual.hex(' ')}")
    exe[off:off+8]=w(it(0x0f,0,REG["a0"],hi(h)))+w(it(0x09,REG["a0"],REG["a0"],s16(lo(h))))
    return h


# PLANET/Garden screen music-title RAM hook.
#
# v70 stops relying on live payload addresses for Garden music-title safety.
# The Tree screen can upload a page with the same 0x380/0 256x256 shape, so
# broad rectangle matching can paint music-title rows into the Pixy Tree
# title/logo texture.  Instead, this version checks the source page contents: a
# candidate page must contain nonzero pixels in the repeated title/note strip
# positions for each of the seven music-title rows before the English overlay is
# applied.  Optional payload address whitelisting remains available for debugging,
# but the default is content-pattern guarded matching so all five Garden pages can
# be patched without touching unrelated TREE uploads.
#
# This hook lives in the same appended MAIN.EXE payload area as the ASCII
# renderer hook. That is important: earlier standalone experiments placed code
# inside the old BSS/zero area, which the game clears before normal execution.
# By appending this hook before the final sector padding and then moving the heap
# after the enlarged EXE payload, the hook is loaded as part of MAIN.EXE and is
# outside the cleared BSS/heap region.
PLANET_TITLE_HOOK_SITE = 0x8009F884
PLANET_TITLE_HOOK_RETURN = PLANET_TITLE_HOOK_SITE + 8
PLANET_TITLE_EXPECT = bytes.fromhex("04 00 25 86 04 00 23 96")  # lh a1,4(s1); lhu v1,4(s1)
PLANET_TITLE_PAGE_ROW_BYTES = 0x80  # 256 px 4bpp page = 128 bytes per row
PLANET_TITLE_DEFAULT_TITLES = ["Relief", "Affection", "Material", "Harvest", "Undine", "Windmill", "Ignition"]
PLANET_TITLE_FONT_5X7 = {
    " ": ["...", "...", "...", "...", "...", "...", "..."],
    "!": [".#.", ".#.", ".#.", ".#.", ".#.", "...", ".#."],
    "'": [".#.", ".#.", ".#.", "...", "...", "...", "..."],
    "(": [".#.", "#..", "#..", "#..", "#..", "#..", ".#."],
    ")": [".#.", "..#", "..#", "..#", "..#", "..#", ".#."],
    "+": ["...", ".#.", ".#.", "###", ".#.", ".#.", "..."],
    ",": ["...", "...", "...", "...", "...", ".#.", "#.."],
    "-": ["...", "...", "...", "###", "...", "...", "..."],
    ".": ["...", "...", "...", "...", "...", "...", ".#."],
    "/": ["..#", "..#", ".#.", ".#.", ".#.", "#..", "#.."],
    ":": ["...", ".#.", "...", "...", "...", ".#.", "..."],
    "&": [".##.", "#..#", "#..#", ".##.", "#.#.", "#..#", ".###"],
    "?": [".###.", "#...#", "....#", "...#.", "..#..", ".....", "..#.."],
    "0": [".###.", "#...#", "#..##", "#.#.#", "##..#", "#...#", ".###."],
    "1": ["..#..", ".##..", "..#..", "..#..", "..#..", "..#..", ".###."],
    "2": [".###.", "#...#", "....#", "...#.", "..#..", ".#...", "#####"],
    "3": ["####.", "....#", "...#.", "..##.", "....#", "#...#", ".###."],
    "4": ["...#.", "..##.", ".#.#.", "#..#.", "#####", "...#.", "...#."],
    "5": ["#####", "#....", "####.", "....#", "....#", "#...#", ".###."],
    "6": [".###.", "#...#", "#....", "####.", "#...#", "#...#", ".###."],
    "7": ["#####", "....#", "...#.", "..#..", ".#...", ".#...", ".#..."],
    "8": [".###.", "#...#", "#...#", ".###.", "#...#", "#...#", ".###."],
    "9": [".###.", "#...#", "#...#", ".####", "....#", "#...#", ".###."],
    "A": [".###.", "#...#", "#...#", "#####", "#...#", "#...#", "#...#"],
    "B": ["####.", "#...#", "#...#", "####.", "#...#", "#...#", "####."],
    "C": [".###.", "#...#", "#....", "#....", "#....", "#...#", ".###."],
    "D": ["####.", "#...#", "#...#", "#...#", "#...#", "#...#", "####."],
    "E": ["#####", "#....", "#....", "####.", "#....", "#....", "#####"],
    "F": ["#####", "#....", "#....", "####.", "#....", "#....", "#...."],
    "G": [".###.", "#...#", "#....", "#.###", "#...#", "#...#", ".###."],
    "H": ["#...#", "#...#", "#...#", "#####", "#...#", "#...#", "#...#"],
    "I": ["#", "#", "#", "#", "#", "#", "#"],
    "J": ["..###", "...#.", "...#.", "...#.", "...#.", "#..#.", ".##.."],
    "K": ["#...#", "#..#.", "#.#..", "##...", "#.#..", "#..#.", "#...#"],
    "L": ["#....", "#....", "#....", "#....", "#....", "#....", "#####"],
    "M": ["#...#", "##.##", "#.#.#", "#.#.#", "#...#", "#...#", "#...#"],
    "N": ["#...#", "##..#", "#.#.#", "#..##", "#...#", "#...#", "#...#"],
    "O": [".###.", "#...#", "#...#", "#...#", "#...#", "#...#", ".###."],
    "P": ["####.", "#...#", "#...#", "####.", "#....", "#....", "#...."],
    "Q": [".###.", "#...#", "#...#", "#...#", "#.#.#", "#..#.", ".##.#"],
    "R": ["####.", "#...#", "#...#", "####.", "#.#..", "#..#.", "#...#"],
    "S": [".####", "#....", "#....", ".###.", "....#", "....#", "####."],
    "T": ["#####", "..#..", "..#..", "..#..", "..#..", "..#..", "..#.."],
    "U": ["#...#", "#...#", "#...#", "#...#", "#...#", "#...#", ".###."],
    "V": ["#...#", "#...#", "#...#", "#...#", "#...#", ".#.#.", "..#.."],
    "W": ["#...#", "#...#", "#...#", "#.#.#", "#.#.#", "##.##", "#...#"],
    "X": ["#...#", "#...#", ".#.#.", "..#..", ".#.#.", "#...#", "#...#"],
    "Y": ["#...#", "#...#", ".#.#.", "..#..", "..#..", "..#..", "..#.."],
    "Z": ["#####", "....#", "...#.", "..#..", ".#...", "#....", "#####"],
    # Lowercase title glyphs. These let the PLANET music labels use Title Case
    # while still taking their advances and draw-shifts from the same renderer
    # metric tables used by the normal text hook.
    "a": [".....", ".###.", "....#", ".####", "#...#", "#...#", ".####"],
    "b": ["#....", "#....", "####.", "#...#", "#...#", "#...#", "####."],
    "c": [".....", ".###.", "#...#", "#....", "#....", "#...#", ".###."],
    "d": ["....#", "....#", ".####", "#...#", "#...#", "#...#", ".####"],
    "e": [".....", ".###.", "#...#", "#####", "#....", "#...#", ".###."],
    "f": ["..##.", ".#..#", ".#...", "###..", ".#...", ".#...", ".#..."],
    "g": [".....", ".####", "#...#", "#...#", ".####", "....#", ".###."],
    "h": ["#....", "#....", "####.", "#...#", "#...#", "#...#", "#...#"],
    "i": ["#", ".", "#", "#", "#", "#", "#"],
    "j": ["..#", "...", "..#", "..#", "..#", "#.#", ".#."],
    "k": ["#....", "#..#.", "#.#..", "##...", "#.#..", "#..#.", "#...#"],
    "l": ["#", "#", "#", "#", "#", "#", "#"],
    "m": [".....", "##.#.", "#.#.#", "#.#.#", "#.#.#", "#.#.#", "#.#.#"],
    "n": [".....", "####.", "#...#", "#...#", "#...#", "#...#", "#...#"],
    "o": [".....", ".###.", "#...#", "#...#", "#...#", "#...#", ".###."],
    "p": [".....", "####.", "#...#", "#...#", "####.", "#....", "#...."],
    "q": [".....", ".####", "#...#", "#...#", ".####", "....#", "....#"],
    "r": [".....", "#.##.", "##..#", "#....", "#....", "#....", "#...."],
    "s": [".....", ".####", "#....", ".###.", "....#", "....#", "####."],
    "t": [".#...", ".#...", "###..", ".#...", ".#...", ".#..#", "..##."],
    "u": [".....", "#...#", "#...#", "#...#", "#...#", "#...#", ".####"],
    "v": [".....", "#...#", "#...#", "#...#", "#...#", ".#.#.", "..#.."],
    "w": [".....", "#...#", "#...#", "#.#.#", "#.#.#", "#.#.#", ".#.#."],
    "x": [".....", "#...#", ".#.#.", "..#..", ".#.#.", "#...#", "#...#"],
    "y": [".....", "#...#", "#...#", "#...#", ".####", "....#", ".###."],
    "z": [".....", "#####", "...#.", "..#..", ".#...", "#....", "#####"],
}

def planet_title_parse_titles(s):
    # Preserve case here. The title renderer now has lowercase glyphs, and the
    # advance/x-offset lookup uses the exact ASCII character so later spacing
    # tweaks to the normal text metrics also affect these labels.
    return [x.strip() for x in str(s).split("|") if x.strip()]

def planet_title_parse_int_list(s, count=None, default=0):
    if s is None or str(s).strip() == "":
        vals = []
    else:
        vals = [int(x.strip(), 0) for x in str(s).split(",") if x.strip()]
    if count is not None:
        if len(vals) > count:
            raise ValueError(f"Expected at most {count} PLANET title Y offsets, got {len(vals)}")
        vals += [default] * (count - len(vals))
    return vals

def planet_title_signed_byte(v):
    v &= 0xFF
    return v - 0x100 if v & 0x80 else v

def planet_title_glyph_rows(ch):
    return PLANET_TITLE_FONT_5X7.get(ch, PLANET_TITLE_FONT_5X7.get(ch.upper(), PLANET_TITLE_FONT_5X7.get("?")))

def planet_title_glyph_width(ch):
    rows = planet_title_glyph_rows(ch)
    return max(len(r) for r in rows)

def planet_title_glyph_bounds(ch):
    """Return first visible column, last visible column, visible width for the
    embedded PLANET/Garden 5x7 title bitmap glyph.

    The normal menu text is rendered from FONT11Z0.TIM and now uses v80
    CP932/bitplane-derived metrics. The Garden music titles are different:
    they are pre-rendered into a VRAM upload page using this embedded 5x7
    bitmap font. Therefore they need their own bitmap-derived metrics instead
    of inheriting the FONT11Z0 x-offset/advance tables.
    """
    rows = planet_title_glyph_rows(ch)
    max_w = max(len(r) for r in rows) if rows else 0
    first = None
    last = None
    for xx in range(max_w):
        hit = False
        for row in rows:
            if xx < len(row) and row[xx] not in (".", " "):
                hit = True
                break
        if hit:
            if first is None:
                first = xx
            last = xx
    if first is None:
        return 0, -1, 0
    return first, last, last - first + 1

def planet_title_safe_adv(ch, adv_table, guard=True):
    o = ord(ch) if len(ch) == 1 else 0
    adv = adv_table[o] if 0 <= o < 128 else 6
    # The embedded 5x7 bitmap is not the real game glyph sheet, so without this
    # small guard very narrow renderer advances such as I=1 can visibly overlap.
    # The value still comes from the main renderer table whenever it is wide
    # enough for the bitmap.
    if guard and ch != " ":
        adv = max(adv, planet_title_glyph_width(ch) + 1)
    return adv

def planet_title_use_bitmap_metrics(args):
    return not getattr(args, "disable_planet_title_bitmap_metrics", False)

def planet_title_bitmap_advance_map(args):
    """Optional per-character overrides layered on top of bitmap-derived title
    advances, reusing the existing title-only CLI options.
    """
    deltas = {}
    overrides = {}
    for item in getattr(args, "planet_title_advance_delta", []) or []:
        chars, delta = item
        delta = int(delta)
        for ch in chars:
            o = ord(ch)
            if not 0 <= o < 128:
                raise ValueError(f"Non-ASCII PLANET title advance-delta character: {ch!r}")
            deltas[ch] = deltas.get(ch, 0) + delta
    for item in getattr(args, "planet_title_advance_override", []) or []:
        chars, adv = item
        adv = int(adv)
        if not 0 <= adv <= 255:
            raise ValueError(f"PLANET title advance override {adv} for {chars!r} must be 0..255")
        for ch in chars:
            o = ord(ch)
            if not 0 <= o < 128:
                raise ValueError(f"Non-ASCII PLANET title advance override character: {ch!r}")
            overrides[ch] = adv
    return deltas, overrides

def planet_title_bitmap_adv(ch, args):
    if ch == " ":
        space_adv = int(getattr(args, "planet_title_space_advance", -1))
        return max(0, space_adv if space_adv >= 0 else int(getattr(args, "space_advance", 3)))
    first, last, width = planet_title_glyph_bounds(ch)
    tracking = int(getattr(args, "planet_title_bitmap_tracking", 1))
    # Existing --planet-title-tracking remains useful as a global title-only
    # adjustment. It is added after the default one-pixel bitmap tracking.
    tracking += int(getattr(args, "planet_title_tracking", 0) or 0)
    adv = max(0, width + tracking)
    deltas, overrides = planet_title_bitmap_advance_map(args)
    if ch in deltas:
        adv = max(0, min(255, adv + int(deltas[ch])))
    if ch in overrides:
        adv = max(0, min(255, int(overrides[ch])))
    return adv

def planet_title_pair_delta(ch, next_ch, pair_kern_map):
    if not ch or not next_ch or not pair_kern_map:
        return 0
    return int(pair_kern_map.get(ch + next_ch, 0))

def planet_title_text_width(text, adv_table, guard=True, pair_kern_map=None, args=None):
    total = 0
    use_bitmap = args is not None and planet_title_use_bitmap_metrics(args)
    for i,ch in enumerate(text):
        if use_bitmap:
            adv = planet_title_bitmap_adv(ch, args)
        else:
            adv = planet_title_safe_adv(ch, adv_table, guard=guard)
        if i + 1 < len(text):
            adv = max(0, adv + planet_title_pair_delta(ch, text[i+1], pair_kern_map))
        total += adv
    return max(0, total - 1) if text else 0

def planet_title_build_advance_table(args):
    """Build the PLANET-title advance table from the normal renderer metrics,
    then apply optional title-only tracking/tuning.

    This keeps PLANET titles linked to the main text spacing by default, while
    allowing small corrections when the embedded title bitmap font reads a bit
    wider/tighter than the regular menu glyphs.
    """
    tbl = list(build_advance_table(args))

    # Title-only tracking: applied to printable non-space ASCII. Negative values
    # tighten the PLANET title labels; positive values loosen them.
    tracking = int(getattr(args, "planet_title_tracking", 0) or 0)
    if tracking:
        for i in range(0x21, 0x7F):
            tbl[i] = max(0, min(255, tbl[i] + tracking))

    # Optional title-only space override. -1 means inherit normal --space-advance.
    space_adv = int(getattr(args, "planet_title_space_advance", -1))
    if space_adv >= 0:
        tbl[0x20] = max(0, min(255, space_adv))

    # Repeatable title-only per-character deltas.
    # Example: --planet-title-advance-delta "fi" -1
    for item in getattr(args, "planet_title_advance_delta", []) or []:
        chars, delta = item
        delta = int(delta)
        for ch in chars:
            o = ord(ch)
            if not 0 <= o < 128:
                raise ValueError(f"Non-ASCII PLANET title advance-delta character: {ch!r}")
            tbl[o] = max(0, min(255, tbl[o] + delta))

    # Repeatable title-only absolute per-character overrides.
    # Example: --planet-title-advance-override "Relief" 5
    for item in getattr(args, "planet_title_advance_override", []) or []:
        chars, adv = item
        adv = int(adv)
        if not 0 <= adv <= 255:
            raise ValueError(f"PLANET title advance override {adv} for {chars!r} must be 0..255")
        for ch in chars:
            o = ord(ch)
            if not 0 <= o < 128:
                raise ValueError(f"Non-ASCII PLANET title advance override character: {ch!r}")
            tbl[o] = adv

    # Keep the PLANET bitmap-title path under the same v74 locked metrics too.
    # This catches old title-only overrides/commands that would otherwise make
    # the generated title overlay look unchanged.
    for ch, adv in iter_v74_final_advance_map(args).items():
        o = ord(ch)
        if 0 <= o < 128:
            tbl[o] = max(0, min(255, int(adv)))

    return bytes(tbl)

def planet_title_set_4bpp(pix, w, x, y, idx):
    if not (0 <= x < w and 0 <= y < (len(pix) * 2 // w)):
        return
    off = y * (w // 2) + (x // 2)
    old = pix[off]
    if x & 1:
        pix[off] = (old & 0x0F) | ((idx & 0x0F) << 4)
    else:
        pix[off] = (old & 0xF0) | (idx & 0x0F)

def planet_title_render_row(title, text_w, slot_h, text_index, shadow_index, clear_index, adv_table, xoff_table, guard=True, pair_kern_map=None, args=None):
    if text_w % 2:
        raise ValueError("PLANET title text width must be even")
    pix = bytearray([clear_index & 0x0F] * (text_w * slot_h // 2))
    # Do not force uppercase: Title Case labels use lowercase glyphs.
    use_bitmap = args is not None and planet_title_use_bitmap_metrics(args)
    tw = planet_title_text_width(title, adv_table, guard=guard, pair_kern_map=pair_kern_map, args=args)
    x0 = max(0, (text_w - tw) // 2)
    y0 = max(0, (slot_h - 7) // 2)

    def draw(dx, dy, idx):
        x = x0
        for i,ch in enumerate(title):
            rows = planet_title_glyph_rows(ch)
            o = ord(ch) if len(ch) == 1 else 0
            if use_bitmap:
                first, last, width = planet_title_glyph_bounds(ch)
                # Trim the embedded bitmap to its visible columns. This makes
                # the VRAM title data obey the same visible-width + tracking
                # model as v80's normal FONT11Z0 text, without applying the
                # unrelated FONT11Z0 draw-bearing table to the 5x7 title font.
                draw_xoff = -first if width else 0
            else:
                draw_xoff = planet_title_signed_byte(xoff_table[o]) if 0 <= o < 128 else 0
            for yy, row in enumerate(rows):
                for xx, c in enumerate(row):
                    if c not in (".", " "):
                        planet_title_set_4bpp(pix, text_w, x + draw_xoff + xx + dx, y0 + yy + dy, idx)
            if use_bitmap:
                adv = planet_title_bitmap_adv(ch, args)
            else:
                adv = planet_title_safe_adv(ch, adv_table, guard=guard)
            if i + 1 < len(title):
                adv = max(0, adv + planet_title_pair_delta(ch, title[i+1], pair_kern_map))
            x += adv
    draw(1, 1, shadow_index)
    draw(0, 0, text_index)
    return bytes(pix)

def planet_title_parse_addr_list(s):
    if s is None:
        return []
    text = str(s).strip()
    if not text:
        return []
    vals = []
    for part in text.replace("|", ",").replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        vals.append(int(part, 0) & 0xFFFFFFFF)
    # Preserve order but remove duplicates.
    out = []
    seen = set()
    for v in vals:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out

def planet_title_payload_whitelist(args):
    if getattr(args, "planet_title_allow_broad", False):
        return []
    vals = planet_title_parse_addr_list(getattr(args, "planet_title_payload_addrs", ""))
    single = int(getattr(args, "planet_title_payload_addr", 0) or 0)
    if single:
        vals.append(single & 0xFFFFFFFF)
    # Preserve order but remove duplicates again after the legacy single addr.
    out = []
    seen = set()
    for v in vals:
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out

def planet_title_pattern_ints(s, default_vals):
    if s is None or str(s).strip() == "":
        return list(default_vals)
    return [int(x.strip(), 0) for x in str(s).replace("|", ",").replace(";", ",").split(",") if x.strip()]

def planet_title_pattern_slots(args):
    """Return per-slot byte offsets used to verify that a candidate 0x380/0
    page actually contains the original repeated music-title/note strip.

    The guard is intentionally based on relative pixel density rather than exact
    palette values: the five Garden pages share the same title/note layout, while
    unrelated TREE/logo pages may share the upload rectangle but should not have
    nonzero bytes in these repeated seven title-strip marker positions.
    """
    if getattr(args, "planet_title_no_pattern_guard", False):
        return []
    xs = planet_title_pattern_ints(getattr(args, "planet_title_pattern_x", ""), [58, 62, 70, 90, 110, 130, 150, 154])
    ys = planet_title_pattern_ints(getattr(args, "planet_title_pattern_y", ""), [2, 5, 8])
    slots = []
    for slot in range(7):
        base_y = int(args.planet_title_y) + slot * int(args.planet_title_slot_pitch)
        slot_offsets = []
        for ry in ys:
            y = base_y + int(ry)
            if not 0 <= y < 256:
                continue
            for x in xs:
                x = int(x)
                if not 0 <= x < 256:
                    continue
                slot_offsets.append(y * PLANET_TITLE_PAGE_ROW_BYTES + (x // 2))
        # Preserve order but drop duplicates caused by nearby odd/even X values.
        dedup = []
        seen = set()
        for off in slot_offsets:
            if off not in seen:
                seen.add(off)
                dedup.append(off)
        slots.append(dedup)
    return slots

def planet_title_make_records(args):
    titles = planet_title_parse_titles(args.planet_title_titles)
    if len(titles) != 7:
        raise ValueError(f"PLANET title hook expects exactly 7 titles, got {len(titles)}")
    text_x = args.planet_title_text_x
    text_w = args.planet_title_text_w
    if text_x % 2 or text_w % 2:
        raise ValueError("PLANET title text x/width must both be even")
    adv_table = planet_title_build_advance_table(args)
    xoff_table = build_x_offset_table(args)
    pair_kern_map = iter_v75_pair_kern_map(args)
    row_len = text_w // 2
    slot_y_offsets = planet_title_parse_int_list(getattr(args, "planet_title_slot_y_offsets", ""), count=len(titles), default=0)
    records = bytearray()
    for slot, title in enumerate(titles):
        block = planet_title_render_row(
            title, text_w, args.planet_title_slot_h,
            args.planet_title_text_index & 0x0F,
            args.planet_title_shadow_index & 0x0F,
            args.planet_title_clear_index & 0x0F,
            adv_table, xoff_table,
            guard=(args.planet_title_bitmap_guard and not args.planet_title_no_bitmap_guard),
            pair_kern_map=pair_kern_map,
            args=args,
        )
        y_base = args.planet_title_y + slot * args.planet_title_slot_pitch + slot_y_offsets[slot]
        for row in range(args.planet_title_slot_h):
            page_off = (y_base + row) * PLANET_TITLE_PAGE_ROW_BYTES + (text_x // 2)
            records += struct.pack("<I", page_off)
            records += block[row * row_len:(row + 1) * row_len]
    return bytes(records), len(titles) * args.planet_title_slot_h, row_len, titles

def planet_title_build_hook(hook_ram, data_ram, record_count, row_byte_len, payload_addrs=None, min_payload_addr=0, pattern_slots=None, pattern_min_per_slot=0):
    a = A(hook_ram)
    a.addiu("sp", "sp", -0x20)
    for i, r in enumerate(["t0", "t1", "t2", "t3", "t4", "t5", "t6", "t7"]):
        a.sw(r, i * 4, "sp")

    a.lhu("t0", 0, "s1"); a.ori("t1", "zero", 0x0380); a.bne("t0", "t1", "done"); a.nop()
    a.lhu("t0", 2, "s1"); a.bne("t0", "zero", "done"); a.nop()
    a.lhu("t0", 4, "s1"); a.ori("t1", "zero", 0x0040); a.bne("t0", "t1", "done"); a.nop()
    a.lhu("t0", 6, "s1"); a.ori("t1", "zero", 0x0100); a.bne("t0", "t1", "done"); a.nop()
    if min_payload_addr:
        # Skip low transient buffers/command-list work areas.  This keeps the
        # music-title overlay off the Tree title-logo page while still allowing
        # the high decompressed Garden title pages.  sltu t1,s2,t0; if s2 < min, done.
        a.lui("t0", hi(min_payload_addr)); a.addiu("t0", "t0", s16(lo(min_payload_addr)))
        a.word(rt(REG["s2"], REG["t0"], REG["t1"], 0, 0x2B))
        a.bne("t1", "zero", "done"); a.nop()
    payload_addrs = list(payload_addrs or [])
    if payload_addrs:
        for addr in payload_addrs:
            a.lui("t1", hi(addr)); a.addiu("t1", "t1", s16(lo(addr)))
            a.beq("s2", "t1", "payload_ok"); a.nop()
        a.j("done"); a.nop()
        a.lab("payload_ok")

    pattern_slots = list(pattern_slots or [])
    pattern_min_per_slot = int(pattern_min_per_slot or 0)
    if pattern_slots and pattern_min_per_slot > 0:
        for slot_i, offsets in enumerate(pattern_slots):
            a.ori("t6", "zero", 0)
            for sample_i, off in enumerate(offsets):
                if not -0x8000 <= int(off) <= 0x7FFF:
                    raise ValueError(f"PLANET title pattern offset 0x{int(off):X} does not fit signed 16-bit MIPS load offset")
                skip = f"pt_pat_skip_{slot_i}_{sample_i}"
                a.lbu("t5", int(off), "s2")
                a.beq("t5", "zero", skip); a.nop()
                a.addiu("t6", "t6", 1)
                a.lab(skip)
            a.sltiu("t1", "t6", pattern_min_per_slot)
            a.bne("t1", "zero", "done"); a.nop()

    a.lui("t0", hi(data_ram)); a.addiu("t0", "t0", s16(lo(data_ram)))
    a.ori("t7", "zero", record_count)
    a.lab("record_loop")
    a.lw("t2", 0, "t0")
    a.addiu("t0", "t0", 4)
    a.addu("t4", "s2", "t2")
    a.ori("t3", "zero", row_byte_len)
    a.lab("copy_loop")
    a.lbu("t5", 0, "t0")
    a.word(it(0x28, REG["t4"], REG["t5"], 0))  # sb t5,0(t4)
    a.addiu("t0", "t0", 1)
    a.addiu("t4", "t4", 1)
    a.addiu("t3", "t3", -1)
    a.bne("t3", "zero", "copy_loop"); a.nop()
    a.addiu("t7", "t7", -1)
    a.bne("t7", "zero", "record_loop"); a.nop()

    a.lab("done")
    for i, r in enumerate(["t0", "t1", "t2", "t3", "t4", "t5", "t6", "t7"]):
        a.lw(r, i * 4, "sp")
    a.addiu("sp", "sp", 0x20)
    # Original overwritten instructions at 8009F884.
    a.lh("a1", 4, "s1")
    a.lhu("v1", 4, "s1")
    a.j(PLANET_TITLE_HOOK_RETURN); a.nop()
    return a.out()

def patch_planet_title_ram_hook(exe, load, args, dry_run=False):
    site_off = ram2off(PLANET_TITLE_HOOK_SITE, load)
    if site_off < 0 or site_off + 8 > len(exe):
        raise RuntimeError(f"PLANET title hook site maps outside EXE: RAM 0x{PLANET_TITLE_HOOK_SITE:08X}, file 0x{site_off:X}")
    cur = bytes(exe[site_off:site_off+8])
    if cur != PLANET_TITLE_EXPECT:
        already = ((read32(exe, site_off) >> 26) == 0x02 and read32(exe, site_off+4) == 0)
        if already and not args.planet_title_force:
            return {"enabled": True, "status": "already_hooked", "site_file": f"0x{site_off:X}"}
        if not args.planet_title_force:
            raise RuntimeError(
                f"PLANET title hook site mismatch at 0x{site_off:X}: {cur.hex(' ')}; "
                f"expected {PLANET_TITLE_EXPECT.hex(' ')}. Use --planet-title-force only if intentional."
            )
    records, record_count, row_len, titles = planet_title_make_records(args)
    # Build once with a placeholder to calculate code length, then rebuild with
    # the final data address after choosing the append position.
    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    payload_addrs = planet_title_payload_whitelist(args)
    pattern_slots = planet_title_pattern_slots(args)
    pattern_min = 0 if not pattern_slots else int(getattr(args, "planet_title_pattern_min_per_slot", 2) or 0)
    temp_code = planet_title_build_hook(hook_ram, hook_ram + 0x100, record_count, row_len, payload_addrs, args.planet_title_min_payload_addr, pattern_slots, pattern_min)
    code_len = len(temp_code)
    data_ram = hook_ram + code_len
    code = planet_title_build_hook(hook_ram, data_ram, record_count, row_len, payload_addrs, args.planet_title_min_payload_addr, pattern_slots, pattern_min)
    assert len(code) == code_len
    total_len = code_len + len(records)
    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[site_off:site_off+8] = w(jt(2, hook_ram)) + w(0)
        exe[hook_off:hook_off+total_len] = code + records
    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "site_ram": f"0x{PLANET_TITLE_HOOK_SITE:08X}",
        "site_file": f"0x{site_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_file": f"0x{hook_off:X}",
        "data_ram": f"0x{data_ram:08X}",
        "code_len": code_len,
        "records_len": len(records),
        "record_count": record_count,
        "row_byte_len": row_len,
        "titles": titles,
        "text_x": args.planet_title_text_x,
        "title_y": args.planet_title_y,
        "text_w": args.planet_title_text_w,
        "slot_h": args.planet_title_slot_h,
        "slot_pitch": args.planet_title_slot_pitch,
        "slot_y_offsets": getattr(args, "planet_title_slot_y_offsets", ""),
        "payload_addr": None if args.planet_title_payload_addr == 0 else f"0x{args.planet_title_payload_addr:08X}",
        "payload_addrs": [f"0x{x:08X}" for x in payload_addrs],
        "allow_broad": bool(getattr(args, "planet_title_allow_broad", False)),
        "min_payload_addr": None if args.planet_title_min_payload_addr == 0 else f"0x{args.planet_title_min_payload_addr:08X}",
        "pattern_guard": bool(pattern_slots and pattern_min > 0),
        "pattern_min_per_slot": pattern_min,
        "pattern_sample_x": getattr(args, "planet_title_pattern_x", ""),
        "pattern_sample_y": getattr(args, "planet_title_pattern_y", ""),
    }

def dec_lui(x):
    if (x>>26)==0x0f and ((x>>21)&31)==0: return (x>>16)&31,x&0xffff
    return None
def dec_addori(x,rtx):
    op=x>>26; rs=(x>>21)&31; rt_=(x>>16)&31; imm=x&0xffff
    if rs==rtx and rt_==rtx:
        if op==0x09: return "addiu",imm
        if op==0x0d: return "ori",imm
    return None
def patch_lui(exe,old,new,maxgap=3):
    oldhi=hi(old); oldlo=lo(old); cnt=0
    for off in range(PSX_HEADER,len(exe)-3,4):
        d=dec_lui(struct.unpack_from("<I",exe,off)[0])
        if not d or d[1]!=oldhi: continue
        rt_,_=d
        for gap in range(1,maxgap+1):
            off2=off+gap*4
            if off2+4>len(exe): break
            d2=dec_addori(struct.unpack_from("<I",exe,off2)[0],rt_)
            if not d2 or d2[1]!=oldlo: continue
            struct.pack_into("<I",exe,off,it(0x0f,0,rt_,hi(new)))
            struct.pack_into("<I",exe,off2,it(0x09,rt_,rt_,s16(lo(new))))
            cnt+=1; break
    return cnt


REG_NAMES=[None]*32
for _name,_idx in REG.items():
    REG_NAMES[_idx]=_name
for _i,_n in enumerate(REG_NAMES):
    if _n is None: REG_NAMES[_i]=f"r{_i}"

def op(x): return (x>>26)&0x3f
def rs_(x): return (x>>21)&31
def rt_(x): return (x>>16)&31
def imm_(x): return x&0xffff
def setimm(x,imm): return (x&0xffff0000)|(imm&0xffff)
def hi_ori(addr): return (addr>>16)&0xffff

def disasm_i(x):
    o=op(x); rs=rs_(x); rt=rt_(x); im=imm_(x)
    if o==0x0f and rs==0: return f"lui {REG_NAMES[rt]},0x{im:04X}"
    if o==0x09: return f"addiu {REG_NAMES[rt]},{REG_NAMES[rs]},0x{im:04X}"
    if o==0x08: return f"addi {REG_NAMES[rt]},{REG_NAMES[rs]},0x{im:04X}"
    if o==0x0d: return f"ori {REG_NAMES[rt]},{REG_NAMES[rs]},0x{im:04X}"
    return f"word 0x{x:08X}"

def writes_reg(word,reg):
    if reg==0: return False
    o=op(word)
    if o==0x00:
        fn=word&0x3f; rd=(word>>11)&31
        if fn in {0x08,0x09,0x0c,0x0d}: return False
        return rd==reg
    if o==0x03: return reg==31
    if o in {0x02,0x04,0x05,0x06,0x07,0x28,0x29,0x2a,0x2b,0x2e}: return False
    return rt_(word)==reg

def parse_row_ranges(spec):
    if not spec: return None
    out=set()
    for part in str(spec).split(','):
        part=part.strip()
        if not part: continue
        if '-' in part:
            a,b=part.split('-',1); a=int(a); b=int(b)
            if b<a: a,b=b,a
            out.update(range(a,b+1))
        else:
            out.add(int(part))
    return out

def direct_old_hi(kind,ptr):
    return hi_ori(ptr) if kind=='lui_ori' else hi(ptr)

def direct_new_hi(kind,ptr):
    return hi_ori(ptr) if kind=='lui_ori' else hi(ptr)

def direct_expected_op(kind):
    return 0x0d if kind=='lui_ori' else (0x08 if kind=='lui_addi' else 0x09)

def scan_direct_mips(orig,rows,code_start=PSX_HEADER,code_end=None,maxgap=8,confidence='high',lifetime=True):
    code_end=len(orig) if code_end is None else min(len(orig),code_end)
    code_start=align(max(0,code_start),4)
    add_targets=collections.defaultdict(lambda: collections.defaultdict(list))
    ori_targets=collections.defaultdict(lambda: collections.defaultdict(list))
    for r in rows:
        add_targets[hi(r.old_ptr)][lo(r.old_ptr)].append(r)
        ori_targets[hi_ori(r.old_ptr)][lo(r.old_ptr)].append(r)
    hits=[]; seen=set()
    for off in range(code_start,code_end-3,4):
        lw=struct.unpack_from('<I',orig,off)[0]
        d=dec_lui(lw)
        if not d: continue
        lr,high=d
        if high not in add_targets and high not in ori_targets: continue
        for gap in range(1,maxgap+1):
            uoff=off+gap*4
            if uoff+4>code_end: break
            uw=struct.unpack_from('<I',orig,uoff)[0]
            # If lifetime-aware mode sees the LUI register overwritten before it is used,
            # later apparent matches are stale and unsafe.
            uses_lui=(rs_(uw)==lr)
            if uses_lui:
                low=imm_(uw); o=op(uw); matches=[]
                if o==0x0d and high in ori_targets:
                    rr=ori_targets[high].get(low,[])
                    if rr: matches.append(('lui_ori',rr))
                elif o in (0x08,0x09) and high in add_targets:
                    rr=add_targets[high].get(low,[])
                    if rr: matches.append(('lui_addi' if o==0x08 else 'lui_addiu',rr))
                if matches:
                    conf='high' if gap<=2 else 'medium'
                    if confidence=='all' or confidence==conf:
                        for kind,rrs in matches:
                            for r in rrs:
                                key=(r.sheet_row,kind,off,uoff)
                                if key in seen: continue
                                seen.add(key)
                                hits.append(DirectMipsHit(r,kind,conf,off,uoff,gap,lr,rt_(uw),lw,uw))
            if lifetime and writes_reg(uw,lr) and not uses_lui:
                break
    return hits

def patch_direct_mips(exe,hits,dry_run=False,strict=True):
    by_lui=collections.defaultdict(list)
    for h in hits: by_lui[h.lui_off].append(h)
    report=[]
    for lui_off,group in sorted(by_lui.items()):
        # Deduplicate identical use sites. Multiple target rows for the same instruction
        # cannot both be correct unless they map to the same old/new pointer.
        by_use={}
        dup_notes=[]
        for h in group:
            key=(h.use_off,h.kind)
            prev=by_use.get(key)
            if prev is None:
                by_use[key]=h
            elif prev.row.old_ptr==h.row.old_ptr and prev.row.new_ptr==h.row.new_ptr:
                dup_notes.append(f"duplicate_same_site_row_{h.row.sheet_row}")
            else:
                # Keep both in the group so the conflicting LUI/high check can reject safely.
                by_use[(h.use_off,h.kind,h.row.sheet_row)]=h
        group=list(by_use.values())
        try:
            kinds={h.kind for h in group}
            new_his={direct_new_hi(h.kind,h.row.new_ptr) for h in group}
            old_his={direct_old_hi(h.kind,h.row.old_ptr) for h in group}
            if len(kinds)>1:
                raise ValueError(f"shared LUI has mixed address-building kinds: {sorted(kinds)}")
            if len(new_his)>1:
                raise ValueError(f"shared LUI would need multiple new high immediates: {[f'0x{x:04X}' for x in sorted(new_his)]}")
            if len(old_his)>1:
                raise ValueError(f"shared LUI has multiple old high immediates: {[f'0x{x:04X}' for x in sorted(old_his)]}")
            h0=group[0]; old_hi=next(iter(old_his)); new_hi=next(iter(new_his))
            lw=struct.unpack_from('<I',exe,lui_off)[0]
            d=dec_lui(lw)
            if not d or d[0]!=h0.lui_reg:
                raise ValueError(f"LUI instruction changed: found {disasm_i(lw)}")
            cur_hi=d[1]
            if cur_hi==old_hi:
                if not dry_run: struct.pack_into('<I',exe,lui_off,setimm(lw,new_hi))
                lui_note='would_patch_lui' if dry_run else 'patched_lui'
            elif cur_hi==new_hi:
                lui_note='lui_already_new'
            else:
                raise ValueError(f"LUI imm mismatch: expected 0x{old_hi:04X} or 0x{new_hi:04X}, found 0x{cur_hi:04X}")
        except Exception as e:
            if strict: raise
            for h in group:
                report.append({'sheet_row':h.row.sheet_row,'status':'error','note':str(e),'kind':h.kind,'confidence':h.confidence,'lui_file_offset':f'0x{h.lui_off:X}','use_file_offset':f'0x{h.use_off:X}','gap':h.gap,'old_pointer':f'0x{h.row.old_ptr:08X}','new_pointer':f'0x{h.row.new_ptr:08X}','old_lui':disasm_i(h.old_lui_word),'old_use':disasm_i(h.old_use_word),'text_preview':h.row.text[:120]})
            continue
        for h in group:
            try:
                uw=struct.unpack_from('<I',exe,h.use_off)[0]
                if op(uw)!=direct_expected_op(h.kind) or rs_(uw)!=h.lui_reg:
                    raise ValueError(f"use instruction changed: found {disasm_i(uw)}")
                old_low=lo(h.row.old_ptr); new_low=lo(h.row.new_ptr); cur_low=imm_(uw)
                if cur_low==old_low:
                    if not dry_run: struct.pack_into('<I',exe,h.use_off,setimm(uw,new_low))
                    status='would_patch' if dry_run else 'patched'; note=f'{lui_note};low_replaced'
                elif cur_low==new_low:
                    status='already_patched'; note=f'{lui_note};low_already_new'
                else:
                    raise ValueError(f"low imm mismatch: expected 0x{old_low:04X} or 0x{new_low:04X}, found 0x{cur_low:04X}")
            except Exception as e:
                if strict: raise
                status='error'; note=str(e)
            report.append({'sheet_row':h.row.sheet_row,'status':status,'note':note,'kind':h.kind,'confidence':h.confidence,'lui_file_offset':f'0x{h.lui_off:X}','use_file_offset':f'0x{h.use_off:X}','gap':h.gap,'old_pointer':f'0x{h.row.old_ptr:08X}','new_pointer':f'0x{h.row.new_ptr:08X}','old_lui':disasm_i(h.old_lui_word),'old_use':disasm_i(h.old_use_word),'text_preview':h.row.text[:120]})
    return report
def collect_ptr32(orig,rows,load):
    out=[]; cid=0
    for row in rows:
        for off in allfind(orig,struct.pack("<I",row.old_ptr&0xffffffff)):
            out.append(Cand(cid,row,off,off2ram(off,load),sec_for(off2ram(off,load)))); cid+=1
    out.sort(key=lambda c:c.off)
    for i,c in enumerate(out): c.cid=i
    return out
def cluster(cands,gap):
    cid=-1; cur=[]; last=None
    def flush():
        if not cur: return
        uq=len({c.row.sheet_row for c in cur})
        for c in cur: c.cluster=cid; c.csize=len(cur); c.cuniq=uq
    for c in cands:
        if last is None or c.off-last>gap:
            flush(); cid+=1; cur=[c]
        else: cur.append(c)
        last=c.off
    flush()
def hexset(path):
    if not path: return set()
    s=Path(path).read_text(encoding="utf-8-sig")
    return {int(tok,0 if tok.lower().startswith("0x") else 16) for tok in re.findall(r"0x[0-9A-Fa-f]+|[0-9A-Fa-f]{4,}",s)}
def ranges(path):
    if not path: return []
    data=json.loads(Path(path).read_text(encoding="utf-8"))
    out=[]
    for x in data:
        if "start" in x and "end" in x: a=int(str(x["start"]),0); b=int(str(x["end"]),0)
        else: a=int(str(x["offset"]),0); b=a+int(str(x["length"]),0)
        out.append((a,b))
    return out
def inranges(off,rs): return any(a<=off<b for a,b in rs)
def select(cands,args):
    allowed={x.strip() for x in args.ptr32_sections.split(",") if x.strip()}
    inc=hexset(args.ptr32_include_offsets); exc=hexset(args.ptr32_exclude_offsets); rs=ranges(args.ptr32_ranges_json)
    prelim=[]
    for c in cands:
        sel=False; reason=""
        if args.ptr32_policy=="none": reason="policy_none"
        elif args.ptr32_policy=="all": sel=True; reason="policy_all"
        elif args.ptr32_policy=="section": sel=c.section in allowed; reason="section_allowed" if sel else f"section_{c.section}_not_allowed"
        elif args.ptr32_policy=="clustered":
            sel=c.section in allowed and c.csize>=args.ptr32_cluster_min and c.cuniq>=args.ptr32_cluster_unique_min
            reason="clustered_selected" if sel else f"cluster_reject_size={c.csize}_unique={c.cuniq}_section={c.section}"
        elif args.ptr32_policy=="ranges": sel=inranges(c.off,rs); reason="range_selected" if sel else "not_in_ranges"
        if inc and c.off in inc: sel=True; reason+=";forced_include"
        if exc and c.off in exc: sel=False; reason+=";forced_exclude"
        c.selected=sel; c.reason=reason
        if sel: prelim.append(c)
    if args.ptr32_slice_count>1:
        ss=sorted(prelim,key=lambda c:c.off); n=len(ss); a=(n*args.ptr32_slice_index)//args.ptr32_slice_count; b=(n*(args.ptr32_slice_index+1))//args.ptr32_slice_count
        keep={id(c) for c in ss[a:b]}
        for c in cands:
            if c.selected and id(c) not in keep: c.selected=False; c.reason+=f";slice_excluded_{args.ptr32_slice_index}_of_{args.ptr32_slice_count}"
            elif c.selected: c.reason+=f";slice_included_{args.ptr32_slice_index}_of_{args.ptr32_slice_count}"
def apply_ptr32(exe,cands):
    cnt=0
    for c in cands:
        if c.selected:
            exe[c.off:c.off+4]=struct.pack("<I",c.row.new_ptr&0xffffffff); cnt+=1
    return cnt


def parse_int_set(spec):
    """Parse comma/space separated integers such as '0x3E48, 0x1234'."""
    if not spec:
        return set()
    out=set()
    for tok in re.findall(r"0x[0-9A-Fa-f]+|[0-9A-Fa-f]{4,}|\d+", str(spec)):
        out.add(int(tok,0 if tok.lower().startswith('0x') else 16 if re.fullmatch(r"[0-9A-Fa-f]{4,}", tok) else 10))
    return out

def apply_auto_ptr32_excludes(cands,offsets,label="auto_exclude"):
    offsets=set(offsets or [])
    changed=[]
    if not offsets:
        return changed
    for c in cands:
        if c.off in offsets:
            was_selected=c.selected
            c.selected=False
            c.reason+=(";" if c.reason else "")+label
            changed.append({
                'candidate_id':c.cid,
                'was_selected':int(was_selected),
                'file_offset':f'0x{c.off:X}',
                'ram_address':f'0x{c.ram:08X}',
                'section':c.section,
                'sheet_row':c.row.sheet_row,
                'old_pointer':f'0x{c.row.old_ptr:08X}',
                'new_pointer':f'0x{c.row.new_ptr:08X}',
                'text_preview':c.row.text[:120],
            })
    return changed

def patch_inplace_rows(exe,rows,row_nums,fill_mode="normal_zero",body_encoding="ascii",dry_run=False):
    """Patch selected spreadsheet rows at their original string slots.

    This is intentionally optional. The confirmed name-screen-critical fix is
    excluding ptr32 file offset 0x3E48 so the Aries/name table keeps its
    original pointer. In-place row patching is only for experiments where you
    want that original slot translated too.

    body_encoding values:
      ascii              use the normal one-byte ASCII insertion encoding
      cp932_fullwidth    encode ASCII letters/numbers as CP932 full-width glyphs

    fill_mode values:
      normal_zero        write encoded text including 00, then 00-fill rest
      normal_space       write encoded text including 00, then ASCII-space-fill rest
      fixed_ascii_space  write body, ASCII-space-fill through slot_len-1, final 00
      fixed_cp932_space  write body, then pad with CP932 full-width spaces (81 40),
                         using one ASCII space first if needed for odd byte count,
                         final 00
    """
    if not row_nums:
        return []
    by_row={r.sheet_row:r for r in rows}
    report=[]
    for n in sorted(row_nums):
        r=by_row.get(n)
        if r is None:
            raise RuntimeError(f"In-place row {n} was not loaded from the spreadsheet")
        if r.old_off<0 or r.old_off+r.old_len>len(exe):
            raise RuntimeError(f"Row {n}: original slot maps outside EXE at file offset 0x{r.old_off:X}")
        old_preview=bytes(exe[r.old_off:r.old_off+min(r.old_len,32)]).hex(' ')
        if body_encoding == "ascii":
            row_enc = r.enc
        elif body_encoding == "cp932_fullwidth":
            row_enc = enc_text_fullwidth_cp932(r.text)
        else:
            raise RuntimeError(f"Unknown in-place body encoding: {body_encoding}")
        if fill_mode in {"normal_zero","normal_space"}:
            if len(row_enc)>r.old_len:
                raise RuntimeError(
                    f"Row {n}: encoded text does not fit original slot: new {len(row_enc)} bytes > slot {r.old_len} bytes; text={r.text!r}; body_encoding={body_encoding}"
                )
            fill_byte=0x20 if fill_mode=="normal_space" else 0x00
            new_slot=bytearray([fill_byte])*r.old_len
            new_slot[:len(row_enc)]=row_enc
        elif fill_mode in {"fixed_ascii_space","fixed_cp932_space"}:
            body=row_enc[:-1] if row_enc.endswith(b"\0") else row_enc
            if len(body)>r.old_len-1:
                raise RuntimeError(
                    f"Row {n}: encoded text body does not fit fixed-width slot: body {len(body)} bytes > slot body {r.old_len-1} bytes; text={r.text!r}; body_encoding={body_encoding}"
                )
            rem=(r.old_len-1)-len(body)
            if fill_mode=="fixed_ascii_space":
                pad=b"\x20"*rem
            else:
                # Prefer the original CP932 full-width space padding style. If
                # the byte count is odd, add one ASCII space first, then 81 40 pairs.
                pad=(b"\x20" if rem%2 else b"") + (b"\x81\x40"*(rem//2))
                if len(pad)!=rem:
                    raise AssertionError("internal CP932 padding length mismatch")
            new_slot=bytearray(body+pad+b"\0")
        else:
            raise RuntimeError(f"Unknown in-place fill mode: {fill_mode}")
        if len(new_slot)!=r.old_len:
            raise AssertionError(f"Row {n}: new slot size {len(new_slot)} != original slot size {r.old_len}")
        if not dry_run:
            exe[r.old_off:r.old_off+r.old_len]=new_slot
        report.append({
            'sheet_row':n,
            'status':'would_patch' if dry_run else 'patched',
            'old_pointer':f'0x{r.old_ptr:08X}',
            'old_offset':f'0x{r.old_off:X}',
            'slot_len':r.old_len,
            'encoded_len':len(row_enc),
            'body_encoding':body_encoding,
            'fill':fill_mode,
            'text_preview':r.text[:180],
            'old_preview':old_preview,
            'new_preview':bytes(new_slot[:min(r.old_len,32)]).hex(' '),
        })
    return report


# Shared text primitive-buffer capacity patch.
#
# The shared text renderer allocates 0x2400 bytes and splits it into two
# 0x1200-byte primitive buffers. Each rendered glyph consumes 0x24 bytes in
# each half, so the original per-call capacity is 0x1200 / 0x24 = 128 glyphs.
# TREE.CDF ASCII tests confirmed freezes above this range. Raising this buffer
# capacity allows long TREE/EVENT/etc. text pages to render without corrupting
# the primitive workspace.
TEXT_PRIM_GLYPH_SIZE = 0x24
TEXT_PRIM_PATCH_SITES = {
    "initA_alloc_size": 0x18094,
    "initA_second_half": 0x180E4,
    "initA_init_count": 0x18138,
    "initB_alloc_size": 0x18B70,
    "initB_second_half": 0x18B94,
    "initB_init_count": 0x18BE8,
    "flush_second_half": 0x18A30,
    "reset1_second_half": 0x18D64,
    "reset2_second_half": 0x18DCC,
    "print_second_half": 0x18E44,
}
TEXT_PRIM_ORIGINAL_WORDS = {
    "initA_alloc_size": 0x24042400,
    "initA_second_half": 0x24A21200,
    "initA_init_count": 0x28C20100,
    "initB_alloc_size": 0x24042400,
    "initB_second_half": 0x24A21200,
    "initB_init_count": 0x28C20100,
    "flush_second_half": 0x26321200,
    "reset1_second_half": 0x24631200,
    "reset2_second_half": 0x24421200,
    "print_second_half": 0x24421200,
}

def i_shape_word(word):
    return ((word >> 26) & 0x3F, (word >> 21) & 31, (word >> 16) & 31)

def set_i_imm(word, imm):
    return (word & 0xFFFF0000) | (imm & 0xFFFF)

def text_prim_word_shape_ok(name, cur):
    exp = TEXT_PRIM_ORIGINAL_WORDS[name]
    return cur == exp or i_shape_word(cur) == i_shape_word(exp)

def patch_text_primitive_capacity(exe, capacity_glyphs=192, force=False, dry_run=False):
    if capacity_glyphs < 128:
        raise ValueError("text primitive capacity should be >= 128 glyphs")
    half_size = capacity_glyphs * TEXT_PRIM_GLYPH_SIZE
    alloc_size = half_size * 2
    init_count = capacity_glyphs * 2
    for label, value in (("half_size", half_size), ("alloc_size", alloc_size), ("init_count", init_count)):
        if value > 0x7FFF:
            raise ValueError(f"{label}=0x{value:X} is too large for signed 16-bit immediates")
    new_imms = {
        "initA_alloc_size": alloc_size,
        "initA_second_half": half_size,
        "initA_init_count": init_count,
        "initB_alloc_size": alloc_size,
        "initB_second_half": half_size,
        "initB_init_count": init_count,
        "flush_second_half": half_size,
        "reset1_second_half": half_size,
        "reset2_second_half": half_size,
        "print_second_half": half_size,
    }
    report = []
    for name, off in TEXT_PRIM_PATCH_SITES.items():
        cur = read32(exe, off)
        if not force and not text_prim_word_shape_ok(name, cur):
            raise RuntimeError(
                f"Text primitive capacity site {name} at 0x{off:X} has unexpected word 0x{cur:08X}; "
                f"expected original/compatible shape. Use --text-primitive-force to patch anyway."
            )
        new_word = set_i_imm(cur, new_imms[name])
        if not dry_run:
            exe[off:off+4] = struct.pack("<I", new_word & 0xFFFFFFFF)
        report.append({
            "site": name,
            "file_offset": f"0x{off:X}",
            "old_word": f"0x{cur:08X}",
            "new_word": f"0x{new_word:08X}",
            "new_imm": f"0x{new_imms[name]:X}",
            "status": "would_patch" if dry_run else "patched",
        })
    return {
        "enabled": True,
        "capacity_glyphs": capacity_glyphs,
        "glyph_primitive_size": TEXT_PRIM_GLYPH_SIZE,
        "half_size": half_size,
        "alloc_size": alloc_size,
        "init_count": init_count,
        "report": report,
    }







# ---------------------------------------------------------------------------
# Full-screen memory-card / Loading/Saving text centering (v95)
# ---------------------------------------------------------------------------
# Rows 2-38 in MAIN_EXE_Text are drawn by the memory-card/status routines with
# explicit FUN_80038498(x, y, text) calls.  These are not 195px textbox strings;
# they are screen-centered status strings.  The English workbook now keeps the
# text unpadded, so this patch adjusts the caller's a0 x-coordinate to roughly
# -width/2 in the game's center-origin coordinate system.

MC_FUN_DRAW_TEXT_XY = 0x80038498
MC_ROW_FIRST = 2
MC_ROW_LAST = 38
MC_REG_ZERO = 0
MC_REG_A0 = 4
MC_REG_A2 = 6

@dataclass(frozen=True)
class MCTextRow:
    excel_row: int
    pointer: int
    text: str
    qa_width_px: Optional[int]

@dataclass(frozen=True)
class MCPatchSite:
    name: str
    ram_a0: int
    pointers: tuple[int, ...]
    note: str = ""

# Known a0 load sites in the stock memory-card/status state machine.
# Dynamic scanning below may find additional direct calls whose a2 pointer still
# refers to one of rows 2-38.
MC_KNOWN_A0_SITES: tuple[MCPatchSite, ...] = (
    MCPatchSite("row02_memory_card_missing_line1", 0x8003A900, (0x800200FC,)),
    MCPatchSite("row03_memory_card_slot_line",   0x8003A914, (0x80020128,)),
    MCPatchSite("row04_checking_memory_card",    0x8003A9BC, (0x8002013C,)),
    MCPatchSite("row09_checking_save_data",      0x8003B034, (0x80020238,)),
    MCPatchSite("row10_do_not_remove_card",      0x8003B048, (0x80020254,)),
    MCPatchSite("row12_use_card_contains",       0x8003B3A4, (0x800202BC,)),
    MCPatchSite("row13_pixy_save_data",          0x8003B3B8, (0x800202E8,)),
    MCPatchSite("row14_save_data_corrupted",     0x8003B4E4, (0x80020308,)),
    MCPatchSite("row15_switch_card",             0x8003B50C, (0x80020338,)),
    MCPatchSite("rows16_18_load_save_overwrite_shared", 0x8003BC4C,
                (0x80020374, 0x80020394, 0x800203B4),
                "shared runtime-selected draw call; centered using widest string"),
    MCPatchSite("row32_is_this_okay",             0x8003D3D0, (0x80020564,)),
    MCPatchSite("row35_creating_new_file",        0x8003D1E0, (0x800205E4,)),
    MCPatchSite("rows37_38_cancel_load_save_shared", 0x8003D3A8,
                (0x8002061C, 0x80020640),
                "shared runtime-selected draw call; centered using widest string"),
)

def mc_parse_pointer(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s, 0)
    except ValueError:
        return None

def mc_parse_intish(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None

def mc_norm_header(value) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())

def mc_choose_sheet(wb):
    for name in ("MAIN_EXE_Text", "MAIN.EXE", "MAIN", "Sheet1"):
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]

def mc_load_text_rows(workbook_path) -> dict[int, MCTextRow]:
    wb = load_workbook(xlsx_path(workbook_path), data_only=True)
    ws = mc_choose_sheet(wb)
    headers = {mc_norm_header(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}
    ptr_col = headers.get("pointer")
    text_col = (headers.get("qa clean translation") or
                headers.get("better translation") or
                headers.get("english translation") or
                headers.get("insertion text") or
                headers.get("string decoded"))
    width_col = headers.get("qa max line width px") or headers.get("max width px")
    if ptr_col is None:
        raise RuntimeError("Could not find a pointer column for memory-card centering.")
    if text_col is None:
        raise RuntimeError("Could not find a translation/text column for memory-card centering.")
    rows: dict[int, MCTextRow] = {}
    for excel_row in range(MC_ROW_FIRST, MC_ROW_LAST + 1):
        ptr = mc_parse_pointer(ws.cell(row=excel_row, column=ptr_col).value)
        if ptr is None:
            continue
        text_value = ws.cell(row=excel_row, column=text_col).value
        text = "" if text_value is None else str(text_value)
        width = mc_parse_intish(ws.cell(row=excel_row, column=width_col).value) if width_col else None
        rows[ptr] = MCTextRow(excel_row=excel_row, pointer=ptr, text=text, qa_width_px=width)
    return rows

def mc_decode_i(word: int) -> tuple[int, int, int, int]:
    return ((word >> 26) & 0x3F, (word >> 21) & 31, (word >> 16) & 31, word & 0xFFFF)

def mc_is_addiu_reg_zero(word: int, rt_reg: int) -> bool:
    op, rs, wrt, _imm = mc_decode_i(word)
    return op == 0x09 and rs == MC_REG_ZERO and wrt == rt_reg

def mc_is_lui_reg(word: int, rt_reg: int) -> bool:
    op, rs, wrt, _imm = mc_decode_i(word)
    return op == 0x0F and rs == MC_REG_ZERO and wrt == rt_reg

def mc_is_addiu_same_reg(word: int, rt_reg: int) -> bool:
    op, rs, wrt, _imm = mc_decode_i(word)
    return op == 0x09 and rs == rt_reg and wrt == rt_reg

def mc_is_zero_to_a0(word: int) -> bool:
    return word in (0x00002021, 0x00002025)  # addu/or a0,zero,zero

def mc_write32(buf: bytearray, off: int, word: int) -> None:
    struct.pack_into("<I", buf, off, word & 0xFFFFFFFF)

def mc_make_li_a0(x: int) -> int:
    if x < -32768 or x > 32767:
        raise ValueError(f"memory-card centered x={x} is outside signed 16-bit range")
    return it(0x09, MC_REG_ZERO, MC_REG_A0, x)

def mc_jal_word(addr: int) -> int:
    return 0x0C000000 | ((addr >> 2) & 0x03FFFFFF)

def mc_center_x_for_width(width_px: int, bias: int = 0) -> int:
    # Center-origin coordinates.  Odd widths are biased left by one pixel so the
    # visual center straddles x=0.
    return -int(math.ceil(width_px / 2.0)) + int(bias)

def mc_group_width(rows_by_ptr: dict[int, MCTextRow], pointers: tuple[int, ...]):
    group_rows = [rows_by_ptr[p] for p in pointers if p in rows_by_ptr]
    widths = [r.qa_width_px for r in group_rows if r.qa_width_px is not None]
    if widths:
        return max(widths), group_rows
    return None, group_rows

def mc_find_loaded_addr_for_reg(buf: bytes | bytearray, search_start: int, search_end: int, reg_num: int):
    out = []
    for lui_off in range(search_start, search_end + 1, 4):
        if lui_off < 0 or lui_off + 4 > len(buf):
            continue
        w_lui = read32(buf, lui_off)
        if not mc_is_lui_reg(w_lui, reg_num):
            continue
        hi_imm = w_lui & 0xFFFF
        for add_off in range(lui_off + 4, min(search_end + 1, lui_off + 28), 4):
            if add_off < 0 or add_off + 4 > len(buf):
                continue
            w_add = read32(buf, add_off)
            if mc_is_addiu_same_reg(w_add, reg_num):
                lo_imm = w_add & 0xFFFF
                addr = ((hi_imm << 16) + s16(lo_imm)) & 0xFFFFFFFF
                out.append((addr, lui_off, add_off))
                break
    return out

def mc_reconstruct_a2_targets(buf: bytes | bytearray, call_off: int):
    start = max(0, call_off - 0x50)
    end = min(len(buf) - 4, call_off + 4)  # include jal delay slot
    return mc_find_loaded_addr_for_reg(buf, start, end, MC_REG_A2)

def mc_find_a0_site_before(buf: bytes | bytearray, call_off: int):
    start = max(0, call_off - 0x60)
    for off in range(call_off - 4, start - 1, -4):
        if off < 0 or off + 4 > len(buf):
            continue
        word = read32(buf, off)
        if mc_is_addiu_reg_zero(word, MC_REG_A0) or mc_is_zero_to_a0(word):
            return off
    return None

def mc_patch_a0_site(buf: bytearray, off: int, new_x: int, name: str, source: str, force: bool, dry_run: bool, load_addr: int, extra: str = "") -> dict:
    old = read32(buf, off)
    expected_shape = mc_is_addiu_reg_zero(old, MC_REG_A0) or mc_is_zero_to_a0(old)
    new_word = mc_make_li_a0(new_x)
    if old == new_word:
        status = "already_applied"
    elif expected_shape or force:
        status = "would_patch" if dry_run else "patched"
        if not dry_run:
            mc_write32(buf, off, new_word)
    else:
        status = "skipped_unexpected_word"
    old_x = ""
    if mc_is_addiu_reg_zero(old, MC_REG_A0):
        old_x = str(s16(old & 0xFFFF))
    elif mc_is_zero_to_a0(old):
        old_x = "0"
    return {
        "site": name,
        "source": source,
        "status": status,
        "file_offset": f"0x{off:X}",
        "ram": f"0x{off2ram(off, load_addr):08X}",
        "old_word": f"0x{old:08X}",
        "new_word": f"0x{new_word:08X}",
        "old_x": old_x,
        "new_x": str(new_x),
        "note": extra,
    }

def mc_apply_known_sites(buf: bytearray, rows_by_ptr: dict[int, MCTextRow], load_addr: int, bias: int, force: bool, dry_run: bool):
    report = []
    patched_a0_offsets = set()
    for site in MC_KNOWN_A0_SITES:
        width, group_rows = mc_group_width(rows_by_ptr, site.pointers)
        if width is None or not group_rows:
            report.append({"site": site.name, "source": "known_site", "status": "skipped_no_workbook_row", "file_offset": "", "ram": f"0x{site.ram_a0:08X}", "old_word": "", "new_word": "", "old_x": "", "new_x": "", "note": site.note})
            continue
        off = ram2off(site.ram_a0, load_addr)
        if off < 0 or off + 4 > len(buf):
            report.append({"site": site.name, "source": "known_site", "status": "skipped_offset_out_of_range", "file_offset": f"0x{off:X}", "ram": f"0x{site.ram_a0:08X}", "old_word": "", "new_word": "", "old_x": "", "new_x": "", "note": site.note})
            continue
        new_x = mc_center_x_for_width(width, bias=bias)
        rows_txt = ",".join(str(r.excel_row) for r in group_rows)
        extra = f"rows={rows_txt}; width_px={width}"
        if site.note:
            extra += f"; {site.note}"
        row = mc_patch_a0_site(buf, off, new_x, site.name, "known_site", force, dry_run, load_addr, extra)
        report.append(row)
        if row["status"] in ("patched", "would_patch", "already_applied"):
            patched_a0_offsets.add(off)
    return report, patched_a0_offsets

def mc_apply_dynamic_scan(buf: bytearray, rows_by_ptr: dict[int, MCTextRow], load_addr: int, bias: int, force: bool, dry_run: bool, already: set[int]):
    report = []
    draw_jal = mc_jal_word(MC_FUN_DRAW_TEXT_XY)
    for call_off in range(0, len(buf) - 8, 4):
        word = read32(buf, call_off)
        if word != draw_jal:
            continue
        targets = mc_reconstruct_a2_targets(buf, call_off)
        matching = [(ptr, lui_off, add_off) for ptr, lui_off, add_off in targets if ptr in rows_by_ptr]
        if not matching:
            continue
        a0_off = mc_find_a0_site_before(buf, call_off)
        if a0_off is None:
            ptrs = sorted({m[0] for m in matching})
            report.append({"site": f"dynamic_call_{call_off:X}", "source": "dynamic_scan", "status": "skipped_no_a0_site_found", "file_offset": "", "ram": f"0x{off2ram(call_off, load_addr):08X}", "old_word": "", "new_word": "", "old_x": "", "new_x": "", "note": "pointers=" + ",".join(f"0x{p:08X}" for p in ptrs)})
            continue
        if a0_off in already:
            continue
        ptrs = tuple(sorted({m[0] for m in matching}))
        width, group_rows = mc_group_width(rows_by_ptr, ptrs)
        if width is None or not group_rows:
            continue
        new_x = mc_center_x_for_width(width, bias=bias)
        rows_txt = ",".join(str(r.excel_row) for r in group_rows)
        extra = f"rows={rows_txt}; width_px={width}; call_ram=0x{off2ram(call_off, load_addr):08X}; pointers=" + ",".join(f"0x{p:08X}" for p in ptrs)
        row = mc_patch_a0_site(buf, a0_off, new_x, f"dynamic_call_{call_off:X}", "dynamic_scan", force, dry_run, load_addr, extra)
        report.append(row)
        if row["status"] in ("patched", "would_patch", "already_applied"):
            already.add(a0_off)
    return report

def write_memory_card_centering_report_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["site", "source", "status", "file_offset", "ram", "old_word", "new_word", "old_x", "new_x", "note"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fields})

def patch_memory_card_centering(exe: bytearray, workbook_path, load_addr: int, *, bias: int = 0, force: bool = False, dry_run: bool = False, use_known_sites: bool = True, use_dynamic_scan: bool = True):
    rows_by_ptr = mc_load_text_rows(workbook_path)
    report = []
    patched_offsets = set()
    if use_known_sites:
        known_report, patched_offsets = mc_apply_known_sites(exe, rows_by_ptr, load_addr, bias, force, dry_run)
        report.extend(known_report)
    if use_dynamic_scan:
        report.extend(mc_apply_dynamic_scan(exe, rows_by_ptr, load_addr, bias, force, dry_run, patched_offsets))
    patched = sum(1 for r in report if r.get("status") in {"patched", "would_patch"})
    already = sum(1 for r in report if r.get("status") == "already_applied")
    skipped = sum(1 for r in report if str(r.get("status", "")).startswith("skipped"))
    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "rows_loaded": len(rows_by_ptr),
        "patched_sites": patched,
        "already_applied_sites": already,
        "skipped_sites": skipped,
        "x_bias": bias,
        "report": report,
    }


# ---------------------------------------------------------------------------
# TUTO full grouped-textbox patch (v67 stock tutorial text after XA subtitles + v66 close timing)
# ---------------------------------------------------------------------------
# Confirmed findings through v60:
#   * The stock tutorial remains an 11 top-level-group state machine.
#   * Each visible group owns one textbox lifecycle: flag0/pre-open, flag1/open,
#     one or more text/XA updates inside the already-open box, then mode8/close.
#   * Stock Group 0 was audio-only, so our inserted Group 0 textbox needs the
#     next visible group (Group 1) to perform the same flag0 pre-open cleanup
#     later visible groups use. This fixed the Group0->Group1 snap in v60.
#   * Group 0 and final Group 10 are special no-TYUTO-textbox-owner cases in the
#     original game. v61 gives them subtitle textboxes while keeping Group 0 free
#     of TYUTO obj helpers. Group 10 is handled as a final no-image subtitle group.
#   * Use --xa-text-xlsx to load/replace the Preferred English TUTO lines without
#     editing this patcher; if omitted, the built-in v61 defaults are used.

TUTO_CASE_JAL_RAMS = [0x80055E60 + 0x1C * i for i in range(11)]
TUTO_ORIG_HANDLER_RAMS = [
    0x80056E4C, 0x80057014, 0x8005738C, 0x8005765C, 0x80057924,
    0x80057CB4, 0x80057F7C, 0x80058244, 0x800585D4, 0x80058AD0,
    0x80058E9C,
]
TUTO_CASE_JAL_EXPECTS = [0x0C000000 | ((ram >> 2) & 0x03FFFFFF) for ram in TUTO_ORIG_HANDLER_RAMS]

TUTO_FN_PLAY_XA     = 0x80052190
TUTO_FN_CHECK_XA    = 0x8005234C
TUTO_FN_STOP_XA     = 0x8005228C
TUTO_FN_SET_TIMER   = 0x80051814
TUTO_FN_CHECK_TIMER = 0x8005183C
TUTO_FN_SET_BOX     = 0x8005073C
TUTO_FN_CLEAR_TEXT  = 0x800381FC
TUTO_FN_DRAW_TEXT   = 0x8003861C
TUTO_FN_ANIM_A      = 0x800590E0
TUTO_FN_ANIM_B      = 0x80059200
TUTO_FN_TUTO_MODE   = 0x8006E16C

TUTO_RAM_INPUT_WORD    = 0x800DAE7C
TUTO_RAM_CALLBACK_BASE = 0x800DA9A0
TUTO_RAM_TUTO01_PATH   = 0x8002134C
TUTO_RAM_TUTO02_PATH   = 0x800213D0

# Stock open/pre-open timers are 30 frames.  v66 leaves the open timer alone,
# but makes the previous-helper close/pre-open wait tunable because v65 proved
# the animation logic correct while the close felt slightly fast.
TUTO_OPEN_WAIT_FRAMES_DEFAULT = 30
TUTO_CLOSE_WAIT_FRAMES_DEFAULT = 40

# Group layout after the original 11 top-level handlers. Each entry owns one
# textbox lifecycle, except Group 0/final which are subtitle insertions over the
# simple Neredy/background presentation.
TUTO_GROUP_LAYOUT = [
    {"index": 0, "label": "intro_audio_subtitle", "xa": "TUTO01", "clips": [0, 1, 2], "obj": None, "prev_obj": None, "kind": "intro"},
    {"index": 1, "label": "tuto01_3_6",          "xa": "TUTO01", "clips": [3, 4, 5, 6], "obj": 0, "prev_obj": None, "kind": "group1"},
    {"index": 2, "label": "tuto01_7",            "xa": "TUTO01", "clips": [7], "obj": 1, "prev_obj": 0, "kind": "visual"},
    {"index": 3, "label": "tuto01_8",            "xa": "TUTO01", "clips": [8], "obj": 2, "prev_obj": 1, "kind": "visual"},
    {"index": 4, "label": "tuto01_9_10",         "xa": "TUTO01", "clips": [9, 10], "obj": 3, "prev_obj": 2, "kind": "visual"},
    {"index": 5, "label": "tuto02_0",            "xa": "TUTO02", "clips": [0], "obj": 4, "prev_obj": 3, "kind": "visual"},
    {"index": 6, "label": "tuto02_1",            "xa": "TUTO02", "clips": [1], "obj": 7, "prev_obj": 4, "kind": "visual"},
    {"index": 7, "label": "tuto02_2_3",          "xa": "TUTO02", "clips": [2, 3], "obj": 5, "prev_obj": 7, "kind": "visual"},
    {"index": 8, "label": "tuto02_4_6",          "xa": "TUTO02", "clips": [4, 5, 6], "obj": 1, "prev_obj": 5, "kind": "visual"},
    {"index": 9, "label": "tuto02_7_8",          "xa": "TUTO02", "clips": [7, 8], "obj": 6, "prev_obj": 1, "kind": "visual"},
    {"index": 10,"label": "tuto02_9_final",      "xa": "TUTO02", "clips": [9], "obj": None, "prev_obj": 6, "kind": "final"},
]

# Stock tutorial text pointer slots used by the original Japanese visible groups.
# These slots are already translated/relocated by the normal MAIN_EXE_Text patcher
# before the TUTO hook is appended, so v67 reloads the live slot value at runtime
# instead of embedding another copy of the English strings here.
TUTO_STOCK_TEXT_SLOTS_BY_GROUP = {
    1: [0x800CD92C],
    2: [0x800CD930],
    3: [0x800CD934],
    4: [0x800CD938],
    5: [0x800CD93C],
    6: [0x800CD940],
    7: [0x800CD944],
    8: [0x800CD948, 0x800CD94C, 0x800CD950],
    9: [0x800CD954, 0x800CD958],
}

DEFAULT_TUTO_XA_TEXT = {
    "TUTO01.XA[0]": "I will explain this project and\nthe development curriculum\nfor the Hectas star system.",
    "TUTO01.XA[1]": "The project's directive is to\nprepare the Hectas system's six\nplanets for human settlement.",
    "TUTO01.XA[2]": "Apologies for not introducing\nmyself sooner. My name is\nNeredy, this ship's navigator.",
    "TUTO01.XA[3]": "According to my data, each planet\nin the Hectas system appears to\nbe severely lacking in Spirit Energy.",
    "TUTO01.XA[4]": "The best method appears to be\nraising the planets' Spirit Energy\nthrough the power of Pixies.",
    "TUTO01.XA[5]": "As things stand, it is impossible to\nraise the Spirit Energy to a safe\nlevel before the colonists arrive.",
    "TUTO01.XA[6]": "It will be necessary to raise\nPixies and evolve them into\nhigher species.",
    "TUTO01.XA[7]": "Pixies are raised inside Gardens, of\nwhich we have prepared five. Each\nGarden is suited to different Pixies.",
    "TUTO01.XA[8]": "Pixies evolve by absorbing the Earth,\nWater, Wind, and Fire Spirit Energy\nfound in Objects placed in a Garden.",
    "TUTO01.XA[9]": "The Spirit Energy in Objects cannot\nbe given directly to Pixies. It is\nsupplied to them via Transfers.",
    "TUTO01.XA[10]": "Place a Pixy and an Object it is\nlikely to like within a Transfer's\neffective range.",
    "TUTO02.XA[0]": "Temperature is important, too. Use\na Garden's thermometer to find\neach species' preferred temperature.",
    "TUTO02.XA[1]": "By setting the Music in a Garden,\nyou can make raising Pixies easier.\nLearn the tracks and use them well.",
    "TUTO02.XA[2]": "When the environment in the Garden\nis suitable, a Pixy may produce a\nModus.",
    "TUTO02.XA[3]": "When placed on a planet, a Modus can\nproduce a development effect equal\nto that of the Pixy that created it.",
    "TUTO02.XA[4]": "A mysterious power called a Blessing\ncan sometimes grant additional power\nto Objects.",
    "TUTO02.XA[5]": "If the environment in the Garden is\nnot suitable, Pixies may behave\nstrangely, so please be careful.",
    "TUTO02.XA[6]": "Once a Pixy's Spirit Energy has\nreached a certain level, please\nevolve it when ready, Master.",
    "TUTO02.XA[7]": "Once a Pixy has evolved, deploy it\nto a planet and raise that planet's\nSpirit Energy.",
    "TUTO02.XA[8]": "If any phenomenon occurs, I will\nreport it. We have much to learn, so\nit is wise to monitor the Pixies.",
    "TUTO02.XA[9]": "This concludes the project briefing.\nThe colonists will arrive soon.\nMaster, please do your best.",
}

TUTO_REQUIRED_CLIP_IDS = []
for _g in TUTO_GROUP_LAYOUT:
    for _clip in _g["clips"]:
        TUTO_REQUIRED_CLIP_IDS.append(f'{_g["xa"]}.XA[{_clip}]')


def normalize_tuto_xa_text(s: str) -> str:
    s = norm(s)
    # Keep runtime renderer input ASCII-safe where possible. The renderer can
    # pass CP932 for non-ASCII, but these punctuation marks are not needed and
    # can turn into replacement glyphs on the PS1 side.
    return (s.replace("\u2018", "'").replace("\u2019", "'")
             .replace("\u201c", '"').replace("\u201d", '"')
             .replace("\u2013", "-").replace("\u2014", "-")
             .replace("\u2026", "..."))


def load_tuto_xa_texts_from_xlsx(path, sheet_name="Updated Translations", clip_id_column="Clip ID", text_column="Preferred English"):
    wb = load_workbook(xlsx_path(path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"XA text sheet {sheet_name!r} not found in {path}; available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    clip_col = hdr(ws, [clip_id_column])
    text_col = hdr(ws, [text_column])
    group_col = hdr(ws, ["Group"])
    if not clip_col:
        raise RuntimeError(f"XA text sheet {sheet_name!r} has no {clip_id_column!r} column")
    if not text_col:
        raise RuntimeError(f"XA text sheet {sheet_name!r} has no {text_column!r} column")
    out = dict(DEFAULT_TUTO_XA_TEXT)
    loaded = 0
    for row in range(2, ws.max_row + 1):
        clip_id = ws.cell(row, clip_col).value
        if clip_id is None:
            continue
        clip_id = str(clip_id).strip()
        if clip_id not in TUTO_REQUIRED_CLIP_IDS:
            continue
        if group_col and str(ws.cell(row, group_col).value or "").strip().upper() not in {"TUTO", ""}:
            continue
        txt = ws.cell(row, text_col).value
        if txt is None or str(txt).strip() == "":
            raise RuntimeError(f"Missing XA Preferred English text for {clip_id} on sheet row {row}")
        out[clip_id] = normalize_tuto_xa_text(str(txt))
        loaded += 1
    missing = [cid for cid in TUTO_REQUIRED_CLIP_IDS if cid not in out or not out[cid]]
    if missing:
        raise RuntimeError("Missing required TUTO XA text rows: " + ", ".join(missing))
    return out, loaded


def resolve_tuto_xa_texts(xa_text_xlsx=None, sheet_name="Updated Translations", clip_id_column="Clip ID", text_column="Preferred English"):
    if xa_text_xlsx:
        return load_tuto_xa_texts_from_xlsx(xa_text_xlsx, sheet_name, clip_id_column, text_column)
    return {k: normalize_tuto_xa_text(v) for k, v in DEFAULT_TUTO_XA_TEXT.items()}, 0


def enc_tuto_text(s: str) -> bytes:
    # Stored as plain text with raw LF 0x0A line breaks, then NUL terminated by
    # the blob builder. This matches the confirmed working MAIN text path.
    return enc_text(normalize_tuto_xa_text(s))


TUTO_REG = {
    'zero':0,'at':1,'v0':2,'v1':3,'a0':4,'a1':5,'a2':6,'a3':7,
    't0':8,'t1':9,'t2':10,'t3':11,'t4':12,'t5':13,'t6':14,'t7':15,
    's0':16,'s1':17,'s2':18,'s3':19,'s4':20,'s5':21,'s6':22,'s7':23,
    't8':24,'t9':25,'k0':26,'k1':27,'gp':28,'sp':29,'fp':30,'ra':31,
}

def tuto_sx16(x:int)->int:
    x &= 0xFFFF
    return x - 0x10000 if x & 0x8000 else x

def tuto_r(rs, rt, rd, sh, fn): return (rs<<21)|(rt<<16)|(rd<<11)|(sh<<6)|fn
def tuto_i(op, rs, rt, imm): return (op<<26)|(rs<<21)|(rt<<16)|(imm & 0xFFFF)
def tuto_j(op, addr): return (op<<26)|((addr>>2)&0x03FFFFFF)
def tuto_W(x): return struct.pack('<I', x & 0xFFFFFFFF)
def tuto_addu(rd, rs, rt): return tuto_r(TUTO_REG[rs], TUTO_REG[rt], TUTO_REG[rd], 0, 0x21)
def tuto_sll(rd, rt, sh): return tuto_r(0, TUTO_REG[rt], TUTO_REG[rd], sh, 0x00)
def tuto_jr(rs): return tuto_r(TUTO_REG[rs], 0, 0, 0, 0x08)
def tuto_jalr(rd, rs): return tuto_r(TUTO_REG[rs], 0, TUTO_REG[rd], 0, 0x09)
def tuto_addiu(rt, rs, imm): return tuto_i(0x09, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_lui(rt, imm): return tuto_i(0x0F, 0, TUTO_REG[rt], imm)
def tuto_andi(rt, rs, imm): return tuto_i(0x0C, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_sltiu(rt, rs, imm): return tuto_i(0x0B, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_beq(rs, rt, off): return tuto_i(0x04, TUTO_REG[rs], TUTO_REG[rt], off)
def tuto_bne(rs, rt, off): return tuto_i(0x05, TUTO_REG[rs], TUTO_REG[rt], off)
def tuto_lw(rt, imm, rs): return tuto_i(0x23, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_lhu(rt, imm, rs): return tuto_i(0x25, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_lh(rt, imm, rs): return tuto_i(0x21, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_sw(rt, imm, rs): return tuto_i(0x2B, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_sh(rt, imm, rs): return tuto_i(0x29, TUTO_REG[rs], TUTO_REG[rt], imm)
def tuto_J(addr): return tuto_j(0x02, addr)
def tuto_JAL(addr): return tuto_j(0x03, addr)
TUTO_NOP = 0

def tuto_hi_lo(addr:int):
    hi_part = (addr + 0x8000) >> 16
    lo_part = tuto_sx16(addr)
    return hi_part & 0xFFFF, lo_part

def tuto_li(reg:str, value:int):
    hi_part, lo_part = tuto_hi_lo(value)
    return [tuto_lui(reg, hi_part), tuto_addiu(reg, reg, lo_part)]

class TutoAsm:
    """Assembler for the custom tutorial handlers.

    Like class A above, this automatically protects generated code from PS1
    R3000A load-delay hazards.  Many TUTO paths already emitted manual NOPs;
    this keeps those intact and only inserts a NOP where the next instruction
    was not already an explicit delay NOP.
    """
    LOAD_OPS = {0x20, 0x21, 0x23, 0x24, 0x25}  # lb, lh, lw, lbu, lhu

    def __init__(self, base:int):
        self.base = base
        self.words = []
        self.labels = {}
        self.fixups = []
        self._load_delay_pending = False
    @property
    def pc(self): return self.base + len(self.words)*4
    @classmethod
    def _is_load_word(cls, word:int) -> bool:
        return ((word >> 26) & 0x3F) in cls.LOAD_OPS
    def _before_instruction(self, next_word=None):
        if not self._load_delay_pending:
            return
        if next_word == TUTO_NOP:
            self._load_delay_pending = False
        else:
            self.words.append(TUTO_NOP)
            self._load_delay_pending = False
    def label(self, name):
        if name in self.labels: raise ValueError(f'duplicate TUTO label {name}')
        # If the previous emitted instruction was a load, this label will point
        # at the automatically inserted delay NOP that precedes the next real
        # instruction. That is intentional and keeps branch targets safe.
        self.labels[name] = self.pc
    def emit(self, word):
        word &= 0xFFFFFFFF
        self._before_instruction(word)
        self.words.append(word)
        self._load_delay_pending = self._is_load_word(word)
    def emit_li(self, reg, value):
        for word in tuto_li(reg, value): self.emit(word)
    def emit_j(self, addr): self.emit(tuto_J(addr))
    def emit_jal(self, addr): self.emit(tuto_JAL(addr))
    def branch(self, kind, rs, rt, label):
        self._before_instruction()
        pos = len(self.words); self.words.append(0); self.fixups.append((pos, kind, rs, rt, label))
    def resolve(self):
        for pos, kind, rs, rt, label in self.fixups:
            target = self.labels[label]
            pc = self.base + pos*4
            off = (target - (pc + 4)) // 4
            if not -32768 <= off <= 32767:
                raise ValueError(f'TUTO branch to {label} out of range')
            self.words[pos] = (tuto_beq if kind == 'beq' else tuto_bne)(rs, rt, off)
        return b''.join(tuto_W(word) for word in self.words)


def _emit_set_state(a:TutoAsm, value:int):
    a.emit(tuto_addiu('v0','zero',int(value)))
    a.emit(tuto_sh('v0',0x06B8,'gp'))

def _emit_inc_state(a:TutoAsm):
    a.emit(tuto_lhu('v0',0x06B8,'gp')); a.emit(TUTO_NOP)
    a.emit(tuto_addiu('v0','v0',1)); a.emit(tuto_sh('v0',0x06B8,'gp'))

def _emit_return_active(a:TutoAsm):
    a.emit(tuto_addiu('v0','zero',1)); a.branch('beq','zero','zero','epilogue'); a.emit(TUTO_NOP)

def _emit_return_complete(a:TutoAsm):
    a.emit(tuto_addu('v0','zero','zero')); a.branch('beq','zero','zero','epilogue'); a.emit(TUTO_NOP)

def _emit_check_timer_active(a:TutoAsm, active_label='ret_active'):
    a.emit_jal(TUTO_FN_CHECK_TIMER); a.emit(TUTO_NOP)
    a.branch('bne','v0','zero',active_label); a.emit(TUTO_NOP)

def _emit_textbox_rect(a:TutoAsm, flag16:int=1):
    # Same rectangle used by the stock visible tutorial routines.
    a.emit(tuto_addiu('a0','zero',0x01C0))
    a.emit(tuto_addiu('a1','zero',0x0100))
    a.emit(tuto_addu('a2','a0','zero'))
    a.emit(tuto_addiu('a3','zero',0x0140))
    a.emit(tuto_addiu('v0','zero',1))
    a.emit(tuto_sw('v0' if flag16 else 'zero',16,'sp'))
    a.emit(tuto_sw('v0',20,'sp'))
    a.emit_jal(TUTO_FN_SET_BOX); a.emit(TUTO_NOP)

def _emit_obj_helpers(a:TutoAsm, obj_id:int, active:int=1):
    a.emit(tuto_addiu('a0','zero',int(obj_id)))
    a.emit(tuto_addiu('a1','zero',int(active)))
    a.emit_jal(TUTO_FN_ANIM_A); a.emit(TUTO_NOP)
    a.emit(tuto_addiu('a0','zero',int(obj_id)))
    a.emit(tuto_addiu('a1','zero',int(active)))
    a.emit_jal(TUTO_FN_ANIM_B); a.emit(TUTO_NOP)

def _emit_active_group_refresh(a:TutoAsm, obj_id:int):
    _emit_obj_helpers(a, obj_id, 1)
    _emit_textbox_rect(a, 1)
    a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)

def _emit_draw_textvar(a:TutoAsm, textvar:int):
    a.emit_li('t0', textvar)
    a.emit(tuto_lw('a0',0,'t0'))
    label = 'skip_draw_%08X_%d' % (textvar, len(a.words))
    a.branch('beq','a0','zero',label); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_DRAW_TEXT); a.emit(TUTO_NOP)
    a.label(label)

def _emit_draw_stock_textslot(a:TutoAsm, slot_ram:int):
    # slot_ram points to the original tutorial text pointer word, such as
    # 0x800CD92C.  The main string patcher relocates the word to English text,
    # so this reloads the current patched pointer at runtime.
    a.emit_li('t0', int(slot_ram))
    a.emit(tuto_lw('a0',0,'t0'))
    label = 'skip_stock_draw_%08X_%d' % (slot_ram, len(a.words))
    a.branch('beq','a0','zero',label); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_DRAW_TEXT); a.emit(TUTO_NOP)
    a.label(label)

def _emit_play_xa(a:TutoAsm, xa_name:str, clip:int):
    a.emit_li('a0', TUTO_RAM_TUTO01_PATH if xa_name == 'TUTO01' else TUTO_RAM_TUTO02_PATH)
    a.emit(tuto_addiu('a1','zero',int(clip)))
    a.emit_jal(TUTO_FN_PLAY_XA); a.emit(TUTO_NOP)

def _emit_confirm_check(a:TutoAsm, confirm_label:str):
    a.emit_li('t0', TUTO_RAM_INPUT_WORD)
    a.emit(tuto_lhu('t0',0,'t0')); a.emit(TUTO_NOP)
    a.emit(tuto_andi('t0','t0',0x2000))
    a.branch('beq','t0','zero',confirm_label); a.emit(TUTO_NOP)

def _emit_confirm_callback(a:TutoAsm):
    a.emit_li('v0', TUTO_RAM_CALLBACK_BASE)
    a.emit(tuto_lw('v0',0,'v0')); a.emit(TUTO_NOP)
    a.emit(tuto_lw('v0',12,'v0')); a.emit(TUTO_NOP)
    a.emit(tuto_jalr('ra','v0'))
    a.emit(tuto_addiu('a0','zero',1))


def _emit_intro_keepalive(a:TutoAsm):
    _emit_textbox_rect(a, 1)
    a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)


def build_tuto_intro_handler(base:int, group_text_vars:list[int]) -> bytes:
    """Replacement for original audio-only Group 0, with subtitles.

    This keeps the confirmed v60 Group0->Group1 behavior: Group 0 issues mode8
    and returns active for one frame; Group 1 then performs flag0 pre-open cleanup.
    """
    a = TutoAsm(base)
    a.emit(tuto_addiu('sp','sp',-32)); a.emit(tuto_sw('ra',24,'sp'))
    a.emit(tuto_lh('v1',0x06B8,'gp'))
    for st in range(0, 11):
        a.emit(tuto_addiu('v0','zero',st)); a.branch('beq','v1','v0',f'state{st}'); a.emit(TUTO_NOP)
    a.branch('beq','zero','zero','ret_complete'); a.emit(TUTO_NOP)

    def wait_xa_then_confirm(next_state:int, *, clear_on_confirm=True, extra_after_confirm=None):
        _emit_intro_keepalive(a)
        local_confirm = f'g0_confirm_{next_state}_{len(a.words)}'
        _emit_confirm_check(a, local_confirm)
        a.emit_jal(TUTO_FN_CHECK_XA); a.emit(TUTO_NOP)
        a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0007))
        _emit_return_active(a)
        a.label(local_confirm)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        _emit_confirm_callback(a)
        if clear_on_confirm:
            a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)
        if extra_after_confirm:
            extra_after_confirm()
        _emit_set_state(a, next_state); _emit_return_active(a)

    a.label('state0')
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x0014))
    _emit_set_state(a, 1); _emit_return_active(a)

    a.label('state1')
    _emit_textbox_rect(a, 0)
    _emit_check_timer_active(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0001))
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x001E))
    _emit_set_state(a, 2); _emit_return_active(a)

    a.label('state2')
    _emit_textbox_rect(a, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    _emit_check_timer_active(a)
    _emit_play_xa(a, 'TUTO01', 0)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    _emit_draw_textvar(a, group_text_vars[0])
    _emit_set_state(a, 3); _emit_return_active(a)

    a.label('state3')
    wait_xa_then_confirm(4, clear_on_confirm=True)

    a.label('state4')
    _emit_textbox_rect(a, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    _emit_play_xa(a, 'TUTO01', 1)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    _emit_draw_textvar(a, group_text_vars[1])
    _emit_set_state(a, 5); _emit_return_active(a)

    a.label('state5')
    def after_clip1_confirm():
        a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0007))
        a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x005A))
    wait_xa_then_confirm(6, clear_on_confirm=True, extra_after_confirm=after_clip1_confirm)

    a.label('state6')
    _emit_intro_keepalive(a)
    _emit_check_timer_active(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x000A))
    _emit_set_state(a, 7); _emit_return_active(a)

    a.label('state7')
    _emit_textbox_rect(a, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)
    _emit_check_timer_active(a)
    _emit_play_xa(a, 'TUTO01', 2)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    _emit_draw_textvar(a, group_text_vars[2])
    _emit_set_state(a, 8); _emit_return_active(a)

    a.label('state8')
    wait_xa_then_confirm(9, clear_on_confirm=False)

    a.label('state9')
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0008))
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_set_state(a, 10); _emit_return_active(a)

    a.label('state10')
    _emit_return_complete(a)

    a.label('ret_active'); _emit_return_active(a)
    a.label('ret_complete'); _emit_return_complete(a)
    a.label('epilogue')
    a.emit(tuto_lw('ra',24,'sp'))
    a.emit(tuto_jr('ra')); a.emit(tuto_addiu('sp','sp',32))
    return a.resolve()


def build_tuto_visual_group_handler(base:int, xa_name:str, clips:list[int], text_vars:list[int], obj_id:int, prev_obj:int|None=None, *, first_group_after_intro:bool=False, close_wait_frames:int=TUTO_CLOSE_WAIT_FRAMES_DEFAULT, stock_text_slots:list[int]|None=None) -> bytes:
    """Generic visible-group handler: XA subtitles first, then original stock tutorial text, then close.

    v67 keeps v66's confirmed textbox/object-helper timing, but adds a post-speech
    phase inside each visible stock group. After the last Neredy/XA subtitle in
    the group, the handler draws the original tutorial text pointer(s) that the
    Japanese routine used. Those pointer slots are already handled by the normal
    MAIN_EXE_Text patcher, so they become the current English tutorial text.
    """
    if len(clips) != len(text_vars):
        raise ValueError(f"TUTO visual group clip/text length mismatch: {xa_name} {clips} vs {len(text_vars)} text vars")
    stock_text_slots = list(stock_text_slots or [])
    n = len(clips)
    m = len(stock_text_slots)
    close_wait_frames = max(1, min(int(close_wait_frames), 300))

    # State numbering:
    #   0,1,2 = open/pre-open/draw first XA subtitle
    #   each XA clip has 4 states as in v66
    #   after the final XA post-confirm, v67 inserts 2 states per stock text:
    #       draw stock text, wait Confirm
    #   then the existing close/complete states run.
    final_start = 3 + (n - 1) * 4
    stock_start = final_start + 3
    close_state = stock_start + (m * 2)
    complete_state = close_state + 1
    max_state = complete_state

    a = TutoAsm(base)
    a.emit(tuto_addiu('sp','sp',-32)); a.emit(tuto_sw('ra',24,'sp'))
    a.emit(tuto_lh('v1',0x06B8,'gp'))
    for st in range(0, max_state + 1):
        a.emit(tuto_addiu('v0','zero',st)); a.branch('beq','v1','v0',f'state{st}'); a.emit(TUTO_NOP)
    a.branch('beq','zero','zero','ret_complete'); a.emit(TUTO_NOP)

    a.label('state0')
    # If this group follows a previous helper image, this timer is the old
    # image's close/pre-open wait.  If not, keep the stock 30-frame open wait.
    pre_open_wait = close_wait_frames if prev_obj is not None else TUTO_OPEN_WAIT_FRAMES_DEFAULT
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',pre_open_wait))
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0003))
    _emit_textbox_rect(a, 0)
    if prev_obj is not None:
        _emit_obj_helpers(a, int(prev_obj), 0)
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_set_state(a, 1); _emit_return_active(a)

    a.label('state1')
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_textbox_rect(a, 0)
    if prev_obj is not None:
        _emit_obj_helpers(a, int(prev_obj), 0)
    _emit_check_timer_active(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addu('a0','zero','zero'))
    _emit_obj_helpers(a, obj_id, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0001))
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x001E))
    _emit_set_state(a, 2); _emit_return_active(a)

    a.label('state2')
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_obj_helpers(a, obj_id, 1)
    _emit_textbox_rect(a, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    _emit_draw_textvar(a, text_vars[0])
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x000A))
    _emit_set_state(a, 3); _emit_return_active(a)

    def emit_wait_playing(playing_state:int, post_state:int, next_state:int):
        confirm_label = f'confirm_{playing_state}_{len(a.words)}'
        _emit_active_group_refresh(a, obj_id)
        _emit_confirm_check(a, confirm_label)
        a.emit_jal(TUTO_FN_CHECK_XA); a.emit(TUTO_NOP)
        a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0007))
        _emit_set_state(a, post_state); _emit_return_active(a)
        a.label(confirm_label)
        _emit_confirm_callback(a)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        _emit_set_state(a, next_state); _emit_return_active(a)

    def emit_wait_post(post_state:int, next_state:int):
        confirm_label = f'post_confirm_{post_state}_{len(a.words)}'
        _emit_active_group_refresh(a, obj_id)
        _emit_confirm_check(a, confirm_label)
        _emit_return_active(a)
        a.label(confirm_label)
        _emit_confirm_callback(a)
        _emit_set_state(a, next_state); _emit_return_active(a)

    for i, clip in enumerate(clips):
        start_state = 3 + i * 4
        playing_state = start_state + 1
        post_state = start_state + 2
        is_final = i == n - 1
        next_state = stock_start if (is_final and m) else (close_state if is_final else start_state + 3)

        a.label(f'state{start_state}')
        _emit_active_group_refresh(a, obj_id)
        _emit_check_timer_active(a)
        _emit_play_xa(a, xa_name, clip)
        _emit_set_state(a, playing_state); _emit_return_active(a)

        a.label(f'state{playing_state}')
        emit_wait_playing(playing_state, post_state, next_state)

        a.label(f'state{post_state}')
        emit_wait_post(post_state, next_state)

        if not is_final:
            update_state = start_state + 3
            a.label(f'state{update_state}')
            _emit_active_group_refresh(a, obj_id)
            a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)
            a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
            _emit_draw_textvar(a, text_vars[i + 1])
            a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x000A))
            _emit_set_state(a, update_state + 1); _emit_return_active(a)

    for j, slot in enumerate(stock_text_slots):
        draw_state = stock_start + j * 2
        wait_state = draw_state + 1
        next_state = close_state if j == m - 1 else draw_state + 2

        a.label(f'state{draw_state}')
        _emit_active_group_refresh(a, obj_id)
        a.emit_jal(TUTO_FN_CLEAR_TEXT); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
        _emit_draw_stock_textslot(a, slot)
        _emit_set_state(a, wait_state); _emit_return_active(a)

        a.label(f'state{wait_state}')
        confirm_label = f'stock_confirm_{j}_{len(a.words)}'
        _emit_active_group_refresh(a, obj_id)
        _emit_confirm_check(a, confirm_label)
        _emit_return_active(a)
        a.label(confirm_label)
        _emit_confirm_callback(a)
        _emit_set_state(a, next_state); _emit_return_active(a)

    a.label(f'state{close_state}')
    _emit_active_group_refresh(a, obj_id)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0008))
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_set_state(a, complete_state); _emit_return_active(a)

    a.label(f'state{complete_state}')
    _emit_return_complete(a)

    a.label('ret_active'); _emit_return_active(a)
    a.label('ret_complete'); _emit_return_complete(a)
    a.label('epilogue')
    a.emit(tuto_lw('ra',24,'sp'))
    a.emit(tuto_jr('ra')); a.emit(tuto_addiu('sp','sp',32))
    return a.resolve()

def build_tuto_final_subtitle_handler(base:int, text_vars:list[int]) -> bytes:
    """Replacement for original ending Group 10, with a subtitle textbox but no TYUTO image.

    v61 treated the final Neredy-only clip as another visual helper group using
    helper id 6, which could make the previous image pop back onscreen.  v64 keeps
    the v63 final no-image fix and is
    deliberately conservative after the v62 active=0 helper experiment stalled
    drawing: it does not call any obj_helpers here.  Group 9 owns/finishes its
    image close; Group 10 only owns the final textbox and Neredy/audio timing.
    """
    if len(text_vars) != 1:
        raise ValueError('Final TUTO subtitle handler expects exactly one text var')
    a = TutoAsm(base)
    a.emit(tuto_addiu('sp','sp',-32)); a.emit(tuto_sw('ra',24,'sp'))
    a.emit(tuto_lh('v1',0x06B8,'gp'))
    for st in range(0, 9):
        a.emit(tuto_addiu('v0','zero',st)); a.branch('beq','v1','v0',f'state{st}'); a.emit(TUTO_NOP)
    a.branch('beq','zero','zero','ret_complete'); a.emit(TUTO_NOP)

    a.label('state0')
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x001E))
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0003))
    _emit_textbox_rect(a, 0)
    _emit_set_state(a, 1); _emit_return_active(a)

    a.label('state1')
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_textbox_rect(a, 0)
    _emit_check_timer_active(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addu('a0','zero','zero'))
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0001))
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x001E))
    _emit_set_state(a, 2); _emit_return_active(a)

    a.label('state2')
    _emit_textbox_rect(a, 1)
    a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0002))
    _emit_draw_textvar(a, text_vars[0])
    a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x000A))
    _emit_set_state(a, 3); _emit_return_active(a)

    a.label('state3')
    _emit_intro_keepalive(a)
    _emit_check_timer_active(a)
    _emit_play_xa(a, 'TUTO02', 9)
    _emit_set_state(a, 4); _emit_return_active(a)

    def final_confirm_or_wait(next_state:int):
        confirm_label = f'final_confirm_{len(a.words)}'
        _emit_intro_keepalive(a)
        _emit_confirm_check(a, confirm_label)
        a.emit_jal(TUTO_FN_CHECK_XA); a.emit(TUTO_NOP)
        a.branch('bne','v0','zero','ret_active'); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0007))
        a.emit_jal(TUTO_FN_SET_TIMER); a.emit(tuto_addiu('a0','zero',0x005A))
        _emit_set_state(a, next_state); _emit_return_active(a)
        a.label(confirm_label)
        _emit_confirm_callback(a)
        a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
        _emit_set_state(a, next_state); _emit_return_active(a)

    a.label('state4')
    final_confirm_or_wait(5)

    a.label('state5')
    _emit_intro_keepalive(a)
    _emit_check_timer_active(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addu('a0','zero','zero'))
    _emit_set_state(a, 6); _emit_return_active(a)

    a.label('state6')
    _emit_intro_keepalive(a)
    a.emit_jal(TUTO_FN_TUTO_MODE); a.emit(tuto_addiu('a0','zero',0x0008))
    a.emit_jal(TUTO_FN_STOP_XA); a.emit(TUTO_NOP)
    _emit_set_state(a, 7); _emit_return_active(a)

    a.label('state7')
    _emit_return_complete(a)

    a.label('state8')
    _emit_return_complete(a)

    a.label('ret_active'); _emit_return_active(a)
    a.label('ret_complete'); _emit_return_complete(a)
    a.label('epilogue')
    a.emit(tuto_lw('ra',24,'sp'))
    a.emit(tuto_jr('ra')); a.emit(tuto_addiu('sp','sp',32))
    return a.resolve()


def build_tuto_group_3_6_handler(base:int, group1_text_vars:list[int]) -> bytes:
    # Keep Group 1 on the generic visible lifecycle, but retain the v60-critical
    # flag0 pre-open cleanup because Group 0 now owns an inserted textbox.
    return build_tuto_visual_group_handler(
        base, 'TUTO01', [3, 4, 5, 6], group1_text_vars, obj_id=0, prev_obj=None, first_group_after_intro=True,
        stock_text_slots=TUTO_STOCK_TEXT_SLOTS_BY_GROUP.get(1, [])
    )


def tuto_patch_word(exe:bytearray, load:int, ram:int, word:int, label:str, expected:int|None=None, force:bool=False, dry_run:bool=False):
    off = ram2off(ram, load)
    if off < 0 or off+4 > len(exe):
        raise RuntimeError(f'TUTO patch address outside EXE: RAM 0x{ram:08X}, file 0x{off:X}, site {label}')
    old = read32(exe, off)
    if expected is not None and old != expected and old != word and not force:
        raise RuntimeError(
            f'TUTO patch site {label} mismatch at RAM 0x{ram:08X}/file 0x{off:X}: '
            f'found 0x{old:08X}, expected 0x{expected:08X}. Use --tuto21-force only if intentional.'
        )
    if not dry_run:
        exe[off:off+4] = tuto_W(word)
    return {
        'site': label,
        'ram': f'0x{ram:08X}',
        'file_offset': f'0x{off:X}',
        'old_word': f'0x{old:08X}',
        'new_word': f'0x{word & 0xFFFFFFFF:08X}',
        'status': 'already_applied' if old == word else ('would_patch' if dry_run else 'patched'),
    }


def _group_clip_id(group:dict, clip:int) -> str:
    return f'{group["xa"]}.XA[{clip}]'


def _build_tuto_text_blob(start_ram:int, text_by_clip:dict[str,str]):
    """Build pointer table + strings for all 21 TUTO XA text entries.

    Returns (blob, group_text_vars), where group_text_vars[i] is the pointer-word
    list for top-level group i.
    """
    ordered_texts = []
    group_vars = []
    for group in TUTO_GROUP_LAYOUT:
        vars_for_group = []
        for clip in group['clips']:
            cid = _group_clip_id(group, clip)
            if cid not in text_by_clip or not text_by_clip[cid]:
                raise RuntimeError(f'Missing TUTO XA text for {cid}')
            ordered_texts.append((cid, text_by_clip[cid], vars_for_group))
        group_vars.append(vars_for_group)

    ptr_table_size = len(ordered_texts) * 4
    blob = bytearray(b'\0' * ptr_table_size)
    string_addrs = []

    for idx, (cid, text, dest_vars) in enumerate(ordered_texts):
        ptr_word_ram = start_ram + idx * 4
        dest_vars.append(ptr_word_ram)
        while len(blob) % 4:
            blob.append(0)
        text_ram = start_ram + len(blob)
        string_addrs.append(text_ram)
        blob.extend(enc_tuto_text(text) + b'\0')

    for i, addr in enumerate(string_addrs):
        struct.pack_into('<I', blob, i * 4, addr & 0xFFFFFFFF)
    while len(blob) % 4:
        blob.append(0)
    return bytes(blob), group_vars


def patch_tuto21_flow(exe:bytearray, load:int, *, xa_text_xlsx=None, xa_text_sheet='Updated Translations', xa_clip_id_column='Clip ID', xa_text_column='Preferred English', tuto_close_wait_frames:int=TUTO_CLOSE_WAIT_FRAMES_DEFAULT, force:bool=False, dry_run:bool=False):
    """Apply v66 grouped TUTO patch: all 11 top-level groups preserved/replaced."""
    if load != 0x80020000 and not force:
        raise RuntimeError(f'TUTO state-split patch expected load address 0x80020000, found 0x{load:08X}. Use --tuto21-force only if intentional.')

    text_by_clip, external_rows = resolve_tuto_xa_texts(xa_text_xlsx, xa_text_sheet, xa_clip_id_column, xa_text_column)
    tuto_close_wait_frames = max(1, min(int(tuto_close_wait_frames), 300))

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)

    text_blob, group_text_vars = _build_tuto_text_blob(hook_ram, text_by_clip)
    payload = bytearray(text_blob)
    handler_addrs = []
    handler_blobs = []

    addr = hook_ram + len(payload)
    for group in TUTO_GROUP_LAYOUT:
        if group['kind'] == 'intro':
            code = build_tuto_intro_handler(addr, group_text_vars[group['index']])
        elif group['kind'] == 'group1':
            code = build_tuto_group_3_6_handler(addr, group_text_vars[group['index']])
        elif group['kind'] == 'final':
            code = build_tuto_final_subtitle_handler(addr, group_text_vars[group['index']])
        else:
            code = build_tuto_visual_group_handler(
                addr, group['xa'], group['clips'], group_text_vars[group['index']],
                obj_id=group['obj'], prev_obj=group.get('prev_obj'),
                close_wait_frames=tuto_close_wait_frames,
                stock_text_slots=TUTO_STOCK_TEXT_SLOTS_BY_GROUP.get(group['index'], [])
            )
        handler_addrs.append(addr)
        handler_blobs.append(code)
        payload += code
        addr += len(code)

    while len(payload) % 4:
        payload += b'\0'

    report = []
    for i, handler_addr in enumerate(handler_addrs):
        report.append(tuto_patch_word(
            exe, load, TUTO_CASE_JAL_RAMS[i], tuto_JAL(handler_addr),
            f'dispatcher_case_{i}_jal_{TUTO_GROUP_LAYOUT[i]["label"]}_v67',
            TUTO_CASE_JAL_EXPECTS[i], force, dry_run
        ))

    if not dry_run:
        if len(exe) < hook_off:
            exe += b'\0' * (hook_off - len(exe))
        exe[hook_off:hook_off+len(payload)] = payload

    return {
        'enabled': True,
        'status': 'would_patch' if dry_run else 'patched',
        'hook_file': f'0x{hook_off:X}',
        'hook_ram': f'0x{hook_ram:08X}',
        'hook_len': len(payload),
        'strategy': f'v67 full grouped TUTO expansion: v66 stable previous-helper active=0 close service with tunable close wait={tuto_close_wait_frames} frames; v60 Group0->Group1 transition; final no-image subtitle group; original stock tutorial text shown after XA subtitles; XA text can be loaded with --xa-text-xlsx',
        'xa_text_source': str(xa_text_xlsx) if xa_text_xlsx else 'built_in_v61_defaults',
        'xa_text_rows_loaded': external_rows,
        'close_wait_frames': tuto_close_wait_frames,
        'text_entries': len(TUTO_REQUIRED_CLIP_IDS),
        'stock_text_slots': {str(k): [f'0x{x:08X}' for x in v] for k, v in TUTO_STOCK_TEXT_SLOTS_BY_GROUP.items()},
        'handlers_replaced': [
            {
                'top_index': g['index'],
                'original': f'0x{TUTO_ORIG_HANDLER_RAMS[g["index"]]:08X}',
                'new_ram': f'0x{handler_addrs[g["index"]]:08X}',
                'xa': g['xa'],
                'clips': g['clips'],
                'kind': g['kind'],
                'obj': g.get('obj'),
                'prev_obj': g.get('prev_obj'),
            } for g in TUTO_GROUP_LAYOUT
        ],
        'patch_sites': report,
    }




# Guarded small grey Yes/No / ON/OFF selector box hook (v109).
#
# v105 confirmed the correct rendering path: small grey selector boxes are
# runtime rectangles drawn through FUN_80051780. v108 keeps that hook and
# lets the hook distinguish the Vibration Function ON/OFF screen from normal
# Yes/No prompts by the original caller RA, with separate vertical tuning.
#
# Original Japanese selector-box shapes:
#   left  option: x=-12, w=24, h=13   (はい / ON side)
#   right option: x= 24, w=36, h=13   (いいえ / OFF side)
#
# Default English target shapes, centered on the original box centers plus final nudges:
#   Yes: x=-10, w=22      No:  x=36, w=16
#   ON:  x= -8, w=18      OFF: x=31, w=25
#   Yes/No height: 12, y shift: -2
#   ON/OFF height: 13, y shift: 0
FUN_80051780_RAM = 0x80051780
SELECTOR_GSBOX_HOOK_MARKER = b"PGSGB114"
SELECTOR_GSBOX_LEFT_ORIG_X = -12
SELECTOR_GSBOX_LEFT_ORIG_W = 24
SELECTOR_GSBOX_RIGHT_ORIG_X = 24
SELECTOR_GSBOX_RIGHT_ORIG_W = 36
SELECTOR_GSBOX_ORIG_H = 13
SELECTOR_GSBOX_DEFAULT_YES_W = 22
SELECTOR_GSBOX_DEFAULT_NO_W = 16
SELECTOR_GSBOX_DEFAULT_ON_W = 18
SELECTOR_GSBOX_DEFAULT_OFF_W = 25
SELECTOR_GSBOX_DEFAULT_YESNO_H = 12
SELECTOR_GSBOX_DEFAULT_YESNO_Y_SHIFT = -2
SELECTOR_GSBOX_DEFAULT_ONOFF_H = 13
SELECTOR_GSBOX_DEFAULT_ONOFF_Y_SHIFT = 0
SELECTOR_GSBOX_DEFAULT_YES_X_SHIFT = 1
SELECTOR_GSBOX_DEFAULT_NO_X_SHIFT = 2
SELECTOR_GSBOX_DEFAULT_ON_X_SHIFT = 1
SELECTOR_GSBOX_DEFAULT_OFF_X_SHIFT = 2
# The Vibration Function routine starts around 0x80066F04 in the analyzed build.
# The hook tests RA in this range and uses ON/OFF widths there; all other matching
# small boxes use Yes/No widths.
SELECTOR_GSBOX_DEFAULT_VIB_START = 0x80066F00
SELECTOR_GSBOX_DEFAULT_VIB_END = 0x80067600
# v64: Proceed without saving? selector entries caught at the hook entry:
#   Yes highlighted: x=-12, y=0x41, w=24, h=13, caller seen as 0x8006CB04
#   No highlighted:  x= 24, y=0x41, w=36, h=13, caller seen as 0x8006E164
# Save Clear Data uses the same shapes but y=0x35, so match y=0x41 exactly
# instead of RA alone. Default extra shift -12 makes final y match Save Clear
# after the normal Yes/No -2px shift.
SELECTOR_GSBOX_DEFAULT_PROCEED_Y = 0x41
SELECTOR_GSBOX_DEFAULT_PROCEED_EXTRA_Y_SHIFT = -12


def _centered_x(orig_x: int, orig_w: int, new_w: int) -> int:
    center2 = orig_x * 2 + orig_w
    return (center2 - int(new_w)) // 2


def _emit_li32(a, reg: str, value: int) -> None:
    a.lui(reg, hi(value))
    a.ori(reg, reg, lo(value))


def _emit_sltu(a, rd: str, rs: str, rt_: str) -> None:
    a.word(rt(REG[rs], REG[rt_], REG[rd], 0, 0x2B))


def build_selector_gsbox_hook(
    hook_ram,
    *,
    original_ram=FUN_80051780_RAM,
    yes_width=SELECTOR_GSBOX_DEFAULT_YES_W,
    no_width=SELECTOR_GSBOX_DEFAULT_NO_W,
    on_width=SELECTOR_GSBOX_DEFAULT_ON_W,
    off_width=SELECTOR_GSBOX_DEFAULT_OFF_W,
    yesno_height=SELECTOR_GSBOX_DEFAULT_YESNO_H,
    yesno_y_shift=SELECTOR_GSBOX_DEFAULT_YESNO_Y_SHIFT,
    onoff_height=SELECTOR_GSBOX_DEFAULT_ONOFF_H,
    onoff_y_shift=SELECTOR_GSBOX_DEFAULT_ONOFF_Y_SHIFT,
    yes_x_shift=SELECTOR_GSBOX_DEFAULT_YES_X_SHIFT,
    no_x_shift=SELECTOR_GSBOX_DEFAULT_NO_X_SHIFT,
    on_x_shift=SELECTOR_GSBOX_DEFAULT_ON_X_SHIFT,
    off_x_shift=SELECTOR_GSBOX_DEFAULT_OFF_X_SHIFT,
    vibration_start=SELECTOR_GSBOX_DEFAULT_VIB_START,
    vibration_end=SELECTOR_GSBOX_DEFAULT_VIB_END,
    proceed_y=SELECTOR_GSBOX_DEFAULT_PROCEED_Y,
    proceed_extra_y_shift=SELECTOR_GSBOX_DEFAULT_PROCEED_EXTRA_Y_SHIFT,
):
    yes_width = int(yes_width)
    no_width = int(no_width)
    on_width = int(on_width)
    off_width = int(off_width)
    yesno_height = int(yesno_height)
    yesno_y_shift = int(yesno_y_shift)
    onoff_height = int(onoff_height)
    onoff_y_shift = int(onoff_y_shift)
    yes_x_shift = int(yes_x_shift)
    no_x_shift = int(no_x_shift)
    on_x_shift = int(on_x_shift)
    off_x_shift = int(off_x_shift)
    vibration_start = int(vibration_start)
    vibration_end = int(vibration_end)
    proceed_y = int(proceed_y)
    proceed_extra_y_shift = int(proceed_extra_y_shift)

    yes_x = _centered_x(SELECTOR_GSBOX_LEFT_ORIG_X, SELECTOR_GSBOX_LEFT_ORIG_W, yes_width) + yes_x_shift
    no_x = _centered_x(SELECTOR_GSBOX_RIGHT_ORIG_X, SELECTOR_GSBOX_RIGHT_ORIG_W, no_width) + no_x_shift
    on_x = _centered_x(SELECTOR_GSBOX_LEFT_ORIG_X, SELECTOR_GSBOX_LEFT_ORIG_W, on_width) + on_x_shift
    off_x = _centered_x(SELECTOR_GSBOX_RIGHT_ORIG_X, SELECTOR_GSBOX_RIGHT_ORIG_W, off_width) + off_x_shift

    a = A(hook_ram)
    # Preserve the original caller RA.  The original 9-argument call has five
    # stack-passed arguments at old_sp+0x10..0x20. After allocating 0x40 bytes,
    # old_sp is new_sp+0x40, so copy new_sp+0x50..0x60 down to new_sp+0x10..0x20.
    a.addiu("sp", "sp", -0x40)
    a.sw("ra", 0x3C, "sp")
    for off in (0x10, 0x14, 0x18, 0x1C, 0x20):
        a.lw("t0", 0x40 + off, "sp")
        a.sw("t0", off, "sp")

    # Only adjust original small selector boxes. Other rectangles pass through.
    a.addiu("t0", "zero", SELECTOR_GSBOX_ORIG_H)
    a.bne("a3", "t0", "call_original")
    a.nop()

    # v64: Proceed without saving? boxes are the same original Yes/No shapes
    # but arrive at y=0x41. Shift only that Y before the usual shape handling.
    # This avoids the too-broad v63 RA-only approach that also affected Vibration.
    if proceed_extra_y_shift:
        a.addiu("t0", "zero", proceed_y)
        a.bne("a1", "t0", "after_proceed_exact_y")
        a.nop()
        a.addiu("t0", "zero", SELECTOR_GSBOX_LEFT_ORIG_X)
        a.beq("a0", "t0", "proceed_check_left_w")
        a.nop()
        a.addiu("t0", "zero", SELECTOR_GSBOX_RIGHT_ORIG_X)
        a.beq("a0", "t0", "proceed_check_right_w")
        a.nop()
        a.j("after_proceed_exact_y")
        a.nop()
        a.lab("proceed_check_left_w")
        a.addiu("t0", "zero", SELECTOR_GSBOX_LEFT_ORIG_W)
        a.bne("a2", "t0", "after_proceed_exact_y")
        a.nop()
        a.addiu("a1", "a1", proceed_extra_y_shift)
        a.j("after_proceed_exact_y")
        a.nop()
        a.lab("proceed_check_right_w")
        a.addiu("t0", "zero", SELECTOR_GSBOX_RIGHT_ORIG_W)
        a.bne("a2", "t0", "after_proceed_exact_y")
        a.nop()
        a.addiu("a1", "a1", proceed_extra_y_shift)
    a.lab("after_proceed_exact_y")

    # Classify whether the call came from the Vibration Function ON/OFF routine.
    # t7 = 1 if vibration_start <= original_ra < vibration_end, else 0.
    _emit_li32(a, "t1", vibration_start)
    _emit_sltu(a, "t2", "ra", "t1")       # t2 = ra < start
    a.bne("t2", "zero", "not_vibration")
    a.nop()
    _emit_li32(a, "t1", vibration_end)
    _emit_sltu(a, "t2", "ra", "t1")       # t2 = ra < end
    a.beq("t2", "zero", "not_vibration")
    a.nop()
    a.addiu("t7", "zero", 1)
    a.j("after_vibration_test")
    a.nop()
    a.lab("not_vibration")
    a.addiu("t7", "zero", 0)
    a.lab("after_vibration_test")

    # Left option: original x=-12, w=24.  Use ON width only on Vibration screen;
    # otherwise use Yes width.
    a.addiu("t0", "zero", SELECTOR_GSBOX_LEFT_ORIG_X)
    a.bne("a0", "t0", "check_right")
    a.nop()
    a.addiu("t0", "zero", SELECTOR_GSBOX_LEFT_ORIG_W)
    a.bne("a2", "t0", "check_right")
    a.nop()
    a.bne("t7", "zero", "left_on")
    a.nop()
    a.addiu("a0", "zero", yes_x)
    if yesno_y_shift:
        a.addiu("a1", "a1", yesno_y_shift)
    a.addiu("a2", "zero", yes_width)
    a.addiu("a3", "zero", yesno_height)
    a.j("call_original")
    a.nop()
    a.lab("left_on")
    a.addiu("a0", "zero", on_x)
    if onoff_y_shift:
        a.addiu("a1", "a1", onoff_y_shift)
    a.addiu("a2", "zero", on_width)
    a.addiu("a3", "zero", onoff_height)
    a.j("call_original")
    a.nop()

    # Right option: original x=24, w=36.  Use OFF width only on Vibration screen;
    # otherwise use No width.
    a.lab("check_right")
    a.addiu("t0", "zero", SELECTOR_GSBOX_RIGHT_ORIG_X)
    a.bne("a0", "t0", "call_original")
    a.nop()
    a.addiu("t0", "zero", SELECTOR_GSBOX_RIGHT_ORIG_W)
    a.bne("a2", "t0", "call_original")
    a.nop()
    a.bne("t7", "zero", "right_off")
    a.nop()
    a.addiu("a0", "zero", no_x)
    if yesno_y_shift:
        a.addiu("a1", "a1", yesno_y_shift)
    a.addiu("a2", "zero", no_width)
    a.addiu("a3", "zero", yesno_height)
    a.j("call_original")
    a.nop()
    a.lab("right_off")
    a.addiu("a0", "zero", off_x)
    if onoff_y_shift:
        a.addiu("a1", "a1", onoff_y_shift)
    a.addiu("a2", "zero", off_width)
    a.addiu("a3", "zero", onoff_height)

    a.lab("call_original")
    a.word(jt(3, original_ram))
    a.nop()
    a.lw("ra", 0x3C, "sp")
    a.addiu("sp", "sp", 0x40)
    a.word(rt(REG["ra"], 0, 0, 0, 0x08))  # jr ra
    a.nop()
    return a.out(), {
        "yes_x": yes_x, "yes_w": yes_width,
        "no_x": no_x, "no_w": no_width,
        "on_x": on_x, "on_w": on_width,
        "off_x": off_x, "off_w": off_width,
        "yesno_height": yesno_height,
        "yesno_y_shift": yesno_y_shift,
        "onoff_height": onoff_height,
        "onoff_y_shift": onoff_y_shift,
        "yes_x_shift": yes_x_shift,
        "no_x_shift": no_x_shift,
        "on_x_shift": on_x_shift,
        "off_x_shift": off_x_shift,
        "vibration_start": f"0x{vibration_start:08X}",
        "vibration_end": f"0x{vibration_end:08X}",
        "proceed_y": proceed_y,
        "proceed_extra_y_shift": proceed_extra_y_shift,
    }


def patch_selector_gsbox_hook(
    exe,
    load,
    *,
    yes_width=SELECTOR_GSBOX_DEFAULT_YES_W,
    no_width=SELECTOR_GSBOX_DEFAULT_NO_W,
    on_width=SELECTOR_GSBOX_DEFAULT_ON_W,
    off_width=SELECTOR_GSBOX_DEFAULT_OFF_W,
    yesno_height=SELECTOR_GSBOX_DEFAULT_YESNO_H,
    yesno_y_shift=SELECTOR_GSBOX_DEFAULT_YESNO_Y_SHIFT,
    onoff_height=SELECTOR_GSBOX_DEFAULT_ONOFF_H,
    onoff_y_shift=SELECTOR_GSBOX_DEFAULT_ONOFF_Y_SHIFT,
    yes_x_shift=SELECTOR_GSBOX_DEFAULT_YES_X_SHIFT,
    no_x_shift=SELECTOR_GSBOX_DEFAULT_NO_X_SHIFT,
    on_x_shift=SELECTOR_GSBOX_DEFAULT_ON_X_SHIFT,
    off_x_shift=SELECTOR_GSBOX_DEFAULT_OFF_X_SHIFT,
    vibration_start=SELECTOR_GSBOX_DEFAULT_VIB_START,
    vibration_end=SELECTOR_GSBOX_DEFAULT_VIB_END,
    proceed_y=SELECTOR_GSBOX_DEFAULT_PROCEED_Y,
    proceed_extra_y_shift=SELECTOR_GSBOX_DEFAULT_PROCEED_EXTRA_Y_SHIFT,
    force=False,
    dry_run=False,
):
    yes_width = int(yes_width)
    no_width = int(no_width)
    on_width = int(on_width)
    off_width = int(off_width)
    yesno_height = int(yesno_height)
    yesno_y_shift = int(yesno_y_shift)
    onoff_height = int(onoff_height)
    onoff_y_shift = int(onoff_y_shift)
    yes_x_shift = int(yes_x_shift)
    no_x_shift = int(no_x_shift)
    on_x_shift = int(on_x_shift)
    off_x_shift = int(off_x_shift)
    vibration_start = int(vibration_start)
    vibration_end = int(vibration_end)
    proceed_y = int(proceed_y)
    proceed_extra_y_shift = int(proceed_extra_y_shift)
    if min(yes_width, no_width, on_width, off_width, yesno_height, onoff_height) < 4:
        raise ValueError("selector grey box width/height values are suspiciously small")
    if max(yes_width, no_width, on_width, off_width) > 80 or max(yesno_height, onoff_height) > 40:
        raise ValueError("selector grey box width/height values are suspiciously large")
    if vibration_start >= vibration_end:
        raise ValueError("selector vibration RA range must have start < end")

    if SELECTOR_GSBOX_HOOK_MARKER in bytes(exe):
        return {
            "enabled": True,
            "status": "already_applied_marker_found",
            "hook_file": None,
            "hook_ram": None,
            "hook_len": 0,
            "patched_calls": 0,
            "scanned_calls": 0,
            "yes_width": yes_width,
            "no_width": no_width,
            "on_width": on_width,
            "off_width": off_width,
            "yesno_height": yesno_height,
            "yesno_y_shift": yesno_y_shift,
            "onoff_height": onoff_height,
            "onoff_y_shift": onoff_y_shift,
            "yes_x_shift": yes_x_shift,
            "no_x_shift": no_x_shift,
            "on_x_shift": on_x_shift,
            "off_x_shift": off_x_shift,
            "vibration_start": f"0x{vibration_start:08X}",
            "vibration_end": f"0x{vibration_end:08X}",
            "proceed_y": proceed_y,
            "proceed_extra_y_shift": proceed_extra_y_shift,
            "report": [],
        }

    original_jal = jt(3, FUN_80051780_RAM)
    old_payload = struct.unpack_from('<I', exe, 0x1c)[0] if len(exe) >= 0x20 else len(exe) - PSX_HEADER
    scan_end = min(len(exe) - 4, PSX_HEADER + max(0, old_payload))
    call_offsets = []
    for off in range(PSX_HEADER, scan_end + 1, 4):
        if read32(exe, off) == original_jal:
            call_offsets.append(off)

    if not call_offsets and not force:
        return {
            "enabled": True,
            "status": "no_fun_80051780_calls_found",
            "hook_file": None,
            "hook_ram": None,
            "hook_len": 0,
            "patched_calls": 0,
            "scanned_calls": 0,
            "yes_width": yes_width,
            "no_width": no_width,
            "on_width": on_width,
            "off_width": off_width,
            "yesno_height": yesno_height,
            "yesno_y_shift": yesno_y_shift,
            "onoff_height": onoff_height,
            "onoff_y_shift": onoff_y_shift,
            "yes_x_shift": yes_x_shift,
            "no_x_shift": no_x_shift,
            "on_x_shift": on_x_shift,
            "off_x_shift": off_x_shift,
            "vibration_start": f"0x{vibration_start:08X}",
            "vibration_end": f"0x{vibration_end:08X}",
            "proceed_y": proceed_y,
            "proceed_extra_y_shift": proceed_extra_y_shift,
            "report": [],
        }

    hook_start = align(len(exe), 4)
    marker_off = hook_start
    code_off = align(marker_off + len(SELECTOR_GSBOX_HOOK_MARKER), 4)
    code_ram = off2ram(code_off, load)
    hook, dims = build_selector_gsbox_hook(
        code_ram,
        yes_width=yes_width,
        no_width=no_width,
        on_width=on_width,
        off_width=off_width,
        yesno_height=yesno_height,
        yesno_y_shift=yesno_y_shift,
        onoff_height=onoff_height,
        onoff_y_shift=onoff_y_shift,
        yes_x_shift=yes_x_shift,
        no_x_shift=no_x_shift,
        on_x_shift=on_x_shift,
        off_x_shift=off_x_shift,
        vibration_start=vibration_start,
        vibration_end=vibration_end,
        proceed_y=proceed_y,
        proceed_extra_y_shift=proceed_extra_y_shift,
    )
    hook_jal = jt(3, code_ram)

    report = []
    for off in call_offsets:
        cur = read32(exe, off)
        status = "would_patch" if dry_run else "patched"
        if cur == hook_jal:
            status = "already_patched"
        elif cur != original_jal and not force:
            status = "skipped_mismatch"
        report.append({
            "call_ram": f"0x{off2ram(off, load):08X}",
            "call_file_offset": f"0x{off:X}",
            "old_word": f"0x{cur:08X}",
            "new_word": f"0x{hook_jal:08X}",
            "status": status,
        })

    if not dry_run:
        if len(exe) < marker_off:
            exe += b"\0" * (marker_off - len(exe))
        exe[marker_off:marker_off + len(SELECTOR_GSBOX_HOOK_MARKER)] = SELECTOR_GSBOX_HOOK_MARKER
        if len(exe) < code_off:
            exe += b"\0" * (code_off - len(exe))
        exe[code_off:code_off + len(hook)] = hook
        for rec in report:
            if rec["status"] in {"patched", "would_patch"}:
                off = int(rec["call_file_offset"], 16)
                struct.pack_into('<I', exe, off, hook_jal & 0xFFFFFFFF)

    patched = sum(1 for r in report if r["status"] in {"patched", "would_patch"})
    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "hook_file": f"0x{code_off:X}",
        "hook_ram": f"0x{code_ram:08X}",
        "hook_len": len(hook),
        "marker_file": f"0x{marker_off:X}",
        "patched_calls": patched,
        "scanned_calls": len(call_offsets),
        "yes_width": yes_width,
        "no_width": no_width,
        "on_width": on_width,
        "off_width": off_width,
        "yesno_height": yesno_height,
        "yesno_y_shift": yesno_y_shift,
        "onoff_height": onoff_height,
        "onoff_y_shift": onoff_y_shift,
        "yes_x_shift": yes_x_shift,
        "no_x_shift": no_x_shift,
        "on_x_shift": on_x_shift,
        "off_x_shift": off_x_shift,
        **dims,
        "report": report,
    }


def write_selector_gsbox_hook_report_csv(path: Path, rows: list[dict]) -> None:
    fields = ["call_ram", "call_file_offset", "old_word", "new_word", "status"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        wri = csv.DictWriter(f, fieldnames=fields)
        wri.writeheader()
        for rec in rows:
            wri.writerow({k: rec.get(k, "") for k in fields})


# Pixy action/created-Modus message suffix spacing fix (v97).
#
# The action confirmation messages for moving Pixies are built in a local text
# buffer by concatenating a fixed-width/copied Pixy name and then a suffix such
# as "has been put into the Garden."  The copied name can retain padding intended
# for Japanese fixed-width names, while longer English names can touch the suffix.
#
# Confirmed standalone test v2 fix: replace the second strcat call for the three
# core action messages plus the two created-Modus messages by default with a small hook.  The hook trims trailing ASCII spaces
# and CP932 full-width spaces from the destination buffer, inserts exactly one
# ASCII space, then calls the original strcat to append the suffix.
ORIGINAL_STRCAT = 0x8009C724
ORIGINAL_STRCAT_JAL = 0x0C0271C9  # jal 0x8009C724
PIXY_SUFFIX_CORE_SITES = [
    ("case8_placed_on_planet_suffix", 0x80098A7C),
    ("case9_housed_inside_ship_suffix", 0x80098AD4),
    ("case10_put_into_garden_suffix", 0x80098B2C),
]
PIXY_SUFFIX_CREATED_MODUS_SITES = [
    ("case12_created_modus_recover_suffix", 0x80098BB8),
    ("case12_created_modus_discard_suffix", 0x80098C2C),
]


def build_pixy_name_suffix_strcat_hook(hook_ram, strcat_ram=ORIGINAL_STRCAT):
    a = A(hook_ram)
    # Called instead of strcat(dest, suffix). The original delay-slot addiu a1
    # that loads the suffix pointer remains at the callsite, so a0/a1 are valid.
    a.addiu("sp", "sp", -0x18)
    a.sw("ra", 0x14, "sp")
    a.sw("a0", 0x10, "sp")
    a.sw("a1", 0x0C, "sp")
    a.addu("t0", "a0", "zero")  # dest/base
    a.addu("t1", "a0", "zero")  # scan pointer

    a.lab("scan_nul")
    a.lbu("t2", 0, "t1")
    a.beq("t2", "zero", "found_nul")
    a.addiu("t1", "t1", 1)       # delay slot: advance past current byte
    a.j("scan_nul")
    a.nop()

    a.lab("found_nul")
    a.addiu("t1", "t1", -1)      # t1 = current NUL position

    a.lab("trim_loop")
    a.beq("t1", "t0", "insert_space")
    a.nop()
    a.addiu("t3", "t1", -1)
    a.lbu("t2", 0, "t3")
    a.addiu("t4", "zero", 0x20)
    a.beq("t2", "t4", "trim_ascii_space")
    a.nop()

    # Check for CP932 full-width space 0x81 0x40 immediately before NUL.
    a.addiu("t3", "t1", -2)
    a.word(rt(REG["t3"], REG["t0"], REG["t4"], 0, 0x2B))  # sltu t4,t3,t0
    a.bne("t4", "zero", "insert_space")
    a.nop()
    a.lbu("t2", 0, "t3")
    a.addiu("t4", "zero", 0x81)
    a.bne("t2", "t4", "insert_space")
    a.nop()
    a.lbu("t2", 1, "t3")
    a.addiu("t4", "zero", 0x40)
    a.bne("t2", "t4", "insert_space")
    a.nop()
    a.addu("t1", "t3", "zero")
    a.j("trim_loop")
    a.nop()

    a.lab("trim_ascii_space")
    a.addu("t1", "t3", "zero")
    a.j("trim_loop")
    a.nop()

    a.lab("insert_space")
    a.addiu("t2", "zero", 0x20)
    a.word(it(0x28, REG["t1"], REG["t2"], 0))     # sb t2,0(t1)
    a.word(it(0x28, REG["t1"], REG["zero"], 1))   # sb zero,1(t1)
    a.lw("a0", 0x10, "sp")
    a.lw("a1", 0x0C, "sp")
    a.word(jt(3, strcat_ram))
    a.nop()
    a.lw("ra", 0x14, "sp")
    a.addiu("sp", "sp", 0x18)
    a.word(rt(REG["ra"], 0, 0, 0, 0x08))           # jr ra
    a.nop()
    return a.out()


def patch_pixy_name_suffix_spacing(exe, load, *, include_created_modus=True, force=False, dry_run=False):
    sites = list(PIXY_SUFFIX_CORE_SITES)
    if include_created_modus:
        sites.extend(PIXY_SUFFIX_CREATED_MODUS_SITES)

    rows = []
    needs_hook = False
    for name, ram in sites:
        off = ram2off(ram, load)
        if off < 0 or off + 4 > len(exe):
            raise RuntimeError(f"Pixy suffix spacing site {name} maps outside EXE: RAM 0x{ram:08X}, file 0x{off:X}")
        cur = read32(exe, off)
        if cur == ORIGINAL_STRCAT_JAL:
            status = "would_patch" if dry_run else "patched"
            needs_hook = True
        elif (cur >> 26) == 0x03 and cur != ORIGINAL_STRCAT_JAL:
            status = "already_hooked_existing_jal"
        elif force:
            status = "would_patch_force" if dry_run else "patched_force"
            needs_hook = True
        else:
            raise RuntimeError(
                f"Pixy suffix spacing site {name} mismatch at RAM 0x{ram:08X} / file 0x{off:X}: "
                f"found 0x{cur:08X}, expected jal strcat 0x{ORIGINAL_STRCAT_JAL:08X}. "
                "Use --pixy-name-suffix-force only if this is intentionally a compatible MAIN.EXE."
            )
        rows.append({
            "name": name,
            "site_ram": f"0x{ram:08X}",
            "file_offset": f"0x{off:X}",
            "old_word": f"0x{cur:08X}",
            "new_word": "",
            "status": status,
        })

    if not needs_hook:
        return {
            "enabled": True,
            "status": "already_hooked",
            "hook_file": None,
            "hook_ram": None,
            "hook_len": 0,
            "include_created_modus": include_created_modus,
            "report": rows,
        }

    hook_off = align(len(exe), 4)
    hook_ram = off2ram(hook_off, load)
    hook = build_pixy_name_suffix_strcat_hook(hook_ram)
    hook_jal = jt(3, hook_ram)
    for row in rows:
        if row["status"] in {"patched", "patched_force", "would_patch", "would_patch_force"}:
            row["new_word"] = f"0x{hook_jal:08X}"
        else:
            row["new_word"] = row["old_word"]

    if not dry_run:
        if len(exe) < hook_off:
            exe += b"\0" * (hook_off - len(exe))
        exe[hook_off:hook_off + len(hook)] = hook
        for row in rows:
            if row["status"] in {"patched", "patched_force"}:
                off = int(row["file_offset"], 16)
                exe[off:off+4] = w(hook_jal)

    return {
        "enabled": True,
        "status": "would_patch" if dry_run else "patched",
        "hook_file": f"0x{hook_off:X}",
        "hook_ram": f"0x{hook_ram:08X}",
        "hook_len": len(hook),
        "include_created_modus": include_created_modus,
        "report": rows,
    }


def write_pixy_name_suffix_report_csv(path: Path, rows: list[dict]) -> None:
    fields=["name","site_ram","file_offset","old_word","new_word","status"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w",encoding="utf-8-sig",newline="") as f:
        wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
        for rec in rows:
            wri.writerow({k:rec.get(k,"") for k in fields})

def main():
    ap=argparse.ArgumentParser(description="PixyGarden MAIN.EXE patcher v108 based on working v108/safe v102/v97, with tuned guarded small grey selector-box FUN_80051780 hook and Pixy-name action/Modus suffix spacing fix enabled by default, rows 2-38 full-screen memory-card/status centering, locked Latin renderer metrics, Garden music-title hook, PLMENU01/PLSEL/TREE/REPORT fixes, and full grouped TUTO XA text patch")
    ap.add_argument("--exe",required=True); ap.add_argument("--xlsx",required=True); ap.add_argument("--out",required=True)
    ap.add_argument("--sheet",default="MAIN_EXE_Text"); ap.add_argument("--pointer-column"); ap.add_argument("--original-bytes-column"); ap.add_argument("--text-column")
    ap.add_argument("--dry-run",action="store_true")
    ap.add_argument("--disable-clear-data-selector-y-fix", action="store_true", help="Disable v34 Results/Clear Data Save clear data? Yes/No selector-box Y adjustment")
    ap.add_argument("--clear-data-selector-y", type=lambda x:int(x,0), default=0x35, help="Y coordinate for the Save clear data? Yes/No grey selector boxes. Default 0x35, 3px lower than stock 0x32.")
    ap.add_argument("--clear-data-selector-y-force", action="store_true", help="Apply the clear-data selector Y patch even if the current words are already non-stock addiu a1,zero,imm instructions")
    ap.add_argument("--disable-memory-card-centering", action="store_true", help="Disable v95 rows 2-38 full-screen memory-card/status centering patch")
    ap.add_argument("--memory-card-centering-force", action="store_true", help="Patch memory-card/status a0 sites even if the current word is not a normal li/clear-a0 shape")
    ap.add_argument("--memory-card-centering-x-bias", type=int, default=0, help="Add this many pixels to every computed centered x coordinate for rows 2-38")
    ap.add_argument("--memory-card-centering-no-known-sites", action="store_true", help="Disable the known-site table for rows 2-38 centering")
    ap.add_argument("--memory-card-centering-no-dynamic-scan", action="store_true", help="Disable the dynamic FUN_80038498 scan for rows 2-38 centering")
    ap.add_argument("--disable-pixy-name-suffix-spacing", action="store_true", help="Disable v97 Pixy-name + action/Modus-suffix spacing fix. Default: enabled.")
    ap.add_argument("--pixy-name-suffix-force", action="store_true", help="Patch Pixy-name suffix strcat callsites even if they are not the expected original jal strcat word")
    ap.add_argument("--disable-pixy-name-suffix-created-modus", action="store_true", help="Disable the two Pixy-name + created-Modus suffix cases. Default: included.")
    ap.add_argument("--pixy-name-suffix-include-created-modus", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--disable-selector-gsbox-hook", action="store_true", help="Disable v110 guarded FUN_80051780 hook for small grey Yes/No/ON/OFF selector boxes. Default: enabled.")
    ap.add_argument("--selector-gsbox-yes-width", type=int, default=SELECTOR_GSBOX_DEFAULT_YES_W, help="Yes grey box width. Default: 22px.")
    ap.add_argument("--selector-gsbox-no-width", type=int, default=SELECTOR_GSBOX_DEFAULT_NO_W, help="No grey box width. Default: 16px.")
    ap.add_argument("--selector-gsbox-on-width", type=int, default=SELECTOR_GSBOX_DEFAULT_ON_W, help="ON grey box width on Vibration Function screen. Default: 18px.")
    ap.add_argument("--selector-gsbox-off-width", type=int, default=SELECTOR_GSBOX_DEFAULT_OFF_W, help="OFF grey box width on Vibration Function screen. Default: 25px.")
    ap.add_argument("--selector-gsbox-yesno-height", type=int, default=SELECTOR_GSBOX_DEFAULT_YESNO_H, help="Yes/No grey box height. Default: 12px.")
    ap.add_argument("--selector-gsbox-yesno-y-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_YESNO_Y_SHIFT, help="Vertical shift for Yes/No grey selector boxes. Default: -2px (up two pixels).")
    ap.add_argument("--selector-gsbox-onoff-height", type=int, default=SELECTOR_GSBOX_DEFAULT_ONOFF_H, help="ON/OFF grey box height. Default: 13px.")
    ap.add_argument("--selector-gsbox-onoff-y-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_ONOFF_Y_SHIFT, help="Vertical shift for ON/OFF grey selector boxes. Default: 0px (one pixel up from v108).")
    ap.add_argument("--selector-gsbox-yes-x-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_YES_X_SHIFT, help="Extra horizontal shift for Yes grey box after centering. Default: +1px.")
    ap.add_argument("--selector-gsbox-no-x-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_NO_X_SHIFT, help="Extra horizontal shift for No grey box after centering. Default: +2px.")
    ap.add_argument("--selector-gsbox-on-x-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_ON_X_SHIFT, help="Extra horizontal shift for ON grey box after centering. Default: +1px.")
    ap.add_argument("--selector-gsbox-off-x-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_OFF_X_SHIFT, help="Extra horizontal shift for OFF grey box after centering. Default: +2px.")
    # Backward-compatible v105/v106 options: if supplied, apply to both Yes/No and ON/OFF.
    ap.add_argument("--selector-gsbox-height", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--selector-gsbox-y-shift", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--selector-gsbox-vibration-start", type=lambda s:int(s,0), default=SELECTOR_GSBOX_DEFAULT_VIB_START, help="Start RAM address for classifying ON/OFF Vibration Function selector calls. Default: 0x80066F00.")
    ap.add_argument("--selector-gsbox-vibration-end", type=lambda s:int(s,0), default=SELECTOR_GSBOX_DEFAULT_VIB_END, help="End RAM address for classifying ON/OFF Vibration Function selector calls. Default: 0x80067600.")
    ap.add_argument("--selector-gsbox-proceed-y", type=lambda s:int(s,0), default=SELECTOR_GSBOX_DEFAULT_PROCEED_Y, help="v64 exact incoming Y for Proceed without saving? original Yes/No selector boxes. Default: 0x41 from hook-entry breakpoint.")
    ap.add_argument("--selector-gsbox-proceed-extra-y-shift", type=int, default=SELECTOR_GSBOX_DEFAULT_PROCEED_EXTRA_Y_SHIFT, help="v64 extra vertical shift applied only to original Yes/No selector boxes at --selector-gsbox-proceed-y. Default: -12px.")
    ap.add_argument("--selector-gsbox-force", action="store_true", help="Allow v106 hook setup even if no original FUN_80051780 calls are detected.")
    # Backward-compatible aliases from v105. They now map to Yes/No widths when supplied.
    ap.add_argument("--selector-gsbox-left-width", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--selector-gsbox-right-width", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--disable-tuto-21-step-flow",action="store_true",help="Disable full TUTO grouped-textbox patch (keeps original 11 top-level tutorial groups)")
    ap.add_argument("--tuto21-force",action="store_true",help="Patch TUTO state-split sites even if current words are not the expected original instructions")
    ap.add_argument("--xa-text-xlsx",default=None,help="Optional XA translation workbook for TUTO text. Reads Clip ID + Preferred English from the Updated Translations sheet.")
    ap.add_argument("--xa-text-sheet",default="Updated Translations",help="Sheet in --xa-text-xlsx containing XA rows. Default: Updated Translations")
    ap.add_argument("--xa-clip-id-column",default="Clip ID",help="Clip ID column in --xa-text-xlsx. Default: Clip ID")
    ap.add_argument("--xa-text-column",default="Preferred English",help="Text column in --xa-text-xlsx. Default: Preferred English")
    ap.add_argument("--tuto-close-wait-frames",type=int,default=TUTO_CLOSE_WAIT_FRAMES_DEFAULT,help="Frames to keep servicing the previous TUTO helper image/textbox close before opening the next helper image. Default: 40; v65 effectively used 30.")
    ap.add_argument("--reports-dir", help="Optional directory for CSV/JSON patch reports")
    ap.add_argument("--disable-report-text-fix",action="store_true",help="Disable the REPORT.CDF label text block fix at MAIN.EXE offset 0x3504")
    ap.add_argument("--report-text-force",action="store_true",help="Apply the REPORT text patch even if the sanity check does not recognize the original bytes")
    ap.add_argument("--disable-plsel-graphic-draw-patch", action="store_true", help="Disable the PLSEL TIM14 combined-strip draw patch")
    ap.add_argument("--plsel-main-right-u", type=lambda x:int(x,0), default=162, help="PLSEL combined strip right texture U. Default: 162")
    ap.add_argument("--plsel-main-right-x", type=lambda x:int(x,0), default=47, help="PLSEL combined strip right local screen X. Default: 47")
    ap.add_argument("--plsel-main-uv-only", action="store_true", help="Patch PLSEL UVs only; do not patch local screen X")
    ap.add_argument("--plsel-left-u", type=lambda x:int(x,0), default=None, help="Optional PLSEL combined strip left texture U override")
    ap.add_argument("--plsel-left-x", type=lambda x:int(x,0), default=None, help="Optional PLSEL combined strip left local screen X override")
    ap.add_argument("--plsel-top-v", type=lambda x:int(x,0), default=None, help="Optional PLSEL combined strip top texture V override")
    ap.add_argument("--plsel-bottom-v", type=lambda x:int(x,0), default=None, help="Optional PLSEL combined strip bottom texture V override")
    ap.add_argument("--disable-plsel-following-graphic-patch", action="store_true", help="Disable the PLSEL following/planet/circle graphic source patch")
    ap.add_argument("--disable-plsel-following-dynamic-patch", action="store_true", help="Disable the additional dynamic/state-dependent PLSEL following/circle graphic patch")
    ap.add_argument("--plsel-following-u-shift", type=lambda x:int(x,0), default=20, help="Atlas U shift for the PLSEL following/planet graphic. Default: +20")
    # PLSEL circlefix v2: screen X shift default is intentionally 0.
    ap.add_argument("--plsel-following-a-left-u", type=int, default=180, help="PLSEL TIM14 following/circle graphic strip A left U. Default: 180.")
    ap.add_argument("--plsel-following-a-right-u", type=int, default=205, help="PLSEL TIM14 following/circle graphic strip A right U. Default: 205.")
    ap.add_argument("--plsel-following-b-left-u", type=int, default=218, help="PLSEL TIM14 following/circle graphic strip B left U. Default: 218.")
    ap.add_argument("--plsel-following-b-right-u", type=int, default=244, help="PLSEL TIM14 following/circle graphic strip B right U. Default: 244.")
    ap.add_argument("--plsel-following-top-v", type=int, default=46, help="PLSEL TIM14 following/circle graphic top V for dynamic draw path. Default: 46.")
    ap.add_argument("--plsel-following-dynamic-a-right-u", type=int, default=206, help="PLSEL TIM14 dynamic following/circle graphic strip A right U. Default: 206.")
    ap.add_argument("--plsel-following-x-shift", type=lambda x:int(x,0), default=0, help="Optional local screen X shift for the PLSEL following/circle graphic. Default: 0 keeps original screen position.")
    ap.add_argument("--plsel-following-uv-only", action="store_true", help="Patch following/planet graphic UVs only; do not move its local screen X values")
    ap.add_argument("--plsel-patch-json", help="Optional JSON containing extra verified PLSEL exe_patches entries")
    ap.add_argument("--plsel-force", action="store_true", help="Force PLSEL patching even if expected instruction words do not match")
    ap.add_argument("--disable-planet-info-element-draw-patch", action="store_true", help="Disable the PLMENU01 element-info draw hook for Earth/Water/Wind/Fire/None labels")
    ap.add_argument("--planet-info-force", action="store_true", help="Force PLMENU01 element-info patching even if expected instruction words do not match")
    ap.add_argument("--planet-info-dash-u", type=lambda x:int(x,0), default=0x75, help="PLMENU01 None/Dash source U. Default: 0x75")
    ap.add_argument("--planet-info-dash-w", type=lambda x:int(x,0), default=0x10, help="PLMENU01 None/Dash sprite width. Default: 0x10")
    ap.add_argument("--planet-info-earth-u", type=lambda x:int(x,0), default=0x19, help="PLMENU01 Earth source U. Default: 0x19 (English word starts at x=25)")
    ap.add_argument("--planet-info-earth-w", type=lambda x:int(x,0), default=0x17, help="PLMENU01 Earth sprite width. Default: 0x17 / 23px")
    ap.add_argument("--planet-info-water-u", type=lambda x:int(x,0), default=0x32, help="PLMENU01 Water source U. Default: 0x32 (English word starts at x=50)")
    ap.add_argument("--planet-info-water-w", type=lambda x:int(x,0), default=0x18, help="PLMENU01 Water sprite width. Default: 0x18 / 24px")
    ap.add_argument("--planet-info-wind-u", type=lambda x:int(x,0), default=0x4C, help="PLMENU01 Wind source U. Default: 0x4C (English word starts at x=76)")
    ap.add_argument("--planet-info-wind-w", type=lambda x:int(x,0), default=0x13, help="PLMENU01 Wind sprite width. Default: 0x13 / 19px")
    ap.add_argument("--planet-info-fire-u", type=lambda x:int(x,0), default=0x61, help="PLMENU01 Fire source U. Default: 0x61 (English word starts at x=97)")
    ap.add_argument("--planet-info-fire-w", type=lambda x:int(x,0), default=0x12, help="PLMENU01 Fire sprite width. Default: 0x12 / 18px")
    ap.add_argument("--planet-info-element-x-shift", type=lambda x:int(x,0), default=-8, help="Legacy shared PLMENU01 element X shift. Deprecated; use per-element shifts below.")
    ap.add_argument("--planet-info-earth-x-shift", type=lambda x:int(x,0), default=-6, help="PLMENU01 Earth destination X shift in pixels. Default: -6")
    ap.add_argument("--planet-info-water-x-shift", type=lambda x:int(x,0), default=-6, help="PLMENU01 Water destination X shift in pixels. Default: -6")
    ap.add_argument("--planet-info-wind-x-shift", type=lambda x:int(x,0), default=-4, help="PLMENU01 Wind destination X shift in pixels. Default: -4")
    ap.add_argument("--planet-info-fire-x-shift", type=lambda x:int(x,0), default=-3, help="PLMENU01 Fire destination X shift in pixels. Default: -3")
    ap.add_argument("--disable-planet-info-planet-icon-patch", action="store_true", help="Disable the PLMENU01 planet icon source-coordinate patch")
    ap.add_argument("--enable-shared-planet-icon-patch", action="store_true", help="Legacy/test mode: directly patch the shared planet source site. Normally leave off so stage uses U=0x68 and copy-slot hook uses --planet-info-planet-u.")
    ap.add_argument("--disable-planet-copy-slot-split-hook", action="store_true", help="Disable hook at 8007787C that writes copied planet slot U at BASE+0x152")
    ap.add_argument("--planet-info-planet-u", type=lambda x:int(x,0), default=0x86, help="PLMENU01 planet icon source U. Default: 0x86")
    ap.add_argument("--planet-info-planet-v", type=lambda x:int(x,0), default=0xB8, help="PLMENU01 planet icon source V written to the copied PLANET info slot. Default: 0xB8")
    ap.add_argument("--disable-planet-info-planet-clut-patch", action="store_true", help="Disable the PLMENU01 planet icon palette/CLUT selector patch")
    ap.add_argument("--planet-info-planet-clut", type=lambda x:int(x,0), default=0x01E9, help="PLMENU01 planet icon palette/CLUT selector. Default: 0x01E9, the original planet-icon palette")
    ap.add_argument("--disable-stage-tim03-moved-uv-patch", action="store_true", help="Disable the STAGE moved planet/N-A source UV patch")
    ap.add_argument("--stage-tim03-planet-u", type=lambda x:int(x,0), default=0xB9, help="STAGE TIM04 moved planet source U written narrowly by the copy-slot hook. Default: 0xB9 / 185")
    ap.add_argument("--stage-tim03-planet-v", type=lambda x:int(x,0), default=0x78, help="STAGE TIM04 moved planet source V written narrowly by the copy-slot hook. Default: 0x78 / 120")
    ap.add_argument("--stage-tim03-na-line-u", "--stage-tim04-na-line-slot-u", dest="stage_tim03_na_line_u", type=lambda x:int(x,0), default=0x75, help="STAGE TIM04 Object-overlay N/A 16x16 symbol source U. Default: 0x75 / 117 (visible line begins at 118).")
    ap.add_argument("--disable-stage-tim04-na-static-seed", action="store_true", help="Disable v120 first-frame N/A static seed and restore the old static candidate to U=0x58 instead.")
    ap.add_argument("--stage-tim03-uv-force", action="store_true", help="Force STAGE TIM03 moved UV patching even if expected instruction words do not match")
    ap.add_argument("--disable-stage-tim04-na-packet-filter-hook", action="store_true", help="Compatibility flag; the experimental N/A packet filter is already disabled by default in v93")
    ap.add_argument("--enable-stage-tim04-na-packet-filter-hook", action="store_true", help="Experimental only: enable the v87 packet-base N/A filter. Not recommended; v87 caused SML flicker.")
    ap.add_argument("--stage-tim04-na-packet-filter-force", action="store_true", help="Force-install the STAGE TIM04 N/A packet filter hook even if the expected store instruction does not match")
    ap.add_argument("--stage-tim04-na-packet-base", type=lambda x:int(x,0), default=0x800F8710, help="Runtime packet base for the Object-overlay N/A 16x16 symbol. Default: 0x800F8710")
    ap.add_argument("--stage-tim04-na-bad-word", type=lambda x:int(x,0), default=0x7852C058, help="Legacy/report-only bad source/CLUT word. v93 filters by packet base and uses --stage-tim04-na-good-word.")
    ap.add_argument("--stage-tim04-na-good-word", type=lambda x:int(x,0), default=0x7852C075, help="Correct source/CLUT word. Default: 0x7852C075")
    ap.add_argument("--stage-tim04-object-element-packet-prefix", type=lambda x:int(x,0), default=0x78520000, help="High 16/source-word prefix for STAGE TIM04 source-word repairs. Default: 0x78520000")
    ap.add_argument("--disable-stage-tim04-na-ft4-source-filter-hook", action="store_true", help="Disable the v130 POLY_FT4 transition-frame N/A source-word hook at 800A3DAC")
    ap.add_argument("--stage-tim04-na-ft4-source-filter-force", action="store_true", help="Force-install the v130 FT4 source-word hook even if expected instruction words do not match")
    ap.add_argument("--stage-tim04-na-ft4-packet-cmd", type=lambda x:int(x,0), default=STAGE_TIM04_NA_FT4_PACKET_CMD, help="Gray POLY_FT4 command/color word matched by the v130 N/A hook. Default: 0x2C808080")
    ap.add_argument("--disable-stage-tim04-na-ft4-edge-filter-hook", action="store_true", help="Disable the v132 early POLY_FT4 N/A source-edge remap hook at 800A3CDC")
    ap.add_argument("--stage-tim04-na-ft4-edge-filter-force", action="store_true", help="Force-install the v132 FT4 edge hook even if expected instruction words do not match")
    ap.add_argument("--disable-stage-tim04-na-template-u-hook", action="store_true", help="Disable the v111 template-level Object-overlay element/N-A U+width remap hook")
    ap.add_argument("--stage-tim04-na-template-u-force", action="store_true", help="Force-install the v111 template-level Object-overlay element/N-A hook even if the expected instruction does not match")
    ap.add_argument("--stage-tim04-na-template-old-u", type=lambda x:int(x,0), default=0x58, help="Old TIM04 N/A 16x16 slot U. Default: 0x58")
    ap.add_argument("--stage-tim04-na-template-v", type=lambda x:int(x,0), default=0xC0, help="TIM04 element/N-A template V match. Default: 0xC0")
    ap.add_argument("--stage-tim04-na-template-size", type=lambda x:int(x,0), default=0x10, help="Width/height match for element/N-A templates before remap. Default: 0x10")
    ap.add_argument("--stage-tim04-na-template-color", type=lambda x:int(x,0), default=0x80, help="RGB byte match for gray TexBlend packet. Default: 0x80")
    ap.add_argument("--stage-tim04-object-overlay-marker", type=lambda x:int(x,0), default=0x34, help="STAGE Object/Pixy overlay marker byte at template +0x17. Default: 0x34; PLANET Modus snapshot used 0x23.")
    ap.add_argument("--stage-tim04-object-overlay-base-x", type=lambda x:int(x,0), default=0x80, help="Unshifted destination X for STAGE Object-overlay element templates. Default: 0x80; v23 writes base_x + per-element shift to avoid accumulating delta shifts.")
    ap.add_argument("--stage-tim04-object-overlay-na-x", type=lambda x:int(x,0), default=0x80, help="Absolute destination X for STAGE Object-overlay N/A-line value. Default: 0x80.")
    ap.add_argument("--stage-tim04-object-overlay-na-slot-ram", type=lambda x:int(x,0), default=0x8012BEE4, help="RAM address of the STAGE Object/Pixy shared value slot. Default from snapshots: 0x8012BEE4; use 0 to disable exact Object-slot repair.")
    ap.add_argument("--stage-tim04-object-overlay-object-guard-halfword", type=lambda x:int(x,0), default=0x1000, help="Halfword at Object value slot +0x1E that distinguishes Object overlay from Pixy overlay. Default: 0x1000.")
    ap.add_argument("--stage-tim04-dash-w", type=lambda x:int(x,0), default=0x10, help="STAGE TIM04 Object-overlay N/A sprite width after remap. Default: 0x10")
    ap.add_argument("--stage-tim04-earth-u", type=lambda x:int(x,0), default=0x19, help="STAGE TIM04 Object-overlay Earth source U. Default: 0x19 (English word starts at x=25)")
    ap.add_argument("--stage-tim04-earth-w", type=lambda x:int(x,0), default=0x17, help="STAGE TIM04 Object-overlay Earth sprite width. Default: 0x17 / 23px")
    ap.add_argument("--stage-tim04-water-u", type=lambda x:int(x,0), default=0x32, help="STAGE TIM04 Object-overlay Water source U. Default: 0x32 (English word starts at x=50)")
    ap.add_argument("--stage-tim04-water-w", type=lambda x:int(x,0), default=0x18, help="STAGE TIM04 Object-overlay Water sprite width. Default: 0x18 / 24px")
    ap.add_argument("--stage-tim04-wind-u", type=lambda x:int(x,0), default=0x4C, help="STAGE TIM04 Object-overlay Wind source U. Default: 0x4C (English word starts at x=76)")
    ap.add_argument("--stage-tim04-wind-w", type=lambda x:int(x,0), default=0x13, help="STAGE TIM04 Object-overlay Wind sprite width. Default: 0x13 / 19px")
    ap.add_argument("--stage-tim04-fire-u", type=lambda x:int(x,0), default=0x61, help="STAGE TIM04 Object-overlay Fire source U. Default: 0x61 (English word starts at x=97)")
    ap.add_argument("--stage-tim04-fire-w", type=lambda x:int(x,0), default=0x12, help="STAGE TIM04 Object-overlay Fire sprite width. Default: 0x12 / 18px")
    ap.add_argument("--stage-tim04-element-x-shift", type=lambda x:int(x,0), default=-8, help="Legacy shared STAGE TIM04 element X shift. Deprecated; use per-element shifts below.")
    ap.add_argument("--stage-tim04-earth-x-shift", type=lambda x:int(x,0), default=-6, help="STAGE TIM04 Earth destination X shift in pixels. Default: -6")
    ap.add_argument("--stage-tim04-water-x-shift", type=lambda x:int(x,0), default=-6, help="STAGE TIM04 Water destination X shift in pixels. Default: -6")
    ap.add_argument("--stage-tim04-wind-x-shift", type=lambda x:int(x,0), default=-4, help="STAGE TIM04 Wind destination X shift in pixels. Default: -4")
    ap.add_argument("--stage-tim04-fire-x-shift", type=lambda x:int(x,0), default=-3, help="STAGE TIM04 Fire destination X shift in pixels. Default: -3")
    ap.add_argument("--disable-modus-stage-template-element-fix", action="store_true", help="Disable Modus subscreen STAGE.TIM Earth/Water/Wind/Fire 24x16 template-width fix. Default: enabled; still required because the PLANET.CDF edit fixed only Elem.")
    ap.add_argument("--modus-stage-element-x", type=lambda x:int(x,0), default=0x7A, help="Destination X guard for Modus Earth/Water/Wind/Fire word template. Default: 0x7A from packet trace")
    ap.add_argument("--modus-stage-element-bad-x", type=lambda x:int(x,0), default=0x7D, help="Observed bad PLANET/Modus element template X to repair narrowly. Default: 0x7D from broken Eart/ndFi snapshots.")
    ap.add_argument("--modus-stage-element-bad-w", type=lambda x:int(x,0), default=0x12, help="Observed bad PLANET/Modus element template width paired with bad-x. Default: 0x12 from broken Eart/ndFi snapshots.")
    ap.add_argument("--modus-stage-element-y", type=lambda x:int(x,0), default=-37, help="Destination Y guard for Modus Earth/Water/Wind/Fire word template. Default: -37 from packet trace")
    ap.add_argument("--modus-stage-element-old-w", type=lambda x:int(x,0), default=0x17, help="Old Modus element word width to match. Default: 0x17 / 23px")
    ap.add_argument("--modus-stage-element-w", type=lambda x:int(x,0), default=0x18, help="New Modus element word width. Default: 0x18 / 24px")
    ap.add_argument("--modus-stage-element-marker", type=lambda x:int(x,0), default=0x23, help="PLANET/Modus element template marker byte at +0x17. Default: 0x23.")
    ap.add_argument("--disable-modus-local-record-early-fix", action="store_true", help="Disable v59 early local PLANET/Modus record-read flicker fix. Default: enabled.")
    ap.add_argument("--modus-local-record-early-force", action="store_true", help="Force v59 early local PLANET/Modus record-read hook if the site already differs.")
    ap.add_argument("--disable-modus-local-record-fix", action="store_true", help="Disable v58 later local PLANET/Modus record flicker fix. Default: enabled.")
    ap.add_argument("--modus-local-record-fix-force", action="store_true", help="Force v58 later local PLANET/Modus record hook if the site already differs.")
    ap.add_argument("--disable-modus-exact-ft4-edge-read-hooks", action="store_true", help="Disable v60 exact active-record FT4 edge-read U/W hooks at 800A3CAC and 800A3CC8. Default: enabled.")
    ap.add_argument("--modus-exact-ft4-edge-read-force", action="store_true", help="Force v60 exact active-record FT4 edge-read hooks if the sites already differ.")
    ap.add_argument("--modus-exact-ft4-edge-record-ram", type=lambda x:int(x,0), default=0x8012A3BC, help="Exact active Modus record RAM address for v60 FT4 edge-read hooks. Default: 0x8012A3BC from No$PSX breakpoints.")
    ap.add_argument("--modus-stage-earth-u", type=lambda x:int(x,0), default=0x18, help="PLANET/Modus Earth source U from PLMENU01.TIM padded rect. Default: 0x18 / x=24; visible ink begins at x=25.")
    ap.add_argument("--modus-stage-earth-w", type=lambda x:int(x,0), default=0x19, help="PLANET/Modus Earth padded source width from PLMENU01.TIM. Default: 0x19 / 25px; visible ink is 23px.")
    ap.add_argument("--modus-stage-water-u", type=lambda x:int(x,0), default=0x31, help="PLANET/Modus Water source U from PLMENU01.TIM padded rect. Default: 0x31 / x=49; visible ink begins at x=50.")
    ap.add_argument("--modus-stage-water-w", type=lambda x:int(x,0), default=0x1A, help="PLANET/Modus Water padded source width from PLMENU01.TIM. Default: 0x1A / 26px; visible ink is 24px.")
    ap.add_argument("--modus-stage-wind-u", type=lambda x:int(x,0), default=0x4B, help="PLANET/Modus Wind source U from PLMENU01.TIM padded rect. Default: 0x4B / x=75; visible ink begins at x=76.")
    ap.add_argument("--modus-stage-wind-w", type=lambda x:int(x,0), default=0x15, help="PLANET/Modus Wind padded source width from PLMENU01.TIM. Default: 0x15 / 21px; visible ink is 19px.")
    ap.add_argument("--modus-stage-fire-u", type=lambda x:int(x,0), default=0x60, help="PLANET/Modus Fire source U from PLMENU01.TIM padded rect. Default: 0x60 / x=96; visible ink begins at x=97.")
    ap.add_argument("--modus-stage-fire-w", type=lambda x:int(x,0), default=0x14, help="PLANET/Modus Fire padded source width from PLMENU01.TIM. Default: 0x14 / 20px; visible ink is 18px.")
    ap.add_argument("--modus-stage-ft4-earth-u", type=lambda x:int(x,0), default=0x19, help="PLANET/Modus FT4/opening Earth visible-ink source U. Default: 0x19 / x=25.")
    ap.add_argument("--modus-stage-ft4-earth-w", type=lambda x:int(x,0), default=0x17, help="PLANET/Modus FT4/opening Earth visible-ink width. Default: 0x17 / 23px.")
    ap.add_argument("--modus-stage-ft4-water-u", type=lambda x:int(x,0), default=0x32, help="PLANET/Modus FT4/opening Water visible-ink source U. Default: 0x32 / x=50.")
    ap.add_argument("--modus-stage-ft4-water-w", type=lambda x:int(x,0), default=0x18, help="PLANET/Modus FT4/opening Water visible-ink width. Default: 0x18 / 24px.")
    ap.add_argument("--modus-stage-ft4-wind-u", type=lambda x:int(x,0), default=0x4C, help="PLANET/Modus FT4/opening Wind visible-ink source U. Default: 0x4C / x=76.")
    ap.add_argument("--modus-stage-ft4-wind-w", type=lambda x:int(x,0), default=0x13, help="PLANET/Modus FT4/opening Wind visible-ink width. Default: 0x13 / 19px.")
    ap.add_argument("--modus-stage-ft4-fire-u", type=lambda x:int(x,0), default=0x61, help="PLANET/Modus FT4/opening Fire visible-ink source U. Default: 0x61 / x=97.")
    ap.add_argument("--modus-stage-ft4-fire-w", type=lambda x:int(x,0), default=0x12, help="PLANET/Modus FT4/opening Fire visible-ink width. Default: 0x12 / 18px.")
    ap.add_argument("--modus-stage-ft4-x-tolerance", type=lambda x:int(x,0), default=4, help="Accept this many pixels of X drift for PLANET/Modus FT4/opening element repair. Default: 4; use 0 for exact old behavior.")
    ap.add_argument("--modus-stage-ft4-x-adjust", type=lambda x:int(x,0), default=0, help="Destination X adjustment written to repaired PLANET/Modus FT4/opening element frames. Default: 0; try -1, 1, or 2 only if the opening frame is visibly offset.")
    ap.add_argument("--modus-stage-element-h", type=lambda x:int(x,0), default=0x10, help="Modus element word height guard. Default: 0x10 / 16px")
    ap.add_argument("--disable-modus-crest-text-x-patch", action="store_true", help="Disable Modus PLANET/PLMENU Crest-name text-window X-position candidate patch.")
    ap.add_argument("--modus-crest-text-site", default="detail_800796f0", choices=["none","detail_800796f0","list_8007cd18_crest","list_80098dc8_crest","list_8009932c_a","list_8009932c_b","list_8009932c_c"], help="Crest text X candidate to patch. Default: detail_800796f0, the currently selected Crest-name box.")
    ap.add_argument("--modus-crest-text-x", type=lambda x:int(x,0), default=-118, help="Override new Crest text-window X for the selected candidate. Default sentinel keeps each candidate's built-in value; for detail_800796f0 this is 0x10, i.e. 14px left.")
    ap.add_argument("--modus-crest-text-x-force", action="store_true", help="Force selected Modus Crest-name text-window X candidate if guards differ.")
    ap.add_argument("--disable-modus-crest-template-x-fix", action="store_true", help="Disable Modus subscreen Crest-label template X-position fix")
    ap.add_argument("--modus-crest-template-old-x", type=lambda x:int(x,0), default=0x0A, help="Old Modus Crest template X. Default: 0x0A from packet trace")
    ap.add_argument("--modus-crest-template-new-x", type=lambda x:int(x,0), default=0x0F, help="New Modus Crest template X. Default: 0x0F, 12px after the left panel cap at x=0x03")
    ap.add_argument("--modus-crest-template-y", type=lambda x:int(x,0), default=0x36, help="Modus Crest template Y guard. Default: 0x36")
    ap.add_argument("--modus-crest-template-u", type=lambda x:int(x,0), default=0x18, help="Modus Crest template source U guard. Default: 0x18")
    ap.add_argument("--modus-crest-template-v", type=lambda x:int(x,0), default=0x00, help="Modus Crest template source V guard. Default: 0x00")
    ap.add_argument("--modus-crest-template-w", type=lambda x:int(x,0), default=0x58, help="Modus Crest template width guard. Default: 0x58 / 88px")
    ap.add_argument("--modus-crest-template-h", type=lambda x:int(x,0), default=0x10, help="Modus Crest template height guard. Default: 0x10 / 16px")
    ap.add_argument("--modus-crest-template-color", type=lambda x:int(x,0), default=0x80, help="Modus Crest template RGB/color-byte guard. Default: 0x80")
    ap.add_argument("--ascii-map-json")
    ap.add_argument("--ascii-advance",type=int,default=12,help="Fallback/default advance for printable one-byte ASCII not assigned by an advance group. v73 default 12 keeps unspecified ASCII at original/Japanese-style spacing.")
    ap.add_argument("--two-byte-advance",type=int,default=12,help="Legacy compatibility option. Non-Latin CP932 now uses the original parser/spacing by default, so this is not used unless you modify the hook behavior.")
    ap.add_argument("--space-advance",type=int,default=3)
    ap.add_argument("--punct-advance",type=int,default=3)
    ap.add_argument("--punct-chars",default=DEFAULT_PUNCT_CHARS,help="ASCII punctuation characters that use --punct-advance; full-width CP932 versions are mapped too")
    ap.add_argument("--xoff-left2-chars",default="Iil",help="ASCII characters drawn 2 pixels left without changing cursor advance. v73 default: I/i/l, aligning their thin strokes to the line cursor.")
    ap.add_argument("--xoff-left1-chars",default="abcdefghjkmnopqrstuvwxyz1",help="ASCII characters drawn 1 pixel left without changing cursor advance. v71 default: lowercase letters except i/l plus digit 1, for line-left alignment.")
    ap.add_argument("--xoff-right3-chars",default="'",help="ASCII characters drawn 3 pixels right without changing cursor advance. Default: apostrophe")
    ap.add_argument("--xoff-right2-chars",default=",.",help="ASCII characters drawn 2 pixels right without changing cursor advance")
    ap.add_argument("--xoff-right1-chars",default="",help="ASCII characters drawn 1 pixel right without changing cursor advance. v73 default: empty so line starts do not drift right.")
    ap.add_argument("--draw-shift",action="append",nargs=2,metavar=("CHARS","PIXELS"),default=[],help="Override/add draw X offset for any ASCII characters. Repeatable. Example: --draw-shift \"ABC\" -1 or --draw-shift \"A1le\" 2")
    ap.add_argument("--capital-r-advance-delta",type=int,default=0,help="Optional extra advance delta for capital R only. Default 0; R otherwise uses the same --uppercase-advance-delta as other capitals.")
    ap.add_argument("--uppercase-advance-delta",type=int,default=0,help="Final advance delta applied to every capital A-Z after grouped advances. v73 default: 0 to preserve 1 px spacing after capitals.")
    ap.add_argument("--hyphen-advance-delta",type=int,default=3,help="Final advance delta applied to hyphen after punctuation/group calculation. Default: +3")
    ap.add_argument("--slash-advance-delta",type=int,default=0,help="Legacy delta applied to forward slash before --slash-fixed-advance. v73 default: 0 because slash is fixed below.")
    ap.add_argument("--paren-advance",type=int,default=10,help="Fixed advance for ASCII/full-width parentheses (). v73 default: 10.")
    ap.add_argument("--slash-fixed-advance",type=int,default=11,help="Fixed advance for ASCII/full-width slash /. v73 default: 11. Use -1 to disable; --advance-override still applies last.")
    ap.add_argument("--advance7-chars",default=DEFAULT_ADVANCE_7_CHARS)
    ap.add_argument("--advance6-chars",default=DEFAULT_ADVANCE_6_CHARS)
    ap.add_argument("--advance5-chars",default=DEFAULT_ADVANCE_5_CHARS)
    ap.add_argument("--advance9-chars",default=DEFAULT_ADVANCE_9_CHARS)
    ap.add_argument("--advance8-chars",default=DEFAULT_ADVANCE_8_CHARS)
    ap.add_argument("--advance3-chars",default=DEFAULT_ADVANCE_3_CHARS)
    ap.add_argument("--advance2-chars",default=DEFAULT_ADVANCE_2_CHARS)
    ap.add_argument("--advance10-chars",default=DEFAULT_ADVANCE_10_CHARS)
    ap.add_argument("--advance4-chars",default=DEFAULT_ADVANCE_4_CHARS)
    ap.add_argument("--disable-v74-final-spacing-lock",action="store_true",help="Compatibility alias: disable the final locked advance pass.")
    ap.add_argument("--disable-v75-final-spacing-lock",action="store_true",help="Disable v75 final locked advances/bearing-tuned advance values. Normally leave enabled.")
    ap.add_argument("--disable-v75-final-xoff-lock",action="store_true",help="Disable v75 final draw-bearing locks such as P left and comma left-bear adjustment.")
    ap.add_argument("--disable-v75-pair-kern",action="store_true",help="Disable v75 context-aware pair kerning. Default includes wi=-1.")
    ap.add_argument("--pair-kern",action="append",nargs=2,metavar=("PAIR","PIXELS"),default=[],help="Context-aware pair kerning delta applied to the first glyph only when followed by the second. Example: --pair-kern wi -2")
    ap.add_argument("--font-tim",help="Optional FONT11Z0.TIM path for v76 build-time metric extraction. If omitted, the patcher auto-detects FONT11Z0.TIM in the current directory or beside the script.")
    ap.add_argument("--font-map-json",help="Optional JSON mapping of ASCII characters to FONT11Z0.TIM cell positions. Format: {\"A\":[col,row,plane], \"B\":{\"col\":1,\"row\":2,\"plane\":3}}. Usually not needed for alphanumerics because v77 has the built-in CP932 modulo-4 printable ASCII map.")
    ap.add_argument("--font-ascii-base-abs",type=int,default=147,help="Absolute FONT11Z0 bitplane slot for full-width ASCII digit 0. Default 147 = row 1, col 15, plane 3.")
    ap.add_argument("--font-cell-width",type=int,default=11,help="FONT11Z0.TIM glyph cell width in pixels. Default: 11")
    ap.add_argument("--font-cell-height",type=int,default=11,help="FONT11Z0.TIM glyph cell height in pixels. Default: 11")
    ap.add_argument("--font-cell-step-x",type=int,default=12,help="FONT11Z0.TIM horizontal cell pitch in pixels, including separator column. Default: 12")
    ap.add_argument("--font-cell-step-y",type=int,default=11,help="FONT11Z0.TIM vertical cell pitch in pixels. FONT11Z0 is 253 px high = 23 rows * 11 px, so default is 11.")
    ap.add_argument("--font-tracking",type=int,default=1,help="Extra pixels added to the visible glyph width when deriving v76 FONT TIM advances. Default: 1")
    ap.add_argument("--disable-font-tim-metrics",action="store_true",help="Disable v76 FONT11Z0.TIM-derived metrics and use the built-in v75 spacing tables only")
    ap.add_argument("--font-zero-is-occupied",action="store_true",help="Treat palette index 0 as occupied when scanning FONT11Z0.TIM columns. Normally leave off so index 0 stays transparent.")
    ap.add_argument("--parser-force",action="store_true",help="Patch renderer parser hook even if the entry bytes are neither original nor an existing J/NOP hook")
    # Backward-compatible aliases retained for old commands; not used by the grouped advance table.
    ap.add_argument("--narrow-advance",type=int,default=4,help=argparse.SUPPRESS)
    ap.add_argument("--narrow-chars",default="Iil",help=argparse.SUPPRESS)
    ap.add_argument("--no-preserve-legacy-controls",action="store_true"); ap.add_argument("--preserve-u-controls",action="store_true")
    ap.add_argument("--no-strip-controls",dest="strip_controls",action="store_false"); ap.set_defaults(strip_controls=True)
    ap.add_argument("--no-lui",action="store_true",help="Disable the original conservative same-register LUI/ADDIU patch pass")
    ap.add_argument("--skip-direct-mips",action="store_true",help="Disable the newer direct MIPS LUI+ADDIU/ORI patch pass")
    ap.add_argument("--direct-mips-confidence",choices=["high","medium","all"],default="high")
    ap.add_argument("--direct-mips-max-gap",type=int,default=8)
    ap.add_argument("--direct-mips-code-start",type=lambda x:int(x,0),default=PSX_HEADER)
    ap.add_argument("--direct-mips-code-end",type=lambda x:int(x,0),default=None)
    ap.add_argument("--direct-mips-rows",default=None,help="Only direct-MIPS patch these sheet rows/ranges, e.g. 23 or 2-38. All strings/ptr32 still process normally.")
    ap.add_argument("--direct-mips-exclude-rows",default=None)
    ap.add_argument("--direct-mips-no-lifetime-aware",action="store_true")
    ap.add_argument("--non-strict-direct-mips",action="store_true",help="Report direct-MIPS errors instead of aborting")
    ap.add_argument("--ptr32-policy",choices=["clustered","all","none","section","ranges"],default="clustered")
    ap.add_argument("--ptr32-sections",default="rdata,data,sdata")
    ap.add_argument("--ptr32-cluster-gap",type=lambda x:int(x,0),default=0x20); ap.add_argument("--ptr32-cluster-min",type=int,default=3); ap.add_argument("--ptr32-cluster-unique-min",type=int,default=2)
    ap.add_argument("--ptr32-ranges-json"); ap.add_argument("--ptr32-include-offsets"); ap.add_argument("--ptr32-exclude-offsets")
    ap.add_argument("--ptr32-slice-count",type=int,default=1); ap.add_argument("--ptr32-slice-index",type=int,default=0)
    ap.add_argument("--align-strings",type=int,default=4)
    ap.add_argument("--disable-name-screen-fix",action="store_true",help="Disable the confirmed PixyGarden name-change-screen fix")
    ap.add_argument("--name-screen-exclude-offsets",default="0x3E48",help="Inline ptr32 file offsets to force-exclude for the name-screen fix. Default: 0x3E48")
    ap.add_argument("--name-screen-inplace-rows",default="",help="Spreadsheet rows to patch in-place for excluded name-screen pointers. Default: empty (leave Aries original/untranslated). Use 233 to patch Aries.")
    ap.add_argument("--name-screen-inplace-fill",choices=["normal_zero","normal_space","fixed_ascii_space","fixed_cp932_space"],default="fixed_cp932_space",help="How to fill unused bytes for optional in-place name-screen rows. Default: fixed_cp932_space, but no rows are patched unless --name-screen-inplace-rows is set.")
    ap.add_argument("--name-screen-inplace-body-encoding",choices=["ascii","cp932_fullwidth"],default="cp932_fullwidth",help="Encoding for optional in-place name-screen rows. cp932_fullwidth keeps two-byte fixed-width name-table glyphs; default: cp932_fullwidth.")
    ap.add_argument("--disable-tree-text-buffer-patch",action="store_true",help="Disable the TREE.CDF details text stack-buffer/copy-limit patch")
    ap.add_argument("--tree-copy-limit",type=lambda x:int(x,0),default=0x1FA,help="TREE details TXT copy byte limit. Original is 0xFA; default patch value is 0x1FA")
    ap.add_argument("--tree-buffer-force",action="store_true",help="Patch TREE buffer sites even if current words are not original/compatible")
    ap.add_argument("--tree-terminator",type=lambda x:int(x,0),default=0x00,help="TREE details TXT terminator byte. Original is 0x39; default 0x00 keeps ASCII digit 9 safe")
    ap.add_argument("--tree-terminator-force",action="store_true",help="Patch TREE terminator site even if current word is not original/compatible")
    ap.add_argument("--disable-planet-stage-terminator-patch",action="store_true",help="Disable PLANET/STAGE HELP.FAT 0x39->0x00 terminator patch")
    ap.add_argument("--planet-stage-terminator",type=lambda x:int(x,0),default=0x00,help="PLANET/STAGE HELP.FAT terminator byte. Original is 0x39; default 0x00")
    ap.add_argument("--planet-stage-terminator-force",action="store_true",help="Patch PLANET/STAGE terminator sites even if current words are not original/compatible")
    ap.add_argument("--disable-text-primitive-capacity-patch",action="store_true",help="Disable the shared text primitive/glyph capacity patch")
    ap.add_argument("--text-primitive-capacity-glyphs","--text-primitive-capacity",dest="text_primitive_capacity_glyphs",type=lambda x:int(x,0),default=192,help="Shared text primitive capacity per draw call. Original is 128; default is 192")
    ap.add_argument("--text-primitive-force",action="store_true",help="Patch text primitive capacity sites even if current words are not original/compatible")
    ap.add_argument("--disable-planet-title-ram-hook",action="store_true",help="Disable the PLANET music-title RAM hook")
    ap.add_argument("--planet-title-force",action="store_true",help="Patch PLANET title hook site even if it does not contain the expected original instructions")
    ap.add_argument("--planet-title-titles",default="|".join(PLANET_TITLE_DEFAULT_TITLES),help="PLANET music titles separated by |. Default: Relief|Affection|Material|Harvest|Undine|Windmill|Ignition")
    ap.add_argument("--planet-title-y",type=int,default=112,help="Top Y of first PLANET title strip on the 256x256 page. Default: 112")
    ap.add_argument("--planet-title-text-x",type=int,default=66,help="Inner English text-area X. Default: 66, preserving the left and right music notes inside the 112px title strip")
    ap.add_argument("--planet-title-text-w",type=int,default=80,help="Inner English text-area width in pixels. Must be even. Default: 80; narrower than the full 112px strip so the right music note is not touched")
    ap.add_argument("--planet-title-slot-h",type=int,default=10,help="Height of each title slot in pixels. Default: 10")
    ap.add_argument("--planet-title-slot-pitch",type=int,default=10,help="Vertical pitch between title slots in pixels. Default: 10, matching the observed source V spacing")
    ap.add_argument("--planet-title-slot-y-offsets",default="",help="Optional comma-separated per-title Y offsets for fine tuning, e.g. 0,0,0,-1,-1,0,0")
    ap.add_argument("--planet-title-text-index",type=int,default=11,help="4bpp palette index for title foreground pixels. Default: 11; matches the brighter/original title pixels")
    ap.add_argument("--planet-title-shadow-index",type=int,default=5,help="4bpp palette index for title shadow pixels. Default: 5; darker drop-shadow pixel")
    ap.add_argument("--planet-title-clear-index",type=int,default=0,help="4bpp palette index used to clear the inner text area. Default: 0")
    ap.add_argument("--planet-title-payload-addr",type=lambda x:int(x,0),default=0,help="Legacy single exact live payload pointer to patch. Use --planet-title-payload-addrs for multiple Garden pages. Default: 0")
    ap.add_argument("--planet-title-payload-addrs",default="",help="Optional comma/pipe/semicolon-separated whitelist of live Garden payload pointers. Default empty: use content-pattern guard instead of address whitelisting so all five Garden pages can match safely.")
    ap.add_argument("--planet-title-allow-broad",action="store_true",help="Ignore the payload-address whitelist. In v70 this still keeps the content-pattern guard unless --planet-title-no-pattern-guard is also used.")
    ap.add_argument("--planet-title-min-payload-addr",type=lambda x:int(x,0),default=0,help="Optional additional lower bound for s2. Default: 0 because v70 uses source-page content pattern matching instead of low-buffer filtering.")
    ap.add_argument("--planet-title-no-pattern-guard",action="store_true",help="Disable source-page content pattern guard. Risky: can corrupt TREE/title-logo pages if broad matching is active.")
    ap.add_argument("--planet-title-pattern-min-per-slot",type=int,default=2,help="Minimum nonzero marker bytes required in each of the seven original music-title slots before overlaying English titles. Default: 2")
    ap.add_argument("--planet-title-pattern-x",default="",help="Optional comma-separated source-page pixel X positions sampled for the music-title pattern guard. Default samples note/title-strip positions.")
    ap.add_argument("--planet-title-pattern-y",default="",help="Optional comma-separated Y offsets within each title slot sampled for the music-title pattern guard. Default: 2,5,8")
    ap.add_argument("--planet-title-bitmap-guard",action="store_true",help="Legacy mode only: apply the old bitmap minimum-advance guard when --disable-planet-title-bitmap-metrics is used.")
    ap.add_argument("--planet-title-no-bitmap-guard",action="store_true",help=argparse.SUPPRESS)
    ap.add_argument("--disable-planet-title-bitmap-metrics",action="store_true",help="Disable v81 bitmap-derived visible-width metrics for the PLANET/Garden music-title VRAM overlay and use the older normal-renderer metric tables instead.")
    ap.add_argument("--planet-title-bitmap-tracking",type=int,default=1,help="Extra pixels added after each visible embedded bitmap title glyph. Default: 1")
    ap.add_argument("--planet-title-tracking",type=int,default=0,help="PLANET-title-only per-character tracking delta applied after the normal text advance table. Negative tightens, positive loosens. Default: 0, inherit normal spacing exactly")
    ap.add_argument("--planet-title-space-advance",type=int,default=-1,help="PLANET-title-only space advance. Default -1 means inherit normal --space-advance")
    ap.add_argument("--planet-title-advance-delta",action="append",nargs=2,metavar=("CHARS","PIXELS"),default=[],help='PLANET-title-only advance delta for selected ASCII characters. Repeatable. Example: --planet-title-advance-delta "fi" -1')
    ap.add_argument("--planet-title-advance-override",action="append",nargs=2,metavar=("CHARS","PIXELS"),default=[],help="PLANET-title-only absolute advance override for selected ASCII characters. Repeatable")
    args=ap.parse_args()
    # Defensive compatibility: older local copies may have report-dir logic
    # without the argparse option. Treat missing reports_dir as disabled.
    if not hasattr(args, "reports_dir"):
        args.reports_dir = None
    # Defensive defaults for optional patch summaries; overwritten if patches run.
    tree_buffer_info = {"enabled": False, "report": []}
    tree_terminator_info = {"enabled": False, "report": []}
    text_primitive_info = {"enabled": False, "report": []}
    planet_stage_terminator_info = {"enabled": False, "report": []}
    report_text_info = {"enabled": False, "status": "disabled"}
    plsel_graphic_draw_info = {"enabled": False, "report": []}
    planet_title_info = {"enabled": False, "status": "disabled"}
    planet_info_element_info = {"enabled": False, "status": "disabled", "report": []}
    planet_copy_slot_split_info = {"enabled": False, "status": "disabled"}
    stage_tim03_uv_info = {"enabled": False, "status": "disabled"}
    stage_tim04_na_packet_filter_info = {"enabled": False, "status": "disabled"}
    stage_tim04_na_ft4_source_filter_info = {"enabled": False, "status": "not_run"}
    stage_tim04_na_ft4_edge_filter_info = {"enabled": False, "status": "not_run"}
    stage_tim04_na_template_u_info = {"enabled": False, "status": "not_run"}
    modus_local_record_early_fix_info = {"enabled": False, "status": "not_run"}
    tuto21_info = {"enabled": False, "status": "disabled"}
    memory_card_centering_info = {"enabled": False, "status": "disabled", "report": []}
    pixy_name_suffix_info = {"enabled": False, "status": "disabled", "report": []}
    if args.ptr32_slice_count<1 or not (0<=args.ptr32_slice_index<args.ptr32_slice_count): raise ValueError("Bad slice args")
    orig=Path(args.exe).read_bytes(); exe=bytearray(orig)
    if orig[:8]!=b"PS-X EXE": raise RuntimeError("Input is not PS-X EXE")
    load=struct.unpack_from("<I",orig,0x18)[0]; old_payload=struct.unpack_from("<I",orig,0x1c)[0]; pbase=load-PSX_HEADER
    rows=load_rows(Path(args.xlsx),args.sheet,args.pointer_column,args.original_bytes_column,args.text_column,pbase,args.strip_controls,orig)
    if not args.disable_memory_card_centering:
        memory_card_centering_info=patch_memory_card_centering(
            exe, Path(args.xlsx), load,
            bias=args.memory_card_centering_x_bias,
            force=args.memory_card_centering_force,
            dry_run=args.dry_run,
            use_known_sites=not args.memory_card_centering_no_known_sites,
            use_dynamic_scan=not args.memory_card_centering_no_dynamic_scan,
        )
    table=align(BSS_END,4); payload,table,stub,xoff_table=ascii_payload(table,args); table_off=ram2off(table,load)
    if len(exe)<table_off: exe+=b"\0"*(table_off-len(exe))
    exe[table_off:table_off+len(payload)]=payload; patch_parser(exe,load,stub,force=args.parser_force)
    pool=align(table_off+len(payload),args.align_strings)
    if len(exe)<pool: exe+=b"\0"*(pool-len(exe))
    for row in rows:
        if args.align_strings>1:
            pad=align(len(exe),args.align_strings)
            if len(exe)<pad: exe+=b"\0"*(pad-len(exe))
        row.new_off=len(exe); row.new_ptr=off2ram(row.new_off,load); exe+=row.enc
    cands=collect_ptr32(orig,rows,load); cluster(cands,args.ptr32_cluster_gap); select(cands,args)
    name_fix_excludes=[]
    name_fix_inplace_report=[]
    if not args.disable_name_screen_fix:
        name_fix_excludes=apply_auto_ptr32_excludes(cands,parse_int_set(args.name_screen_exclude_offsets),label="name_screen_fix_exclude")
    lui_total=0
    if not args.no_lui:
        for row in rows:
            lui_total+=patch_lui(exe,row.old_ptr,row.new_ptr)
    direct_report=[]
    if not args.skip_direct_mips:
        include_rows=parse_row_ranges(args.direct_mips_rows)
        exclude_rows=parse_row_ranges(args.direct_mips_exclude_rows) or set()
        scan_rows=[r for r in rows if (include_rows is None or r.sheet_row in include_rows) and r.sheet_row not in exclude_rows]
        direct_hits=scan_direct_mips(orig,scan_rows,code_start=args.direct_mips_code_start,code_end=args.direct_mips_code_end,maxgap=args.direct_mips_max_gap,confidence=args.direct_mips_confidence,lifetime=not args.direct_mips_no_lifetime_aware)
        direct_report=patch_direct_mips(exe,direct_hits,dry_run=args.dry_run,strict=not args.non_strict_direct_mips)
    ptr_total=apply_ptr32(exe,cands)
    if not args.disable_name_screen_fix:
        name_fix_inplace_report=patch_inplace_rows(exe,rows,parse_row_ranges(args.name_screen_inplace_rows) or set(),fill_mode=args.name_screen_inplace_fill,body_encoding=args.name_screen_inplace_body_encoding,dry_run=args.dry_run)
    selector_gsbox_hook_info={"enabled":False,"status":"disabled","report":[]}
    clear_data_selector_y_info={"enabled":False,"status":"disabled","report":[]}
    report_text_info={"enabled":False,"status":"disabled"}
    plsel_graphic_draw_info={"enabled":False,"report":[]}
    planet_title_info={"enabled":False,"status":"disabled"}
    planet_info_element_info={"enabled":False,"status":"disabled","report":[]}
    planet_copy_slot_split_info={"enabled":False,"status":"disabled"}
    stage_tim03_uv_info={"enabled":False,"status":"disabled"}
    modus_local_record_early_fix_info={"enabled":False,"status":"disabled"}
    modus_crest_text_x_info={"enabled":False,"status":"disabled"}
    tuto21_info={"enabled":False,"status":"disabled"}
    tree_buffer_info={"enabled":False,"report":[]}
    if not args.disable_tree_text_buffer_patch:
        tree_buffer_info=patch_tree_text_buffer(exe,copy_limit=args.tree_copy_limit,force=args.tree_buffer_force,dry_run=args.dry_run)
    tree_terminator_info={"enabled":False,"terminator":"0x39","report":[]}
    if (args.tree_terminator & 0xFF) != 0x39 or args.tree_terminator_force:
        tree_terminator_info=patch_tree_copy_terminator(exe,terminator=args.tree_terminator,force=args.tree_terminator_force,dry_run=args.dry_run)
        tree_terminator_info["report"]=[tree_terminator_info]
    text_primitive_info={"enabled":False,"report":[]}
    if not args.disable_text_primitive_capacity_patch:
        text_primitive_info=patch_text_primitive_capacity(exe,capacity_glyphs=args.text_primitive_capacity_glyphs,force=args.text_primitive_force,dry_run=args.dry_run)
    if not args.disable_planet_title_ram_hook:
        planet_title_info=patch_planet_title_ram_hook(exe,load,args,dry_run=args.dry_run)
    if not args.disable_planet_info_element_draw_patch:
        planet_info_element_info=patch_planet_info_element_draw(exe,load,args,dry_run=args.dry_run)
    stage_tim03_uv_info=patch_stage_tim03_moved_uvs(exe,args,dry_run=args.dry_run)
    # v93 safety: do NOT install the experimental v87 packet-base hook by default.
    # It caused a flickering artifact near the SML label, meaning packet base 800F8710
    # is not a stable/unique selector for the Object-overlay N/A symbol.
    if getattr(args, "enable_stage_tim04_na_packet_filter_hook", False):
        stage_tim04_na_packet_filter_info=patch_stage_tim04_na_packet_filter_hook(exe,load,args,dry_run=args.dry_run)
    else:
        stage_tim04_na_packet_filter_info={"enabled": False, "status": "disabled_v93_safety_no_packet_hook"}
    stage_tim04_na_ft4_edge_filter_info=patch_stage_tim04_na_ft4_edge_filter_hook(exe,load,args,dry_run=args.dry_run)
    stage_tim04_na_ft4_source_filter_info=patch_stage_tim04_na_ft4_source_filter_hook(exe,load,args,dry_run=args.dry_run)
    stage_tim04_na_template_u_info=patch_stage_tim04_na_template_u_hook(exe,load,args,dry_run=args.dry_run)
    modus_exact_ft4_edge_read_info=patch_modus_exact_ft4_edge_read_hooks(exe,load,args,dry_run=args.dry_run)
    modus_local_record_early_fix_info=patch_modus_local_record_early_fix_hook(exe,load,args,dry_run=args.dry_run)
    modus_local_record_fix_info=patch_modus_local_record_fix_hook(exe,load,args,dry_run=args.dry_run)
    modus_crest_text_x_info=patch_modus_crest_text_x(exe,args,dry_run=args.dry_run)
    planet_copy_slot_split_info=patch_planet_copy_slot_split_hook(exe,load,args,dry_run=args.dry_run)
    if not args.disable_tuto_21_step_flow:
        tuto21_info=patch_tuto21_flow(exe,load,xa_text_xlsx=args.xa_text_xlsx,xa_text_sheet=args.xa_text_sheet,xa_clip_id_column=args.xa_clip_id_column,xa_text_column=args.xa_text_column,tuto_close_wait_frames=args.tuto_close_wait_frames,force=args.tuto21_force,dry_run=args.dry_run)
    if not args.disable_pixy_name_suffix_spacing:
        pixy_name_suffix_info=patch_pixy_name_suffix_spacing(
            exe, load,
            include_created_modus=(args.pixy_name_suffix_include_created_modus or not args.disable_pixy_name_suffix_created_modus),
            force=args.pixy_name_suffix_force,
            dry_run=args.dry_run,
        )
    if not args.disable_selector_gsbox_hook:
        # v105 compatibility: old left/right options override the Yes/No widths only.
        if args.selector_gsbox_left_width is not None:
            args.selector_gsbox_yes_width = args.selector_gsbox_left_width
        if args.selector_gsbox_right_width is not None:
            args.selector_gsbox_no_width = args.selector_gsbox_right_width
        if args.selector_gsbox_height is not None:
            args.selector_gsbox_yesno_height = args.selector_gsbox_height
            args.selector_gsbox_onoff_height = args.selector_gsbox_height
        if args.selector_gsbox_y_shift is not None:
            args.selector_gsbox_yesno_y_shift = args.selector_gsbox_y_shift
            args.selector_gsbox_onoff_y_shift = args.selector_gsbox_y_shift
        selector_gsbox_hook_info=patch_selector_gsbox_hook(
            exe, load,
            yes_width=args.selector_gsbox_yes_width,
            no_width=args.selector_gsbox_no_width,
            on_width=args.selector_gsbox_on_width,
            off_width=args.selector_gsbox_off_width,
            yesno_height=args.selector_gsbox_yesno_height,
            yesno_y_shift=args.selector_gsbox_yesno_y_shift,
            onoff_height=args.selector_gsbox_onoff_height,
            onoff_y_shift=args.selector_gsbox_onoff_y_shift,
            yes_x_shift=args.selector_gsbox_yes_x_shift,
            no_x_shift=args.selector_gsbox_no_x_shift,
            on_x_shift=args.selector_gsbox_on_x_shift,
            off_x_shift=args.selector_gsbox_off_x_shift,
            vibration_start=args.selector_gsbox_vibration_start,
            vibration_end=args.selector_gsbox_vibration_end,
            proceed_y=args.selector_gsbox_proceed_y,
            proceed_extra_y_shift=args.selector_gsbox_proceed_extra_y_shift,
            force=args.selector_gsbox_force,
            dry_run=args.dry_run,
        )
    if not args.disable_clear_data_selector_y_fix:
        clear_data_selector_y_info=patch_results_clear_data_selector_y(exe,load,y_value=args.clear_data_selector_y,force=args.clear_data_selector_y_force,dry_run=args.dry_run)
    unpadded=len(exe); final=align(len(exe),SECTOR)
    if len(exe)<final: exe+=b"\0"*(final-len(exe))
    new_payload=len(exe)-PSX_HEADER; struct.pack_into("<I",exe,0x1c,new_payload)
    loaded_end=load+new_payload; heap=patch_heap(exe,load,loaded_end)
    if not args.disable_report_text_fix:
        report_text_info=patch_report_text_block(exe,force=args.report_text_force,dry_run=args.dry_run)
    if not args.disable_plsel_graphic_draw_patch:
        plsel_graphic_draw_info=patch_plsel_graphic_draw(exe,args,dry_run=args.dry_run)
    out=Path(args.out)
    if args.reports_dir:
        reports_dir=Path(args.reports_dir); reports_dir.mkdir(parents=True, exist_ok=True)
        sr=reports_dir/(out.name+".strings_report.csv"); pr=reports_dir/(out.name+".ptr32_candidates.csv"); dr=reports_dir/(out.name+".direct_mips_report.csv"); nr=reports_dir/(out.name+".name_screen_fix_report.csv"); tr=reports_dir/(out.name+".tree_buffer_patch_report.csv"); ttr=reports_dir/(out.name+".tree_terminator_patch_report.csv"); tpr=reports_dir/(out.name+".text_primitive_capacity_report.csv"); pstr=reports_dir/(out.name+".planet_stage_terminator_report.csv"); suv=reports_dir/(out.name+".stage_tim03_uv_report.csv"); mcr=reports_dir/(out.name+".memory_card_centering_report.csv"); psr=reports_dir/(out.name+".pixy_name_suffix_spacing_report.csv"); sgr=reports_dir/(out.name+".selector_gsbox_hook_report.csv"); sj=reports_dir/(out.name+".summary.json")
        byrow={}
        for c in cands:
            r=byrow.setdefault(c.row.sheet_row,{"total":0,"sel":0}); r["total"]+=1; r["sel"]+=1 if c.selected else 0
        with sr.open("w",encoding="utf-8-sig",newline="") as f:
            wri=csv.DictWriter(f,fieldnames=["sheet_row","old_pointer","old_offset","new_pointer","new_offset","old_slot_len","new_len","ptr32_candidates","ptr32_selected","text_preview"]); wri.writeheader()
            for r in rows:
                rr=byrow.get(r.sheet_row,{"total":0,"sel":0})
                wri.writerow({"sheet_row":r.sheet_row,"old_pointer":f"0x{r.old_ptr:08X}","old_offset":f"0x{r.old_off:X}","new_pointer":f"0x{r.new_ptr:08X}","new_offset":f"0x{r.new_off:X}","old_slot_len":r.old_len,"new_len":len(r.enc),"ptr32_candidates":rr["total"],"ptr32_selected":rr["sel"],"text_preview":r.text[:180]})
        with pr.open("w",encoding="utf-8-sig",newline="") as f:
            wri=csv.DictWriter(f,fieldnames=["candidate_id","selected","reason","file_offset","ram_address","section","cluster_id","cluster_size","cluster_unique_rows","sheet_row","old_pointer","new_pointer","text_preview"]); wri.writeheader()
            for c in cands:
                wri.writerow({"candidate_id":c.cid,"selected":int(c.selected),"reason":c.reason,"file_offset":f"0x{c.off:X}","ram_address":f"0x{c.ram:08X}","section":c.section,"cluster_id":c.cluster,"cluster_size":c.csize,"cluster_unique_rows":c.cuniq,"sheet_row":c.row.sheet_row,"old_pointer":f"0x{c.row.old_ptr:08X}","new_pointer":f"0x{c.row.new_ptr:08X}","text_preview":c.row.text[:120]})
        with dr.open("w",encoding="utf-8-sig",newline="") as f:
            fields=["sheet_row","status","note","kind","confidence","lui_file_offset","use_file_offset","gap","old_pointer","new_pointer","old_lui","old_use","text_preview"]
            wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
            for rec in direct_report:
                wri.writerow({k:rec.get(k,"") for k in fields})
        with nr.open("w",encoding="utf-8-sig",newline="") as f:
            fields=["record_type","candidate_id","was_selected","file_offset","ram_address","section","sheet_row","old_pointer","new_pointer","status","old_offset","slot_len","encoded_len","fill","body_encoding","text_preview","old_preview","new_preview"]
            wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
            for rec in name_fix_excludes:
                row={k:"" for k in fields}; row.update(rec); row["record_type"]="ptr32_exclude"; wri.writerow(row)
            for rec in name_fix_inplace_report:
                row={k:"" for k in fields}; row.update(rec); row["record_type"]="inplace_row"; wri.writerow(row)
        with tr.open("w",encoding="utf-8-sig",newline="") as f:
            fields=["site","file_offset","old_word","new_word","status","copy_limit","clear_size","ra_off","frame_size"]
            wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
            for rec in tree_buffer_info.get("report",[]):
                wri.writerow({k:rec.get(k,"") for k in fields})
        with ttr.open("w",encoding="utf-8-sig",newline="") as f:
            fields=["file_offset","old_word","new_word","terminator","status"]
            wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
            for rec in tree_terminator_info.get("report",[]):
                wri.writerow({k:rec.get(k,"") for k in fields})
        if text_primitive_info.get("enabled"):
            with tpr.open("w",encoding="utf-8-sig",newline="") as f:
                fields=["site","file_offset","old_word","new_word","new_imm","status"]
                wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
                for rec in text_primitive_info.get("report",[]):
                    wri.writerow({k:rec.get(k,"") for k in fields})
        with suv.open("w",encoding="utf-8-sig",newline="") as f:
            fields=["name","description","file_offset","expected_word","old_word","new_word","new_imm","status"]
            wri=csv.DictWriter(f,fieldnames=fields); wri.writeheader()
            for rec in stage_tim03_uv_info.get("report",[]):
                wri.writerow({k:rec.get(k,"") for k in fields})
        if memory_card_centering_info.get("enabled"):
            write_memory_card_centering_report_csv(mcr, memory_card_centering_info.get("report", []))
        if pixy_name_suffix_info.get("enabled"):
            write_pixy_name_suffix_report_csv(psr, pixy_name_suffix_info.get("report", []))
        if selector_gsbox_hook_info.get("enabled"):
            write_selector_gsbox_hook_report_csv(sgr, selector_gsbox_hook_info.get("report", []))
        summary={"input":str(args.exe),"output":str(args.out),"rows":len(rows),"memory_card_centering":{k:v for k,v in memory_card_centering_info.items() if k != "report"},"memory_card_centering_report_rows":len(memory_card_centering_info.get("report",[])),"pixy_name_suffix_spacing":{k:v for k,v in pixy_name_suffix_info.items() if k != "report"},"pixy_name_suffix_spacing_report_rows":len(pixy_name_suffix_info.get("report",[])),"selector_gsbox_hook":{k:v for k,v in selector_gsbox_hook_info.items() if k != "report"},"selector_gsbox_hook_report_rows":len(selector_gsbox_hook_info.get("report",[])),"clear_data_selector_y":{k:v for k,v in clear_data_selector_y_info.items() if k != "report"},"clear_data_selector_y_report":clear_data_selector_y_info.get("report",[]),"old_file_size":len(orig),"new_file_size":len(exe),"old_payload":old_payload,"new_payload":new_payload,"ascii_table_ram":f"0x{table:08X}","ascii_stub_ram":f"0x{stub:08X}","ascii_xoff_table_ram":f"0x{xoff_table:08X}","pool_start_off":f"0x{pool:X}","pool_start_ram":f"0x{off2ram(pool,load):08X}","unpadded_size":unpadded,"loaded_end":f"0x{loaded_end:08X}","heap_base_minus4":f"0x{heap:08X}","legacy_lui_patches":lui_total,"direct_mips_sites":len(direct_report),"direct_mips_patched":sum(1 for r in direct_report if r.get("status") in {"patched","would_patch"}),"direct_mips_already_patched":sum(1 for r in direct_report if r.get("status")=="already_patched"),"direct_mips_errors":sum(1 for r in direct_report if r.get("status")=="error"),"direct_mips_confidence":args.direct_mips_confidence,"ptr32_policy":args.ptr32_policy,"ptr32_candidates_total":len(cands),"ptr32_selected_total":sum(1 for c in cands if c.selected),"ptr32_patches":ptr_total,"ptr32_cluster_gap":args.ptr32_cluster_gap,"ptr32_cluster_min":args.ptr32_cluster_min,"ptr32_cluster_unique_min":args.ptr32_cluster_unique_min,"ptr32_sections":args.ptr32_sections,"ptr32_slice_count":args.ptr32_slice_count,"ptr32_slice_index":args.ptr32_slice_index,"name_screen_fix_enabled":not args.disable_name_screen_fix,"name_screen_ptr32_excludes":len(name_fix_excludes),"name_screen_inplace_rows":len(name_fix_inplace_report),"advance_groups":{"7":args.advance7_chars,"6":args.advance6_chars,"5":args.advance5_chars,"9":args.advance9_chars,"8":args.advance8_chars,"3":args.advance3_chars,"2":args.advance2_chars,"10":args.advance10_chars,"4":args.advance4_chars},"v75_final_spacing_lock":None if (args.disable_v74_final_spacing_lock or args.disable_v75_final_spacing_lock) else DEFAULT_V75_FINAL_ADVANCE_OVERRIDES,"effective_sample_advances":{"I":build_advance_table(args)[ord("I")],"i":build_advance_table(args)[ord("i")],"l":build_advance_table(args)[ord("l")],"p":build_advance_table(args)[ord("p")],"m":build_advance_table(args)[ord("m")],"M":build_advance_table(args)[ord("M")],"R":build_advance_table(args)[ord("R")],"P":build_advance_table(args)[ord("P")],"u":build_advance_table(args)[ord("u")],"w":build_advance_table(args)[ord("w")],"period":build_advance_table(args)[ord(".")],"z":build_advance_table(args)[ord("z")]} ,"font_tim_metrics":load_v76_font_tim_metrics(args),"x_offsets":{"left2":args.xoff_left2_chars,"left1":args.xoff_left1_chars,"right3":args.xoff_right3_chars,"right2":args.xoff_right2_chars,"right1":args.xoff_right1_chars,"draw_shift":args.draw_shift,"v75_final":iter_v75_final_xoff_map(args),"pair_kern":iter_v75_pair_kern_map(args)},"capital_r_advance_delta":args.capital_r_advance_delta,"uppercase_advance_delta":args.uppercase_advance_delta,"hyphen_advance_delta":args.hyphen_advance_delta,"slash_advance_delta":args.slash_advance_delta,"paren_advance":args.paren_advance,"slash_fixed_advance":args.slash_fixed_advance,"space_advance":args.space_advance,"punct_advance":args.punct_advance,"two_byte_advance_legacy_unused":args.two_byte_advance,"nonlatin_cp932_spacing":"original_parser","tree_text_buffer_patch_enabled":tree_buffer_info.get("enabled",False),"tree_copy_limit":f"0x{tree_buffer_info.get('copy_limit',0):X}" if tree_buffer_info.get("enabled") else None,"tree_clear_size":f"0x{tree_buffer_info.get('clear_size',0):X}" if tree_buffer_info.get("enabled") else None,"tree_frame_size":f"0x{tree_buffer_info.get('frame_size',0):X}" if tree_buffer_info.get("enabled") else None,"tree_terminator_patch_enabled":tree_terminator_info.get("enabled",False),"tree_terminator":tree_terminator_info.get("terminator") if tree_terminator_info.get("enabled") else "0x39","planet_stage_terminator_patch_enabled":planet_stage_terminator_info.get("enabled",False),"planet_stage_terminator":planet_stage_terminator_info.get("terminator") if planet_stage_terminator_info.get("enabled") else "0x39","plsel_graphic_draw_patch_enabled":plsel_graphic_draw_info.get("enabled",False),"plsel_graphic_draw_patch_sites":len(plsel_graphic_draw_info.get("report",[])),"plsel_graphic_draw_patch_report":plsel_graphic_draw_info.get("report",[]),"planet_title_ram_hook_enabled":planet_title_info.get("enabled",False),"planet_title_ram_hook_status":planet_title_info.get("status"),"planet_title_ram_hook_site":planet_title_info.get("site_file"),"planet_title_ram_hook_addr":planet_title_info.get("hook_ram"),"planet_title_ram_hook_titles":" | ".join(planet_title_info.get("titles",[])) if planet_title_info.get("titles") else None,"planet_title_ram_hook_text_area":{"x":planet_title_info.get("text_x"),"y":planet_title_info.get("title_y"),"w":planet_title_info.get("text_w"),"slot_h":planet_title_info.get("slot_h"),"slot_pitch":planet_title_info.get("slot_pitch")} if planet_title_info.get("enabled") else None,"planet_info_element_draw_patch_enabled":planet_info_element_info.get("enabled",False),"planet_info_element_draw_patch_status":planet_info_element_info.get("status"),"planet_info_element_draw_patch_hook":planet_info_element_info.get("hook_ram"),"planet_info_element_draw_patch_report":planet_info_element_info.get("report",[]),"planet_copy_slot_split_hook":planet_copy_slot_split_info,"stage_tim03_moved_uv_patch":stage_tim03_uv_info,"stage_tim04_na_packet_filter_hook":stage_tim04_na_packet_filter_info,"stage_tim04_na_ft4_edge_filter_hook":stage_tim04_na_ft4_edge_filter_info,"stage_tim04_na_ft4_source_filter_hook":stage_tim04_na_ft4_source_filter_info,"stage_tim04_na_template_u_hook":stage_tim04_na_template_u_info,"modus_local_record_early_fix_hook":modus_local_record_early_fix_info,"modus_local_record_fix_hook":modus_local_record_fix_info,"stage_tim04_object_overlay_na_x":args.stage_tim04_object_overlay_na_x,"modus_plmenu01_uvs":{"Earth":"visible x=25 w=23; padded U=0x18 W=0x19","Water":"visible x=50 w=24; padded U=0x31 W=0x1A","Wind":"visible x=76 w=19; padded U=0x4B W=0x15","Fire":"visible x=97 w=18; padded U=0x60 W=0x14"},"modus_plmenu01_ft4_uvs":{"Earth":"U=0x19 W=0x17","Water":"U=0x32 W=0x18","Wind":"U=0x4C W=0x13","Fire":"U=0x61 W=0x12"},"modus_ft4_x_tolerance":args.modus_stage_ft4_x_tolerance,"modus_ft4_source_aliases":"v35 matches stock, steady padded, and FT4 visible U values","ft4_opening_repairs_v43":"Object Overlay relocation-only old-U->edited-U/W/X; PLANET/Modus element first-UV/edge repairs mirror stable template values","modus_exact_ft4_edge_read_hooks":modus_exact_ft4_edge_read_info,"stage_tim04_object_overlay_base_x":args.stage_tim04_object_overlay_base_x,"tuto21_flow":tuto21_info,"planet_info_planet_clut":args.planet_info_planet_clut,"planet_title_tracking":args.planet_title_tracking,"planet_title_bitmap_metrics_enabled":planet_title_use_bitmap_metrics(args),"planet_title_bitmap_tracking":args.planet_title_bitmap_tracking,"planet_title_space_advance":args.planet_title_space_advance,"planet_title_advance_delta":args.planet_title_advance_delta,"planet_title_advance_override":args.planet_title_advance_override,"report_text_patch_enabled":report_text_info.get("enabled",False),"report_text_patch_status":report_text_info.get("status"),"report_text_patch_offset":report_text_info.get("file_offset"),"text_primitive_capacity_patch_enabled":text_primitive_info.get("enabled",False),"text_primitive_capacity_glyphs":text_primitive_info.get("capacity_glyphs"),"text_primitive_half_size":f"0x{text_primitive_info.get('half_size',0):X}" if text_primitive_info.get("enabled") else None,"text_primitive_alloc_size":f"0x{text_primitive_info.get('alloc_size',0):X}" if text_primitive_info.get("enabled") else None}
        sj.write_text(json.dumps(summary,indent=2),encoding="utf-8")
    print("PixyGarden conservative namefix patch summary")
    print("----------------------------------------")
    print(f"Output EXE:              {args.out}")
    if args.reports_dir:
        print(f"Reports directory:       {args.reports_dir}")
    print(f"Rows loaded:             {len(rows)}")
    print(f"Legacy LUI patches:      {lui_total}")
    print(f"Direct MIPS patched:     {sum(1 for r in direct_report if r.get('status') in {'patched','would_patch'})} / {len(direct_report)}")
    print(f"Direct MIPS errors:      {sum(1 for r in direct_report if r.get('status')=='error')}")
    print(f"PTR32 policy:            {args.ptr32_policy}")
    print(f"PTR32 candidates total:  {len(cands)}")
    print(f"PTR32 selected/patched:  {ptr_total}")
    if not args.disable_name_screen_fix:
        print(f"Name fix ptr32 excludes: {len(name_fix_excludes)}")
        print(f"Name fix in-place rows:  {len(name_fix_inplace_report)}")
    if tree_buffer_info.get("enabled"):
        print(f"TREE copy limit:         0x{tree_buffer_info['copy_limit']:X}")
        print(f"TREE frame size:         0x{tree_buffer_info['frame_size']:X}")
    if tree_terminator_info.get("enabled"):
        print(f"TREE TXT terminator:     {tree_terminator_info['terminator']}")
    if text_primitive_info.get("enabled"):
        print(f"Text primitive capacity: {text_primitive_info['capacity_glyphs']} glyphs (half=0x{text_primitive_info['half_size']:X}, alloc=0x{text_primitive_info['alloc_size']:X})")
    if planet_stage_terminator_info.get("enabled"):
        print(f"PLANET/STAGE terminator:{planet_stage_terminator_info['terminator']}")
    else:
        print("TREE TXT terminator:     0x39 (original)")
    if report_text_info.get("enabled"):
        print(f"REPORT text fix:         {report_text_info.get('status')} at {report_text_info.get('file_offset')}")
    if plsel_graphic_draw_info.get("enabled"):
        statuses = ", ".join(f"{r.get('name')}={r.get('status')}" for r in plsel_graphic_draw_info.get("report",[]))
        print(f"PLSEL graphic draw fix:  {statuses}")
    if planet_title_info.get("enabled"):
        print(f"PLANET title RAM hook:   {planet_title_info.get('status')} at {planet_title_info.get('hook_ram')} / file {planet_title_info.get('hook_file')}")
        if planet_title_info.get("titles"):
            print(f"PLANET title text area:  x={planet_title_info.get('text_x')}, y={planet_title_info.get('title_y')}, w={planet_title_info.get('text_w')}, pitch={planet_title_info.get('slot_pitch')}")
            print(f"PLANET title spacing:    tracking={args.planet_title_tracking}, space={args.planet_title_space_advance}, deltas={args.planet_title_advance_delta}, overrides={args.planet_title_advance_override}")
    if planet_info_element_info.get("enabled"):
        print(f"PLANET info elem hook:   {planet_info_element_info.get('status')} at {planet_info_element_info.get('hook_ram')} / file {planet_info_element_info.get('hook_file')}")
        print(f"PLANET info elements:    Earth U=0x{args.planet_info_earth_u:X}/W=0x{args.planet_info_earth_w:X}, Water U=0x{args.planet_info_water_u:X}/W=0x{args.planet_info_water_w:X}, Wind U=0x{args.planet_info_wind_u:X}/W=0x{args.planet_info_wind_w:X}, Fire U=0x{args.planet_info_fire_u:X}/W=0x{args.planet_info_fire_w:X}, Earth X={args.planet_info_earth_x_shift}, Water X={args.planet_info_water_x_shift}, Wind X={args.planet_info_wind_x_shift}, Fire X={args.planet_info_fire_x_shift}, Dash U=0x{args.planet_info_dash_u:X}/W=0x{args.planet_info_dash_w:X}, Planet U=0x{args.planet_info_planet_u:X}, Planet V=0x{args.planet_info_planet_v:X}, Planet palette=0x{args.planet_info_planet_clut:X} (original unless overridden)")
    if stage_tim03_uv_info.get("enabled"):
        print(f"STAGE moved UVs:         {stage_tim03_uv_info.get('status')} planet=({stage_tim03_uv_info.get('planet_u')},{stage_tim03_uv_info.get('planet_v')}) N/A slot U={stage_tim03_uv_info.get('na_line_u')}")
    if modus_crest_text_x_info.get("enabled"):
        print(f"Modus Crest text X:      {modus_crest_text_x_info.get('status')} at {modus_crest_text_x_info.get('site_file')} -> x={modus_crest_text_x_info.get('new_x')}")
    if modus_local_record_early_fix_info.get("enabled"):
        print(f"Modus early record fix:  {modus_local_record_early_fix_info.get('status')} at {modus_local_record_early_fix_info.get('site_file')} -> {modus_local_record_early_fix_info.get('hook_ram')}")
    if planet_copy_slot_split_info.get("enabled"):
        print(f"PLANET copy-slot split:  {planet_copy_slot_split_info.get('status')} at {planet_copy_slot_split_info.get('hook_ram')} / file {planet_copy_slot_split_info.get('hook_file')} ; copied info slot U=0x{args.planet_info_planet_u:X}")
    if tuto21_info.get("enabled"):
        print(f"TUTO state split:       {tuto21_info.get('status')} at {tuto21_info.get('hook_ram')} / file {tuto21_info.get('hook_file')} ; {tuto21_info.get('strategy')}")
    if memory_card_centering_info.get("enabled"):
        print(f"Memory-card centering:  {memory_card_centering_info.get('patched_sites')} patched, {memory_card_centering_info.get('already_applied_sites')} already, {memory_card_centering_info.get('skipped_sites')} skipped; rows loaded={memory_card_centering_info.get('rows_loaded')}")
    if pixy_name_suffix_info.get("enabled"):
        patched_suffix=sum(1 for r in pixy_name_suffix_info.get("report",[]) if r.get("status") in {"patched","would_patch","patched_force","would_patch_force"})
        already_suffix=sum(1 for r in pixy_name_suffix_info.get("report",[]) if str(r.get("status","")).startswith("already"))
        print(f"Pixy suffix spacing:    {pixy_name_suffix_info.get('status')} at {pixy_name_suffix_info.get('hook_ram')} ; patched={patched_suffix}, already={already_suffix}")
    if selector_gsbox_hook_info.get("enabled"):
        print(f"Selector grey-box hook: {selector_gsbox_hook_info.get('status')} at {selector_gsbox_hook_info.get('hook_ram')} ; calls={selector_gsbox_hook_info.get('patched_calls')}/{selector_gsbox_hook_info.get('scanned_calls')} ; Yes={selector_gsbox_hook_info.get('yes_x')},{selector_gsbox_hook_info.get('yes_w')} No={selector_gsbox_hook_info.get('no_x')},{selector_gsbox_hook_info.get('no_w')} ON={selector_gsbox_hook_info.get('on_x')},{selector_gsbox_hook_info.get('on_w')} OFF={selector_gsbox_hook_info.get('off_x')},{selector_gsbox_hook_info.get('off_w')} YesNo h={selector_gsbox_hook_info.get('yesno_height')} y={selector_gsbox_hook_info.get('yesno_y_shift')} ONOFF h={selector_gsbox_hook_info.get('onoff_height')} y={selector_gsbox_hook_info.get('onoff_y_shift')} ProceedExactY={selector_gsbox_hook_info.get('proceed_y')} extra={selector_gsbox_hook_info.get('proceed_extra_y_shift')}")
    _v75_adv = build_advance_table(args)
    _v75_xoff = build_x_offset_table(args)
    def _sx(v): return v-256 if v & 0x80 else v
    print(f"Effective Latin advances:I={_v75_adv[ord('I')]}, i={_v75_adv[ord('i')]}, l={_v75_adv[ord('l')]}, p={_v75_adv[ord('p')]}, m={_v75_adv[ord('m')]}, M={_v75_adv[ord('M')]}, R={_v75_adv[ord('R')]}, P={_v75_adv[ord('P')]}, u={_v75_adv[ord('u')]}, w={_v75_adv[ord('w')]}, .= {_v75_adv[ord('.')]}, z={_v75_adv[ord('z')]}")
    print(f"Effective draw offsets:  P={_sx(_v75_xoff[ord('P')])}, comma={_sx(_v75_xoff[ord(',')])}; pair kerns={iter_v75_pair_kern_map(args)}")
    _v76_font_info = load_v76_font_tim_metrics(args)
    if _v76_font_info.get('enabled'):
        print(f"v80 FONT TIM metrics active from {_v76_font_info.get('tim_path')} using map {_v76_font_info.get('map_path')}; mapped chars={len(_v76_font_info.get('metrics',{}))}")
    else:
        msg=_v76_font_info.get('warning')
        if msg:
            print(f"v80 FONT TIM metrics: {msg}")
    print(f"v81 PLANET title bitmap metrics enabled={planet_title_use_bitmap_metrics(args)}, tracking={args.planet_title_bitmap_tracking}")
    print(f"New file size:           0x{len(exe):X} ({len(exe)} bytes)")
    if clear_data_selector_y_info.get("enabled"):
        print(f"Clear Data selector Y fix: {clear_data_selector_y_info.get('patched_or_already')}/{clear_data_selector_y_info.get('sites')} sites at y=0x{int(clear_data_selector_y_info.get('new_y')) & 0xFFFF:02X}")
    if args.dry_run:
        print("Dry run: output EXE not written."); return
    out.write_bytes(exe); print("Patched EXE written.")
if __name__=="__main__": main()
