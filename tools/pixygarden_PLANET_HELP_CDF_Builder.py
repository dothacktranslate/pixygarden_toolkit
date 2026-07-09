#!/usr/bin/env python3
from __future__ import annotations

"""
PixyGarden PLANET/STAGE HELP.FAT text patch helper.

Default English TXT rules:
  * workbook newlines -> literal bytes 5C 6E ("\\n")
  * text -> CP932, preserving one-byte English ASCII
  * TXT terminator -> 0x00
  * TXT padding/alignment -> 0xFF to 4 bytes

This repacks nested HELP.FAT containers instead of requiring each TXT file to fit
its old individual slot.
"""

import argparse
import csv
import hashlib
import re
import struct
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("This script requires openpyxl. Install with: python -m pip install openpyxl") from exc

SECTOR = 0x800
FAT_ENTRY_SIZE = 0x14
TOP_ENTRY_SIZE = 0x20

@dataclass
class FatEntry:
    path: str
    index: int
    name: str
    table_off: int
    rel_off: int
    abs_off: int
    size: int
    level: int = 0
    first_data: int = 0
    table_count: int = 0

@dataclass
class TopEntry:
    path: str
    index: int
    name: str
    table_off: int
    abs_off: int
    size: int
    alloc: int
    sector: int
    sector_count: int

@dataclass
class TxtBuild:
    excel_row: int
    row_id: int
    archive_path: str
    translation_column_used: str
    body: bytes
    payload: bytes
    terminator: int
    text: str
    notes: list[str]


def norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def normalize_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def normalize_archive_path(p: str) -> str:
    return "/".join(part for part in p.replace("\\", "/").split("/") if part and part != ".").upper()


def align(value: int, n: int) -> int:
    if n <= 1:
        return value
    return ((value + n - 1) // n) * n


def parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(s, 0)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return default


def parse_byte(v: Any, default: Optional[int] = None, bare_two_digit_hex: bool = True) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v) & 0xFF
    if isinstance(v, int):
        return v & 0xFF
    if isinstance(v, float):
        return int(v) & 0xFF
    s = str(v).strip()
    if not s:
        return default
    if s.lower().startswith("0x"):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    if bare_two_digit_hex and re.fullmatch(r"[0-9A-Fa-f]{2}", s):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    try:
        return int(s, 0) & 0xFF
    except ValueError:
        return default


def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column letter: {col!r}")
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_col(headers: list[str], requested: Optional[str], aliases: Iterable[str], required: bool = True) -> Optional[int]:
    raw_lower = {h.strip().lower(): i + 1 for i, h in enumerate(headers) if h}
    norm_map = {normalize_header(h): i + 1 for i, h in enumerate(headers) if h}
    if requested:
        s = requested.strip()
        if s.lower() in raw_lower:
            return raw_lower[s.lower()]
        ns = normalize_header(s)
        if ns in norm_map:
            return norm_map[ns]
        if re.fullmatch(r"[A-Za-z]+", s):
            return col_letter_to_index(s)
        raise ValueError(f"Could not resolve column {requested!r}. Use a column letter or exact header.")
    norm_aliases = [normalize_header(a) for a in aliases]
    for alias in norm_aliases:
        if alias in norm_map:
            return norm_map[alias]
    for alias in norm_aliases:
        for h_norm, idx in norm_map.items():
            if alias and (alias in h_norm or h_norm in alias):
                return idx
    if required:
        raise ValueError(f"Could not find any of these columns: {', '.join(aliases)}")
    return None


def choose_sheet(wb, requested: Optional[str]) -> str:
    if requested:
        if requested not in wb.sheetnames:
            raise SystemExit(f"Sheet {requested!r} not found. Available: {', '.join(wb.sheetnames)}")
        return requested
    for candidate in ("Strings", "PLANET Strings", "Planet Strings", "QA 184px", "QA 154px", "Sheet1"):
        if candidate in wb.sheetnames:
            return candidate
    return wb.sheetnames[0]


def parse_row_set(spec: Optional[str]) -> Optional[set[int]]:
    if not spec:
        return None
    out: set[int] = set()
    for part in re.split(r"[, ]+", spec.strip()):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            ia = int(a, 0); ib = int(b, 0)
            if ib < ia:
                ia, ib = ib, ia
            out.update(range(ia, ib + 1))
        else:
            out.add(int(part, 0))
    return out


def encode_linebreak(mode: str) -> bytes:
    if mode == "backslash_n":
        return b"\\n"
    if mode == "lf":
        return b"\n"
    if mode == "crlf":
        return b"\r\n"
    if mode == "none":
        return b""
    raise ValueError(f"Unknown linebreak mode: {mode}")


def to_fullwidth_ascii(ch: str) -> str:
    if ch == " ":
        return "　"
    o = ord(ch)
    if 0x21 <= o <= 0x7E:
        return chr(0xFF00 + o - 0x20)
    return ch


def encode_body(text: str, text_mode: str, linebreak_mode: str, encoding: str, errors: str) -> bytes:
    text = norm_text(text)
    lb = encode_linebreak(linebreak_mode)
    out = bytearray()
    for i, line in enumerate(text.split("\n")):
        if i:
            out += lb
        if text_mode == "cp932_fullwidth":
            line = "".join(to_fullwidth_ascii(ch) for ch in line)
        elif text_mode not in ("ascii", "cp932"):
            raise ValueError(f"Unknown text mode: {text_mode}")
        out += line.encode(encoding, errors=errors)
    return bytes(out)


def cstr(raw: bytes) -> str:
    return raw.split(b"\0", 1)[0].decode("ascii", errors="replace")


def encode_name(name: str) -> bytes:
    raw = name.encode("ascii", errors="strict")
    if len(raw) > 0x10:
        raise ValueError(f"FAT name too long: {name!r}")
    return raw + b"\0" * (0x10 - len(raw))


def parse_top_cdf(data: bytes, container_name: str = "CDF") -> tuple[dict[str, int], list[TopEntry]]:
    if len(data) < 0x10:
        raise ValueError("File too small for CDF header")
    data_start, count, unk1, unk2 = struct.unpack_from("<IIII", data, 0)
    out: list[TopEntry] = []
    for i in range(count):
        off = 0x10 + i * TOP_ENTRY_SIZE
        if off + TOP_ENTRY_SIZE > len(data):
            raise ValueError(f"Top-level CDF entry {i} exceeds file size")
        name = cstr(data[off:off + 0x14])
        sector, sector_count, size = struct.unpack_from("<III", data, off + 0x14)
        abs_off = sector * SECTOR
        alloc = sector_count * SECTOR
        out.append(TopEntry(name, i, name, off, abs_off, size, alloc, sector, sector_count))
    return {"data_start": data_start, "entry_count": count, "unknown1": unk1, "unknown2": unk2}, out


def parse_fat_table_raw(seg: bytes) -> tuple[int, list[dict[str, Any]]]:
    if len(seg) < FAT_ENTRY_SIZE:
        raise ValueError("FAT segment too small")
    first_data = struct.unpack_from("<I", seg, 0x10)[0]
    if first_data <= 0 or first_data > len(seg) or first_data % FAT_ENTRY_SIZE:
        raise ValueError(f"Bad FAT first_data/table size: 0x{first_data:X}")
    table_count = first_data // FAT_ENTRY_SIZE
    rows: list[dict[str, Any]] = []
    for i in range(table_count):
        off = i * FAT_ENTRY_SIZE
        raw = bytes(seg[off:off + FAT_ENTRY_SIZE])
        name = cstr(raw[:0x10])
        rel = struct.unpack_from("<I", raw, 0x10)[0]
        rows.append({"index": i, "off": off, "raw": raw, "name": name, "rel": rel})
    return first_data, rows


def old_size_for_table_row(row: dict[str, Any], rows: list[dict[str, Any]], seg_len: int) -> int:
    rel = int(row["rel"])
    if not row["name"] or rel <= 0 or rel > seg_len:
        return 0
    later = sorted(int(r["rel"]) for r in rows if r["name"] and int(r["rel"]) > rel and int(r["rel"]) <= seg_len)
    next_rel = later[0] if later else seg_len
    return next_rel - rel


def parse_fat_segment(seg: bytes, base_abs: int, container_path: str, level: int) -> list[FatEntry]:
    try:
        first_data, rows = parse_fat_table_raw(seg)
    except Exception:
        return []
    out: list[FatEntry] = []
    for row in rows:
        name = row["name"]
        rel = int(row["rel"])
        if not name or rel <= 0 or rel > len(seg):
            continue
        old_size = old_size_for_table_row(row, rows, len(seg))
        p = f"{container_path}/{name}"
        out.append(FatEntry(
            path=p,
            index=int(row["index"]),
            name=name,
            table_off=base_abs + int(row["off"]),
            rel_off=rel,
            abs_off=base_abs + rel,
            size=old_size,
            level=level,
            first_data=first_data,
            table_count=len(rows),
        ))
    return out


def recursive_fat_manifest(data: bytes, root_path: str = "HELP.FAT") -> list[FatEntry]:
    out: list[FatEntry] = []
    def rec(seg: bytes, base_abs: int, container_path: str, level: int) -> None:
        entries = parse_fat_segment(seg, base_abs, container_path, level)
        out.extend(entries)
        for e in entries:
            if e.name.upper().endswith(".FAT"):
                rec(data[e.abs_off:e.abs_off + e.size], e.abs_off, e.path, level + 1)
    rec(data, 0, root_path, 0)
    return out


def recursive_cdf_manifest(data: bytes, tops: list[TopEntry]) -> list[Any]:
    out: list[Any] = list(tops)
    def rec(seg: bytes, base_abs: int, container_path: str, level: int) -> None:
        entries = parse_fat_segment(seg, base_abs, container_path, level)
        out.extend(entries)
        for e in entries:
            if e.name.upper().endswith(".FAT"):
                rec(data[e.abs_off:e.abs_off + e.size], e.abs_off, e.path, level + 1)
    for e in tops:
        if e.name.upper().endswith(".FAT"):
            rec(data[e.abs_off:e.abs_off + e.size], e.abs_off, e.path, 1)
    return out


def write_manifest_csv(path: Path, entries: list[Any]) -> None:
    fields = ["path", "name", "index", "level", "table_off", "rel_off", "abs_off", "size", "alloc", "sector", "sector_count"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for e in entries:
            w.writerow({
                "path": e.path,
                "name": e.name,
                "index": e.index,
                "level": getattr(e, "level", 0),
                "table_off": f"0x{getattr(e, 'table_off', 0):X}",
                "rel_off": "" if isinstance(e, TopEntry) else f"0x{e.rel_off:X}",
                "abs_off": f"0x{e.abs_off:X}",
                "size": f"0x{e.size:X}",
                "alloc": "" if not hasattr(e, "alloc") else f"0x{e.alloc:X}",
                "sector": "" if not hasattr(e, "sector") else e.sector,
                "sector_count": "" if not hasattr(e, "sector_count") else e.sector_count,
            })


def load_txt_builds(args: argparse.Namespace) -> tuple[list[TxtBuild], str]:
    wb = load_workbook(args.xlsx, data_only=True)
    sheet_name = choose_sheet(wb, args.sheet)
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    row_col = resolve_col(headers, args.row_column, ["row", "id", "index"], required=False)
    path_col = resolve_col(headers, args.path_column, ["archive_path", "archive path", "path", "cdf_path", "file_path", "file"], required=True)
    trans_col = resolve_col(headers, args.translation_column, [
        "English Translation", "english_translation", "Final Translation", "QA Clean Translation",
        "Better Translation", "Clean Translation", "Translation", "Insertion Text", "Final Insertion Text", "text",
    ], required=True)
    fallback_col = resolve_col(headers, args.fallback_translation_column, [
        "English Translation", "english_translation", "Final Translation", "Translation", "Clean Translation",
    ], required=False)

    term = parse_byte(args.terminator, 0x00)
    if term is None:
        term = 0x00
    term &= 0xFF
    pad_byte = args.pad_byte & 0xFF
    row_filter = parse_row_set(args.rows)

    out: list[TxtBuild] = []
    for r in range(2, ws.max_row + 1):
        raw_path = norm_text(ws.cell(r, path_col).value).strip().replace("\\", "/")
        if not raw_path:
            continue
        if args.only_path_contains and args.only_path_contains.lower() not in raw_path.lower():
            continue
        if args.include_regex and not re.search(args.include_regex, raw_path):
            continue
        if args.exclude_regex and re.search(args.exclude_regex, raw_path):
            continue
        row_id = parse_int(ws.cell(r, row_col).value, r - 1) if row_col else r - 1
        if row_id is None:
            row_id = r - 1
        if row_filter is not None and row_id not in row_filter:
            continue
        text = norm_text(ws.cell(r, trans_col).value)
        used_col = headers[trans_col - 1] or str(trans_col)
        notes: list[str] = []
        if not text and fallback_col:
            text = norm_text(ws.cell(r, fallback_col).value)
            used_col = headers[fallback_col - 1] or str(fallback_col)
        if not text:
            notes.append("empty_translation")
            if not args.allow_empty:
                text = ""
        body = encode_body(text, args.text_mode, args.linebreak_mode, args.encoding, args.encoding_errors)
        if b"\x00" in body:
            notes.append("body_contains_00")
        payload = bytearray(body)
        payload.append(term)
        min_size = align(len(payload), args.align)
        if len(payload) < min_size:
            payload.extend(bytes([pad_byte]) * (min_size - len(payload)))
        out.append(TxtBuild(r, row_id, raw_path, used_col, body, bytes(payload), term, text, notes))
    return out, sheet_name


def make_rebuilt_fat(old_seg: bytes, container_path: str, replacements_by_path: dict[str, bytes], build_meta_by_path: dict[str, TxtBuild], args: argparse.Namespace) -> tuple[bytes, list[dict[str, Any]], set[str]]:
    first_data, table_rows = parse_fat_table_raw(old_seg)
    table = bytearray(old_seg[:first_data])
    data_out = bytearray()
    report_rows: list[dict[str, Any]] = []
    used: set[str] = set()

    for row in table_rows:
        name = row["name"]
        old_rel = int(row["rel"])
        if not name or old_rel <= 0 or old_rel > len(old_seg):
            continue
        archive_path = f"{container_path}/{name}"
        key = normalize_archive_path(archive_path)
        old_size = old_size_for_table_row(row, table_rows, len(old_seg))
        old_payload = bytes(old_seg[old_rel:old_rel + old_size])
        payload = replacements_by_path.get(key, old_payload)
        action = "replaced" if key in replacements_by_path else "kept_original"
        notes: list[str] = []
        tb = build_meta_by_path.get(key)
        if tb:
            notes.extend(tb.notes)
        if key in replacements_by_path:
            used.add(key)
        else:
            notes.append("kept_original_no_replacement")
        new_rel = first_data + len(data_out)
        if new_rel % args.align:
            pad = align(new_rel, args.align) - new_rel
            data_out.extend(bytes([args.pad_byte & 0xFF]) * pad)
            new_rel += pad
        table[row["off"]:row["off"] + 0x10] = encode_name(name)
        struct.pack_into("<I", table, row["off"] + 0x10, new_rel)
        data_out.extend(payload)
        report_rows.append({
            "container": container_path,
            "archive_path": archive_path,
            "action": action,
            "excel_row": tb.excel_row if tb else "",
            "row": tb.row_id if tb else "",
            "old_rel_off": f"0x{old_rel:X}",
            "old_size": f"0x{old_size:X}",
            "new_rel_off": f"0x{new_rel:X}",
            "new_payload_size": f"0x{len(payload):X}",
            "delta_vs_old_slot": len(payload) - old_size,
            "terminator_byte": f"0x{tb.terminator:02X}" if tb else "",
            "encoded_text_len": len(tb.body) if tb else "",
            "generated_min_size": len(tb.payload) if tb else "",
            "line_count": (0 if not tb or tb.text == "" else tb.text.count("\n") + 1),
            "glyph_count_no_newlines": (len(tb.text.replace("\n", "")) if tb else ""),
            "translation_column_used": tb.translation_column_used if tb else "",
            "notes": ";".join(notes),
        })
    return bytes(table + data_out), report_rows, used


def locate_child_segments(root_seg: bytes, root_path: str) -> dict[str, tuple[bytes, FatEntry]]:
    entries = parse_fat_segment(root_seg, 0, root_path, 0)
    out: dict[str, tuple[bytes, FatEntry]] = {}
    for e in entries:
        if e.name.upper().endswith(".FAT"):
            out[normalize_archive_path(e.path)] = (root_seg[e.abs_off:e.abs_off + e.size], e)
    return out


def rebuild_help_fat(root_seg: bytes, root_path: str, builds: list[TxtBuild], args: argparse.Namespace) -> tuple[bytes, list[dict[str, Any]], dict[str, Any]]:
    builds_by_path: dict[str, TxtBuild] = {}
    duplicate_paths: list[str] = []
    for tb in builds:
        key = normalize_archive_path(tb.archive_path)
        if key in builds_by_path:
            duplicate_paths.append(tb.archive_path)
        builds_by_path[key] = tb
    if duplicate_paths:
        raise SystemExit("Duplicate archive paths in workbook: " + ", ".join(duplicate_paths[:10]))

    child_segments = locate_child_segments(root_seg, root_path)
    child_replacements_for_root: dict[str, bytes] = {}
    all_report_rows: list[dict[str, Any]] = []
    used_build_keys: set[str] = set()

    for child_key, (child_seg, child_entry) in child_segments.items():
        child_path = child_entry.path
        child_prefix = normalize_archive_path(child_path) + "/"
        txt_replacements: dict[str, bytes] = {}
        txt_meta: dict[str, TxtBuild] = {}
        for key, tb in builds_by_path.items():
            if key.startswith(child_prefix):
                txt_replacements[key] = tb.payload
                txt_meta[key] = tb
        rebuilt_child, child_rows, used = make_rebuilt_fat(child_seg, child_path, txt_replacements, txt_meta, args)
        used_build_keys.update(used)
        all_report_rows.extend(child_rows)
        child_replacements_for_root[child_key] = rebuilt_child

    # Rebuild the root HELP.FAT with the rebuilt child HELP*.FAT payloads.
    root_rebuilt, root_rows, root_used = make_rebuilt_fat(root_seg, root_path, child_replacements_for_root, {}, args)
    all_report_rows = root_rows + all_report_rows

    missing = sorted(set(builds_by_path) - used_build_keys)
    meta = {
        "old_root_size": len(root_seg),
        "new_root_used_size": len(root_rebuilt),
        "child_count": len(child_segments),
        "build_count": len(builds),
        "missing_build_keys": missing,
        "entries_replaced": sum(1 for r in all_report_rows if r["action"] == "replaced" and str(r["archive_path"]).upper().endswith(".TXT")),
        "entries_kept_original": sum(1 for r in all_report_rows if r["action"] == "kept_original" and str(r["archive_path"]).upper().endswith(".TXT")),
    }
    return root_rebuilt, all_report_rows, meta


def write_report(path: Path, rows: list[dict[str, Any]], extra_fields: Optional[list[str]] = None) -> None:
    fields = [
        "container", "archive_path", "action", "excel_row", "row",
        "old_rel_off", "old_size", "new_rel_off", "new_payload_size", "delta_vs_old_slot",
        "terminator_byte", "encoded_text_len", "generated_min_size", "line_count",
        "glyph_count_no_newlines", "translation_column_used",
    ]
    if extra_fields:
        fields += extra_fields
    fields.append("notes")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def verify_help_texts(root_seg: bytes, builds: list[TxtBuild], root_path: str = "HELP.FAT", terminator: int = 0x00) -> list[str]:
    entries = recursive_fat_manifest(root_seg, root_path)
    lookup = {normalize_archive_path(e.path): e for e in entries if e.name.upper().endswith(".TXT")}
    errors: list[str] = []
    for tb in builds:
        key = normalize_archive_path(tb.archive_path)
        e = lookup.get(key)
        if not e:
            errors.append(f"missing after rebuild: {tb.archive_path}")
            continue
        raw = root_seg[e.abs_off:e.abs_off + e.size]
        idx = raw.find(bytes([terminator]))
        if idx < 0:
            errors.append(f"no terminator 0x{terminator:02X}: {tb.archive_path}")
            continue
        actual = raw[:idx]
        if actual != tb.body:
            errors.append(f"body mismatch: {tb.archive_path}")
    return errors


def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Patch/repack PLANET.CDF HELP.FAT text from a PixyGarden Planet text workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--xlsx", required=True, help="Workbook containing HELP.FAT archive paths and English Translation text")
    ap.add_argument("--sheet", default=None, help="Worksheet name; default auto-picks Strings or first sheet")
    ap.add_argument("--source-cdf", required=True, help="Source/current PLANET.CDF")
    ap.add_argument("--out-cdf", required=True, help="Patched PLANET.CDF output path")
    ap.add_argument("--report", required=True, help="CSV patch report path")
    ap.add_argument("--write-manifest", help="Optional recursive manifest CSV before patching")
    ap.add_argument("--out-help-fat", help="Optionally write rebuilt HELP.FAT bytes for inspection")

    ap.add_argument("--target-container", default="HELP.FAT", help="Top-level HELP.FAT container to rebuild. Default HELP.FAT")
    ap.add_argument("--path-column", default=None)
    ap.add_argument("--translation-column", default=None)
    ap.add_argument("--fallback-translation-column", default=None)
    ap.add_argument("--row-column", default=None)
    ap.add_argument("--text-mode", choices=["ascii", "cp932", "cp932_fullwidth"], default="ascii")
    ap.add_argument("--encoding", default="cp932")
    ap.add_argument("--encoding-errors", choices=["strict", "replace", "ignore"], default="strict")
    ap.add_argument("--linebreak-mode", choices=["backslash_n", "lf", "crlf", "none"], default="backslash_n")
    ap.add_argument("--terminator", default="0x00", help="Terminator byte for generated TXT files. Default 0x00.")
    ap.add_argument("--pad-byte", type=lambda x: int(x, 0), default=0xFF, help="Pad byte. Default 0xFF.")
    ap.add_argument("--align", type=lambda x: int(x, 0), default=4, help="TXT/FAT payload alignment. Default 4.")
    ap.add_argument("--rows", help="Only build these workbook row IDs, e.g. 1,3,5-8")
    ap.add_argument("--only-path-contains", help="Only process workbook rows whose archive_path contains this text")
    ap.add_argument("--include-regex", help="Only process paths matching this regex")
    ap.add_argument("--exclude-regex", help="Skip paths matching this regex")
    ap.add_argument("--allow-empty", action="store_true")
    ap.add_argument("--dry-run", action="store_true", help="Write report/manifest only; do not write binary outputs")
    ap.add_argument("--no-fail-on-warnings", action="store_true", help="Return success even with warnings")
    args = ap.parse_args(argv)

    source = Path(args.source_cdf)
    cdf = bytearray(source.read_bytes())
    original_size = len(cdf)
    top_info, tops = parse_top_cdf(bytes(cdf), source.name)
    all_entries = recursive_cdf_manifest(bytes(cdf), tops)
    if args.write_manifest:
        write_manifest_csv(Path(args.write_manifest), all_entries)

    target = None
    for e in tops:
        if normalize_archive_path(e.path) == normalize_archive_path(args.target_container):
            target = e
            break
    if target is None:
        raise SystemExit(f"Top-level target container {args.target_container!r} not found in {args.source_cdf}")

    builds, sheet_name = load_txt_builds(args)
    old_help = bytes(cdf[target.abs_off:target.abs_off + target.size])
    rebuilt_used, rows, meta = rebuild_help_fat(old_help, args.target_container, builds, args)

    warnings: list[str] = []
    if meta["missing_build_keys"]:
        warnings.append(f"{len(meta['missing_build_keys'])} workbook rows did not match entries in {args.target_container}")

    new_used = int(meta["new_root_used_size"])
    if new_used > target.alloc:
        raise SystemExit(f"Rebuilt HELP.FAT uses 0x{new_used:X}, exceeding PLANET.CDF allocation 0x{target.alloc:X}")
    rebuilt_write = rebuilt_used + bytes([args.pad_byte & 0xFF]) * (target.alloc - new_used)

    # Update top-level CDF entry size to actual rebuilt HELP.FAT used size.
    struct.pack_into("<I", cdf, target.table_off + 0x1C, new_used)

    for r in rows:
        r["top_old_size"] = f"0x{target.size:X}"
        r["top_alloc"] = f"0x{target.alloc:X}"
        r["top_new_used_size"] = f"0x{new_used:X}"
        r["top_slack_after_repack"] = f"0x{target.alloc - new_used:X}"
    write_report(Path(args.report), rows, ["top_old_size", "top_alloc", "top_new_used_size", "top_slack_after_repack"])

    verify_errors = verify_help_texts(rebuilt_used, builds, args.target_container, parse_byte(args.terminator, 0x00) or 0x00)
    if verify_errors:
        warnings.append(f"verification errors: {len(verify_errors)}")
        for err in verify_errors[:10]:
            warnings.append(err)

    if not args.dry_run:
        cdf[target.abs_off:target.abs_off + len(rebuilt_write)] = rebuilt_write
        out = Path(args.out_cdf)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_bytes(bytes(cdf))
        if out.stat().st_size != original_size:
            raise SystemExit(f"Output CDF size changed unexpectedly: source 0x{original_size:X}, output 0x{out.stat().st_size:X}")
        if args.out_help_fat:
            p = Path(args.out_help_fat); p.parent.mkdir(parents=True, exist_ok=True); p.write_bytes(rebuilt_used)

    print("PixyGarden PLANET.CDF HELP.FAT text patcher")
    print("--------------------------------------------")
    print(f"Workbook:             {args.xlsx}")
    print(f"Sheet:                {sheet_name}")
    print(f"Rows loaded:          {len(builds)}")
    print(f"Source CDF:           {args.source_cdf}")
    print(f"Target:               {args.target_container}")
    print(f"Old HELP.FAT size:    0x{target.size:X}")
    print(f"HELP.FAT allocation:  0x{target.alloc:X}")
    print(f"Rebuilt used size:    0x{new_used:X}")
    print(f"Slack after repack:   0x{target.alloc - new_used:X}")
    print(f"Entries replaced:     {meta['entries_replaced']}")
    print(f"Entries kept original:{meta['entries_kept_original']}")
    print(f"Report:               {args.report}")
    print(f"Text mode:            {args.text_mode}")
    print(f"Linebreak mode:       {args.linebreak_mode}")
    print(f"Terminator:           0x{(parse_byte(args.terminator, 0) or 0):02X}")
    print(f"Pad byte:             0x{args.pad_byte & 0xFF:02X}")
    print(f"Verification errors:  {len(verify_errors)}")
    if args.dry_run:
        print("Dry run: no binary outputs written.")
    else:
        print(f"Patched CDF:          {args.out_cdf}")
        print(f"Output SHA-256:       {hashlib.sha256(Path(args.out_cdf).read_bytes()).hexdigest()}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  - {w}")
        if not args.no_fail_on_warnings:
            return 2
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
