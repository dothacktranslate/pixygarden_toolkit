#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import struct
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

# PixyGarden SCR Repacker
# v11: EVENT.CDF-safe template mode; uses original CDF SCR structure/control bytes and replaces text only.
#
# Known/assumed SCR structure from E001.SCR:
#   byte 0: opcode
#   byte 1: total record length, including opcode and length bytes
#   byte 2..length-1: opcode-specific payload
#
# Text records:
#   opcode 0x11 appears to carry text/control strings.
#   Dialogue strings usually begin with ASCII prefix "c7".
#   Line breaks are literal backslash-n bytes: 5C 6E.
#
# Branch/target records:
#   opcode 0x06, 0x09, 0x3B: payload[0:2] = little-endian absolute SCR offset
#   opcode 0x07: payload[0:2] = little-endian absolute SCR offset,
#                payload[2:]  = ASCII condition/expression
#
# This tool supports:
#   1) Extract a translation CSV from an SCR.
#   2) Rebuild/repack an SCR from that CSV.
#   3) Recalculate branch targets after record sizes change.
#
# It preserves all non-text payloads byte-for-byte except target fields in
# known branch opcodes.


@dataclass
class Record:
    index: int
    old_offset: int
    opcode: int
    length: int
    payload: bytes

    @property
    def total_len(self) -> int:
        return 2 + len(self.payload)

    def to_bytes(self) -> bytes:
        total = self.total_len
        if total > 0xFF:
            raise ValueError(
                f"Record {self.index} at old offset 0x{self.old_offset:04X} "
                f"would be 0x{total:X} bytes, but SCR record length is one byte."
            )
        if total < 2:
            raise ValueError("Internal error: SCR record shorter than 2 bytes")
        return bytes([self.opcode & 0xFF, total & 0xFF]) + self.payload


@dataclass
class Replacement:
    key: str
    text: str
    enabled: bool = True


@dataclass
class RepackReportRow:
    record_index: int
    old_offset: int
    new_offset: int
    opcode: int
    old_length: int
    new_length: int
    status: str
    notes: str = ""


@dataclass
class CdfEntry:
    index: int
    name: str
    sector: int
    sector_count: int
    size: int
    table_offset: int

    @property
    def abs_offset(self) -> int:
        return self.sector * 0x800

    @property
    def alloc_size(self) -> int:
        return self.sector_count * 0x800


def parse_byte_list(spec: str) -> set[int]:
    out: set[int] = set()
    if not spec:
        return out
    for part in re.split(r"[, ]+", spec.strip()):
        if part:
            out.add(int(part, 0) & 0xFF)
    return out


def norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def to_fullwidth_ascii(ch: str) -> str:
    if ch == " ":
        return "　"
    o = ord(ch)
    if 0x21 <= o <= 0x7E:
        return chr(0xFF00 + o - 0x20)
    return ch


def decode_cp932_lossy(b: bytes) -> str:
    return b.decode("cp932", errors="replace")


def split_text_payload(payload: bytes, terminator: int = 0x00) -> tuple[bytes, bytes, bytes]:
    pos = payload.find(bytes([terminator & 0xFF]))
    if pos < 0:
        return payload, b"", b""
    return payload[:pos], payload[pos:pos + 1], payload[pos + 1:]


def encode_linebreak(mode: str) -> bytes:
    if mode == "backslash_n":
        return b"\\n"
    if mode == "lf":
        return b"\n"
    if mode == "crlf":
        return b"\r\n"
    if mode == "none":
        return b""
    raise ValueError(f"Unknown linebreak mode: {mode}")


def encode_text_body(
    text: str,
    text_mode: str,
    linebreak_mode: str,
    encoding: str = "cp932",
    errors: str = "replace",
    fullwidth_digits: bool = False,
) -> bytes:
    text = norm_text(text)
    lb = encode_linebreak(linebreak_mode)
    out = bytearray()

    def encode_ascii_mixed_digit_line(line: str) -> bytes:
        b = bytearray()
        for ch in line:
            if fullwidth_digits and "0" <= ch <= "9":
                # Emit CP932 full-width digit. MAIN.EXE's CP932 direct path maps
                # these back to the same digit advance groups already in use.
                b += to_fullwidth_ascii(ch).encode("cp932", errors=errors)
            else:
                b += ch.encode(encoding, errors=errors)
        return bytes(b)

    for i, line in enumerate(text.split("\n")):
        if i:
            out += lb
        if text_mode == "ascii":
            out += encode_ascii_mixed_digit_line(line)
        elif text_mode == "cp932":
            out += encode_ascii_mixed_digit_line(line)
        elif text_mode == "cp932_fullwidth":
            full = "".join(to_fullwidth_ascii(ch) for ch in line)
            out += full.encode("cp932", errors=errors)
        else:
            raise ValueError(f"Unknown text mode: {text_mode}")
    return bytes(out)


def parse_scr(data: bytes) -> tuple[list[Record], bytes]:
    records: list[Record] = []
    pos = 0
    idx = 0
    while pos + 2 <= len(data):
        opcode = data[pos]
        length = data[pos + 1]
        # FF FF or any impossible length ends normal parsing and becomes tail.
        if length < 2 or pos + length > len(data):
            break
        payload = data[pos + 2:pos + length]
        records.append(Record(idx, pos, opcode, length, payload))
        pos += length
        idx += 1
    tail = data[pos:]
    return records, tail


def record_key(rec: Record) -> str:
    return f"{rec.index}"


def offset_key(rec: Record) -> str:
    return f"0x{rec.old_offset:04X}"


def extract_text_csv(records: list[Record], out_csv: Path, text_opcode: int, terminator: int) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "record_index", "offset_hex", "opcode_hex", "record_length", "payload_length",
        "body_bytes_before_terminator", "has_null_terminator", "prefix",
        "original_decoded", "editable_text_without_c7_prefix", "translation", "notes",
    ]
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for rec in records:
            if rec.opcode != text_opcode:
                continue
            body, term, tail = split_text_payload(rec.payload, terminator)
            decoded = decode_cp932_lossy(body)
            prefix = ""
            editable = decoded
            if decoded.startswith("c7"):
                prefix = "c7"
                editable = decoded[2:]
            notes = []
            notes.append(f"terminator=0x{terminator:02X}" if term else "no_terminator_found")
            if tail:
                notes.append(f"tail_after_terminator_len=0x{len(tail):X}")
            w.writerow({
                "record_index": rec.index,
                "offset_hex": f"0x{rec.old_offset:04X}",
                "opcode_hex": f"0x{rec.opcode:02X}",
                "record_length": rec.length,
                "payload_length": len(rec.payload),
                "body_bytes_before_terminator": len(body),
                "has_null_terminator": "YES" if term else "NO",
                "prefix": prefix,
                "original_decoded": decoded,
                "editable_text_without_c7_prefix": editable,
                "translation": "",
                "notes": ";".join(notes),
            })



def _iter_replacement_rows_table(path: Path):
    """Yield dict rows from CSV/TSV/XLSX replacement files."""
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xlsx replacement files requires openpyxl. "
                "Install it with: python -m pip install openpyxl"
            ) from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return

        fieldnames = [("" if h is None else str(h).strip()) for h in header]
        while fieldnames and fieldnames[-1] == "":
            fieldnames.pop()
        if not fieldnames:
            return

        for vals in rows:
            vals = list(vals[:len(fieldnames)])
            if len(vals) < len(fieldnames):
                vals.extend([None] * (len(fieldnames) - len(vals)))
            row = {
                fieldnames[i]: ("" if vals[i] is None else vals[i])
                for i in range(len(fieldnames))
                if fieldnames[i] != ""
            }
            if any(str(v).strip() != "" for v in row.values()):
                yield row
        return

    last_exc = None
    for enc in ("utf-8-sig", "cp932", "utf-16"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
                except csv.Error:
                    dialect = csv.excel_tab if "\t" in sample and "," not in sample else csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                if not reader.fieldnames:
                    return
                for row in reader:
                    yield row
            return
        except UnicodeDecodeError as exc:
            last_exc = exc
            continue

    raise UnicodeError(f"Could not decode replacement file {path}. Last error: {last_exc}")


def _field_picker(fieldnames):
    fields_lower = {str(name).lower(): name for name in fieldnames if name is not None}

    def pick(*names: str) -> Optional[str]:
        for n in names:
            if n.lower() in fields_lower:
                return fields_lower[n.lower()]
        return None

    return pick


def load_replacements(csv_path: Optional[Path], blank_means_keep: bool = True) -> dict[str, Replacement]:
    if not csv_path:
        return {}
    replacements: dict[str, Replacement] = {}

    rows = list(_iter_replacement_rows_table(csv_path))
    if not rows:
        return replacements

    pick = _field_picker(rows[0].keys())
    rec_col = pick("record_index", "index", "record")
    off_col = pick("offset_hex", "offset", "old_offset")
    trans_col = pick("translation", "new_text", "english", "english_translation", "Better Translation")
    enable_col = pick("enabled", "use", "selected")
    if not trans_col:
        raise ValueError("Replacement file must contain a 'translation' column, or equivalent.")

    for row in rows:
        trans = norm_text(row.get(trans_col, ""))
        if blank_means_keep and trans == "":
            continue
        enabled = True
        if enable_col:
            flag = str(row.get(enable_col, "")).strip().lower()
            if flag in ("0", "no", "false", "skip", "disabled"):
                enabled = False
        keys = []
        if rec_col and str(row.get(rec_col, "")).strip() != "":
            keys.append(str(int(str(row.get(rec_col)).strip(), 0)))
        if off_col and str(row.get(off_col, "")).strip() != "":
            off = int(str(row.get(off_col)).strip(), 0)
            keys.append(f"0x{off:04X}")
        for k in keys:
            replacements[k] = Replacement(k, trans, enabled=enabled)
    return replacements


def find_replacement(replacements: dict[str, Replacement], rec: Record) -> Optional[Replacement]:
    keys = (record_key(rec), offset_key(rec), offset_key(rec).lower())
    for k in keys:
        if k in replacements and replacements[k].enabled:
            return replacements[k]
    return None


def has_c7_prefix(payload: bytes, terminator: int = 0x00) -> bool:
    body, _, _ = split_text_payload(payload, terminator)
    return body.startswith(b"c7")


def make_text_payload(
    rec: Record,
    new_text: str,
    *,
    text_mode: str,
    linebreak_mode: str,
    encoding: str,
    encoding_errors: str,
    terminator: int,
    auto_c7: bool,
    keep_original_tail: bool,
    pad_to_original_payload_len: bool,
    fullwidth_digits: bool,
) -> tuple[bytes, str]:
    notes = []
    _body, _term, tail = split_text_payload(rec.payload, terminator)
    text = norm_text(new_text)

    # Keep c7 as literal ASCII control bytes. Do not allow --fullwidth-digits
    # to turn the "7" into CP932 fullwidth ７.
    literal_prefix = b""
    if auto_c7 and has_c7_prefix(rec.payload, terminator):
        if text.startswith("c7"):
            text = text[2:]
            notes.append("kept_user_c7_prefix_as_literal_ascii")
        else:
            notes.append("auto_added_c7_prefix_as_literal_ascii")
        literal_prefix = b"c7"

    encoded = bytearray(literal_prefix)
    encoded.extend(
        encode_text_body(
            text,
            text_mode,
            linebreak_mode,
            encoding,
            encoding_errors,
            fullwidth_digits=fullwidth_digits,
        )
    )

    if fullwidth_digits and any("0" <= ch <= "9" for ch in text):
        notes.append("encoded_ascii_digits_as_cp932_fullwidth")

    # Original EVENT opcode 0x11 text/control payloads are double-null
    # terminated: body 00 00. The second 00 is not padding outside the record;
    # it is part of the record payload. If an older translated EVENT.CDF already
    # lost that byte, restore it here.
    encoded.append(terminator & 0xFF)
    if tail:
        encoded.extend(tail)
        notes.append(f"preserved_original_tail_len=0x{len(tail):X}")
    else:
        encoded.append(terminator & 0xFF)
        notes.append("added_original_style_second_null_terminator")

    if pad_to_original_payload_len and len(encoded) <= len(rec.payload):
        if len(encoded) < len(rec.payload):
            encoded.extend(b"\x00" * (len(rec.payload) - len(encoded)))
            notes.append("padded_text_payload_to_original_len")
    elif pad_to_original_payload_len and len(encoded) > len(rec.payload):
        raise ValueError(
            f"Replacement for record {rec.index} at 0x{rec.old_offset:04X} does not fit original payload: "
            f"new=0x{len(encoded):X}, old=0x{len(rec.payload):X}"
        )

    if len(encoded) + 2 > 0xFF:
        raise ValueError(
            f"Replacement for record {rec.index} at 0x{rec.old_offset:04X} makes record too long: "
            f"total=0x{len(encoded)+2:X}, max=0xFF"
        )
    return bytes(encoded), ";".join(notes)



def build_records_with_replacements(
    records: list[Record],
    replacements: dict[str, Replacement],
    *,
    text_opcode: int,
    text_mode: str,
    linebreak_mode: str,
    encoding: str,
    encoding_errors: str,
    terminator: int,
    auto_c7: bool,
    keep_original_tail: bool,
    preserve_record_lengths: bool,
    fullwidth_digits: bool,
) -> tuple[list[Record], list[RepackReportRow]]:
    new_records: list[Record] = []
    report: list[RepackReportRow] = []
    cur_new_off = 0
    for rec in records:
        new_payload = rec.payload
        status = "unchanged"
        notes = ""
        if rec.opcode == text_opcode:
            repl = find_replacement(replacements, rec)
            if repl is not None:
                new_payload, notes = make_text_payload(
                    rec, repl.text,
                    text_mode=text_mode,
                    linebreak_mode=linebreak_mode,
                    encoding=encoding,
                    encoding_errors=encoding_errors,
                    terminator=terminator,
                    auto_c7=auto_c7,
                    keep_original_tail=keep_original_tail,
                    pad_to_original_payload_len=preserve_record_lengths,
                    fullwidth_digits=fullwidth_digits,
                )
                status = "text_replaced"
        new_rec = Record(rec.index, rec.old_offset, rec.opcode, 2 + len(new_payload), new_payload)
        new_records.append(new_rec)
        report.append(RepackReportRow(
            record_index=rec.index,
            old_offset=rec.old_offset,
            new_offset=cur_new_off,
            opcode=rec.opcode,
            old_length=rec.length,
            new_length=new_rec.total_len,
            status=status,
            notes=notes,
        ))
        cur_new_off += new_rec.total_len
    return new_records, report


def _append_report_note(row: Optional[RepackReportRow], msg: str) -> None:
    if row is None:
        return
    row.notes = ";".join(x for x in [row.notes, msg] if x)


def _mark_report_status(row: Optional[RepackReportRow], status_suffix: str) -> None:
    if row is None:
        return
    if row.status == "unchanged":
        row.status = status_suffix
    elif status_suffix not in row.status.split("+"):
        row.status = row.status + "+" + status_suffix


def _nearest_boundary(target: int, boundaries: set[int], max_delta: int) -> Optional[int]:
    """Return a nearby valid command boundary, preferring +1 over -1.

    The E004 failure showed targets one byte before the intended command.  The
    positive-side preference fixes that specific class before trying the reverse.
    """
    for delta in range(1, max_delta + 1):
        plus = target + delta
        if plus in boundaries:
            return plus
        minus = target - delta
        if minus in boundaries:
            return minus
    return None


def _record_new_offsets(records: list[Record]) -> tuple[dict[int, int], int]:
    offsets: dict[int, int] = {}
    cur = 0
    for rec in records:
        offsets[rec.index] = cur
        cur += rec.total_len
    return offsets, cur


def _branch_target_opcodes(jump_opcodes: set[int], conditional_jump_opcodes: set[int]) -> set[int]:
    return set(jump_opcodes) | set(conditional_jump_opcodes)


def update_branch_targets(
    old_records: list[Record],
    new_records: list[Record],
    report: list[RepackReportRow],
    *,
    jump_opcodes: set[int],
    conditional_jump_opcodes: set[int],
    strict_targets: bool,
    auto_fix_near_targets: bool = True,
    near_target_delta: int = 1,
) -> None:
    """Recalculate branch targets using old command boundaries.

    v5 change:
      Older builds could silently preserve a target that was not in the old
      boundary map.  That is dangerous because the target may be only one byte
      before/after a command after text insertion.  This version can repair
      small near-boundary errors before mapping to the new boundary.
    """
    old_to_new = {old.old_offset: row.new_offset for old, row in zip(old_records, report)}
    old_boundaries = set(old_to_new)
    report_by_idx = {r.record_index: r for r in report}
    target_opcodes = _branch_target_opcodes(jump_opcodes, conditional_jump_opcodes)

    for rec in new_records:
        if rec.opcode not in target_opcodes:
            continue
        if len(rec.payload) < 2:
            continue

        old_target = int.from_bytes(rec.payload[:2], "little")
        row = report_by_idx.get(rec.index)

        mapped_from = old_target
        if mapped_from not in old_to_new:
            fixed = None
            if auto_fix_near_targets:
                fixed = _nearest_boundary(mapped_from, old_boundaries, near_target_delta)
            if fixed is not None:
                _append_report_note(row, f"old_branch_target_autofixed_0x{old_target:04X}_to_old_boundary_0x{fixed:04X}")
                _mark_report_status(row, "branch_target_autofixed")
                mapped_from = fixed
            else:
                msg = f"branch_target_0x{old_target:04X}_not_in_record_map"
                _append_report_note(row, msg)
                if strict_targets:
                    raise ValueError(
                        f"Record {rec.index} at old offset 0x{rec.old_offset:04X} has target 0x{old_target:04X}, "
                        "which is not a parsed record start. Rebuild aborted so the SCR cannot desynchronize."
                    )
                continue

        new_target = old_to_new[mapped_from]
        if new_target > 0xFFFF:
            raise ValueError(f"New branch target 0x{new_target:X} exceeds 16-bit SCR offset range.")

        if new_target != old_target:
            payload = bytearray(rec.payload)
            struct.pack_into("<H", payload, 0, new_target)
            rec.payload = bytes(payload)
            rec.length = 2 + len(rec.payload)
            if row:
                msg = f"branch_target_updated_0x{old_target:04X}_to_0x{new_target:04X}"
                _mark_report_status(row, "branch_target_updated")
                _append_report_note(row, msg)
        else:
            _append_report_note(row, f"branch_target_unchanged_0x{old_target:04X}")


def validate_and_repair_branch_targets(
    records: list[Record],
    report: list[RepackReportRow],
    *,
    jump_opcodes: set[int],
    conditional_jump_opcodes: set[int],
    tail_start: int,
    strict_targets: bool,
    auto_fix_near_targets: bool = True,
    near_target_delta: int = 1,
    context: str = "",
) -> int:
    """Final post-rebuild branch validation.

    This validates the finished SCR, not just the old-to-new mapping. Every
    branch target should land on a command boundary or the parsed tail/terminator
    boundary. If a target is off by a tiny amount, it can be repaired in place.

    Returns the number of final-stage target repairs.
    """
    offsets_by_idx, end_of_records = _record_new_offsets(records)
    boundaries = set(offsets_by_idx.values())
    boundaries.add(tail_start)
    boundaries.add(end_of_records)
    report_by_idx = {r.record_index: r for r in report}
    target_opcodes = _branch_target_opcodes(jump_opcodes, conditional_jump_opcodes)
    repairs = 0

    for rec in records:
        if rec.opcode not in target_opcodes or len(rec.payload) < 2:
            continue
        target = int.from_bytes(rec.payload[:2], "little")
        row = report_by_idx.get(rec.index)
        if target in boundaries:
            continue

        fixed = None
        if auto_fix_near_targets:
            fixed = _nearest_boundary(target, boundaries, near_target_delta)

        if fixed is not None:
            payload = bytearray(rec.payload)
            struct.pack_into("<H", payload, 0, fixed)
            rec.payload = bytes(payload)
            rec.length = 2 + len(rec.payload)
            repairs += 1
            _mark_report_status(row, "final_branch_target_autofixed")
            _append_report_note(row, f"final_branch_target_autofixed_0x{target:04X}_to_0x{fixed:04X}")
            continue

        msg = f"FINAL_BRANCH_TARGET_NOT_ON_COMMAND_BOUNDARY_0x{target:04X}"
        if context:
            msg = f"{context}:{msg}"
        _append_report_note(row, msg)
        if strict_targets:
            raise ValueError(
                f"{context + ': ' if context else ''}Record {rec.index} at old offset 0x{rec.old_offset:04X} "
                f"has final target 0x{target:04X}, which is not a parsed command/tail boundary."
            )

    return repairs



def write_report(path: Path, rows: list[RepackReportRow], tail: bytes, new_size: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "record_index", "old_offset_hex", "new_offset_hex", "opcode_hex",
        "old_length_hex", "new_length_hex", "delta", "status", "notes",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({
                "record_index": r.record_index,
                "old_offset_hex": f"0x{r.old_offset:04X}",
                "new_offset_hex": f"0x{r.new_offset:04X}",
                "opcode_hex": f"0x{r.opcode:02X}",
                "old_length_hex": f"0x{r.old_length:02X}",
                "new_length_hex": f"0x{r.new_length:02X}",
                "delta": r.new_length - r.old_length,
                "status": r.status,
                "notes": r.notes,
            })
        w.writerow({})
        w.writerow({"record_index": "SUMMARY", "old_offset_hex": "tail_hex", "new_offset_hex": tail.hex(" ")})
        w.writerow({"record_index": "SUMMARY", "old_offset_hex": "new_size_hex", "new_offset_hex": f"0x{new_size:X}"})


def rebuild_scr_bytes(records: list[Record], tail: bytes) -> bytes:
    out = bytearray()
    for rec in records:
        out += rec.to_bytes()
    out += tail
    return bytes(out)





def parse_event_cdf(data: bytes) -> tuple[int, int, int, int, list[CdfEntry]]:
    """Parse the simple EVENT/TREE-style CDF root directory.

    Entry layout:
      name[0x14], sector u32, sector_count u32, size u32
    Header:
      data_start_or_first_data_offset u32, entry_count u32, unk1 u32, unk2 u32
    """
    if len(data) < 0x10:
        raise ValueError("CDF too small")
    data_start, count, unk1, unk2 = struct.unpack_from("<IIII", data, 0)
    if count <= 0 or count > 4096:
        raise ValueError(f"Implausible CDF entry count: {count}")
    table_end = 0x10 + count * 0x20
    if table_end > len(data):
        raise ValueError("CDF directory table exceeds file size")

    entries: list[CdfEntry] = []
    for i in range(count):
        off = 0x10 + i * 0x20
        name = data[off:off + 0x14].split(b"\x00", 1)[0].decode("ascii", errors="replace")
        sector, sector_count, size = struct.unpack_from("<III", data, off + 0x14)
        e = CdfEntry(i, name, sector, sector_count, size, off)
        if e.abs_offset + e.size > len(data):
            raise ValueError(
                f"CDF entry {i} {name} points outside file: off=0x{e.abs_offset:X}, size=0x{e.size:X}, file=0x{len(data):X}"
            )
        entries.append(e)
    return data_start, count, unk1, unk2, entries


def cdf_entry_data(data: bytes, e: CdfEntry) -> bytes:
    return data[e.abs_offset:e.abs_offset + e.size]


def write_event_cdf_in_place(
    original: bytes,
    entries: list[CdfEntry],
    new_entry_data: dict[str, bytes],
    *,
    pad_byte: int = 0xFF,
) -> bytes:
    """Write changed entries into their existing sector allocations.

    Fails if any replacement no longer fits the original allocation.
    """
    out = bytearray(original)
    for e in entries:
        if e.name not in new_entry_data:
            continue
        blob = new_entry_data[e.name]
        if len(blob) > e.alloc_size:
            raise ValueError(
                f"{e.name} grew to 0x{len(blob):X}, which exceeds its original allocation 0x{e.alloc_size:X}. "
                "Use --repack-cdf to rebuild the CDF layout."
            )
        out[e.abs_offset:e.abs_offset + len(blob)] = blob
        out[e.abs_offset + len(blob):e.abs_offset + e.alloc_size] = bytes([pad_byte & 0xFF]) * (e.alloc_size - len(blob))
        struct.pack_into("<I", out, e.table_offset + 0x1C, len(blob))
    return bytes(out)


def write_event_cdf_repacked(
    original: bytes,
    entries: list[CdfEntry],
    new_entry_data: dict[str, bytes],
    *,
    data_start: int,
    pad_byte: int = 0xFF,
) -> bytes:
    """Rebuild the CDF data region and update sectors/sizes.

    This preserves the header and names, but recalculates sector, sector_count,
    and byte size for every entry. Files are written in directory order and
    padded to 0x800-byte sector boundaries.
    """
    if data_start <= 0 or data_start % 0x800 != 0:
        # EVENT.CDF uses a byte offset such as 0x1800. If another CDF's first
        # header value is odd, fall back to the original first entry sector.
        data_start = min(e.abs_offset for e in entries)

    table_end = 0x10 + len(entries) * 0x20
    if data_start < table_end:
        raise ValueError(f"CDF data_start 0x{data_start:X} overlaps directory table end 0x{table_end:X}")

    out = bytearray(original[:data_start])
    if len(out) < data_start:
        out.extend(bytes([pad_byte & 0xFF]) * (data_start - len(out)))

    cur = data_start
    for e in entries:
        blob = new_entry_data.get(e.name)
        if blob is None:
            blob = cdf_entry_data(original, e)

        if cur % 0x800:
            pad = 0x800 - (cur % 0x800)
            out.extend(bytes([pad_byte & 0xFF]) * pad)
            cur += pad

        sector = cur // 0x800
        sectors = (len(blob) + 0x7FF) // 0x800
        alloc = sectors * 0x800

        # Update directory table.
        struct.pack_into("<III", out, e.table_offset + 0x14, sector, sectors, len(blob))

        out.extend(blob)
        out.extend(bytes([pad_byte & 0xFF]) * (alloc - len(blob)))
        cur += alloc

    return bytes(out)


def write_combined_repack_report(
    path: Path,
    scr_reports: list[tuple[str, int, int, int, int, list[RepackReportRow]]],
    cdf_rows: list[dict[str, object]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "section", "scr_file", "record_index", "old_offset_hex", "new_offset_hex", "opcode_hex",
        "old_length_hex", "new_length_hex", "delta", "status", "notes",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for scr_name, old_size, new_size, replaced, branches, rows in scr_reports:
            w.writerow({
                "section": "SCR_SUMMARY",
                "scr_file": scr_name,
                "record_index": "SUMMARY",
                "old_offset_hex": f"old_size=0x{old_size:X}",
                "new_offset_hex": f"new_size=0x{new_size:X}",
                "delta": new_size - old_size,
                "status": f"replaced={replaced}; branch_changes={branches}",
            })
            for r in rows:
                w.writerow({
                    "section": "SCR_RECORD",
                    "scr_file": scr_name,
                    "record_index": r.record_index,
                    "old_offset_hex": f"0x{r.old_offset:04X}",
                    "new_offset_hex": f"0x{r.new_offset:04X}",
                    "opcode_hex": f"0x{r.opcode:02X}",
                    "old_length_hex": f"0x{r.old_length:02X}",
                    "new_length_hex": f"0x{r.new_length:02X}",
                    "delta": r.new_length - r.old_length,
                    "status": r.status,
                    "notes": r.notes,
                })
        for row in cdf_rows:
            w.writerow({
                "section": "CDF_ENTRY",
                "scr_file": row.get("name", ""),
                "record_index": row.get("index", ""),
                "old_offset_hex": row.get("old_size_hex", ""),
                "new_offset_hex": row.get("new_size_hex", ""),
                "delta": row.get("delta", ""),
                "status": row.get("status", ""),
                "notes": row.get("notes", ""),
            })


def classify_text_body(decoded: str) -> tuple[str, str]:
    """Classify opcode 0x11 payloads for extraction.

    In EVENT SCRs, opcode 0x11 can be real dialogue/text or compact control
    strings such as u1v3f67f7f50f77f7f7ff9202020. Those control strings are
    needed by the script but should normally not be translated.
    """
    if decoded.startswith("c7"):
        return "dialogue_c7", "c7 dialogue/text"
    # Heuristic: these are compact ASCII control strings made of command-ish
    # letters/numbers and no Japanese/full-width chars or spaces except padding.
    stripped = decoded.strip("\x00 ").strip()
    if stripped and re.fullmatch(r"[A-Za-z][A-Za-z0-9@_=+\-.,:;#%&/\\]*", stripped):
        return "control_ascii", "likely non-dialogue script/text-control payload"
    # Any payload containing Japanese/full-width/non-ASCII is worth reviewing.
    if any(ord(ch) >= 0x80 for ch in decoded):
        return "review_text", "non-ASCII text-like payload"
    return "review_ascii", "ASCII payload; review before translating"


def should_extract_text_row(kind: str, extract_filter: str) -> bool:
    if extract_filter == "all":
        return True
    if extract_filter == "c7":
        return kind == "dialogue_c7"
    if extract_filter == "no-control":
        return kind != "control_ascii"
    if extract_filter == "review":
        return kind in {"dialogue_c7", "review_text", "review_ascii"}
    raise ValueError(f"Unknown extract filter: {extract_filter}")


def iter_text_rows_for_csv(records: list[Record], scr_file: str, text_opcode: int, terminator: int, extract_filter: str = "all"):
    for rec in records:
        if rec.opcode != text_opcode:
            continue
        body, term, tail = split_text_payload(rec.payload, terminator)
        decoded = decode_cp932_lossy(body)
        kind, kind_note = classify_text_body(decoded)
        if not should_extract_text_row(kind, extract_filter):
            continue

        prefix = ""
        editable = decoded
        if decoded.startswith("c7"):
            prefix = "c7"
            editable = decoded[2:]

        notes = [kind_note]
        if term:
            notes.append(f"terminator=0x{terminator:02X}")
        else:
            notes.append("no_terminator_found")
        if tail:
            notes.append(f"tail_after_terminator_len=0x{len(tail):X}")

        yield {
            "scr_file": scr_file,
            "record_index": rec.index,
            "offset_hex": f"0x{rec.old_offset:04X}",
            "opcode_hex": f"0x{rec.opcode:02X}",
            "record_length": rec.length,
            "payload_length": len(rec.payload),
            "body_bytes_before_terminator": len(body),
            "has_null_terminator": "YES" if term else "NO",
            "record_kind": kind,
            "prefix": prefix,
            "original_decoded": decoded,
            "editable_text_without_c7_prefix": editable,
            "translation": "",
            "notes": ";".join(notes),
        }


def extract_text_csv_multi(items: list[tuple[Path, str, list[Record]]], out_csv: Path, text_opcode: int, terminator: int, extract_filter: str = "all") -> int:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "scr_file", "record_index", "offset_hex", "opcode_hex", "record_length", "payload_length",
        "body_bytes_before_terminator", "has_null_terminator", "record_kind", "prefix", "original_decoded",
        "editable_text_without_c7_prefix", "translation", "notes",
    ]
    count = 0
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for _path, scr_name, records in items:
            for row in iter_text_rows_for_csv(records, scr_name, text_opcode, terminator, extract_filter):
                w.writerow(row)
                count += 1
    return count


def load_replacements_multi(csv_path: Optional[Path], blank_means_keep: bool = True) -> tuple[dict[str, dict[str, Replacement]], dict[str, Replacement]]:
    """Load replacements from a combined CSV/XLSX."""
    if not csv_path:
        return {}, {}

    by_scr: dict[str, dict[str, Replacement]] = {}
    global_replacements: dict[str, Replacement] = {}

    rows = list(_iter_replacement_rows_table(csv_path))
    if not rows:
        return by_scr, global_replacements

    pick = _field_picker(rows[0].keys())
    scr_col = pick("scr_file", "file", "filename", "scr")
    rec_col = pick("record_index", "index", "record")
    off_col = pick("offset_hex", "offset", "old_offset")
    trans_col = pick("translation", "new_text", "english", "english_translation", "Better Translation")
    enable_col = pick("enabled", "use", "selected")

    if not trans_col:
        raise ValueError("Replacement file must contain a 'translation' column, or equivalent.")

    for row in rows:
        trans = norm_text(row.get(trans_col, ""))
        if blank_means_keep and trans == "":
            continue

        enabled = True
        if enable_col:
            flag = str(row.get(enable_col, "")).strip().lower()
            if flag in ("0", "no", "false", "skip", "disabled"):
                enabled = False

        keys = []
        if rec_col and str(row.get(rec_col, "")).strip() != "":
            keys.append(str(int(str(row.get(rec_col)).strip(), 0)))
        if off_col and str(row.get(off_col, "")).strip() != "":
            off = int(str(row.get(off_col)).strip(), 0)
            keys.append(f"0x{off:04X}")
            keys.append(f"0x{off:04x}")

        scr_name = norm_text(row.get(scr_col, "")).strip().replace("\\", "/") if scr_col else ""
        target = by_scr.setdefault(scr_name, {}) if scr_name else global_replacements
        for k in keys:
            target[k] = Replacement(k, trans, enabled=enabled)

    return by_scr, global_replacements


def replacements_for_scr(scr_name: str, by_scr: dict[str, dict[str, Replacement]], global_repl: dict[str, Replacement]) -> dict[str, Replacement]:
    scr_name = scr_name.replace("\\", "/")
    basename = Path(scr_name).name
    merged: dict[str, Replacement] = {}
    merged.update(global_repl)
    # Exact relative path match first, then basename fallback.
    if scr_name in by_scr:
        merged.update(by_scr[scr_name])
    if basename in by_scr:
        merged.update(by_scr[basename])
    return merged


def discover_scr_inputs(args) -> tuple[list[tuple[Path, str]], Optional[Path]]:
    """Return [(path, scr_file_name_for_csv)], base_dir_for_relative_names."""
    items: list[Path] = []
    base: Optional[Path] = None

    if args.scr:
        items.append(Path(args.scr))
    if args.scr_dir:
        base = Path(args.scr_dir)
        pattern = args.scr_pattern or "*.SCR"
        if args.recursive:
            items.extend(sorted(base.rglob(pattern)))
        else:
            items.extend(sorted(base.glob(pattern)))
    if args.scr_glob:
        items.extend(sorted(Path().glob(args.scr_glob)))

    # De-duplicate while preserving order.
    seen = set()
    unique: list[Path] = []
    for p in items:
        rp = p.resolve()
        if rp not in seen:
            seen.add(rp)
            unique.append(p)

    out: list[tuple[Path, str]] = []
    for p in unique:
        if base is not None:
            try:
                name = p.resolve().relative_to(base.resolve()).as_posix()
            except Exception:
                name = p.name
        else:
            name = p.name
        out.append((p, name))
    return out, base


def repack_scr_data(
    data: bytes,
    scr_name: str,
    replacements: dict[str, Replacement],
    args,
) -> tuple[bytes, list[RepackReportRow], int, int]:
    records, tail = parse_scr(data)
    if not records:
        raise ValueError(f"No SCR records parsed from {scr_name}")

    new_records, report_rows = build_records_with_replacements(
        records,
        replacements,
        text_opcode=args.text_opcode & 0xFF,
        text_mode=args.text_mode,
        fullwidth_digits=getattr(args, "fullwidth_digits", False),
        linebreak_mode=args.linebreak_mode,
        encoding=args.encoding,
        encoding_errors=args.encoding_errors,
        terminator=args.terminator & 0xFF,
        auto_c7=args.auto_c7,
        keep_original_tail=args.keep_original_tail,
        preserve_record_lengths=args.preserve_record_lengths,
    )

    jump_opcodes = parse_byte_list(args.jump_opcodes)
    conditional_jump_opcodes = parse_byte_list(args.conditional_jump_opcodes)

    # v5: run branch update even when preserving record lengths, because a
    # previously built SCR can already contain a near-boundary target error.
    update_branch_targets(
        records,
        new_records,
        report_rows,
        jump_opcodes=jump_opcodes,
        conditional_jump_opcodes=conditional_jump_opcodes,
        strict_targets=args.strict_targets,
        auto_fix_near_targets=getattr(args, "auto_fix_near_targets", True),
        near_target_delta=getattr(args, "near_target_delta", 1),
    )

    # Final target validation on the finished record list. Include the parsed
    # tail/terminator offset as a valid landing point.
    offsets_by_idx, tail_start = _record_new_offsets(new_records)
    validate_and_repair_branch_targets(
        new_records,
        report_rows,
        jump_opcodes=jump_opcodes,
        conditional_jump_opcodes=conditional_jump_opcodes,
        tail_start=tail_start,
        strict_targets=getattr(args, "final_strict_targets", True),
        auto_fix_near_targets=getattr(args, "auto_fix_near_targets", True),
        near_target_delta=getattr(args, "near_target_delta", 1),
        context=scr_name,
    )

    out_bytes = rebuild_scr_bytes(new_records, tail)

    # Parse once more and validate the actual serialized bytes. This catches any
    # accidental length/payload mismatch introduced during repair.
    final_records, final_tail = parse_scr(out_bytes)
    final_offsets, final_tail_start = _record_new_offsets(final_records)
    final_boundaries = set(final_offsets.values()) | {final_tail_start}
    target_opcodes = _branch_target_opcodes(jump_opcodes, conditional_jump_opcodes)
    final_bad = []
    for rec in final_records:
        if rec.opcode in target_opcodes and len(rec.payload) >= 2:
            target = int.from_bytes(rec.payload[:2], "little")
            if target < len(out_bytes) and target not in final_boundaries:
                final_bad.append((rec.index, rec.old_offset, rec.opcode, target))
    if final_bad and getattr(args, "final_strict_targets", True):
        sample = ", ".join(f"rec{idx}@0x{off:04X}->0x{target:04X}" for idx, off, op, target in final_bad[:8])
        raise ValueError(f"{scr_name}: final serialized SCR has branch targets not on command boundaries: {sample}")

    replaced_count = sum(1 for r in report_rows if "text_replaced" in r.status)
    branch_count = sum(
        1 for r in report_rows
        if "branch_target_updated" in r.status
        or "branch_target_autofixed" in r.status
        or "final_branch_target_autofixed" in r.status
    )
    return out_bytes, report_rows, replaced_count, branch_count


def repack_one_scr(
    src: Path,
    scr_name: str,
    out_path: Path,
    report_path: Optional[Path],
    replacements: dict[str, Replacement],
    args,
) -> tuple[int, int, int, int]:
    data = src.read_bytes()
    out_bytes, report_rows, replaced_count, branch_count = repack_scr_data(data, scr_name, replacements, args)
    if report_path:
        # Recompute tail for reporting only.
        _records, tail = parse_scr(out_bytes)
        write_report(report_path, report_rows, tail, len(out_bytes))
    if not args.dry_run:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(out_bytes)
    return len(data), len(out_bytes), replaced_count, branch_count


def handle_event_cdf_mode(args) -> int:
    """EVENT.CDF mode.

    v11 template model:
      --event-cdf is the base/output source. Non-SCR entries are copied from it.
      --template-event-cdf, when supplied, is used only for SCR parsing/rebuild
      so original command/control structure and post-string bytes are preserved.

    This avoids using a previously translated/repacked EVENT.CDF as the
    structural source after it has already lost bytes or shifted targets.
    """
    base_event_path = Path(args.event_cdf)
    base_data = base_event_path.read_bytes()
    base_data_start, base_count, base_unk1, base_unk2, base_entries = parse_event_cdf(base_data)
    base_by_name = {e.name: e for e in base_entries}

    template_event_path = Path(args.template_event_cdf) if getattr(args, "template_event_cdf", None) else base_event_path
    template_data = template_event_path.read_bytes()
    template_data_start, template_count, template_unk1, template_unk2, template_entries = parse_event_cdf(template_data)

    template_scr_entries = [
        e for e in template_entries
        if e.name.upper().endswith(".SCR") and e.name in base_by_name
    ]

    parsed_items = []
    for e in template_scr_entries:
        records, tail = parse_scr(cdf_entry_data(template_data, e))
        if not records:
            print(f"WARNING: no SCR records parsed from template CDF entry {e.name}; skipping")
            continue
        parsed_items.append((template_event_path, e.name, records))

    print("PixyGarden SCR tool v11 EVENT.CDF-safe mode")
    print("-------------------------------------------")
    print(f"Base EVENT.CDF:     {base_event_path}")
    print(f"Template EVENT.CDF: {template_event_path}")
    print(f"EVENT.CDF entries:  {len(base_entries)}")
    print(f"SCR entries parsed from template: {len(parsed_items)}")

    if args.extract_text_csv:
        count_rows = extract_text_csv_multi(
            parsed_items,
            Path(args.extract_text_csv),
            args.text_opcode & 0xFF,
            args.terminator & 0xFF,
            args.extract_filter,
        )
        print(f"Wrote combined text template: {args.extract_text_csv}")
        print(f"Text rows written: {count_rows}")

    if args.replacements_csv:
        if not args.out_event_cdf:
            raise SystemExit("--out-event-cdf is required when rebuilding an EVENT.CDF.")

        by_scr, global_repl = load_replacements_multi(
            Path(args.replacements_csv),
            blank_means_keep=not args.blank_means_empty,
        )

        new_entry_data: dict[str, bytes] = {}
        scr_reports: list[tuple[str, int, int, int, int, list[RepackReportRow]]] = []
        cdf_rows: list[dict[str, object]] = []

        for template_e in template_scr_entries:
            old_blob = cdf_entry_data(template_data, template_e)
            base_e = base_by_name[template_e.name]
            replacements = replacements_for_scr(template_e.name, by_scr, global_repl)

            out_blob, report_rows, replaced, branches = repack_scr_data(
                old_blob,
                template_e.name,
                replacements,
                args,
            )
            new_entry_data[template_e.name] = out_blob

            scr_reports.append((template_e.name, len(old_blob), len(out_blob), replaced, branches, report_rows))

            fits = len(out_blob) <= base_e.alloc_size
            cdf_rows.append({
                "index": base_e.index,
                "name": template_e.name,
                "old_size_hex": f"0x{len(old_blob):X}",
                "new_size_hex": f"0x{len(out_blob):X}",
                "delta": len(out_blob) - len(old_blob),
                "status": "fits_base_allocation" if fits else "requires_cdf_repack",
                "notes": (
                    f"template_allocation=0x{template_e.alloc_size:X}; "
                    f"base_allocation=0x{base_e.alloc_size:X}; "
                    f"replaced={replaced}; branch_changes={branches}"
                ),
            })
            print(
                f"{template_e.name}: template_old=0x{len(old_blob):X} "
                f"new=0x{len(out_blob):X} delta={len(out_blob)-len(old_blob)} "
                f"replaced={replaced} branch_changes={branches}"
            )

        # Output starts from the base CDF, not the template CDF. This preserves
        # any non-SCR translated images/resources already present in --event-cdf.
        if args.repack_cdf:
            out_cdf = write_event_cdf_repacked(
                base_data,
                base_entries,
                new_entry_data,
                data_start=base_data_start,
                pad_byte=args.cdf_pad_byte,
            )
            print("CDF packing: rebuilt base CDF directory sectors/sizes from scratch")
        else:
            out_cdf = write_event_cdf_in_place(
                base_data,
                base_entries,
                new_entry_data,
                pad_byte=args.cdf_pad_byte,
            )
            print("CDF packing: in-place within base CDF original sector allocations")

        if not args.dry_run:
            Path(args.out_event_cdf).write_bytes(out_cdf)
            print(f"Wrote EVENT.CDF: {args.out_event_cdf}")
        else:
            print("Dry run: no EVENT.CDF written.")

        report_path = Path(args.event_cdf_report) if args.event_cdf_report else Path(args.out_event_cdf).with_suffix(".repack_report.csv")
        write_combined_repack_report(report_path, scr_reports, cdf_rows)
        print(f"Wrote report: {report_path}")

        print(f"Total text records replaced:  {sum(x[3] for x in scr_reports)}")
        print(f"Total branch target changes:  {sum(x[4] for x in scr_reports)}")

    if not args.extract_text_csv and not args.replacements_csv:
        print("No action requested. Use --extract-text-csv and/or --replacements-csv.")
    return 0



def main() -> int:
    ap = argparse.ArgumentParser(description="PixyGarden SCR text extractor/repacker with multi-SCR support, extraction filtering, and branch-target fixups.")
    ap.add_argument("--scr", help="Single input .SCR file")
    ap.add_argument("--scr-dir", help="Directory containing SCR files")
    ap.add_argument("--scr-pattern", default="*.SCR", help="Pattern used with --scr-dir. Default *.SCR")
    ap.add_argument("--scr-glob", help="Glob pattern for SCR files, e.g. 'EVENT_extracted/**/*.SCR'")
    ap.add_argument("--recursive", action="store_true", help="Use recursive search with --scr-dir")

    ap.add_argument("--out", help="Output rebuilt .SCR file for single-SCR mode")
    ap.add_argument("--out-dir", help="Output directory for multi-SCR rebuild mode")
    ap.add_argument("--extract-text-csv", help="Write extracted text template CSV. Multi mode writes one combined CSV with scr_file column.")
    ap.add_argument("--replacements-csv", help="CSV/XLSX containing translations to insert")
    ap.add_argument("--report", help="Write rebuild report CSV for single-SCR mode")
    ap.add_argument("--report-dir", help="Directory for per-SCR reports in multi-SCR mode")

    ap.add_argument("--event-cdf", help="Input EVENT.CDF. In this mode, .SCR entries are read directly from the CDF.")
    ap.add_argument("--template-event-cdf", help="Original/template EVENT.CDF used only for SCR record structure/control bytes. Non-SCR files still come from --event-cdf.")
    ap.add_argument("--out-event-cdf", help="Output EVENT.CDF after applying replacements directly to .SCR entries.")
    ap.add_argument("--event-cdf-report", help="Combined EVENT.CDF rebuild report CSV.")
    ap.add_argument("--repack-cdf", action="store_true",
                    help="Rebuild the CDF sector layout and update every entry's sector/sector_count/size. Without this, changed files must fit their original allocations.")
    ap.add_argument("--cdf-pad-byte", type=lambda x: int(x, 0), default=0xFF,
                    help="Padding byte for CDF sector slack. Default 0xFF.")

    ap.add_argument("--text-opcode", type=lambda x: int(x, 0), default=0x11)
    ap.add_argument("--extract-filter", choices=["all", "c7", "no-control", "review"], default="all",
                    help="Filter opcode 0x11 rows during extraction. Use c7 for dialogue-only EVENT templates.")
    ap.add_argument("--jump-opcodes", default="0x06,0x09,0x3B")
    ap.add_argument("--conditional-jump-opcodes", default="0x07")
    ap.add_argument("--text-mode", choices=["ascii", "cp932", "cp932_fullwidth"], default="ascii")
    ap.add_argument("--fullwidth-digits", action="store_true",
                    help="Encode ASCII digits 0-9 as CP932 full-width digits inside text records. Useful when EVENT control parsing garbles raw digits.")
    ap.add_argument("--encoding", default="cp932", help="Encoding for ascii/cp932 text modes; default cp932")
    ap.add_argument("--encoding-errors", choices=["strict", "replace", "ignore"], default="replace")
    ap.add_argument("--linebreak-mode", choices=["backslash_n", "lf", "crlf", "none"], default="backslash_n")
    ap.add_argument("--terminator", type=lambda x: int(x, 0), default=0x00)
    ap.add_argument("--no-auto-c7", dest="auto_c7", action="store_false")
    ap.set_defaults(auto_c7=True)
    ap.add_argument("--keep-original-tail", action="store_true")
    ap.add_argument("--preserve-record-lengths", action="store_true")
    ap.add_argument("--blank-means-empty", action="store_true")
    ap.add_argument("--strict-targets", action="store_true",
                    help="Abort when an original branch target cannot be mapped to a parsed command boundary.")
    ap.add_argument("--strict-final-targets", dest="final_strict_targets", action="store_true",
                    help="Abort if the rebuilt SCR still has branch-like targets that miss command boundaries after near-target auto-fix.")
    ap.set_defaults(final_strict_targets=False)
    ap.add_argument("--no-auto-fix-near-targets", dest="auto_fix_near_targets", action="store_false",
                    help="Disable automatic repair of branch targets that are off by a small number of bytes.")
    ap.set_defaults(auto_fix_near_targets=True)
    ap.add_argument("--near-target-delta", type=int, default=1,
                    help="Maximum +/- byte distance for near-boundary branch target auto-repair. Default 1.")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if args.event_cdf:
        return handle_event_cdf_mode(args)

    inputs, base = discover_scr_inputs(args)
    if not inputs:
        raise SystemExit("No SCR input selected. Use --scr, --scr-dir, or --scr-glob.")

    parsed_items = []
    for src, scr_name in inputs:
        data = src.read_bytes()
        records, tail = parse_scr(data)
        if not records:
            print(f"WARNING: no SCR records parsed from {src}; skipping")
            continue
        parsed_items.append((src, scr_name, records))

    if not parsed_items:
        raise SystemExit("No valid SCR files parsed.")

    print("PixyGarden SCR tool v4")
    print("----------------------")
    print(f"SCR files parsed: {len(parsed_items)}")

    if args.extract_text_csv:
        count = extract_text_csv_multi(parsed_items, Path(args.extract_text_csv), args.text_opcode & 0xFF, args.terminator & 0xFF, args.extract_filter)
        print(f"Wrote combined text template: {args.extract_text_csv}")
        print(f"Text rows written: {count}")

    if args.replacements_csv:
        by_scr, global_repl = load_replacements_multi(Path(args.replacements_csv), blank_means_keep=not args.blank_means_empty)
        multi = len(parsed_items) > 1 or args.scr_dir or args.scr_glob
        if multi and not args.out_dir:
            raise SystemExit("--out-dir is required when rebuilding multiple SCR files.")
        if not multi and not args.out:
            raise SystemExit("--out is required when rebuilding a single SCR file.")

        total_replaced = 0
        total_branches = 0
        for src, scr_name, _records in parsed_items:
            replacements = replacements_for_scr(scr_name, by_scr, global_repl)
            if multi:
                out_path = Path(args.out_dir) / scr_name
                report_path = None
                if args.report_dir:
                    report_path = Path(args.report_dir) / (Path(scr_name).with_suffix(".SCR_report.csv"))
            else:
                out_path = Path(args.out)
                report_path = Path(args.report) if args.report else None

            old_size, new_size, replaced, branches = repack_one_scr(src, scr_name, out_path, report_path, replacements, args)
            total_replaced += replaced
            total_branches += branches
            print(f"{scr_name}: old=0x{old_size:X} new=0x{new_size:X} delta={new_size-old_size} replaced={replaced} branch_fixups={branches}")

        print(f"Total text records replaced:  {total_replaced}")
        print(f"Total branch targets updated: {total_branches}")
        if args.dry_run:
            print("Dry run: no rebuilt SCR files written.")

    if not args.extract_text_csv and not args.replacements_csv:
        print("No action requested. Use --extract-text-csv and/or --replacements-csv.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
