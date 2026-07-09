#!/usr/bin/env python3
from __future__ import annotations

"""
PixyGarden TREE.CDF direct TXT patcher, flexible workbook + INFO.FAT repack mode.

This version is intended for TREE.CDF DETAILS.FAT/INFO.FAT English text.
It reads a translation workbook, builds INFO_*.TXT payloads, rebuilds the
INFO.FAT data area with updated per-file offsets, then patches the rebuilt
INFO.FAT back into TREE.CDF.

Why this exists
---------------
Same-slot replacement is unnecessarily strict for INFO.FAT text because every
TXT file lives inside one nested FAT container. Some translated TXT payloads may
be a few bytes larger than their original individual slot while the whole rebuilt
INFO.FAT still fits in the original INFO.FAT container. This script repacks the
TXT entries inside INFO.FAT, preserving the outer TREE.CDF size.

Default TREE English workflow
-----------------------------
  * workbook newlines -> literal bytes 5C 6E ("\\n")
  * text body -> CP932, preserving one-byte English ASCII
  * terminator -> 0x00 by default, so ASCII '9' is safe
  * generated TXT payloads are aligned to 4 bytes with 0xFF
  * rebuilt INFO.FAT is padded back to its original container size with 0xFF
  * TREE.CDF output size remains exactly unchanged

Typical use
-----------
  python pixygarden_TREE_TXT_CDF_Builder_v11_repack_INFO.py ^
    --xlsx pixygarden_tree_cdf_text_QA_fixed_v80.xlsx ^
    --source-cdf TREE.CDF ^
    --out-cdf TREE_text_patched.CDF ^
    --report TREE_text_patch_report.csv

Dry run:
  python pixygarden_TREE_TXT_CDF_Builder_v11_repack_INFO.py ^
    --xlsx pixygarden_tree_cdf_text_QA_fixed_v80.xlsx ^
    --source-cdf TREE.CDF ^
    --out-cdf TREE_text_patched.CDF ^
    --report TREE_text_patch_report.csv ^
    --dry-run
"""

import argparse
import csv
import hashlib
import re
import struct
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("This script requires openpyxl. Install with: python -m pip install openpyxl") from exc

SECTOR = 0x800
FAT_ENTRY_SIZE = 0x14
TOP_ENTRY_SIZE = 0x20

# ---------------------------------------------------------------------------
# Generic/workbook helpers
# ---------------------------------------------------------------------------

def norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(s, 0)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return default


def parse_byte(v: Any, default: Optional[int] = None, bare_two_digit_hex: bool = True) -> Optional[int]:
    """Parse a byte. Bare two hex-looking digits are treated as hex, so '39' -> 0x39."""
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return int(v) & 0xFF
    if isinstance(v, int):
        return v & 0xFF
    if isinstance(v, float):
        return int(v) & 0xFF
    s = str(v).strip()
    if not s:
        return default
    if s.lower().startswith("0x"):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    if bare_two_digit_hex and re.fullmatch(r"[0-9A-Fa-f]{2}", s):
        try:
            return int(s, 16) & 0xFF
        except ValueError:
            return default
    try:
        return int(s, 0) & 0xFF
    except ValueError:
        return default


def col_letter_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column letter: {col!r}")
        n = n * 26 + (ord(ch) - 64)
    return n


def normalize_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def resolve_col(
    headers: list[str],
    requested: Optional[str],
    aliases: Iterable[str],
    required: bool = True,
    fuzzy_contains: bool = True,
) -> Optional[int]:
    """Resolve a 1-based worksheet column from a letter, exact header, alias, or fuzzy header."""
    raw_lower = {h.strip().lower(): i + 1 for i, h in enumerate(headers) if h}
    norm_map = {normalize_header(h): i + 1 for i, h in enumerate(headers) if h}

    if requested:
        s = requested.strip()
        if s.lower() in raw_lower:
            return raw_lower[s.lower()]
        ns = normalize_header(s)
        if ns in norm_map:
            return norm_map[ns]
        # Only treat pure letters as Excel columns after header matching.
        if re.fullmatch(r"[A-Za-z]+", s):
            return col_letter_to_index(s)
        raise ValueError(f"Could not resolve column {requested!r}. Use a column letter or exact header.")

    norm_aliases = [normalize_header(a) for a in aliases]
    for alias in norm_aliases:
        if alias in norm_map:
            return norm_map[alias]
    if fuzzy_contains:
        for alias in norm_aliases:
            if not alias:
                continue
            for h_norm, idx in norm_map.items():
                if alias in h_norm or h_norm in alias:
                    return idx
    if required:
        raise ValueError(f"Could not find any of these columns: {', '.join(aliases)}")
    return None


def choose_sheet(wb, requested: Optional[str]) -> str:
    if requested:
        if requested not in wb.sheetnames:
            raise SystemExit(f"Sheet {requested!r} not found. Available: {', '.join(wb.sheetnames)}")
        return requested
    for candidate in (
        "TREE Strings",
        "Literal Detailed TREE",
        "Linefilled TREE",
        "Retranslated TREE",
        "Retranslated TREE",
        "Retranslated",
        "Sheet1",
    ):
        if candidate in wb.sheetnames:
            return candidate
    return wb.sheetnames[0]


def normalize_archive_path(p: str) -> str:
    return "/".join(part for part in p.replace("\\", "/").split("/") if part and part != ".").upper()


def align(value: int, n: int) -> int:
    if n <= 1:
        return value
    return ((value + n - 1) // n) * n


def parse_row_set(spec: Optional[str]) -> Optional[set[int]]:
    if not spec:
        return None
    out: set[int] = set()
    for part in re.split(r"[, ]+", spec.strip()):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            ia = int(a, 0)
            ib = int(b, 0)
            if ib < ia:
                ia, ib = ib, ia
            out.update(range(ia, ib + 1))
        else:
            out.add(int(part, 0))
    return out


def encode_linebreak(mode: str) -> bytes:
    if mode == "backslash_n":
        return b"\\n"
    if mode == "lf":
        return b"\n"
    if mode == "crlf":
        return b"\r\n"
    if mode == "none":
        return b""
    raise ValueError(f"Unknown linebreak mode: {mode}")


def to_fullwidth_ascii(ch: str) -> str:
    if ch == " ":
        return "　"
    o = ord(ch)
    if 0x21 <= o <= 0x7E:
        return chr(0xFF00 + o - 0x20)
    return ch


def encode_body(text: str, text_mode: str, linebreak_mode: str, encoding: str, errors: str) -> bytes:
    text = norm_text(text)
    lb = encode_linebreak(linebreak_mode)
    out = bytearray()
    for i, line in enumerate(text.split("\n")):
        if i:
            out += lb
        if text_mode == "cp932_fullwidth":
            line = "".join(to_fullwidth_ascii(ch) for ch in line)
        elif text_mode not in ("ascii", "cp932"):
            raise ValueError(f"Unknown text mode: {text_mode}")
        out += line.encode(encoding, errors=errors)
    return bytes(out)

# ---------------------------------------------------------------------------
# CDF / FAT parsing
# ---------------------------------------------------------------------------

@dataclass
class ArchiveEntry:
    level: int
    container: str
    path: str
    index: int
    name: str
    kind: str
    table_off: int
    rel_off: int
    abs_off: int
    size: int
    sector: Optional[int] = None
    sector_count: Optional[int] = None
    alloc: Optional[int] = None
    padding: Optional[int] = None
    first_data: Optional[int] = None
    table_count: Optional[int] = None


def cstr(raw: bytes) -> str:
    return raw.split(b"\0", 1)[0].decode("ascii", errors="replace")


def encode_name(name: str) -> bytes:
    raw = name.encode("ascii", errors="strict")
    if len(raw) > 0x10:
        raise ValueError(f"FAT name too long: {name!r}")
    return raw + b"\0" * (0x10 - len(raw))


def parse_top_cdf(data: bytes, container_name: str = "CDF") -> tuple[dict[str, int], list[ArchiveEntry]]:
    if len(data) < 0x10:
        raise ValueError("File too small for TREE-style CDF header")
    data_start, count, unk1, unk2 = struct.unpack_from("<IIII", data, 0)
    table_start = 0x10
    entries: list[ArchiveEntry] = []
    for i in range(count):
        off = table_start + i * TOP_ENTRY_SIZE
        if off + TOP_ENTRY_SIZE > len(data):
            raise ValueError(f"Top-level CDF entry {i} exceeds file size")
        name = cstr(data[off:off + 0x14])
        sector, sector_count, size = struct.unpack_from("<III", data, off + 0x14)
        abs_off = sector * SECTOR
        alloc = sector_count * SECTOR
        entries.append(ArchiveEntry(
            level=0,
            container=container_name,
            path=name,
            index=i,
            name=name,
            kind="CDF_ENTRY",
            table_off=off,
            rel_off=abs_off,
            abs_off=abs_off,
            sector=sector,
            sector_count=sector_count,
            size=size,
            alloc=alloc,
            padding=alloc - size,
        ))
    return {"data_start": data_start, "entry_count": count, "unknown1": unk1, "unknown2": unk2}, entries


def parse_fat_segment(seg: bytes, base_abs: int, container_path: str, level: int) -> list[ArchiveEntry]:
    if len(seg) < FAT_ENTRY_SIZE:
        return []
    first_data = struct.unpack_from("<I", seg, 0x10)[0]
    if first_data <= 0 or first_data > len(seg) or first_data % 4:
        return []
    table_count = first_data // FAT_ENTRY_SIZE
    if table_count <= 0 or table_count > 10000:
        return []

    raw: list[dict[str, int | str]] = []
    for i in range(table_count):
        off = i * FAT_ENTRY_SIZE
        if off + FAT_ENTRY_SIZE > len(seg):
            break
        name = cstr(seg[off:off + 0x10])
        rel = struct.unpack_from("<I", seg, off + 0x10)[0]
        if name and 0 < rel <= len(seg):
            raw.append({"index": i, "name": name, "rel_off": rel, "table_off": off})

    sorted_raw = sorted(raw, key=lambda e: int(e["rel_off"]))
    rel_to_next: dict[int, int] = {}
    for j, e in enumerate(sorted_raw):
        rel = int(e["rel_off"])
        next_rel = len(seg)
        if j + 1 < len(sorted_raw):
            next_rel = int(sorted_raw[j + 1]["rel_off"])
        rel_to_next[rel] = next_rel

    entries: list[ArchiveEntry] = []
    for e in raw:
        rel = int(e["rel_off"])
        size = rel_to_next.get(rel, len(seg)) - rel
        abs_off = base_abs + rel
        name = str(e["name"])
        path = f"{container_path}/{name}"
        entries.append(ArchiveEntry(
            level=level,
            container=container_path,
            path=path,
            index=int(e["index"]),
            name=name,
            kind="FAT_ENTRY",
            table_off=base_abs + int(e["table_off"]),
            rel_off=rel,
            abs_off=abs_off,
            size=size,
            first_data=first_data,
            table_count=table_count,
        ))
    return entries


def recursive_manifest(data: bytes, top_entries: list[ArchiveEntry], max_depth: int = 10) -> list[ArchiveEntry]:
    out: list[ArchiveEntry] = list(top_entries)

    def rec(seg: bytes, base_abs: int, container_path: str, level: int) -> None:
        if level > max_depth:
            return
        entries = parse_fat_segment(seg, base_abs, container_path, level)
        out.extend(entries)
        for e in entries:
            if e.name.upper().endswith(".FAT"):
                child = data[e.abs_off:e.abs_off + e.size]
                rec(child, e.abs_off, e.path, level + 1)

    for e in top_entries:
        if e.name.upper().endswith(".FAT"):
            seg = data[e.abs_off:e.abs_off + e.size]
            rec(seg, e.abs_off, e.path, 1)
    return out


def build_lookup(entries: list[ArchiveEntry]) -> dict[str, ArchiveEntry]:
    return {normalize_archive_path(e.path): e for e in entries}


def find_entry(path: str, lookup: dict[str, ArchiveEntry], entries: list[ArchiveEntry], allow_suffix: bool = True) -> Optional[ArchiveEntry]:
    key = normalize_archive_path(path)
    if key in lookup:
        return lookup[key]
    if allow_suffix:
        matches = [e for e in entries if normalize_archive_path(e.path).endswith(key)]
        if len(matches) == 1:
            return matches[0]
        base = key.split("/")[-1]
        if base:
            matches = [e for e in entries if normalize_archive_path(e.path).split("/")[-1] == base]
            if len(matches) == 1:
                return matches[0]
    return None


def write_manifest_csv(path: Path, entries: list[ArchiveEntry]) -> None:
    fields = [
        "level", "container", "path", "index", "name", "kind", "table_off", "rel_off", "abs_off",
        "sector", "sector_count", "size", "alloc", "padding", "first_data", "table_count",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for e in entries:
            w.writerow({
                "level": e.level,
                "container": e.container,
                "path": e.path,
                "index": e.index,
                "name": e.name,
                "kind": e.kind,
                "table_off": f"0x{e.table_off:X}",
                "rel_off": f"0x{e.rel_off:X}",
                "abs_off": f"0x{e.abs_off:X}",
                "sector": "" if e.sector is None else e.sector,
                "sector_count": "" if e.sector_count is None else e.sector_count,
                "size": f"0x{e.size:X}",
                "alloc": "" if e.alloc is None else f"0x{e.alloc:X}",
                "padding": "" if e.padding is None else f"0x{e.padding:X}",
                "first_data": "" if e.first_data is None else f"0x{e.first_data:X}",
                "table_count": "" if e.table_count is None else e.table_count,
            })

# ---------------------------------------------------------------------------
# TXT building / INFO.FAT repack
# ---------------------------------------------------------------------------

@dataclass
class TxtBuild:
    excel_row: int
    row_id: int
    archive_path: str
    translation_column_used: str
    body: bytes
    generated_min: bytes
    terminator: int
    text: str
    notes: list[str]


def load_txt_builds(args: argparse.Namespace) -> tuple[list[TxtBuild], str]:
    wb = load_workbook(args.xlsx, data_only=True)
    sheet_name = choose_sheet(wb, args.sheet)
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    row_col = resolve_col(headers, args.row_column, ["row", "id", "index"], required=False)
    path_col = resolve_col(headers, args.path_column, ["archive_path", "archive path", "path", "cdf_path", "file_path", "file"], required=True)
    trans_col = resolve_col(headers, args.translation_column, [
        "english_translation", "English Translation", "Final Translation", "QA Clean Translation",
        "Better Translation", "Clean Translation", "Translation", "Insertion Text", "Final Insertion Text", "text",
    ], required=True)
    fallback_col = resolve_col(headers, args.fallback_translation_column, [
        "english_translation", "English Translation", "Final Translation", "Translation", "Clean Translation",
    ], required=False)

    default_term = parse_byte(args.terminator, 0x00)
    if default_term is None:
        default_term = 0x00
    default_term &= 0xFF
    pad_byte = args.pad_byte & 0xFF
    row_filter = parse_row_set(args.rows)

    out: list[TxtBuild] = []
    for r in range(2, ws.max_row + 1):
        raw_path = norm_text(ws.cell(r, path_col).value).strip().replace("\\", "/")
        if not raw_path:
            continue
        if args.only_path_contains and args.only_path_contains.lower() not in raw_path.lower():
            continue
        if args.include_regex and not re.search(args.include_regex, raw_path):
            continue
        if args.exclude_regex and re.search(args.exclude_regex, raw_path):
            continue

        row_id = parse_int(ws.cell(r, row_col).value, r - 1) if row_col else r - 1
        if row_id is None:
            row_id = r - 1
        if row_filter is not None and row_id not in row_filter:
            continue

        text = norm_text(ws.cell(r, trans_col).value)
        used_col = headers[trans_col - 1] or str(trans_col)
        notes: list[str] = []
        if not text and fallback_col:
            text = norm_text(ws.cell(r, fallback_col).value)
            used_col = headers[fallback_col - 1] or str(fallback_col)
        if not text:
            notes.append("empty_translation")
            if not args.allow_empty:
                text = ""

        body = encode_body(text, args.text_mode, args.linebreak_mode, args.encoding, args.encoding_errors)
        if default_term in body and default_term != 0x00:
            notes.append(f"body_contains_terminator_byte_0x{default_term:02X}")
        if b"\x00" in body:
            notes.append("body_contains_00")
        file_bytes = bytearray(body)
        file_bytes.append(default_term)
        min_size = align(len(file_bytes), args.align)
        if len(file_bytes) < min_size:
            file_bytes.extend(bytes([pad_byte]) * (min_size - len(file_bytes)))

        out.append(TxtBuild(
            excel_row=r,
            row_id=row_id,
            archive_path=raw_path,
            translation_column_used=used_col,
            body=body,
            generated_min=bytes(file_bytes),
            terminator=default_term,
            text=text,
            notes=notes,
        ))
    return out, sheet_name


def parse_fat_table_raw(seg: bytes) -> tuple[int, list[dict[str, Any]]]:
    if len(seg) < FAT_ENTRY_SIZE:
        raise ValueError("FAT segment too small")
    first_data = struct.unpack_from("<I", seg, 0x10)[0]
    if first_data <= 0 or first_data > len(seg) or first_data % FAT_ENTRY_SIZE:
        raise ValueError(f"Bad FAT first_data/table size: 0x{first_data:X}")
    table_count = first_data // FAT_ENTRY_SIZE
    rows: list[dict[str, Any]] = []
    for i in range(table_count):
        off = i * FAT_ENTRY_SIZE
        raw = bytes(seg[off:off + FAT_ENTRY_SIZE])
        name = cstr(raw[:0x10])
        rel = struct.unpack_from("<I", raw, 0x10)[0]
        rows.append({"index": i, "off": off, "raw": raw, "name": name, "rel": rel})
    return first_data, rows


def old_size_for_table_row(row: dict[str, Any], rows: list[dict[str, Any]], seg_len: int) -> int:
    rel = int(row["rel"])
    if not row["name"] or rel <= 0 or rel > seg_len:
        return 0
    later = sorted(int(r["rel"]) for r in rows if r["name"] and int(r["rel"]) > rel and int(r["rel"]) <= seg_len)
    next_rel = later[0] if later else seg_len
    return next_rel - rel


def make_rebuilt_info_fat(
    old_seg: bytes,
    container_path: str,
    builds_by_path: dict[str, TxtBuild],
    args: argparse.Namespace,
) -> tuple[bytes, list[dict[str, Any]], dict[str, Any]]:
    first_data, table_rows = parse_fat_table_raw(old_seg)
    table = bytearray(old_seg[:first_data])
    data_out = bytearray()
    report_rows: list[dict[str, Any]] = []
    used_keys: set[str] = set()

    for row in table_rows:
        name = row["name"]
        old_rel = int(row["rel"])
        if not name or old_rel <= 0 or old_rel > len(old_seg):
            # Preserve empty/invalid sentinel table rows exactly.
            continue
        archive_path = f"{container_path}/{name}"
        key = normalize_archive_path(archive_path)
        old_size = old_size_for_table_row(row, table_rows, len(old_seg))
        old_payload = old_seg[old_rel:old_rel + old_size]
        tb = builds_by_path.get(key)
        notes: list[str] = []
        if tb:
            payload = tb.generated_min
            notes.extend(tb.notes)
            used_keys.add(key)
            action = "replaced"
        else:
            payload = bytes(old_payload)
            notes.append("kept_original_no_workbook_row")
            action = "kept_original"

        new_rel = first_data + len(data_out)
        if new_rel % args.align:
            # This should not happen because payloads are aligned, but keep table offsets aligned.
            pad = align(new_rel, args.align) - new_rel
            data_out.extend(bytes([args.pad_byte & 0xFF]) * pad)
            new_rel += pad
        table[row["off"]:row["off"] + 0x10] = encode_name(name)
        struct.pack_into("<I", table, row["off"] + 0x10, new_rel)
        data_out.extend(payload)

        tb_excel_row = tb.excel_row if tb else ""
        tb_row_id = tb.row_id if tb else ""
        text = tb.text if tb else ""
        report_rows.append({
            "excel_row": tb_excel_row,
            "row": tb_row_id,
            "archive_path": archive_path,
            "action": action,
            "old_rel_off": f"0x{old_rel:X}",
            "old_size": f"0x{old_size:X}",
            "new_rel_off": f"0x{new_rel:X}",
            "new_payload_size": f"0x{len(payload):X}",
            "delta_vs_old_slot": len(payload) - old_size,
            "terminator_byte": f"0x{tb.terminator:02X}" if tb else "",
            "encoded_text_len": len(tb.body) if tb else "",
            "generated_min_size": len(tb.generated_min) if tb else "",
            "line_count": (0 if text == "" else text.count("\n") + 1) if tb else "",
            "glyph_count_no_newlines": len(text.replace("\n", "")) if tb else "",
            "translation_column_used": tb.translation_column_used if tb else "",
            "notes": ";".join(notes),
        })

    missing_from_container = sorted(set(builds_by_path) - used_keys)
    new_used = bytes(table + data_out)
    meta = {
        "first_data": first_data,
        "table_count": len(table_rows),
        "old_info_size": len(old_seg),
        "new_used_size": len(new_used),
        "missing_build_keys": missing_from_container,
    }
    return new_used, report_rows, meta


def update_top_entry_size_for_path(cdf: bytearray, top_entries: list[ArchiveEntry], path: str, new_size: int) -> bool:
    key = normalize_archive_path(path)
    for e in top_entries:
        if normalize_archive_path(e.path) == key:
            if e.alloc is not None and new_size > e.alloc:
                raise ValueError(f"New top-level size 0x{new_size:X} exceeds allocation 0x{e.alloc:X} for {path}")
            struct.pack_into("<I", cdf, e.table_off + 0x1C, new_size)
            return True
    return False

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Flexible PixyGarden TREE.CDF INFO.FAT text direct patcher/repacker.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--xlsx", required=True, help="Workbook containing TREE archive paths and English translation text")
    ap.add_argument("--sheet", default=None, help="Worksheet name; default auto-picks a TREE-like sheet or first sheet")
    ap.add_argument("--source-cdf", required=True, help="Source/current TREE.CDF")
    ap.add_argument("--out-cdf", required=True, help="Patched TREE.CDF output path")
    ap.add_argument("--report", required=True, help="CSV patch report path")
    ap.add_argument("--write-manifest", help="Optional recursive manifest CSV before patching")
    ap.add_argument("--out-info-fat", help="Optionally write rebuilt INFO.FAT bytes for inspection")

    ap.add_argument("--target-container", default="DETAILS.FAT/INFO.FAT", help="Nested FAT container to rebuild. Default DETAILS.FAT/INFO.FAT")
    ap.add_argument("--top-parent", default="DETAILS.FAT", help="Top-level CDF entry containing target container. Default DETAILS.FAT")

    ap.add_argument("--path-column", default=None, help="Archive path column letter/header. Auto-detects archive_path/path.")
    ap.add_argument("--translation-column", default=None, help="Translation column letter/header. Auto-detects english_translation/etc.")
    ap.add_argument("--fallback-translation-column", default=None, help="Optional fallback translation column if primary is empty")
    ap.add_argument("--row-column", default=None, help="Row/id column letter/header. Optional; Excel row-1 used if omitted.")

    ap.add_argument("--text-mode", choices=["ascii", "cp932", "cp932_fullwidth"], default="ascii")
    ap.add_argument("--encoding", default="cp932")
    ap.add_argument("--encoding-errors", choices=["strict", "replace", "ignore"], default="strict")
    ap.add_argument("--linebreak-mode", choices=["backslash_n", "lf", "crlf", "none"], default="backslash_n")
    ap.add_argument("--terminator", default="0x00", help="Terminator byte for generated TXT files. Default 0x00.")
    ap.add_argument("--pad-byte", type=lambda x: int(x, 0), default=0xFF, help="Pad byte. Default 0xFF.")
    ap.add_argument("--align", type=lambda x: int(x, 0), default=4, help="TXT payload alignment. Default 4.")

    ap.add_argument("--rows", help="Only build these workbook row IDs, e.g. 1,3,5-8")
    ap.add_argument("--only-path-contains", help="Only process workbook rows whose archive_path contains this text")
    ap.add_argument("--include-regex", help="Only process paths matching this regex")
    ap.add_argument("--exclude-regex", help="Skip paths matching this regex")
    ap.add_argument("--allow-empty", action="store_true", help="Allow empty translations")
    ap.add_argument("--allow-shrink", action="store_true", default=True, help="Allow rebuilt INFO.FAT used data to be smaller than original; default on")
    ap.add_argument("--allow-container-growth", action="store_true", help="Allow target container to grow if it still fits in top-parent allocation and is the last child")
    ap.add_argument("--dry-run", action="store_true", help="Write report/manifest only; do not write binary outputs")
    ap.add_argument("--no-fail-on-warnings", action="store_true", help="Return success even with warnings")
    args = ap.parse_args(argv)

    src_path = Path(args.source_cdf)
    cdf = bytearray(src_path.read_bytes())
    original_cdf_size = len(cdf)
    top_info, top_entries = parse_top_cdf(bytes(cdf), src_path.name)
    entries = recursive_manifest(bytes(cdf), top_entries)
    lookup = build_lookup(entries)
    if args.write_manifest:
        write_manifest_csv(Path(args.write_manifest), entries)

    target = find_entry(args.target_container, lookup, entries, allow_suffix=True)
    if not target:
        raise SystemExit(f"Target container {args.target_container!r} not found in source CDF")
    old_info = bytes(cdf[target.abs_off:target.abs_off + target.size])

    builds, sheet_name = load_txt_builds(args)
    builds_by_path: dict[str, TxtBuild] = {}
    duplicate_paths: list[str] = []
    for tb in builds:
        key = normalize_archive_path(tb.archive_path)
        if key in builds_by_path:
            duplicate_paths.append(tb.archive_path)
        builds_by_path[key] = tb
    if duplicate_paths:
        raise SystemExit("Duplicate archive paths in workbook: " + ", ".join(duplicate_paths[:10]))

    new_used, rows, meta = make_rebuilt_info_fat(old_info, args.target_container, builds_by_path, args)
    missing = meta["missing_build_keys"]
    warnings: list[str] = []
    if missing:
        warnings.append(f"{len(missing)} workbook rows did not match entries in {args.target_container}")

    old_container_size = int(meta["old_info_size"])
    new_used_size = int(meta["new_used_size"])
    new_container_size = old_container_size
    if new_used_size <= old_container_size:
        rebuilt_info = new_used + bytes([args.pad_byte & 0xFF]) * (old_container_size - new_used_size)
    else:
        if not args.allow_container_growth:
            raise SystemExit(
                f"Rebuilt {args.target_container} uses 0x{new_used_size:X}, exceeding old container 0x{old_container_size:X}. "
                "Use --allow-container-growth if this container is last in its parent and top-level allocation has room."
            )
        # Growth is conservative: only supported if the target is the last child of the top parent.
        parent = find_entry(args.top_parent, lookup, entries, allow_suffix=False)
        if not parent or parent.kind != "CDF_ENTRY":
            raise SystemExit(f"Top parent {args.top_parent!r} not found as a top-level CDF entry")
        if target.abs_off + target.size != parent.abs_off + parent.size:
            raise SystemExit("Container growth is only supported when target container is last inside the top parent")
        grow_by = new_used_size - old_container_size
        new_parent_size = parent.size + grow_by
        if parent.alloc is None or new_parent_size > parent.alloc:
            raise SystemExit(
                f"Growing {args.target_container} by 0x{grow_by:X} would make {args.top_parent} size 0x{new_parent_size:X}, "
                f"which exceeds allocation 0x{(parent.alloc or 0):X}."
            )
        rebuilt_info = new_used
        new_container_size = new_used_size
        update_top_entry_size_for_path(cdf, top_entries, args.top_parent, new_parent_size)
        warnings.append(f"grew {args.target_container} by 0x{grow_by:X} and updated {args.top_parent} size")

    # Add container summary columns to each report row.
    for r in rows:
        r["container_old_size"] = f"0x{old_container_size:X}"
        r["container_new_used_size"] = f"0x{new_used_size:X}"
        r["container_written_size"] = f"0x{new_container_size:X}"
        r["container_slack_after_repack"] = f"0x{max(0, new_container_size - new_used_size):X}"

    report_fields = [
        "excel_row", "row", "archive_path", "action",
        "old_rel_off", "old_size", "new_rel_off", "new_payload_size", "delta_vs_old_slot",
        "terminator_byte", "encoded_text_len", "generated_min_size",
        "line_count", "glyph_count_no_newlines", "translation_column_used",
        "container_old_size", "container_new_used_size", "container_written_size", "container_slack_after_repack",
        "notes",
    ]
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=report_fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    if not args.dry_run:
        cdf[target.abs_off:target.abs_off + len(rebuilt_info)] = rebuilt_info
        # If shrinking/padding to original container size, this write covers exactly the old container.
        if len(rebuilt_info) < old_container_size:
            raise RuntimeError("Internal error: rebuilt_info shorter than selected container write size")
        out_cdf = Path(args.out_cdf)
        out_cdf.parent.mkdir(parents=True, exist_ok=True)
        out_cdf.write_bytes(bytes(cdf))
        if out_cdf.stat().st_size != original_cdf_size:
            raise SystemExit(
                f"Internal error: output CDF size changed. source=0x{original_cdf_size:X}, output=0x{out_cdf.stat().st_size:X}"
            )
        if args.out_info_fat:
            out_info = Path(args.out_info_fat)
            out_info.parent.mkdir(parents=True, exist_ok=True)
            out_info.write_bytes(rebuilt_info)

    print("PixyGarden TREE TXT CDF builder v11 INFO.FAT repack")
    print("----------------------------------------------------")
    print(f"Workbook:               {args.xlsx}")
    print(f"Sheet:                  {sheet_name}")
    print(f"Rows loaded:            {len(builds)}")
    print(f"Source CDF:             {args.source_cdf}")
    print(f"Target container:       {args.target_container}")
    print(f"Target abs/old size:    0x{target.abs_off:X} / 0x{old_container_size:X}")
    print(f"Rebuilt used size:      0x{new_used_size:X}")
    print(f"Written container size: 0x{new_container_size:X}")
    print(f"Slack after repack:     0x{max(0, new_container_size - new_used_size):X}")
    print(f"Report:                 {report_path}")
    print(f"Text mode:              {args.text_mode}")
    print(f"Linebreak mode:         {args.linebreak_mode}")
    print(f"Terminator:             0x{(parse_byte(args.terminator, 0) or 0):02X}")
    print(f"Pad byte:               0x{args.pad_byte & 0xFF:02X}")
    print(f"Entries replaced:       {sum(1 for r in rows if r['action'] == 'replaced')}")
    print(f"Entries kept original:  {sum(1 for r in rows if r['action'] == 'kept_original')}")
    if args.out_cdf:
        print(f"Patched CDF:            {args.out_cdf}")
    if args.out_info_fat:
        print(f"Patched INFO.FAT:       {args.out_info_fat}")
    if args.dry_run:
        print("Dry run: no binary outputs written.")
    else:
        sha = hashlib.sha256(Path(args.out_cdf).read_bytes()).hexdigest()
        print(f"Output SHA-256:         {sha}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  - {w}")
        if not args.no_fail_on_warnings:
            return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
